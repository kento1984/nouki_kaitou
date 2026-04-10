"""report_generator モジュールのテスト"""

import datetime
import tempfile
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from nouki_kaitou.models import (
    BranchSettings,
    BunnoEntry,
    CacheStore,
    ConfirmingRecord,
    HolidayMap,
    HistoryRecord,
    OrderRow,
    ReportResult,
    ReportRow,
    StockoutEntry,
    TrackingEntry,
)
from nouki_kaitou.report_generator import (
    build_report_row,
    create_delivery_report,
    create_delivery_report_by_order_numbers,
    group_order_numbers_by_customer,
    _resolve_manufacturer_name,
    _resolve_product_name,
    _resolve_delivery_place,
    _resolve_price,
    _is_provisional_price,
    _pass_basic_filter,
    _is_excluded,
    _determine_flags,
    _classify_order,
    _collect_stockout_info,
)


# ============================================
# テスト用ヘルパー
# ============================================
TODAY = datetime.date(2026, 2, 16)
EXEC_TIME = datetime.datetime(2026, 2, 16, 10, 0, 0)
BRANCH = BranchSettings(
    name="京葉営業所",
    default_cutoff=15,
    base_center="関東商品センター",
)


def _make_cache(**kwargs) -> CacheStore:
    """テスト用キャッシュを作成する。"""
    return CacheStore(
        mfg_name=kwargs.get("mfg_name", {"D01": "ダイヘン", "Z99": ""}),
        mfg_days=kwargs.get("mfg_days", {"D01": 2}),
        cust_days=kwargs.get("cust_days", {}),
        cust_retention=kwargs.get("cust_retention", {}),
        cust_route=kwargs.get("cust_route", {}),
        confirm=kwargs.get("confirm", {}),
        storage=kwargs.get("storage", {}),
    )


def _make_row(**kwargs) -> OrderRow:
    """テスト用OrderRowを作成する。"""
    defaults = {
        "order_number": "1000001",
        "detail_number": "10",
        "document_type": "【受注】在庫販売",
        "customer_name": "テスト商事",
        "product_name": "溶接棒 3.2mm",
        "ship_status": "未処理",
        "quantity": "100",
        "unit_price": "500",
        "net_amount": "50000",
        "manufacturer_name": "ダイヘン",
        "storage_place": "関東商品センター",
        "customer_order_number": "PO-001",
        "customer_contact": "田中",
        "comment_detail": "",
        "comment_external": "",
        "comment_internal": "",
        "rejection_reason": "",
        "ship_to_name": "テスト商事",
        "registration_date": TODAY,
        "time_value": "10:00:00",
        "order_delivery_date": datetime.date(2026, 2, 20),
        "specified_delivery_date": None,
        "item_group_code": "D01",
    }
    defaults.update(kwargs)
    return OrderRow(**defaults)


# ============================================
# group_order_numbers_by_customer
# ============================================
class TestGroupOrderNumbersByCustomer:
    """注番の顧客別グループ化テスト"""

    def test_basic_grouping(self):
        data = [
            _make_row(order_number="100", customer_name="A社"),
            _make_row(order_number="200", customer_name="B社"),
            _make_row(order_number="300", customer_name="A社"),
        ]
        result = group_order_numbers_by_customer(data, ["100", "200", "300"])
        assert result == {"A社": ["100", "300"], "B社": ["200"]}

    def test_dedup(self):
        data = [
            _make_row(order_number="100", customer_name="A社"),
        ]
        result = group_order_numbers_by_customer(data, ["100", "100"])
        assert result == {"A社": ["100"]}

    def test_not_found(self):
        data = [
            _make_row(order_number="100", customer_name="A社"),
        ]
        result = group_order_numbers_by_customer(data, ["999"])
        assert result == {}

    def test_empty(self):
        result = group_order_numbers_by_customer([], ["100"])
        assert result == {}


# ============================================
# _resolve_manufacturer_name
# ============================================
class TestResolveManufacturerName:
    """メーカー名解決テスト"""

    def test_normal(self):
        row = _make_row(item_group_code="D01")
        cache = _make_cache()
        assert _resolve_manufacturer_name(row, cache) == "ダイヘン"

    def test_z99_with_space(self):
        row = _make_row(
            item_group_code="Z99",
            product_name="ABC商事 特殊溶接棒"
        )
        cache = _make_cache()
        assert _resolve_manufacturer_name(row, cache) == "ABC商事"

    def test_z99_full_width_space(self):
        row = _make_row(
            item_group_code="Z99",
            product_name="ABC商事\u3000特殊溶接棒"
        )
        cache = _make_cache()
        assert _resolve_manufacturer_name(row, cache) == "ABC商事"

    def test_z99_no_space(self):
        row = _make_row(
            item_group_code="Z99",
            product_name="特殊溶接棒"
        )
        cache = _make_cache()
        assert _resolve_manufacturer_name(row, cache) == ""

    def test_fallback_to_manufacturer_field(self):
        row = _make_row(
            item_group_code="X01",
            manufacturer_name="フォールバック社"
        )
        cache = _make_cache(mfg_name={})
        assert _resolve_manufacturer_name(row, cache) == "フォールバック社"

    def test_fallback_to_code(self):
        row = _make_row(
            item_group_code="X01",
            manufacturer_name=""
        )
        cache = _make_cache(mfg_name={})
        assert _resolve_manufacturer_name(row, cache) == "X01"

    def test_numeric_code_cache_hit(self):
        """4桁数字コードでキャッシュヒットする"""
        row = _make_row(item_group_code="0075", manufacturer_name="")
        cache = _make_cache(mfg_name={"0075": "ダイヘン"})
        assert _resolve_manufacturer_name(row, cache) == "ダイヘン"

    def test_numeric_code_without_leading_zero(self):
        """先頭ゼロなし数字コードでも正規化によりヒット"""
        row = _make_row(item_group_code="75", manufacturer_name="")
        cache = _make_cache(mfg_name={"0075": "ダイヘン"})
        assert _resolve_manufacturer_name(row, cache) == "ダイヘン"


# ============================================
# _resolve_product_name
# ============================================
class TestResolveProductName:
    """品名解決テスト"""

    def test_normal(self):
        row = _make_row(item_group_code="D01", product_name="溶接棒 3.2mm")
        assert _resolve_product_name(row, "ダイヘン") == "溶接棒 3.2mm"

    def test_z99(self):
        row = _make_row(
            item_group_code="Z99",
            product_name="ABC商事 特殊溶接棒"
        )
        assert _resolve_product_name(row, "ABC商事") == "特殊溶接棒"

    def test_z97_full_width_space(self):
        row = _make_row(
            item_group_code="Z97",
            product_name="XYZ\u3000特殊品"
        )
        assert _resolve_product_name(row, "XYZ") == "特殊品"


# ============================================
# _resolve_delivery_place
# ============================================
class TestResolveDeliveryPlace:
    """納入先名解決テスト"""

    def test_same_as_customer(self):
        row = _make_row(customer_name="テスト商事", ship_to_name="テスト商事")
        assert _resolve_delivery_place(row, TODAY) == "貴社"

    def test_different_no_sama(self):
        row = _make_row(ship_to_name="東京工場")
        assert _resolve_delivery_place(row, TODAY) == "東京工場様"

    def test_different_with_sama(self):
        row = _make_row(ship_to_name="東京工場様")
        assert _resolve_delivery_place(row, TODAY) == "東京工場様"

    def test_pickup(self):
        row = _make_row(
            comment_external="引取 2/20",
            ship_to_name="テスト商事"
        )
        assert _resolve_delivery_place(row, TODAY) == "お引き取り"

    def test_same_fullwidth_halfwidth_parens(self):
        """全角（有）vs 半角(有) → 正規化して「貴社」"""
        row = _make_row(
            customer_name="（有）三橋機工",
            ship_to_name="(有)三橋機工",
        )
        assert _resolve_delivery_place(row, TODAY) == "貴社"

    def test_same_fullwidth_halfwidth_kabu_space(self):
        """全角（株）+全角スペース vs 半角(株)+半角スペース → 「貴社」"""
        row = _make_row(
            customer_name="テスト（株）\u3000本社",
            ship_to_name="テスト(株) 本社",
        )
        assert _resolve_delivery_place(row, TODAY) == "貴社"


# ============================================
# _resolve_price
# ============================================
class TestResolvePrice:
    """価格解決テスト"""

    def test_normal(self):
        row = _make_row(unit_price="500", net_amount="50000")
        result = _resolve_price(row, "2月20日配達予定", False)
        assert result == ("500", "50000")

    def test_unit_price_1(self):
        row = _make_row(unit_price=1)
        result = _resolve_price(row, "2月20日配達予定", False)
        assert result == ("確認中", "確認中")

    def test_unit_price_str_1(self):
        row = _make_row(unit_price="1")
        result = _resolve_price(row, "2月20日配達予定", False)
        assert result == ("確認中", "確認中")

    def test_confirming_no_dollar(self):
        row = _make_row(comment_internal="")
        result = _resolve_price(row, "確認中", False)
        assert result == ("確認中", "確認中")

    def test_confirming_with_dollar(self):
        row = _make_row(comment_internal="$$")
        result = _resolve_price(row, "確認中", False)
        assert result == ("500", "50000")

    def test_force_delivered_overrides(self):
        row = _make_row(comment_internal="")
        result = _resolve_price(row, "確認中", True)
        assert result == ("500", "50000")

    def test_unit_price_1_00(self):
        """SAPの "1.00" 形式でも確認中になる"""
        row = _make_row(unit_price="1.00", net_amount="5.00")
        result = _resolve_price(row, "3月16日出荷済み", False)
        assert result == ("確認中", "確認中")

    def test_unit_price_1_000(self):
        """SAPの "1.000" 形式でも確認中になる"""
        row = _make_row(unit_price="1.000", net_amount="5.000")
        result = _resolve_price(row, "3月16日出荷済み", False)
        assert result == ("確認中", "確認中")

    def test_dollar_overrides_provisional(self):
        """$$があれば仮単価1でもそのまま表示"""
        row = _make_row(unit_price="1", net_amount="5", comment_internal="$$")
        result = _resolve_price(row, "3月16日出荷済み", False)
        assert result == ("1", "5")

    def test_fullwidth_dollar_overrides_provisional(self):
        """＄＄（全角）でも仮単価をそのまま表示"""
        row = _make_row(unit_price="1.00", net_amount="5.00", comment_internal="＄＄")
        result = _resolve_price(row, "確認中", False)
        assert result == ("1.00", "5.00")

    def test_unit_price_not_1(self):
        """正式単価は確認中にならない"""
        row = _make_row(unit_price="100", net_amount="500")
        result = _resolve_price(row, "3月16日出荷済み", False)
        assert result == ("100", "500")


# ============================================
# _is_provisional_price
# ============================================
class TestIsProvisionalPrice:
    """仮単価判定テスト"""

    def test_str_1(self):
        assert _is_provisional_price("1") is True

    def test_int_1(self):
        assert _is_provisional_price(1) is True

    def test_float_1_00(self):
        assert _is_provisional_price("1.00") is True

    def test_float_1_000(self):
        assert _is_provisional_price("1.000") is True

    def test_not_1(self):
        assert _is_provisional_price("500") is False

    def test_comma_format(self):
        assert _is_provisional_price("1,000") is False

    def test_empty(self):
        assert _is_provisional_price("") is False

    def test_text(self):
        assert _is_provisional_price("確認中") is False


# ============================================
# _pass_basic_filter
# ============================================
class TestPassBasicFilter:
    """基本フィルタテスト"""

    def test_pass(self):
        row = _make_row()
        assert _pass_basic_filter(row) is True

    def test_reject_deleted(self):
        row = _make_row(rejection_reason="明細削除")
        assert _pass_basic_filter(row) is False

    def test_reject_wrong_doc_type(self):
        row = _make_row(document_type="【受注】返品")
        assert _pass_basic_filter(row) is False

    def test_reject_hash_exclude(self):
        row = _make_row(comment_internal="##除外")
        assert _pass_basic_filter(row) is False

    def test_reject_fullwidth_hash(self):
        row = _make_row(comment_internal="＃＃除外")
        assert _pass_basic_filter(row) is False

    def test_chokusouhan(self):
        row = _make_row(document_type="【受注】直送販売")
        assert _pass_basic_filter(row) is True


# ============================================
# _is_excluded
# ============================================
class TestIsExcluded:
    """除外判定テスト"""

    def test_excluded_from_history(self):
        row = _make_row()
        cache = _make_cache()
        assert _is_excluded(row, "除外", cache) is True

    def test_excluded_from_confirming(self):
        row = _make_row(order_number="100", detail_number="10")
        cache = _make_cache(confirm={"100|10": ("除外", "", None)})
        assert _is_excluded(row, "", cache) is True

    def test_not_excluded(self):
        row = _make_row()
        cache = _make_cache()
        assert _is_excluded(row, "", cache) is False


# ============================================
# _determine_flags
# ============================================
class TestDetermineFlags:
    """フラグ判定テスト"""

    def test_stock_normal(self):
        row = _make_row(ship_status="未処理")
        cache = _make_cache()
        fd, hm, bc = _determine_flags(row, cache, {}, "100|10", "")
        assert fd is False
        assert hm is False
        assert bc is False

    def test_chokusouhan_completed_force_delivered(self):
        """直送+処理完了+転送中 → forceDelivered"""
        row = _make_row(
            document_type="【受注】直送販売",
            storage_place="転送中（直送用）",
            ship_status="処理完了",
        )
        cache = _make_cache()
        fd, hm, bc = _determine_flags(row, cache, {}, "100|10", "")
        assert fd is True
        assert hm is False

    def test_himozuki_no_force(self):
        """紐付き（直送+非転送中） → isHimozuki, forceDelivered=False"""
        row = _make_row(
            document_type="【受注】直送販売",
            storage_place="関東商品センター",
            ship_status="処理完了",
        )
        cache = _make_cache()
        fd, hm, bc = _determine_flags(row, cache, {}, "100|10", "")
        assert fd is False
        assert hm is True

    def test_bunno_completed(self):
        """分納+処理完了 → isBunnoCompleted"""
        row = _make_row(
            ship_status="処理完了",
            order_number="100",
            detail_number="10",
        )
        cache = _make_cache(
            confirm={"100|10": ("未", "分納", None)}
        )
        fd, hm, bc = _determine_flags(row, cache, {}, "100|10", "")
        assert bc is True

    def test_previous_confirming_forces(self):
        """前回確認中+処理完了 → forceDelivered"""
        row = _make_row(
            document_type="【受注】在庫販売",
            ship_status="処理完了",
        )
        cache = _make_cache()
        sent = {"1000001|10": "確認中"}
        fd, hm, bc = _determine_flags(
            row, cache, sent, "1000001|10", "確認中"
        )
        assert fd is True

    def test_bunno_in_confirming_blocks_force_delivered(self):
        """進行中の分納 → force_deliveredをブロック

        確認中一覧に「分納」で残っていて、まだ処理完了でない場合、
        force_deliveredはFalse（is_bunno_in_confirmingがブロック）。
        ただし ship_status != 処理完了 なので、そもそも
        force_delivered判定ブロックに入らない。
        """
        row = _make_row(
            document_type="【受注】直送販売",
            storage_place="転送中（直送用）",
            ship_status="未処理",
            order_number="100",
            detail_number="10",
        )
        cache = _make_cache(
            confirm={"100|10": ("未", "分納", None)}
        )
        fd, hm, bc = _determine_flags(row, cache, {}, "100|10", "")
        assert fd is False
        assert bc is False

    def test_bunno_completed_resets_bunno_in_confirming(self):
        """分納+処理完了 → is_bunno_completedでis_bunno_in_confirmingリセット

        分納が処理完了になるとis_bunno_completedがTrueになり、
        is_bunno_in_confirmingはFalseにリセットされる。
        直送の場合force_deliveredもTrueになりうる。
        """
        row = _make_row(
            document_type="【受注】直送販売",
            storage_place="転送中（直送用）",
            ship_status="処理完了",
            order_number="100",
            detail_number="10",
        )
        cache = _make_cache(
            confirm={"100|10": ("未", "分納", None)}
        )
        fd, hm, bc = _determine_flags(row, cache, {}, "100|10", "")
        assert bc is True
        # is_bunno_in_confirmingがリセットされるためブロックされない
        assert fd is True


# ============================================
# build_report_row
# ============================================
class TestBuildReportRow:
    """OrderRow → ReportRow変換テスト"""

    def test_basic(self):
        row = _make_row()
        cache = _make_cache()
        report, status = build_report_row(
            row, cache, {}, BRANCH, EXEC_TIME, False, TODAY
        )
        assert isinstance(report, ReportRow)
        assert report.manufacturer_name == "ダイヘン"
        assert report.order_number == "1000001"
        assert report.delivery_place == "貴社"
        assert "予定" in status or "済み" in status

    def test_stockout_confirming(self):
        """欠品中 + 確認中 → 欠品中"""
        row = _make_row(
            comment_detail="欠品中",
            order_delivery_date=datetime.date(2026, 12, 31),
        )
        cache = _make_cache()
        _, status = build_report_row(
            row, cache, {}, BRANCH, EXEC_TIME, False, TODAY
        )
        assert status == "欠品中"

    def test_stockout_with_date(self):
        """欠品中 + 納期あり → ○月○日配達予定（欠品）"""
        row = _make_row(
            comment_detail="欠品中",
            order_delivery_date=datetime.date(2026, 2, 20),
        )
        cache = _make_cache()
        _, status = build_report_row(
            row, cache, {}, BRANCH, EXEC_TIME, False, TODAY
        )
        assert "（欠品）" in status

    def test_bunno(self):
        """分納コメントあり → 分納"""
        row = _make_row(
            comment_detail="分納:50個 2/20,50個 未定",
            ship_status="未処理",
        )
        cache = _make_cache()
        _, status = build_report_row(
            row, cache, {}, BRANCH, EXEC_TIME, False, TODAY
        )
        assert status == "分納"

    def test_bunno_completed_ignored(self):
        """分納+処理完了 → 分納は無視（通常の納期計算）"""
        row = _make_row(
            comment_detail="分納:50個 2/20,50個 未定",
            ship_status="処理完了",
        )
        cache = _make_cache()
        _, status = build_report_row(
            row, cache, {}, BRANCH, EXEC_TIME, False, TODAY
        )
        assert status != "分納"

    def test_remark_cleaned(self):
        """備考から欠品テキスト・分納テキスト除去"""
        row = _make_row(
            comment_detail="欠品中 1月上旬予定 分納:50個 2/20"
        )
        cache = _make_cache()
        report, _ = build_report_row(
            row, cache, {}, BRANCH, EXEC_TIME, False, TODAY
        )
        assert "欠品中" not in report.remarks
        assert "分納:" not in report.remarks

    def test_stockout_scheduling(self):
        """欠品中 + 日程調整中 → 欠品中"""
        row = _make_row(
            comment_detail="欠品中",
            # 指定納期・受注納期ともに12/31 → 日程調整中になる
            order_delivery_date=datetime.date(2026, 12, 31),
            specified_delivery_date=datetime.date(2026, 12, 31),
            document_type="【受注】在庫販売",
        )
        cache = _make_cache()
        _, status = build_report_row(
            row, cache, {}, BRANCH, EXEC_TIME, False, TODAY
        )
        assert status == "欠品中"

    def test_stockout_with_confirmed_date_skip(self):
        """欠品中 + 確認中一覧に確定日 → 欠品overrideスキップ"""
        row = _make_row(
            comment_detail="欠品中",
            order_delivery_date=datetime.date(2026, 12, 31),
            order_number="100",
            detail_number="10",
        )
        # 確認中一覧に確定日が手入力されている
        cache = _make_cache(
            confirm={"100|10": ("未", "欠品中", datetime.date(2026, 3, 15))}
        )
        _, status = build_report_row(
            row, cache, {}, BRANCH, EXEC_TIME, False, TODAY
        )
        # 欠品overrideスキップ → 確定日に基づく通常の納期計算
        assert status != "欠品中"
        assert "（欠品）" not in status

    def test_bunno_partial_ship_status(self):
        """一部処理済み + 分納 → 分納（未処理と同じ）"""
        row = _make_row(
            comment_detail="分納:50個 2/20,50個 未定",
            ship_status="一部処理済み",
        )
        cache = _make_cache()
        _, status = build_report_row(
            row, cache, {}, BRANCH, EXEC_TIME, False, TODAY
        )
        assert status == "分納"

    def test_price_confirming(self):
        """単価=1 → 確認中"""
        row = _make_row(unit_price=1, net_amount=100)
        cache = _make_cache()
        report, _ = build_report_row(
            row, cache, {}, BRANCH, EXEC_TIME, False, TODAY
        )
        assert report.unit_price == "確認中"
        assert report.net_amount == "確認中"


# ============================================
# _classify_order
# ============================================
class TestClassifyOrder:
    """確定/確認中分類テスト"""

    def test_confirmed(self):
        """通常の確定伝票 → confirmed_orders"""
        confirmed: list[HistoryRecord] = []
        confirming: list[ConfirmingRecord] = []
        row = _make_row()
        cache = _make_cache()
        _classify_order(
            row, "2月20日配達予定", cache, [],
            False, "ダイヘン", "溶接棒",
            confirmed, confirming,
        )
        assert len(confirmed) == 1
        assert len(confirming) == 0
        assert confirmed[0].delivery_answer == "2月20日配達予定"

    def test_confirming(self):
        """確認中 → confirming_orders"""
        confirmed: list[HistoryRecord] = []
        confirming: list[ConfirmingRecord] = []
        row = _make_row()
        cache = _make_cache()
        _classify_order(
            row, "確認中", cache, [],
            False, "ダイヘン", "溶接棒",
            confirmed, confirming,
        )
        assert len(confirmed) == 0
        assert len(confirming) == 1

    def test_stockout(self):
        """欠品中 → confirming_orders with status=欠品中"""
        confirmed: list[HistoryRecord] = []
        confirming: list[ConfirmingRecord] = []
        row = _make_row()
        cache = _make_cache()
        _classify_order(
            row, "欠品中", cache, [],
            False, "ダイヘン", "溶接棒",
            confirmed, confirming,
        )
        assert len(confirming) == 1
        assert confirming[0].status == "欠品中"

    def test_bunno_with_mitei(self):
        """分納（未定あり） → confirming_orders with status=分納"""
        confirmed: list[HistoryRecord] = []
        confirming: list[ConfirmingRecord] = []
        row = _make_row()
        cache = _make_cache()
        bunno = [BunnoEntry("50個", "2/20", ""), BunnoEntry("50個", "未定", "")]
        _classify_order(
            row, "分納", cache, bunno,
            False, "ダイヘン", "溶接棒",
            confirmed, confirming,
        )
        assert len(confirming) == 1
        assert confirming[0].status == "分納"

    def test_bunno_all_confirmed(self):
        """分納（全確定） → confirmed_orders with 分納完了"""
        confirmed: list[HistoryRecord] = []
        confirming: list[ConfirmingRecord] = []
        row = _make_row()
        cache = _make_cache()
        bunno = [BunnoEntry("50個", "2/20", ""), BunnoEntry("50個", "2/25", "")]
        _classify_order(
            row, "分納", cache, bunno,
            False, "ダイヘン", "溶接棒",
            confirmed, confirming,
        )
        assert len(confirmed) == 1
        assert confirmed[0].delivery_answer == "分納完了"

    def test_bunno_completed(self):
        """分納完了 → confirmed_orders with 分納完了"""
        confirmed: list[HistoryRecord] = []
        confirming: list[ConfirmingRecord] = []
        row = _make_row()
        cache = _make_cache()
        _classify_order(
            row, "納品済み", cache, [],
            True, "ダイヘン", "溶接棒",
            confirmed, confirming,
        )
        assert len(confirmed) == 1
        assert confirmed[0].delivery_answer == "分納完了"

    def test_scheduling(self):
        """日程調整中 → confirming_orders"""
        confirmed: list[HistoryRecord] = []
        confirming: list[ConfirmingRecord] = []
        row = _make_row()
        cache = _make_cache()
        _classify_order(
            row, "日程調整中", cache, [],
            False, "ダイヘン", "溶接棒",
            confirmed, confirming,
        )
        assert len(confirming) == 1

    def test_bunno_keep_in_confirming_with_mitei(self):
        """既存分納(keep) + 未定あり → 確認中に留まる"""
        confirmed: list[HistoryRecord] = []
        confirming: list[ConfirmingRecord] = []
        row = _make_row(ship_status="未処理")
        # 前回「分納」で確認中一覧に残っている
        cache = _make_cache(
            confirm={"1000001|10": ("未", "分納", None)}
        )
        bunno = [BunnoEntry("50個", "2/20", ""), BunnoEntry("50個", "未定", "")]
        _classify_order(
            row, "分納", cache, bunno,
            False, "ダイヘン", "溶接棒",
            confirmed, confirming,
        )
        assert len(confirming) == 1
        assert confirming[0].status == "分納"
        assert len(confirmed) == 0

    def test_bunno_keep_all_confirmed(self):
        """既存分納(keep) + 全確定 → 送付履歴「分納完了」"""
        confirmed: list[HistoryRecord] = []
        confirming: list[ConfirmingRecord] = []
        row = _make_row(ship_status="未処理")
        # 前回「分納」で確認中一覧に残っている
        cache = _make_cache(
            confirm={"1000001|10": ("未", "分納", None)}
        )
        # 全エントリに日付あり → 全確定
        bunno = [BunnoEntry("50個", "2/20", ""), BunnoEntry("50個", "2/25", "")]
        _classify_order(
            row, "分納", cache, bunno,
            False, "ダイヘン", "溶接棒",
            confirmed, confirming,
        )
        assert len(confirmed) == 1
        assert confirmed[0].delivery_answer == "分納完了"
        assert len(confirming) == 0

    def test_stockout_partial_marker(self):
        """「（欠品）」付記 → confirming_orders with status=欠品中"""
        confirmed: list[HistoryRecord] = []
        confirming: list[ConfirmingRecord] = []
        row = _make_row()
        cache = _make_cache()
        _classify_order(
            row, "2月20日配達予定（欠品）", cache, [],
            False, "ダイヘン", "溶接棒",
            confirmed, confirming,
        )
        assert len(confirming) == 1
        assert confirming[0].status == "欠品中"

    def test_bunno_partial_ship_status(self):
        """一部処理済み + 分納 → 未処理と同じ扱い（確認中に残す）"""
        confirmed: list[HistoryRecord] = []
        confirming: list[ConfirmingRecord] = []
        row = _make_row(ship_status="一部処理済み")
        cache = _make_cache(
            confirm={"1000001|10": ("未", "分納", None)}
        )
        bunno = [BunnoEntry("50個", "2/20", ""), BunnoEntry("50個", "未定", "")]
        _classify_order(
            row, "分納", cache, bunno,
            False, "ダイヘン", "溶接棒",
            confirmed, confirming,
        )
        # 処理完了でないので確認中に残る
        assert len(confirming) == 1
        assert confirming[0].status == "分納"
        assert len(confirmed) == 0

    def test_price_pending_goes_to_confirming(self):
        """仮単価1円 + 納期確定 → 確認中一覧に「価格確認中」で残る"""
        confirmed: list[HistoryRecord] = []
        confirming: list[ConfirmingRecord] = []
        row = _make_row(unit_price="1.00", net_amount="5.00")
        cache = _make_cache()
        _classify_order(
            row, "3月16日出荷済み", cache, [],
            False, "ダイヘン", "溶接棒",
            confirmed, confirming,
        )
        assert len(confirmed) == 0
        assert len(confirming) == 1
        assert confirming[0].status == "価格確認中"

    def test_price_pending_str_1(self):
        """仮単価 "1" でも確認中一覧に残る"""
        confirmed: list[HistoryRecord] = []
        confirming: list[ConfirmingRecord] = []
        row = _make_row(unit_price="1", net_amount="5")
        cache = _make_cache()
        _classify_order(
            row, "2月20日配達予定", cache, [],
            False, "ダイヘン", "溶接棒",
            confirmed, confirming,
        )
        assert len(confirmed) == 0
        assert len(confirming) == 1
        assert confirming[0].status == "価格確認中"

    def test_price_confirmed_with_dollar(self):
        """$$あり + 仮単価1 → 確定扱い（送付履歴へ）"""
        confirmed: list[HistoryRecord] = []
        confirming: list[ConfirmingRecord] = []
        row = _make_row(
            unit_price="1.00", net_amount="5.00",
            comment_internal="$$",
        )
        cache = _make_cache()
        _classify_order(
            row, "3月16日出荷済み", cache, [],
            False, "ダイヘン", "溶接棒",
            confirmed, confirming,
        )
        assert len(confirmed) == 1
        assert len(confirming) == 0
        assert confirmed[0].delivery_answer == "3月16日出荷済み"

    def test_price_pending_normal_price_goes_to_confirmed(self):
        """正式単価 → 通常通り送付履歴へ"""
        confirmed: list[HistoryRecord] = []
        confirming: list[ConfirmingRecord] = []
        row = _make_row(unit_price="500", net_amount="50000")
        cache = _make_cache()
        _classify_order(
            row, "3月16日出荷済み", cache, [],
            False, "ダイヘン", "溶接棒",
            confirmed, confirming,
        )
        assert len(confirmed) == 1
        assert len(confirming) == 0

    def test_price_pending_with_confirming_delivery(self):
        """納期「確認中」+ 仮単価1 → 既存分岐で「未処理」（価格確認中にならない）"""
        confirmed: list[HistoryRecord] = []
        confirming: list[ConfirmingRecord] = []
        row = _make_row(unit_price="1.00", net_amount="5.00")
        cache = _make_cache()
        _classify_order(
            row, "確認中", cache, [],
            False, "ダイヘン", "溶接棒",
            confirmed, confirming,
        )
        assert len(confirming) == 1
        assert confirming[0].status == "未処理"

    def test_price_pending_with_scheduling_delivery(self):
        """納期「日程調整中」+ 仮単価1 → 既存分岐（価格確認中にならない）"""
        confirmed: list[HistoryRecord] = []
        confirming: list[ConfirmingRecord] = []
        row = _make_row(unit_price="1.00", net_amount="5.00")
        cache = _make_cache()
        _classify_order(
            row, "日程調整中", cache, [],
            False, "ダイヘン", "溶接棒",
            confirmed, confirming,
        )
        assert len(confirming) == 1
        assert confirming[0].status == "未処理"

    def test_bunno_takes_priority_over_price_pending(self):
        """分納 + 仮単価1 → 分納が優先（価格確認中に上書きされない）"""
        confirmed: list[HistoryRecord] = []
        confirming: list[ConfirmingRecord] = []
        row = _make_row(unit_price="1.00", net_amount="5.00")
        cache = _make_cache()
        bunno = [BunnoEntry("50個", "2/20", ""), BunnoEntry("50個", "未定", "")]
        _classify_order(
            row, "分納", cache, bunno,
            False, "ダイヘン", "溶接棒",
            confirmed, confirming,
        )
        assert len(confirming) == 1
        assert confirming[0].status == "分納"

    def test_stockout_takes_priority_over_price_pending(self):
        """欠品 + 仮単価1 → 欠品中が優先（価格確認中に上書きされない）"""
        confirmed: list[HistoryRecord] = []
        confirming: list[ConfirmingRecord] = []
        row = _make_row(unit_price="1.00", net_amount="5.00")
        cache = _make_cache()
        _classify_order(
            row, "欠品中", cache, [],
            False, "ダイヘン", "溶接棒",
            confirmed, confirming,
        )
        assert len(confirming) == 1
        assert confirming[0].status == "欠品中"

    def test_bunno_completed_not_blocked_by_price_pending(self):
        """分納全確定 + 仮単価1 → 分納完了が優先（送付履歴に入る）"""
        confirmed: list[HistoryRecord] = []
        confirming: list[ConfirmingRecord] = []
        row = _make_row(unit_price="1.00", net_amount="5.00")
        cache = _make_cache()
        bunno = [BunnoEntry("50個", "2/20", ""), BunnoEntry("50個", "2/25", "")]
        _classify_order(
            row, "分納", cache, bunno,
            False, "ダイヘン", "溶接棒",
            confirmed, confirming,
        )
        assert len(confirmed) == 1
        assert confirmed[0].delivery_answer == "分納完了"


# ============================================
# create_delivery_report_by_order_numbers
# ============================================
class TestCreateDeliveryReportByOrderNumbers:
    """注番指定モード回答書作成テスト"""

    def test_basic(self, tmp_path):
        """基本的な回答書作成"""
        data = [_make_row(order_number="100")]
        cache = _make_cache()
        result = create_delivery_report_by_order_numbers(
            data, "テスト商事", ["100"],
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is not None
        assert result.customer_name == "テスト商事"
        assert Path(result.file_path).exists()

    def test_no_matching_data(self, tmp_path):
        """該当なし → None"""
        data = [_make_row(order_number="100")]
        cache = _make_cache()
        result = create_delivery_report_by_order_numbers(
            data, "テスト商事", ["999"],
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is None

    def test_rejects_deleted(self, tmp_path):
        """明細削除は除外"""
        data = [_make_row(order_number="100", rejection_reason="明細削除")]
        cache = _make_cache()
        result = create_delivery_report_by_order_numbers(
            data, "テスト商事", ["100"],
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is None

    def test_multiple_orders(self, tmp_path):
        """複数注番"""
        data = [
            _make_row(order_number="100", detail_number="10"),
            _make_row(order_number="200", detail_number="10",
                      product_name="ガス 10L"),
        ]
        cache = _make_cache()
        result = create_delivery_report_by_order_numbers(
            data, "テスト商事", ["100", "200"],
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is not None
        # Excelを読んで2行あることを確認
        wb = load_workbook(result.file_path)
        ws = wb.active
        assert ws.cell(row=7, column=12).value == "100"
        assert ws.cell(row=8, column=12).value == "200"

    def test_tracking_info_collected(self, tmp_path):
        """送り状情報収集"""
        data = [
            _make_row(
                order_number="100",
                comment_external="佐川 1234567890",
            )
        ]
        cache = _make_cache()
        result = create_delivery_report_by_order_numbers(
            data, "テスト商事", ["100"],
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is not None
        assert len(result.tracking_info_list) == 1

    def test_stockout_info_collected(self, tmp_path):
        """欠品情報収集"""
        data = [
            _make_row(
                order_number="100",
                comment_detail="欠品中 3月上旬予定",
                order_delivery_date=datetime.date(2026, 12, 31),
            )
        ]
        cache = _make_cache()
        result = create_delivery_report_by_order_numbers(
            data, "テスト商事", ["100"],
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is not None
        assert len(result.stockout_info_list) == 1


# ============================================
# A案: remarks_mode=external でL列が社外コメントになる
# ============================================
class TestRemarksMode:
    """remarks_mode=external のL列置き換えテスト"""

    def test_external_mode_l_column_header(self, tmp_path):
        """remarks_mode=external → L列ヘッダーが「連絡事項」"""
        branch = BranchSettings(
            name="松本営業所", default_cutoff=15, remarks_mode="external",
        )
        data = [_make_row(order_number="100", comment_external="納品書同封")]
        cache = _make_cache()
        result = create_delivery_report_by_order_numbers(
            data, "テスト商事", ["100"],
            cache, tmp_path, {}, branch, EXEC_TIME, today=TODAY,
        )
        assert result is not None
        wb = load_workbook(result.file_path)
        ws = wb.active
        assert ws.cell(row=6, column=12).value == "連絡事項"

    def test_external_mode_l_column_value(self, tmp_path):
        """remarks_mode=external → L列にクリーニング済み社外コメント"""
        branch = BranchSettings(
            name="松本営業所", default_cutoff=15, remarks_mode="external",
        )
        data = [_make_row(order_number="100", comment_external="納品書同封")]
        cache = _make_cache()
        result = create_delivery_report_by_order_numbers(
            data, "テスト商事", ["100"],
            cache, tmp_path, {}, branch, EXEC_TIME, today=TODAY,
        )
        assert result is not None
        wb = load_workbook(result.file_path)
        ws = wb.active
        assert ws.cell(row=7, column=12).value == "納品書同封"

    def test_external_mode_tracking_stripped(self, tmp_path):
        """remarks_mode=external → 送り状番号は除去される"""
        branch = BranchSettings(
            name="松本営業所", default_cutoff=15, remarks_mode="external",
        )
        data = [_make_row(
            order_number="100",
            comment_external="佐川:1234567890 納品書同封",
        )]
        cache = _make_cache()
        result = create_delivery_report_by_order_numbers(
            data, "テスト商事", ["100"],
            cache, tmp_path, {}, branch, EXEC_TIME, today=TODAY,
        )
        assert result is not None
        wb = load_workbook(result.file_path)
        ws = wb.active
        val = ws.cell(row=7, column=12).value
        assert "佐川" not in str(val)
        assert "納品書同封" in str(val)

    def test_external_mode_tracking_only_empty(self, tmp_path):
        """社外コメントが送り状番号のみ → クリーニング後空文字 → L列は空文字"""
        branch = BranchSettings(
            name="松本営業所", default_cutoff=15, remarks_mode="external",
        )
        data = [_make_row(order_number="100", comment_external="佐川:1234567890")]
        cache = _make_cache()
        result = create_delivery_report_by_order_numbers(
            data, "テスト商事", ["100"],
            cache, tmp_path, {}, branch, EXEC_TIME, today=TODAY,
        )
        assert result is not None
        wb = load_workbook(result.file_path)
        ws = wb.active
        # クリーニング結果が空 → external_comment="" → 空セル（注番にフォールバックしない）
        # openpyxlは保存→再読み込みで空文字をNoneに変換する
        assert ws.cell(row=7, column=12).value is None

    def test_detail_mode_l_column_unchanged(self, tmp_path):
        """remarks_mode=detail → L列は注番のまま"""
        branch = BranchSettings(
            name="京葉営業所", default_cutoff=15, remarks_mode="detail",
        )
        data = [_make_row(order_number="100", comment_external="納品書同封")]
        cache = _make_cache()
        result = create_delivery_report_by_order_numbers(
            data, "テスト商事", ["100"],
            cache, tmp_path, {}, branch, EXEC_TIME, today=TODAY,
        )
        assert result is not None
        wb = load_workbook(result.file_path)
        ws = wb.active
        assert ws.cell(row=6, column=12).value == "弊社注番"
        assert ws.cell(row=7, column=12).value == "100"

    def test_external_mode_period_report(self, tmp_path):
        """期間モードでもremarks_mode=externalが動作する"""
        branch = BranchSettings(
            name="松本営業所", default_cutoff=15, remarks_mode="external",
        )
        data = [_make_row(comment_external="注意事項あり")]
        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト商事", {},
            cache, tmp_path, {}, branch, EXEC_TIME,
            date_from=datetime.date(2026, 2, 1),
            date_to=datetime.date(2026, 2, 28),
            today=TODAY,
        )
        assert result is not None
        wb = load_workbook(result.file_path)
        ws = wb.active
        assert ws.cell(row=6, column=12).value == "連絡事項"
        assert ws.cell(row=7, column=12).value == "注意事項あり"


# ============================================
# create_delivery_report
# ============================================
class TestCreateDeliveryReport:
    """期間指定モード回答書作成テスト"""

    def test_basic(self, tmp_path):
        """基本的な回答書作成"""
        data = [_make_row()]
        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト商事", {},
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            date_from=datetime.date(2026, 2, 1),
            date_to=datetime.date(2026, 2, 28),
            today=TODAY,
        )
        assert result is not None
        assert result.customer_name == "テスト商事"
        assert Path(result.file_path).exists()
        assert len(result.confirmed_orders) == 1

    def test_no_matching_customer(self, tmp_path):
        """顧客名不一致 → None"""
        data = [_make_row(customer_name="別会社")]
        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト商事", {},
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is None

    def test_date_range_filter(self, tmp_path):
        """期間外はフィルタされる"""
        data = [
            _make_row(registration_date=datetime.date(2026, 1, 15)),
        ]
        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト商事", {},
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            date_from=datetime.date(2026, 2, 1),
            date_to=datetime.date(2026, 2, 28),
            today=TODAY,
        )
        assert result is None

    def test_already_sent_skip(self, tmp_path):
        """送付済み伝票はスキップ（登録日が今日より前）"""
        data = [
            _make_row(
                registration_date=datetime.date(2026, 2, 10),
            ),
        ]
        sent = {"1000001|10": "2月15日配達予定"}
        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト商事", sent,
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            date_from=datetime.date(2026, 2, 1),
            date_to=datetime.date(2026, 2, 28),
            today=TODAY,
        )
        assert result is None

    def test_today_registration_resend(self, tmp_path):
        """当日登録の送付済み → 再送OK"""
        data = [
            _make_row(registration_date=TODAY),
        ]
        sent = {"1000001|10": "2月16日配達予定"}
        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト商事", sent,
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            date_from=datetime.date(2026, 2, 1),
            date_to=datetime.date(2026, 2, 28),
            today=TODAY,
        )
        # 当日登録の場合、registration_date < today が False なので再送される
        assert result is not None

    def test_force_delivered(self, tmp_path):
        """直送+処理完了 → 納品済み"""
        data = [
            _make_row(
                document_type="【受注】直送販売",
                storage_place="転送中（直送用）",
                ship_status="処理完了",
                order_delivery_date=datetime.date(2026, 12, 31),
            ),
        ]
        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト商事", {},
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is not None
        # 確定伝票として記録
        assert len(result.confirmed_orders) == 1
        assert result.confirmed_orders[0].delivery_answer == "納品済み"

    def test_confirming_orders(self, tmp_path):
        """確認中 → confirming_orders"""
        data = [
            _make_row(
                order_delivery_date=datetime.date(2026, 12, 31),
                document_type="【受注】直送販売",
                storage_place="転送中（直送用）",
            ),
        ]
        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト商事", {},
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is not None
        assert len(result.confirming_orders) == 1
        assert result.has_confirming is True

    def test_excluded_skip(self, tmp_path):
        """除外伝票はスキップ"""
        data = [_make_row()]
        sent = {"1000001|10": "除外"}
        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト商事", sent,
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is None

    def test_bunno_completed_notification(self, tmp_path):
        """分納完了の通知収集"""
        data = [
            _make_row(
                ship_status="処理完了",
                order_number="100",
                detail_number="10",
                comment_detail="分納:50個 2/10,50個 2/15",
            ),
        ]
        cache = _make_cache(
            confirm={"100|10": ("未", "分納", None)}
        )
        result = create_delivery_report(
            data, "テスト商事", {},
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is not None
        assert len(result.bunno_completed_list) == 1

    def test_hash_exclude(self, tmp_path):
        """##除外"""
        data = [
            _make_row(comment_internal="##この伝票は除外"),
        ]
        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト商事", {},
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is None

    def test_bunno_kanryo_skip(self, tmp_path):
        """分納完了は送付済みとしてスキップ"""
        data = [_make_row(registration_date=datetime.date(2026, 2, 10))]
        sent = {"1000001|10": "分納完了"}
        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト商事", sent,
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            date_from=datetime.date(2026, 2, 1),
            date_to=datetime.date(2026, 2, 28),
            today=TODAY,
        )
        assert result is None


# ============================================
# 結合テスト
# ============================================
class TestIntegration:
    """結合テスト"""

    def test_full_flow_period_mode(self, tmp_path):
        """期間モードの完全フロー"""
        data = [
            # 確定伝票
            _make_row(
                order_number="100", detail_number="10",
                product_name="溶接棒A", quantity="50",
            ),
            # 確認中伝票
            _make_row(
                order_number="200", detail_number="10",
                product_name="ガスB", quantity="10",
                order_delivery_date=datetime.date(2026, 12, 31),
                document_type="【受注】直送販売",
                storage_place="転送中（直送用）",
            ),
            # 別顧客（フィルタされる）
            _make_row(
                order_number="300", detail_number="10",
                customer_name="別会社",
            ),
        ]
        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト商事", {},
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            date_from=datetime.date(2026, 2, 1),
            date_to=datetime.date(2026, 2, 28),
            today=TODAY,
        )
        assert result is not None
        assert len(result.confirmed_orders) == 1
        assert len(result.confirming_orders) == 1

        # Excelファイルを検証
        wb = load_workbook(result.file_path)
        ws = wb.active
        assert ws.cell(row=7, column=12).value == "100"
        assert ws.cell(row=8, column=12).value == "200"

    def test_full_flow_order_number_mode(self, tmp_path):
        """注番モードの完全フロー"""
        data = [
            _make_row(order_number="100", detail_number="10"),
            _make_row(order_number="100", detail_number="20",
                      product_name="溶接ワイヤ"),
            _make_row(order_number="200", detail_number="10",
                      customer_name="別会社"),
        ]
        cache = _make_cache()

        # まず注番をグループ化
        groups = group_order_numbers_by_customer(data, ["100"])
        assert "テスト商事" in groups

        # 回答書作成
        result = create_delivery_report_by_order_numbers(
            data, "テスト商事", groups["テスト商事"],
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is not None
        wb = load_workbook(result.file_path)
        ws = wb.active
        # 2行（同じ注番の2明細）
        assert ws.cell(row=7, column=12).value == "100"
        assert ws.cell(row=8, column=12).value == "100"

    def test_price_pending_lifecycle(self, tmp_path):
        """価格確認中のライフサイクル統合テスト

        ステップ1: 単価1円＋納期確定 → 確認中一覧に「価格確認中」
        ステップ2: 次回実行 → sent_ordersに載らない → 再出力
        ステップ3: 売価確定（単価≠1）→ 正式単価で出力 → 送付履歴へ
        ステップ4: その次の実行 → sent_ordersに載っている → スキップ
        """
        from nouki_kaitou.history import load_delivery_history

        # --- ステップ1: 仮単価で初回実行 ---
        data_step1 = [
            _make_row(
                order_number="500", detail_number="10",
                unit_price="1.00", net_amount="5.00",
                product_name="ザ・硝フッ酸 2C 20KG",
                # 指定納期あり → 納期は確定（「出荷済み」等になる）
                specified_delivery_date=datetime.date(2026, 2, 10),
            ),
        ]
        cache = _make_cache()
        result1 = create_delivery_report(
            data_step1, "テスト商事", {},
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            date_from=datetime.date(2026, 2, 1),
            date_to=datetime.date(2026, 2, 28),
            today=TODAY,
        )
        assert result1 is not None
        # 確認中一覧に「価格確認中」で入る（送付履歴ではない）
        assert len(result1.confirming_orders) == 1
        assert result1.confirming_orders[0].status == "価格確認中"
        assert len(result1.confirmed_orders) == 0
        # 回答書の単価・金額は「確認中」
        wb1 = load_workbook(result1.file_path)
        ws1 = wb1.active
        assert ws1.cell(row=7, column=7).value == "確認中"  # G列: 単価
        assert ws1.cell(row=7, column=8).value == "確認中"  # H列: 金額
        # 納期回答欄は確定した値のまま（A案）
        delivery_val = str(ws1.cell(row=7, column=9).value or "")
        assert "済" in delivery_val  # 「配達済み」or「出荷済み」

        # --- ステップ2: 次回実行（売価まだ未確定） ---
        # 確認中一覧に「価格確認中」がある状態をシミュレート
        # load_delivery_historyは確認中一覧から「除外」しか読まない
        # → 「価格確認中」はsent_ordersに載らない
        wb_hist = Workbook()
        ws_hist = wb_hist.active
        ws_hist.append(["送付日時", "受注日", "顧客名", "注番",
                        "明細", "メーカー", "品名", "ステータス"])

        wb_conf = Workbook()
        ws_conf = wb_conf.active
        ws_conf.append(["送付日時", "受注日", "顧客名", "注番",
                        "明細", "メーカー", "品名", "ステータス"])
        ws_conf.append([
            datetime.datetime(2026, 2, 16, 10, 0),
            datetime.date(2026, 2, 16),
            "テスト商事", "500", "10",
            "ダイヘン", "ザ・硝フッ酸", "価格確認中",
        ])

        sent_orders2 = load_delivery_history(
            ws_hist, ws_conf, cache, {}, TODAY,
        )
        # 「価格確認中」はsent_ordersに入らない
        assert "500|10" not in sent_orders2

        # sent_orders空 → 再出力される
        result2 = create_delivery_report(
            data_step1, "テスト商事", sent_orders2,
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            date_from=datetime.date(2026, 2, 1),
            date_to=datetime.date(2026, 2, 28),
            today=TODAY,
        )
        assert result2 is not None
        assert len(result2.confirming_orders) == 1
        assert result2.confirming_orders[0].status == "価格確認中"

        # --- ステップ3: 売価確定（単価500円に変更） ---
        data_step3 = [
            _make_row(
                order_number="500", detail_number="10",
                unit_price="500", net_amount="2500",
                product_name="ザ・硝フッ酸 2C 20KG",
                specified_delivery_date=datetime.date(2026, 2, 10),
            ),
        ]
        result3 = create_delivery_report(
            data_step3, "テスト商事", sent_orders2,
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            date_from=datetime.date(2026, 2, 1),
            date_to=datetime.date(2026, 2, 28),
            today=TODAY,
        )
        assert result3 is not None
        # 正式単価 → 送付履歴に移動（confirmed_orders）
        assert len(result3.confirmed_orders) == 1
        assert len(result3.confirming_orders) == 0
        # 回答書の単価・金額は正式値
        wb3 = load_workbook(result3.file_path)
        ws3 = wb3.active
        assert ws3.cell(row=7, column=7).value == 500   # G列: 単価
        assert ws3.cell(row=7, column=8).value == 2500   # H列: 金額

        # --- ステップ4: その次の実行 → スキップ ---
        # 送付履歴にステップ3の結果が記録された状態
        delivery_answer = result3.confirmed_orders[0].delivery_answer
        sent_orders4 = {"500|10": delivery_answer}
        data_step4 = [
            _make_row(
                order_number="500", detail_number="10",
                unit_price="500", net_amount="2500",
                product_name="ザ・硝フッ酸 2C 20KG",
                specified_delivery_date=datetime.date(2026, 2, 10),
                # registration_dateがtodayより前でスキップ判定が効く
                registration_date=datetime.date(2026, 2, 10),
            ),
        ]
        result4 = create_delivery_report(
            data_step4, "テスト商事", sent_orders4,
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            date_from=datetime.date(2026, 2, 1),
            date_to=datetime.date(2026, 2, 28),
            today=TODAY,
        )
        # sent_ordersに載っている + registration_date < today → スキップ
        assert result4 is None


# ============================================
# _collect_stockout_info
# ============================================
class TestCollectStockoutInfo:
    """欠品情報収集テスト"""

    def test_normal_stockout(self):
        """通常の欠品 → StockoutEntry収集"""
        row = _make_row(
            comment_detail="欠品中 3月上旬予定",
            ship_status="未処理",
        )
        cache = _make_cache()
        result: list[StockoutEntry] = []
        _collect_stockout_info(row, cache, "欠品中", result)
        assert len(result) == 1
        assert result[0].approx_delivery == "3月上旬入荷予定"

    def test_stockout_completed_excluded(self):
        """欠品+処理完了 → 欠品解消で収集しない"""
        row = _make_row(
            comment_detail="欠品中 3月上旬予定",
            ship_status="処理完了",
        )
        cache = _make_cache()
        result: list[StockoutEntry] = []
        _collect_stockout_info(row, cache, "納品済み", result)
        assert len(result) == 0

    def test_stockout_shipping_excluded(self):
        """送料行の欠品 → 収集しない"""
        row = _make_row(
            comment_detail="欠品中",
            product_name="送料",
            ship_status="未処理",
        )
        cache = _make_cache()
        result: list[StockoutEntry] = []
        _collect_stockout_info(row, cache, "欠品中", result)
        assert len(result) == 0

    def test_stockout_with_bunno_excluded(self):
        """欠品+分納併存 → 欠品セクション除外（分納セクションで表示）"""
        row = _make_row(
            comment_detail="欠品中 分納:50個 2/20,50個 未定",
            ship_status="未処理",
        )
        cache = _make_cache()
        result: list[StockoutEntry] = []
        _collect_stockout_info(row, cache, "分納", result)
        assert len(result) == 0

    def test_no_stockout_marker(self):
        """「欠品中」なし → 収集しない"""
        row = _make_row(
            comment_detail="通常コメント",
            ship_status="未処理",
        )
        cache = _make_cache()
        result: list[StockoutEntry] = []
        _collect_stockout_info(row, cache, "確認中", result)
        assert len(result) == 0


# ============================================
# 統合テスト: 分納・欠品ライフサイクル
# ============================================
class TestBunnoStockoutLifecycle:
    """分納・欠品のライフサイクル統合テスト"""

    def test_stockout_force_delivered(self, tmp_path):
        """欠品+直送処理完了 → force_deliveredで「納品済み」"""
        data = [
            _make_row(
                comment_detail="欠品中",
                document_type="【受注】直送販売",
                storage_place="転送中（直送用）",
                ship_status="処理完了",
                order_delivery_date=datetime.date(2026, 12, 31),
            ),
        ]
        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト商事", {},
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is not None
        assert len(result.confirmed_orders) == 1
        assert result.confirmed_orders[0].delivery_answer == "納品済み"
        # 欠品情報は収集されない（処理完了なので）
        assert len(result.stockout_info_list) == 0

    def test_stockout_confirming_lifecycle(self, tmp_path):
        """欠品 → 確認中一覧に「欠品中」で入る"""
        data = [
            _make_row(
                comment_detail="欠品中 3月上旬予定",
                order_delivery_date=datetime.date(2026, 12, 31),
                document_type="【受注】直送販売",
                storage_place="転送中（直送用）",
            ),
        ]
        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト商事", {},
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is not None
        assert len(result.confirming_orders) == 1
        assert result.confirming_orders[0].status == "欠品中"
        assert len(result.stockout_info_list) == 1

    def test_bunno_with_mitei_lifecycle(self, tmp_path):
        """分納（未定あり） → 確認中一覧に「分納」+ 分納情報収集"""
        data = [
            _make_row(
                comment_detail="分納:50個 2/20,50個 未定",
                ship_status="未処理",
                order_delivery_date=datetime.date(2026, 2, 20),
                document_type="【受注】直送販売",
                storage_place="転送中（直送用）",
            ),
        ]
        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト商事", {},
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is not None
        assert len(result.confirming_orders) == 1
        assert result.confirming_orders[0].status == "分納"
        assert len(result.bunno_info_list) == 1

    def test_bunno_partial_ship_lifecycle(self, tmp_path):
        """一部処理済み + 分納 → 未処理と同じ扱い"""
        data = [
            _make_row(
                comment_detail="分納:50個 2/10,50個 未定",
                ship_status="一部処理済み",
                order_delivery_date=datetime.date(2026, 2, 20),
                document_type="【受注】直送販売",
                storage_place="転送中（直送用）",
            ),
        ]
        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト商事", {},
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is not None
        # 一部処理済みでも分納なら確認中に残る
        assert len(result.confirming_orders) == 1
        assert result.confirming_orders[0].status == "分納"
        assert len(result.bunno_info_list) == 1
