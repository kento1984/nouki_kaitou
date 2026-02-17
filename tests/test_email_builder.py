"""email_builder.py のユニットテスト

メール生成の全関数を網羅的にテスト。
"""

import datetime

import pytest
from openpyxl import Workbook

from nouki_kaitou.email_builder import (
    build_email_body_html,
    build_email_subject,
    create_emails,
    get_rep_email_addresses,
    html_escape,
)
from nouki_kaitou.models import (
    BranchSettings,
    BunnoEntry,
    CacheStore,
    StockoutEntry,
    TrackingEntry,
)


# ============================================
# html_escape
# ============================================
class TestHtmlEscape:
    def test_ampersand(self):
        assert html_escape("A&B") == "A&amp;B"

    def test_less_than(self):
        assert html_escape("<tag>") == "&lt;tag&gt;"

    def test_quote(self):
        assert html_escape('"text"') == "&quot;text&quot;"

    def test_no_escape(self):
        assert html_escape("テスト文字列") == "テスト文字列"

    def test_empty(self):
        assert html_escape("") == ""

    def test_combined(self):
        assert html_escape('A&B<C>"D') == "A&amp;B&lt;C&gt;&quot;D"


# ============================================
# build_email_subject
# ============================================
class TestBuildEmailSubject:
    def test_basic_subject(self):
        """基本的な件名"""
        result = build_email_subject(
            "テスト株式会社", today=datetime.date(2026, 2, 16)
        )
        assert result == "【マツモト産業】納期回答書_02/16受注分_テスト株式会社様"

    def test_with_rep_name(self):
        """担当者名あり"""
        result = build_email_subject(
            "テスト株式会社", rep_name="田中",
            today=datetime.date(2026, 2, 16),
        )
        assert result == (
            "【マツモト産業】納期回答書_02/16受注分_テスト株式会社様"
            "（田中様担当分）"
        )

    def test_other_rep(self):
        """__OTHER__は担当者なし扱い"""
        result = build_email_subject(
            "テスト株式会社", rep_name="__OTHER__",
            today=datetime.date(2026, 2, 16),
        )
        assert "担当分" not in result

    def test_empty_rep(self):
        """空文字の担当者名"""
        result = build_email_subject(
            "テスト株式会社", rep_name="",
            today=datetime.date(2026, 2, 16),
        )
        assert "担当分" not in result

    def test_single_digit_month(self):
        """1桁月のゼロパディング"""
        result = build_email_subject(
            "テスト", today=datetime.date(2026, 3, 5)
        )
        assert "03/05" in result


# ============================================
# get_rep_email_addresses
# ============================================
class TestGetRepEmailAddresses:
    def _make_rep_ws(self, rows):
        """テスト用担当者マスターシートを作成"""
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1).value = "顧客名"
        ws.cell(row=1, column=2).value = "担当者"
        ws.cell(row=1, column=3).value = "メール1"
        ws.cell(row=1, column=4).value = "メール2"

        for row_idx, data in enumerate(rows, 2):
            for col_idx, val in enumerate(data, 1):
                ws.cell(row=row_idx, column=col_idx).value = val

        return ws

    def test_basic(self):
        """基本的なメアド取得"""
        ws = self._make_rep_ws([
            ("テスト株式会社", "田中", "tanaka@example.com", None),
        ])
        result = get_rep_email_addresses("テスト株式会社", "田中", ws)
        assert result == "tanaka@example.com"

    def test_multiple_emails(self):
        """複数メアド"""
        ws = self._make_rep_ws([
            ("テスト株式会社", "田中", "t1@example.com", "t2@example.com"),
        ])
        result = get_rep_email_addresses("テスト株式会社", "田中", ws)
        assert result == "t1@example.com; t2@example.com"

    def test_with_sama(self):
        """担当者名に「様」付き"""
        ws = self._make_rep_ws([
            ("テスト株式会社", "田中様", "tanaka@example.com", None),
        ])
        result = get_rep_email_addresses("テスト株式会社", "田中", ws)
        assert result == "tanaka@example.com"

    def test_not_found(self):
        """該当なし"""
        ws = self._make_rep_ws([
            ("テスト株式会社", "田中", "tanaka@example.com", None),
        ])
        result = get_rep_email_addresses("テスト株式会社", "佐藤", ws)
        assert result == ""

    def test_none_ws(self):
        """シートがNone"""
        result = get_rep_email_addresses("テスト", "田中", None)
        assert result == ""

    def test_different_customer(self):
        """顧客名不一致"""
        ws = self._make_rep_ws([
            ("テスト株式会社", "田中", "tanaka@example.com", None),
        ])
        result = get_rep_email_addresses("別の会社", "田中", ws)
        assert result == ""


# ============================================
# build_email_body_html — 基本構造
# ============================================
class TestBuildEmailBodyHtml:
    def test_basic_structure(self):
        """基本的なHTML構造"""
        branch = BranchSettings(name="京葉営業所")
        result = build_email_body_html("テスト株式会社", branch)
        assert "<html>" in result
        assert "</html>" in result
        assert "テスト株式会社 御中" in result
        assert "マツモト産業㈱京葉営業所" in result

    def test_greeting(self):
        """挨拶文が含まれる"""
        branch = BranchSettings(name="京葉営業所")
        result = build_email_body_html("テスト株式会社", branch)
        assert "いつもお世話になっております" in result
        assert "納期回答書をお送りいたします" in result

    def test_confirming_note(self):
        """確認中の注記が含まれる"""
        branch = BranchSettings(name="京葉営業所")
        result = build_email_body_html("テスト", branch)
        assert "確認中" in result
        assert "メーカー確認後あらためてご連絡" in result

    def test_signature(self):
        """署名が含まれる"""
        branch = BranchSettings(name="京葉営業所")
        result = build_email_body_html("テスト", branch)
        assert "マツモト産業株式会社" in result
        assert "京葉営業所" in result

    def test_html_escape_customer(self):
        """顧客名のHTMLエスケープ"""
        branch = BranchSettings(name="テスト")
        result = build_email_body_html("A&B<株式会社>", branch)
        assert "A&amp;B&lt;株式会社&gt;" in result


# ============================================
# build_email_body_html — 送り状セクション
# ============================================
class TestBuildEmailBodyHtmlTracking:
    def test_tracking_section(self):
        """送り状セクションが含まれる"""
        branch = BranchSettings(name="テスト")
        tracking = [
            ("メーカーA", "製品A", "100",
             TrackingEntry(carrier_name="ヤマト運輸", tracking_number="1234567890123")),
        ]
        result = build_email_body_html(
            "テスト", branch, tracking_info_list=tracking,
        )
        assert "送り状番号のご連絡" in result
        assert "ヤマト運輸" in result
        assert "1234567890123" in result

    def test_direct_tracking_link(self):
        """直接追跡可能な運送会社はハイパーリンク"""
        branch = BranchSettings(name="テスト")
        tracking = [
            ("メーカーA", "製品A", "100",
             TrackingEntry(carrier_name="佐川急便", tracking_number="9876543210")),
        ]
        result = build_email_body_html(
            "テスト", branch, tracking_info_list=tracking,
        )
        assert "href=" in result
        assert "sagawa" in result

    def test_indirect_tracking(self):
        """間接追跡の運送会社は追跡ページリンク"""
        branch = BranchSettings(name="テスト")
        tracking = [
            ("メーカーA", "製品A", "100",
             TrackingEntry(carrier_name="トナミ運輸", tracking_number="1234567890")),
        ]
        result = build_email_body_html(
            "テスト", branch, tracking_info_list=tracking,
        )
        assert "追跡ページ" in result
        assert "番号を入力してください" in result

    def test_no_tracking_no_section(self):
        """送り状なしではセクションなし"""
        branch = BranchSettings(name="テスト")
        result = build_email_body_html("テスト", branch)
        assert "送り状番号のご連絡" not in result

    def test_product_displayed(self):
        """商品名が表示される"""
        branch = BranchSettings(name="テスト")
        tracking = [
            ("メーカーA", "製品ABC", "50",
             TrackingEntry(carrier_name="ヤマト運輸", tracking_number="1234567890123")),
        ]
        result = build_email_body_html(
            "テスト", branch, tracking_info_list=tracking,
        )
        assert "メーカーA" in result
        assert "製品ABC" in result
        assert "x50" in result


# ============================================
# build_email_body_html — 欠品セクション
# ============================================
class TestBuildEmailBodyHtmlStockout:
    def test_stockout_section(self):
        """欠品セクションが含まれる"""
        branch = BranchSettings(name="テスト")
        stockout = [
            StockoutEntry(
                manufacturer_name="メーカーX",
                product_name="製品X",
                quantity="30",
                approx_delivery="3月上旬入荷予定",
            ),
        ]
        result = build_email_body_html(
            "テスト", branch, stockout_info_list=stockout,
        )
        assert "欠品中の商品について" in result
        assert "メーカーX" in result
        assert "3月上旬入荷予定" in result

    def test_stockout_no_approx(self):
        """概算納期なし → 入荷次第ご連絡"""
        branch = BranchSettings(name="テスト")
        stockout = [
            StockoutEntry(
                manufacturer_name="メーカーX",
                product_name="製品X",
                quantity="30",
            ),
        ]
        result = build_email_body_html(
            "テスト", branch, stockout_info_list=stockout,
        )
        assert "入荷次第ご連絡" in result

    def test_no_stockout_no_section(self):
        """欠品なしではセクションなし"""
        branch = BranchSettings(name="テスト")
        result = build_email_body_html("テスト", branch)
        assert "欠品中の商品について" not in result


# ============================================
# build_email_body_html — 分納セクション
# ============================================
class TestBuildEmailBodyHtmlBunno:
    def test_bunno_section(self):
        """分納セクションが含まれる"""
        branch = BranchSettings(name="テスト")
        bunno = [
            {
                "manufacturer": "メーカーA",
                "product": "製品A",
                "quantity": "300",
                "entries": [
                    BunnoEntry(quantity="200m", date_str="3/10"),
                    BunnoEntry(quantity="100m", date_str="3/15"),
                ],
                "calc_details": [
                    ("200m", "3/10", "", "3/10出荷予定"),
                    ("100m", "3/15", "", "3/15出荷予定"),
                ],
                "is_ship_rule": True,
                "days_to_add": 0,
                "order_number": "12345",
                "detail_number": "10",
                "is_rosenbin": False,
            },
        ]
        result = build_email_body_html(
            "テスト", branch, bunno_info_list=bunno,
            today=datetime.date(2026, 2, 16),
        )
        assert "分納のご連絡" in result
        assert "分納にてお届け" in result
        assert "メーカーA" in result

    def test_bunno_circled_numbers(self):
        """丸数字が使われる"""
        branch = BranchSettings(name="テスト")
        bunno = [
            {
                "manufacturer": "メーカーA",
                "product": "製品A",
                "quantity": "300",
                "entries": [
                    BunnoEntry(quantity="200m", date_str="3/10"),
                    BunnoEntry(quantity="100m", date_str="3/15"),
                ],
                "calc_details": [
                    ("200m", "3/10", "", "3/10出荷予定"),
                    ("100m", "3/15", "", "3/15出荷予定"),
                ],
                "is_ship_rule": True,
                "days_to_add": 0,
                "order_number": "12345",
                "detail_number": "10",
                "is_rosenbin": False,
            },
        ]
        result = build_email_body_html(
            "テスト", branch, bunno_info_list=bunno,
            today=datetime.date(2026, 2, 16),
        )
        assert "①" in result
        assert "②" in result

    def test_no_bunno_no_section(self):
        """分納なしではセクションなし"""
        branch = BranchSettings(name="テスト")
        result = build_email_body_html("テスト", branch)
        assert "分納のご連絡" not in result


# ============================================
# build_email_body_html — 分納完了セクション
# ============================================
class TestBuildEmailBodyHtmlBunnoCompleted:
    def test_bunno_completed_section(self):
        """分納完了セクションが含まれる"""
        branch = BranchSettings(name="テスト")
        completed = [("メーカーA", "製品A", "300")]
        result = build_email_body_html(
            "テスト", branch, bunno_completed_list=completed,
        )
        assert "分納完了のご連絡" in result
        assert "全て出荷が完了" in result
        assert "メーカーA" in result

    def test_no_completed_no_section(self):
        """分納完了なしではセクションなし"""
        branch = BranchSettings(name="テスト")
        result = build_email_body_html("テスト", branch)
        assert "分納完了のご連絡" not in result


# ============================================
# create_emails
# ============================================
class TestCreateEmails:
    def _make_customer_ws(self, rows):
        """テスト用顧客マスターシート"""
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1).value = "顧客名"
        ws.cell(row=1, column=2).value = "住所"
        ws.cell(row=1, column=3).value = "電話"
        ws.cell(row=1, column=4).value = "担当"
        ws.cell(row=1, column=5).value = "メール1"
        ws.cell(row=1, column=6).value = "メール2"

        for row_idx, data in enumerate(rows, 2):
            for col_idx, val in enumerate(data, 1):
                ws.cell(row=row_idx, column=col_idx).value = val

        return ws

    def test_basic_create(self):
        """基本的なメール作成"""
        branch = BranchSettings(name="京葉営業所", shared_email="shared@test.com")
        cust_ws = self._make_customer_ws([
            ("テスト株式会社", "", "", "", "test@example.com", None),
        ])

        files = [
            {
                "customer_name": "テスト株式会社",
                "file_path": "/tmp/test.pdf",
            },
        ]

        results = create_emails(
            files, branch, cust_ws, today=datetime.date(2026, 2, 16),
        )

        assert len(results) == 1
        assert results[0]["to"] == "test@example.com"
        assert "テスト株式会社" in results[0]["subject"]
        assert "<html>" in results[0]["html_body"]
        assert results[0]["attachments"] == ["/tmp/test.pdf"]
        assert results[0]["shared_email"] == "shared@test.com"

    def test_skip_no_email(self):
        """メアド未登録はスキップ"""
        branch = BranchSettings(name="テスト")
        cust_ws = self._make_customer_ws([
            ("テスト株式会社", "", "", "", None, None),
        ])

        files = [
            {"customer_name": "テスト株式会社", "file_path": "/tmp/test.pdf"},
        ]

        results = create_emails(files, branch, cust_ws)
        assert len(results) == 0

    def test_multiple_customers(self):
        """複数顧客"""
        branch = BranchSettings(name="テスト")
        cust_ws = self._make_customer_ws([
            ("顧客A", "", "", "", "a@test.com", None),
            ("顧客B", "", "", "", "b@test.com", None),
        ])

        files = [
            {"customer_name": "顧客A", "file_path": "/tmp/a.pdf"},
            {"customer_name": "顧客B", "file_path": "/tmp/b.pdf"},
        ]

        results = create_emails(
            files, branch, cust_ws, today=datetime.date(2026, 2, 16),
        )
        assert len(results) == 2
        assert results[0]["to"] == "a@test.com"
        assert results[1]["to"] == "b@test.com"

    def test_with_rep_master(self):
        """担当者マスターからメアド取得"""
        branch = BranchSettings(name="テスト")
        cust_ws = self._make_customer_ws([
            ("テスト株式会社", "", "", "", "company@test.com", None),
        ])

        rep_wb = Workbook()
        rep_ws = rep_wb.active
        rep_ws.cell(row=1, column=1).value = "顧客名"
        rep_ws.cell(row=1, column=2).value = "担当者"
        rep_ws.cell(row=1, column=3).value = "メール"
        rep_ws.cell(row=2, column=1).value = "テスト株式会社"
        rep_ws.cell(row=2, column=2).value = "田中"
        rep_ws.cell(row=2, column=3).value = "tanaka@test.com"

        files = [
            {
                "customer_name": "テスト株式会社",
                "file_path": "/tmp/test.pdf",
                "rep_name": "田中",
            },
        ]

        results = create_emails(
            files, branch, cust_ws, rep_master_ws=rep_ws,
            today=datetime.date(2026, 2, 16),
        )
        assert len(results) == 1
        assert results[0]["to"] == "tanaka@test.com"
        assert "田中様担当分" in results[0]["subject"]

    def test_with_tracking_and_stockout(self):
        """送り状+欠品情報付きメール"""
        branch = BranchSettings(name="テスト")
        cust_ws = self._make_customer_ws([
            ("テスト", "", "", "", "test@test.com", None),
        ])

        files = [
            {
                "customer_name": "テスト",
                "file_path": "/tmp/test.pdf",
                "tracking_info_list": [
                    ("メーカーA", "製品A", "100",
                     TrackingEntry(carrier_name="ヤマト運輸",
                                  tracking_number="1234567890123")),
                ],
                "stockout_info_list": [
                    StockoutEntry(
                        manufacturer_name="メーカーB",
                        product_name="製品B",
                        quantity="50",
                    ),
                ],
            },
        ]

        results = create_emails(
            files, branch, cust_ws, today=datetime.date(2026, 2, 16),
        )
        assert len(results) == 1
        body = results[0]["html_body"]
        assert "送り状番号のご連絡" in body
        assert "欠品中の商品について" in body

    def test_empty_files(self):
        """空リスト"""
        branch = BranchSettings(name="テスト")
        cust_ws = self._make_customer_ws([])
        results = create_emails([], branch, cust_ws)
        assert results == []

    def test_send_directly_flag(self):
        """send_directlyフラグが結果に反映"""
        branch = BranchSettings(name="テスト")
        cust_ws = self._make_customer_ws([
            ("テスト", "", "", "", "test@test.com", None),
        ])
        files = [{"customer_name": "テスト", "file_path": ""}]

        results = create_emails(files, branch, cust_ws, send_directly=True)
        assert results[0]["send_directly"] is True


# ============================================
# 統合テスト
# ============================================
class TestIntegration:
    def test_full_email_generation(self):
        """全セクション含むメール生成"""
        branch = BranchSettings(name="京葉営業所")
        today = datetime.date(2026, 2, 16)

        tracking = [
            ("メーカーA", "製品A", "100",
             TrackingEntry(carrier_name="佐川急便", tracking_number="9876543210")),
        ]
        stockout = [
            StockoutEntry(
                manufacturer_name="メーカーB",
                product_name="製品B",
                quantity="50",
                approx_delivery="3月上旬",
            ),
        ]
        bunno = [
            {
                "manufacturer": "メーカーC",
                "product": "製品C",
                "quantity": "300",
                "entries": [
                    BunnoEntry(quantity="200m", date_str="3/10"),
                    BunnoEntry(quantity="100m", date_str="未定"),
                ],
                "calc_details": [
                    ("200m", "3/10", "", "3/10出荷予定"),
                    ("100m", "未定", "", "確認中"),
                ],
                "is_ship_rule": True,
                "days_to_add": 0,
                "order_number": "12345",
                "detail_number": "10",
                "is_rosenbin": False,
            },
        ]
        completed = [("メーカーD", "製品D", "500")]

        result = build_email_body_html(
            customer_name="テスト株式会社",
            branch=branch,
            stockout_info_list=stockout,
            tracking_info_list=tracking,
            bunno_info_list=bunno,
            bunno_completed_list=completed,
            today=today,
        )

        # 全セクションが含まれている
        assert "テスト株式会社 御中" in result
        assert "京葉営業所" in result
        assert "送り状番号のご連絡" in result
        assert "欠品中の商品について" in result
        assert "分納のご連絡" in result
        assert "分納完了のご連絡" in result
        assert "メーカー確認後あらためてご連絡" in result
        assert "マツモト産業株式会社" in result
        assert "</html>" in result
