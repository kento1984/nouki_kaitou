"""Phase 10: エラーハンドリングテスト

不正データ・境界値・異常系の耐性テスト:
- 不正な日付・数値
- 欠損列・空のシート
- 空文字列・None値
- 異常なコメント形式
- 破損したキャッシュデータ
"""

import datetime
import tempfile
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from nouki_kaitou.bunno import (
    calculate_bunno_date,
    extract_bunno_info,
    extract_location_from_token,
    has_bunno_mitei,
    normalize_bunno_date,
    split_qty_and_date,
)
from nouki_kaitou.business_days import (
    add_business_days,
    get_next_business_day,
    get_next_delivery_day,
    get_previous_business_day,
    is_holiday,
)
from nouki_kaitou.config import get_branch_settings, load_branch_settings, load_holidays
from nouki_kaitou.confirming import get_confirmed_delivery_date, get_confirming_status
from nouki_kaitou.customer import get_customer_delivery_days, get_email_addresses, is_route_delivery
from nouki_kaitou.data_loader import get_column_positions, load_source_file, parse_order_row
from nouki_kaitou.delivery_calc import (
    calculate_delivery_date,
    extract_arrival_date_from_internal,
    extract_pickup_date,
)
from nouki_kaitou.email_builder import build_email_body_html, build_email_subject
from nouki_kaitou.history import (
    clean_old_confirming_list,
    clean_old_history,
    initialize_delivery_history,
    load_delivery_history,
    save_confirming_list,
    save_delivery_history,
)
from nouki_kaitou.manufacturer import get_delivery_days_to_add, get_manufacturer_name
from nouki_kaitou.models import (
    BranchSettings,
    BunnoEntry,
    CacheStore,
    ConfirmingRecord,
    HolidayMap,
    HistoryRecord,
    OrderRow,
    ReportRow,
)
from nouki_kaitou.report_generator import build_report_row, create_delivery_report
from nouki_kaitou.stockout import extract_approx_delivery, remove_stockout_text
from nouki_kaitou.tracking import extract_tracking_info
from nouki_kaitou.utils import (
    convert_to_half_width,
    extract_date_from_string,
    format_date_japanese,
    is_december_31,
    parse_date,
    parse_time,
)


# ============================================
# テスト用ヘルパー
# ============================================
TODAY = datetime.date(2026, 2, 16)
EXEC_TIME = datetime.datetime(2026, 2, 16, 10, 0, 0)
BRANCH = BranchSettings(
    name="京葉営業所",
    default_cutoff=15,
    base_center="関東商品センター",
)


def _make_cache(**kwargs) -> CacheStore:
    return CacheStore(
        mfg_name=kwargs.get("mfg_name", {"D01": "ダイヘン"}),
        mfg_days=kwargs.get("mfg_days", {"D01": 2}),
        cust_days=kwargs.get("cust_days", {}),
        cust_retention=kwargs.get("cust_retention", {}),
        cust_route=kwargs.get("cust_route", {}),
        confirm=kwargs.get("confirm", {}),
        storage=kwargs.get("storage", {}),
    )


def _make_row(**kwargs) -> OrderRow:
    defaults = {
        "order_number": "1000001",
        "detail_number": "10",
        "document_type": "【受注】在庫販売",
        "customer_name": "テスト商事",
        "product_name": "溶接棒 3.2mm",
        "ship_status": "未処理",
        "quantity": "100",
        "unit_price": "500",
        "net_amount": "50000",
        "manufacturer_name": "ダイヘン",
        "storage_place": "関東商品センター",
        "customer_order_number": "PO-001",
        "customer_contact": "田中",
        "comment_detail": "",
        "comment_external": "",
        "comment_internal": "",
        "rejection_reason": "",
        "ship_to_name": "テスト商事",
        "registration_date": TODAY,
        "time_value": "10:00:00",
        "order_delivery_date": datetime.date(2026, 2, 20),
        "specified_delivery_date": None,
        "item_group_code": "D01",
    }
    defaults.update(kwargs)
    return OrderRow(**defaults)


# ============================================
# utils: 不正入力の耐性
# ============================================
class TestUtilsErrorHandling:
    """utils関数の不正入力テスト"""

    def test_parse_date_none(self):
        assert parse_date(None) is None

    def test_parse_date_empty_string(self):
        assert parse_date("") is None

    def test_parse_date_invalid_format(self):
        assert parse_date("abc") is None

    def test_parse_date_invalid_date(self):
        assert parse_date("2026/13/32") is None

    def test_parse_date_int_value(self):
        """数値が渡された場合"""
        result = parse_date(12345)
        # Noneか有効な日付のどちらか
        assert result is None or isinstance(result, datetime.date)

    def test_parse_time_none(self):
        assert parse_time(None) is None

    def test_parse_time_empty(self):
        assert parse_time("") is None

    def test_parse_time_invalid(self):
        assert parse_time("abc") is None

    def test_convert_to_half_width_empty(self):
        assert convert_to_half_width("") == ""

    def test_convert_to_half_width_none_like(self):
        """通常の文字列のみ渡される前提"""
        result = convert_to_half_width("abc123")
        assert result == "abc123"

    def test_is_december_31_none(self):
        assert is_december_31(None) is False

    def test_format_date_japanese_basic(self):
        result = format_date_japanese(datetime.date(2026, 2, 16))
        assert "2月16日" in result

    def test_extract_date_from_string_none(self):
        assert extract_date_from_string(None) is None

    def test_extract_date_from_string_empty(self):
        assert extract_date_from_string("") is None

    def test_extract_date_from_string_no_date(self):
        assert extract_date_from_string("日付なし") is None


# ============================================
# delivery_calc: 不正入力の耐性
# ============================================
class TestDeliveryCalcErrorHandling:
    """納期計算の不正入力テスト"""

    def test_no_dates_at_all(self):
        """受注納期も指定納期もNone"""
        row = _make_row(
            order_delivery_date=None,
            specified_delivery_date=None,
        )
        cache = _make_cache()
        result = calculate_delivery_date(
            row, cache, {}, BRANCH, EXEC_TIME, TODAY
        )
        assert result == "日程調整中"

    def test_empty_time_value(self):
        """時刻が空文字"""
        row = _make_row(
            time_value="",
            ship_status="処理完了",
        )
        cache = _make_cache()
        result = calculate_delivery_date(
            row, cache, {}, BRANCH, EXEC_TIME, TODAY
        )
        # 時刻パースできないが、処理完了パスに入る前に指定納期チェックで先にresultが出る可能性
        assert isinstance(result, str)

    def test_none_holidays(self):
        """holidays=None"""
        row = _make_row()
        cache = _make_cache()
        result = calculate_delivery_date(
            row, cache, None, BRANCH, EXEC_TIME, TODAY
        )
        assert isinstance(result, str)

    def test_none_branch(self):
        """branch=None → デフォルト値で動作"""
        row = _make_row()
        cache = _make_cache()
        result = calculate_delivery_date(
            row, cache, {}, None, EXEC_TIME, TODAY
        )
        assert isinstance(result, str)

    def test_none_execution_time(self):
        """execution_time=None → 現在時刻で動作"""
        row = _make_row()
        cache = _make_cache()
        result = calculate_delivery_date(
            row, cache, {}, BRANCH, None, TODAY
        )
        assert isinstance(result, str)

    def test_extract_pickup_date_empty(self):
        assert extract_pickup_date("", TODAY) is None

    def test_extract_pickup_date_no_keyword(self):
        assert extract_pickup_date("テスト 2/20", TODAY) is None

    def test_extract_arrival_date_empty(self):
        assert extract_arrival_date_from_internal("", TODAY) is None

    def test_extract_arrival_date_no_at(self):
        assert extract_arrival_date_from_internal("テスト 2/20", TODAY) is None


# ============================================
# bunno: 不正入力の耐性
# ============================================
class TestBunnoErrorHandling:
    """分納処理の不正入力テスト"""

    def test_extract_bunno_empty(self):
        assert extract_bunno_info("") == []

    def test_extract_bunno_no_keyword(self):
        assert extract_bunno_info("テストコメント") == []

    def test_extract_bunno_keyword_only(self):
        """「分納:」のみでデータなし"""
        result = extract_bunno_info("分納:")
        assert result == []

    def test_extract_bunno_invalid_quantity(self):
        """数量が数値でない"""
        result = extract_bunno_info("分納:abc 2/20")
        assert result == []

    def test_normalize_bunno_date_empty(self):
        assert normalize_bunno_date("") == ""

    def test_normalize_bunno_date_stockout(self):
        """欠品中 → 未定"""
        assert normalize_bunno_date("欠品中") == "未定"

    def test_split_qty_and_date_no_split(self):
        assert split_qty_and_date("テスト") is None

    def test_split_qty_and_date_no_after_unit(self):
        """単位の後に何もない"""
        result = split_qty_and_date("10個")
        assert result is None

    def test_extract_location_empty(self):
        token, loc = extract_location_from_token("")
        assert token == ""
        assert loc == ""

    def test_extract_location_no_brackets(self):
        token, loc = extract_location_from_token("テスト")
        assert token == "テスト"
        assert loc == ""

    def test_has_bunno_mitei_empty(self):
        assert has_bunno_mitei([]) is False

    def test_has_bunno_mitei_no_mitei(self):
        entries = [BunnoEntry("50個", "2/20", "")]
        assert has_bunno_mitei(entries) is False

    def test_calculate_bunno_date_invalid_date(self):
        """無効な日付文字列"""
        result = calculate_bunno_date(
            "abc", False, 0, {}, None, "", "", False, TODAY
        )
        assert result == "確認中"

    def test_calculate_bunno_date_invalid_month(self):
        """月が13"""
        result = calculate_bunno_date(
            "13/20", False, 0, {}, None, "", "", False, TODAY
        )
        assert result == "確認中"

    def test_calculate_bunno_date_no_slash(self):
        """スラッシュなし"""
        result = calculate_bunno_date(
            "20260220", False, 0, {}, None, "", "", False, TODAY
        )
        assert result == "確認中"


# ============================================
# tracking: 不正入力の耐性
# ============================================
class TestTrackingErrorHandling:
    """送り状処理の不正入力テスト"""

    def test_extract_tracking_empty(self):
        assert extract_tracking_info("") == []

    def test_extract_tracking_no_match(self):
        result = extract_tracking_info("テストコメント")
        assert result == []

    def test_extract_tracking_short_number(self):
        """番号が短すぎる場合"""
        result = extract_tracking_info("佐川 123")
        assert len(result) == 0


# ============================================
# stockout: 不正入力の耐性
# ============================================
class TestStockoutErrorHandling:
    """欠品処理の不正入力テスト"""

    def test_extract_approx_empty(self):
        result = extract_approx_delivery("")
        assert result == ""

    def test_remove_stockout_empty(self):
        assert remove_stockout_text("") == ""

    def test_remove_stockout_no_keyword(self):
        assert remove_stockout_text("テスト") == "テスト"


# ============================================
# confirming: 不正入力の耐性
# ============================================
class TestConfirmingErrorHandling:
    """確認中一覧参照の不正入力テスト"""

    def test_get_confirmed_date_empty_cache(self):
        cache = CacheStore()
        result = get_confirmed_delivery_date("100", "10", cache)
        assert result is None

    def test_get_confirming_status_empty_cache(self):
        cache = CacheStore()
        result = get_confirming_status("100", "10", cache)
        assert result == ""

    def test_get_confirmed_date_missing_key(self):
        cache = _make_cache(confirm={"999|10": ("未", "分納", None)})
        result = get_confirmed_delivery_date("100", "10", cache)
        assert result is None


# ============================================
# manufacturer: 不正入力の耐性
# ============================================
class TestManufacturerErrorHandling:
    """メーカー一覧参照の不正入力テスト"""

    def test_get_name_empty_code(self):
        cache = _make_cache()
        result = get_manufacturer_name("", cache)
        assert result == ""

    def test_get_name_missing_code(self):
        cache = _make_cache()
        result = get_manufacturer_name("NOTEXIST", cache)
        assert result == ""

    def test_get_days_empty_code(self):
        cache = _make_cache()
        result = get_delivery_days_to_add("", cache)
        assert result == 2  # デフォルト値

    def test_get_days_missing_code(self):
        cache = _make_cache()
        result = get_delivery_days_to_add("NOTEXIST", cache)
        assert result == 2  # デフォルト値


# ============================================
# customer: 不正入力の耐性
# ============================================
class TestCustomerErrorHandling:
    """顧客マスター参照の不正入力テスト"""

    def test_delivery_days_empty_name(self):
        cache = _make_cache()
        result = get_customer_delivery_days("", cache)
        assert result == []

    def test_delivery_days_missing_name(self):
        cache = _make_cache()
        result = get_customer_delivery_days("存在しない", cache)
        assert result == []

    def test_is_route_delivery_empty(self):
        cache = _make_cache()
        result = is_route_delivery("", cache)
        assert result is False

    def test_get_email_none_ws(self):
        result = get_email_addresses("テスト", None)
        assert result == ""

    def test_get_email_empty_sheet(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["顧客名", "データ", "データ", "データ", "メール"])
        result = get_email_addresses("テスト", ws)
        assert result == ""


# ============================================
# business_days: 不正入力の耐性
# ============================================
class TestBusinessDaysErrorHandling:
    """営業日計算の不正入力テスト"""

    def test_add_zero_days(self):
        result = add_business_days(TODAY, 0, {})
        assert result == TODAY

    def test_add_negative_days(self):
        """負の日数"""
        result = add_business_days(TODAY, -1, {})
        # 0以下のときはそのまま返す実装もある
        assert isinstance(result, datetime.date)

    def test_none_holidays(self):
        result = add_business_days(TODAY, 1, None)
        assert isinstance(result, datetime.date)

    def test_is_holiday_none_holidays(self):
        result = is_holiday(TODAY, None)
        assert isinstance(result, bool)

    def test_get_next_delivery_day_empty_list(self):
        """曜日リストが空"""
        result = get_next_delivery_day(TODAY, [], {})
        assert isinstance(result, datetime.date)


# ============================================
# data_loader: 不正入力の耐性
# ============================================
class TestDataLoaderErrorHandling:
    """データ読込の不正入力テスト"""

    def test_get_column_positions_empty(self):
        """空のデータ"""
        result = get_column_positions([])
        assert result is None

    def test_get_column_positions_short_rows(self):
        """5行未満"""
        data = [["a"], ["b"], ["c"]]
        result = get_column_positions(data)
        assert result is None

    def test_get_column_positions_missing_required(self):
        """必須列が足りない"""
        data = [[], [], [], [], ["受発注伝票", "明細"]]
        result = get_column_positions(data)
        assert result is None

    def test_parse_order_row_out_of_range(self):
        """範囲外の行"""
        data = [["a", "b"]]
        cols = {"受発注伝票": 0}
        result = parse_order_row(data, 99, cols)
        assert isinstance(result, OrderRow)
        assert result.order_number == ""

    def test_load_source_file_nonexistent(self, tmp_path):
        """存在しないファイル"""
        with pytest.raises(FileNotFoundError):
            load_source_file(tmp_path / "nonexistent.xls")


# ============================================
# config: 不正入力の耐性
# ============================================
class TestConfigErrorHandling:
    """設定読込の不正入力テスト"""

    def test_load_holidays_no_calendar_sheet(self):
        """特別日カレンダーシートがない"""
        wb = Workbook()
        result = load_holidays(wb)
        assert result == {}

    def test_load_branch_settings_no_sheet(self):
        """営業所設定シートがない"""
        wb = Workbook()
        result = load_branch_settings(wb, [], {})
        assert isinstance(result, BranchSettings)
        assert result.name == ""

    def test_get_branch_settings_none_holidays(self):
        """holidays=None"""
        branch = BranchSettings(default_cutoff=15)
        name, cutoff, sig = get_branch_settings(branch, None, TODAY)
        assert cutoff == 15

    def test_get_branch_settings_special_cutoff(self):
        """特別締切時間"""
        branch = BranchSettings(default_cutoff=15)
        holidays: HolidayMap = {TODAY: 12}
        name, cutoff, sig = get_branch_settings(branch, holidays, TODAY)
        assert cutoff == 12


# ============================================
# history: 不正入力の耐性
# ============================================
class TestHistoryErrorHandling:
    """送付履歴管理の不正入力テスト"""

    def test_save_empty_orders(self, tmp_path):
        """空のリストを保存"""
        path = str(tmp_path / "test.xlsx")
        wb = initialize_delivery_history(path)
        ws = wb["送付履歴"]
        save_delivery_history(ws, [])
        # エラーなし、何も書き込まれない
        assert ws.cell(row=2, column=1).value is None

    def test_save_confirming_empty(self, tmp_path):
        """空の確認中リストを保存"""
        path = str(tmp_path / "test.xlsx")
        wb = initialize_delivery_history(path)
        ws = wb["確認中一覧"]
        save_confirming_list(ws, [])
        assert ws.cell(row=2, column=1).value is None

    def test_load_empty_history(self, tmp_path):
        """空の送付履歴を読み込み"""
        path = str(tmp_path / "test.xlsx")
        wb = initialize_delivery_history(path)
        cache = _make_cache()
        result = load_delivery_history(
            wb["送付履歴"], wb["確認中一覧"], cache, {}, TODAY
        )
        assert result == {}

    def test_clean_old_history_empty(self, tmp_path):
        """空の送付履歴をクリーンアップ"""
        path = str(tmp_path / "test.xlsx")
        wb = initialize_delivery_history(path)
        deleted = clean_old_history(wb["送付履歴"], 365, TODAY)
        assert deleted == 0

    def test_clean_old_confirming_empty(self, tmp_path):
        """空の確認中一覧をクリーンアップ"""
        path = str(tmp_path / "test.xlsx")
        wb = initialize_delivery_history(path)
        deleted = clean_old_confirming_list(wb["確認中一覧"], 365, TODAY)
        assert deleted == 0

    def test_duplicate_order_in_history(self, tmp_path):
        """重複注番の保存（2回に分けて保存）"""
        path = str(tmp_path / "test.xlsx")
        wb = initialize_delivery_history(path)
        ws = wb["送付履歴"]

        # 1回目の保存
        orders1 = [
            HistoryRecord(
                order_date=TODAY,
                customer_name="テスト商事",
                order_number="100",
                detail_number="10",
                delivery_answer="2月20日配達予定",
            ),
        ]
        save_delivery_history(ws, orders1, EXEC_TIME)
        assert ws.cell(row=2, column=4).value == "100"
        assert ws.cell(row=2, column=8).value == "2月20日配達予定"

        # 2回目の保存（同じ注番で納品済み → 更新される）
        orders2 = [
            HistoryRecord(
                order_date=TODAY,
                customer_name="テスト商事",
                order_number="100",
                detail_number="10",
                delivery_answer="納品済み",
            ),
        ]
        save_delivery_history(ws, orders2, EXEC_TIME)
        assert ws.cell(row=2, column=8).value == "納品済み"


# ============================================
# email_builder: 不正入力の耐性
# ============================================
class TestEmailBuilderErrorHandling:
    """メール生成の不正入力テスト"""

    def test_subject_empty_name(self):
        result = build_email_subject("", "", TODAY)
        assert "02/16" in result

    def test_body_empty_name(self):
        result = build_email_body_html("", BRANCH, today=TODAY)
        assert "html" in result.lower()

    def test_body_no_optional_info(self):
        """オプション情報なし"""
        result = build_email_body_html(
            "テスト商事", BRANCH,
            stockout_info_list=None,
            tracking_info_list=None,
            bunno_info_list=None,
            bunno_completed_list=None,
            today=TODAY,
        )
        assert "テスト商事" in result
        assert "欠品" not in result  # 欠品セクションなし

    def test_body_empty_lists(self):
        """空リスト"""
        result = build_email_body_html(
            "テスト商事", BRANCH,
            stockout_info_list=[],
            tracking_info_list=[],
            bunno_info_list=[],
            bunno_completed_list=[],
            today=TODAY,
        )
        assert "テスト商事" in result


# ============================================
# report_generator: 不正入力の耐性
# ============================================
class TestReportGeneratorErrorHandling:
    """回答書生成の不正入力テスト"""

    def test_build_report_row_empty_row(self):
        """ほぼ空のOrderRow"""
        row = OrderRow()
        cache = CacheStore()
        report, status = build_report_row(
            row, cache, {}, BRANCH, EXEC_TIME, False, TODAY
        )
        assert isinstance(report, ReportRow)
        assert isinstance(status, str)

    def test_build_report_row_all_none_dates(self):
        """日付が全てNone"""
        row = _make_row(
            registration_date=None,
            order_delivery_date=None,
            specified_delivery_date=None,
        )
        cache = _make_cache()
        report, status = build_report_row(
            row, cache, {}, BRANCH, EXEC_TIME, False, TODAY
        )
        assert isinstance(report, ReportRow)
        assert status == "日程調整中"

    def test_create_report_none_output_dir(self, tmp_path):
        """出力ディレクトリが存在する場合の正常動作"""
        data = [_make_row()]
        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト商事", {},
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is not None

    def test_special_chars_in_customer_name(self, tmp_path):
        """顧客名に特殊文字"""
        data = [
            _make_row(
                customer_name="テスト＆商事（東京）",
            ),
        ]
        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト＆商事（東京）", {},
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is not None

    def test_very_long_product_name(self, tmp_path):
        """非常に長い品名"""
        long_name = "A" * 500
        data = [_make_row(product_name=long_name)]
        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト商事", {},
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is not None

    def test_unicode_in_comments(self):
        """コメントにUnicode特殊文字"""
        row = _make_row(
            comment_detail="①②③ 特殊文字テスト★",
            comment_external="テスト→確認",
            comment_internal="テスト♪",
        )
        cache = _make_cache()
        report, status = build_report_row(
            row, cache, {}, BRANCH, EXEC_TIME, False, TODAY
        )
        assert isinstance(report, ReportRow)
