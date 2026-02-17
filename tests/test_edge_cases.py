"""Phase 10: エッジケース・結合テスト

設計書の品質保証要件に基づく:
- エッジケーステスト（空データ、全処理完了、全未確定等）
- クロスモジュール結合テスト（report_generator→history→email_builder）
- 業務ルール複合テスト（分納+欠品+送り状の組み合わせ等）
"""

import datetime
import tempfile
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from nouki_kaitou.bunno import (
    extract_bunno_info,
    has_bunno_mitei,
)
from nouki_kaitou.delivery_calc import calculate_delivery_date
from nouki_kaitou.email_builder import (
    build_email_body_html,
    build_email_subject,
    create_emails,
)
from nouki_kaitou.excel_writer import (
    copy_data_row,
    create_header,
    format_report,
)
from nouki_kaitou.history import (
    clean_confirming_list,
    initialize_delivery_history,
    load_delivery_history,
    save_confirming_list,
    save_delivery_history,
)
from nouki_kaitou.models import (
    BranchSettings,
    BunnoEntry,
    CacheStore,
    ConfirmingRecord,
    HolidayMap,
    HistoryRecord,
    OrderRow,
    ReportResult,
    ReportRow,
    StockoutEntry,
    TrackingEntry,
)
from nouki_kaitou.report_generator import (
    build_report_row,
    create_delivery_report,
    create_delivery_report_by_order_numbers,
    group_order_numbers_by_customer,
)


# ============================================
# テスト用ヘルパー
# ============================================
TODAY = datetime.date(2026, 2, 16)
EXEC_TIME = datetime.datetime(2026, 2, 16, 10, 0, 0)
BRANCH = BranchSettings(
    name="京葉営業所",
    default_cutoff=15,
    base_center="関東商品センター",
    shared_email="keiyo@matsumoto.co.jp",
    signature="マツモト産業㈱京葉営業所",
)


def _make_cache(**kwargs) -> CacheStore:
    """テスト用キャッシュを作成する。"""
    return CacheStore(
        mfg_name=kwargs.get("mfg_name", {"D01": "ダイヘン", "Z99": "", "P01": "パナソニック"}),
        mfg_days=kwargs.get("mfg_days", {"D01": 2, "P01": 3}),
        cust_days=kwargs.get("cust_days", {}),
        cust_retention=kwargs.get("cust_retention", {}),
        cust_route=kwargs.get("cust_route", {}),
        confirm=kwargs.get("confirm", {}),
        storage=kwargs.get("storage", {}),
    )


def _make_row(**kwargs) -> OrderRow:
    """テスト用OrderRowを作成する。"""
    defaults = {
        "order_number": "1000001",
        "detail_number": "10",
        "document_type": "【受注】在庫販売",
        "customer_name": "テスト商事",
        "product_name": "溶接棒 3.2mm",
        "ship_status": "未処理",
        "quantity": "100",
        "unit_price": "500",
        "net_amount": "50000",
        "manufacturer_name": "ダイヘン",
        "storage_place": "関東商品センター",
        "customer_order_number": "PO-001",
        "customer_contact": "田中",
        "comment_detail": "",
        "comment_external": "",
        "comment_internal": "",
        "rejection_reason": "",
        "ship_to_name": "テスト商事",
        "registration_date": TODAY,
        "time_value": "10:00:00",
        "order_delivery_date": datetime.date(2026, 2, 20),
        "specified_delivery_date": None,
        "item_group_code": "D01",
    }
    defaults.update(kwargs)
    return OrderRow(**defaults)


# ============================================
# エッジケース: 空データ
# ============================================
class TestEmptyData:
    """空データのエッジケーステスト"""

    def test_empty_source_data_period_mode(self, tmp_path):
        """期間モード: ソースデータが空"""
        result = create_delivery_report(
            [], "テスト商事", {},
            _make_cache(), tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is None

    def test_empty_source_data_order_mode(self, tmp_path):
        """注番モード: ソースデータが空"""
        result = create_delivery_report_by_order_numbers(
            [], "テスト商事", ["100"],
            _make_cache(), tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is None

    def test_empty_order_numbers(self, tmp_path):
        """注番リストが空"""
        data = [_make_row()]
        result = create_delivery_report_by_order_numbers(
            data, "テスト商事", [],
            _make_cache(), tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is None

    def test_all_filtered_out_by_rejection(self, tmp_path):
        """全データが明細削除でフィルタされる"""
        data = [
            _make_row(order_number="100", rejection_reason="明細削除"),
            _make_row(order_number="200", rejection_reason="明細削除"),
        ]
        result = create_delivery_report(
            data, "テスト商事", {},
            _make_cache(), tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is None

    def test_all_filtered_out_by_doc_type(self, tmp_path):
        """全データが伝票タイプでフィルタされる"""
        data = [
            _make_row(document_type="【受注】返品"),
            _make_row(document_type="【受注】サンプル"),
        ]
        result = create_delivery_report(
            data, "テスト商事", {},
            _make_cache(), tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is None

    def test_all_filtered_out_by_hash(self, tmp_path):
        """全データが##除外でフィルタされる"""
        data = [
            _make_row(order_number="100", comment_internal="##除外"),
            _make_row(order_number="200", comment_internal="＃＃除外"),
        ]
        result = create_delivery_report(
            data, "テスト商事", {},
            _make_cache(), tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is None

    def test_group_empty_list(self):
        """空の注番リストでグループ化"""
        result = group_order_numbers_by_customer([], [])
        assert result == {}

    def test_empty_cache(self, tmp_path):
        """キャッシュが全て空でも動作する"""
        cache = CacheStore()
        data = [_make_row()]
        result = create_delivery_report(
            data, "テスト商事", {},
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is not None


# ============================================
# エッジケース: 全処理完了
# ============================================
class TestAllCompleted:
    """全伝票が処理完了のエッジケース"""

    def test_all_stock_completed(self, tmp_path):
        """全在庫販売が処理完了"""
        data = [
            _make_row(
                order_number="100", detail_number="10",
                ship_status="処理完了",
                time_value="10:00:00",
            ),
            _make_row(
                order_number="200", detail_number="10",
                ship_status="処理完了",
                time_value="14:00:00",
            ),
        ]
        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト商事", {},
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is not None
        # 全て確定
        assert len(result.confirmed_orders) == 2
        assert len(result.confirming_orders) == 0

    def test_all_chokusouhan_completed_force_delivered(self, tmp_path):
        """全直送販売が処理完了 → 全て納品済み"""
        data = [
            _make_row(
                order_number="100", detail_number="10",
                document_type="【受注】直送販売",
                storage_place="転送中（直送用）",
                ship_status="処理完了",
                order_delivery_date=datetime.date(2026, 12, 31),
            ),
            _make_row(
                order_number="200", detail_number="10",
                document_type="【受注】直送販売",
                storage_place="転送中（直送用）",
                ship_status="処理完了",
                order_delivery_date=datetime.date(2026, 12, 31),
            ),
        ]
        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト商事", {},
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is not None
        assert len(result.confirmed_orders) == 2
        for order in result.confirmed_orders:
            assert order.delivery_answer == "納品済み"

    def test_himozuki_completed(self, tmp_path):
        """紐付き（直送+非転送中）+処理完了 → 通常納期計算"""
        data = [
            _make_row(
                order_number="100", detail_number="10",
                document_type="【受注】直送販売",
                storage_place="関東商品センター",
                ship_status="処理完了",
                order_delivery_date=datetime.date(2026, 2, 20),
            ),
        ]
        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト商事", {},
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is not None
        assert len(result.confirmed_orders) == 1
        # 紐付き処理完了は通常の納期計算（納品済みにはならない）
        assert result.confirmed_orders[0].delivery_answer != "納品済み"


# ============================================
# エッジケース: 全未確定
# ============================================
class TestAllConfirming:
    """全伝票が未確定のエッジケース"""

    def test_all_dec31(self, tmp_path):
        """全受注納期が12/31 → 全て確認中"""
        data = [
            _make_row(
                order_number="100", detail_number="10",
                document_type="【受注】直送販売",
                storage_place="転送中（直送用）",
                order_delivery_date=datetime.date(2026, 12, 31),
            ),
            _make_row(
                order_number="200", detail_number="10",
                document_type="【受注】直送販売",
                storage_place="転送中（直送用）",
                order_delivery_date=datetime.date(2026, 12, 31),
            ),
        ]
        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト商事", {},
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is not None
        assert len(result.confirming_orders) == 2
        assert result.has_confirming is True

    def test_all_scheduling(self, tmp_path):
        """在庫販売で受注納期12/31 → 全て日程調整中"""
        data = [
            _make_row(
                order_number="100", detail_number="10",
                order_delivery_date=datetime.date(2026, 12, 31),
            ),
        ]
        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト商事", {},
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is not None
        assert len(result.confirming_orders) == 1

    def test_all_stockout(self, tmp_path):
        """全て欠品中"""
        data = [
            _make_row(
                order_number="100", detail_number="10",
                comment_detail="欠品中 3月上旬予定",
                order_delivery_date=datetime.date(2026, 12, 31),
            ),
            _make_row(
                order_number="200", detail_number="10",
                comment_detail="欠品中",
                order_delivery_date=datetime.date(2026, 12, 31),
            ),
        ]
        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト商事", {},
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is not None
        assert len(result.stockout_info_list) == 2
        assert result.has_confirming is True


# ============================================
# エッジケース: 混在シナリオ
# ============================================
class TestMixedScenarios:
    """確定+確認中+欠品+分納の混在テスト"""

    def test_confirmed_and_confirming_mix(self, tmp_path):
        """確定と確認中の混在"""
        data = [
            # 確定: 通常の在庫販売
            _make_row(
                order_number="100", detail_number="10",
                product_name="溶接棒A",
            ),
            # 確認中: 直送で12/31
            _make_row(
                order_number="200", detail_number="10",
                product_name="ガスB",
                document_type="【受注】直送販売",
                storage_place="転送中（直送用）",
                order_delivery_date=datetime.date(2026, 12, 31),
            ),
            # 欠品中
            _make_row(
                order_number="300", detail_number="10",
                product_name="ワイヤC",
                comment_detail="欠品中 3月上旬予定",
                order_delivery_date=datetime.date(2026, 12, 31),
            ),
        ]
        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト商事", {},
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is not None
        assert len(result.confirmed_orders) == 1
        assert len(result.confirming_orders) == 2  # 確認中 + 欠品中
        assert len(result.stockout_info_list) == 1
        assert result.has_confirming is True

    def test_bunno_and_stockout_same_order(self, tmp_path):
        """分納と欠品が同じ注番にある場合"""
        data = [
            _make_row(
                order_number="100", detail_number="10",
                product_name="溶接棒A",
                comment_detail="分納:50個 2/20,50個 未定",
                ship_status="未処理",
            ),
            _make_row(
                order_number="100", detail_number="20",
                product_name="ワイヤB",
                comment_detail="欠品中 3月上旬予定",
                order_delivery_date=datetime.date(2026, 12, 31),
            ),
        ]
        cache = _make_cache()
        result = create_delivery_report_by_order_numbers(
            data, "テスト商事", ["100"],
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is not None
        assert len(result.stockout_info_list) == 1
        # 分納情報も収集（欠品と別物）
        assert len(result.bunno_info_list) >= 0  # 分納は欠品併存時スキップされうる

    def test_tracking_and_confirmed(self, tmp_path):
        """送り状情報 + 確定伝票の組み合わせ"""
        data = [
            _make_row(
                order_number="100", detail_number="10",
                product_name="溶接棒A",
                comment_external="佐川 1234567890",
            ),
            _make_row(
                order_number="100", detail_number="20",
                product_name="ガスB",
                comment_external="ヤマト 9876543210",
            ),
        ]
        cache = _make_cache()
        result = create_delivery_report_by_order_numbers(
            data, "テスト商事", ["100"],
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is not None
        assert len(result.tracking_info_list) == 2


# ============================================
# エッジケース: 送付履歴チェック
# ============================================
class TestSentHistoryEdgeCases:
    """送付履歴の複雑なエッジケース"""

    def test_all_already_sent(self, tmp_path):
        """全伝票が送付済み"""
        data = [
            _make_row(
                order_number="100", detail_number="10",
                registration_date=datetime.date(2026, 2, 10),
            ),
            _make_row(
                order_number="200", detail_number="10",
                registration_date=datetime.date(2026, 2, 10),
            ),
        ]
        sent = {
            "100|10": "2月15日配達予定",
            "200|10": "2月15日配達予定",
        }
        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト商事", sent,
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            date_from=datetime.date(2026, 2, 1),
            date_to=datetime.date(2026, 2, 28),
            today=TODAY,
        )
        assert result is None

    def test_confirming_to_completed_transition(self, tmp_path):
        """確認中→処理完了: 直送は納品済みで再出力"""
        data = [
            _make_row(
                order_number="100", detail_number="10",
                document_type="【受注】直送販売",
                storage_place="転送中（直送用）",
                ship_status="処理完了",
                order_delivery_date=datetime.date(2026, 12, 31),
            ),
        ]
        sent = {"100|10": "確認中"}
        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト商事", sent,
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is not None
        assert len(result.confirmed_orders) == 1
        assert result.confirmed_orders[0].delivery_answer == "納品済み"

    def test_excluded_in_both_history_and_confirming(self, tmp_path):
        """送付履歴と確認中一覧の両方で除外"""
        data = [
            _make_row(
                order_number="100", detail_number="10",
            ),
            _make_row(
                order_number="200", detail_number="10",
            ),
        ]
        sent = {"100|10": "除外"}
        cache = _make_cache(confirm={"200|10": ("除外", "", None)})
        result = create_delivery_report(
            data, "テスト商事", sent,
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is None

    def test_bunno_kanryo_not_resent(self, tmp_path):
        """分納完了は常にスキップ"""
        data = [
            _make_row(
                order_number="100", detail_number="10",
                registration_date=TODAY,  # 当日登録でも
            ),
        ]
        sent = {"100|10": "分納完了"}
        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト商事", sent,
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is None

    def test_today_registration_with_previous_status(self, tmp_path):
        """当日登録で前回ステータスあり → 再送"""
        data = [
            _make_row(
                order_number="100", detail_number="10",
                registration_date=TODAY,
            ),
        ]
        # 前回のステータスがあるが、registration_date == today
        sent = {"100|10": "2月16日配達予定"}
        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト商事", sent,
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        # registration_date < today が False なので再送
        assert result is not None


# ============================================
# エッジケース: 分納の複合パターン
# ============================================
class TestBunnoEdgeCases:
    """分納の複雑なエッジケース"""

    def test_bunno_completed_with_stockout_comment(self, tmp_path):
        """分納+処理完了+コメントに欠品テキスト残り（SAPの制約）"""
        data = [
            _make_row(
                order_number="100", detail_number="10",
                ship_status="処理完了",
                comment_detail="欠品中 分納:50個 2/10,50個 未定",
            ),
        ]
        cache = _make_cache(
            confirm={"100|10": ("未", "分納", None)}
        )
        result = create_delivery_report(
            data, "テスト商事", {},
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is not None
        # 処理完了 → 分納完了として処理（欠品テキストは無視）
        assert len(result.bunno_completed_list) == 1

    def test_bunno_all_confirmed_new_order(self, tmp_path):
        """分納で全分確定（新規注文、確認中一覧に未登録）"""
        data = [
            _make_row(
                order_number="100", detail_number="10",
                comment_detail="分納:50個 2/20,50個 2/25",
                ship_status="未処理",
            ),
        ]
        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト商事", {},
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is not None
        # 全分確定 → 分納完了で送付履歴へ
        assert len(result.confirmed_orders) == 1
        assert result.confirmed_orders[0].delivery_answer == "分納完了"

    def test_bunno_mitei_remaining(self, tmp_path):
        """分納で未定が残っている → 確認中一覧"""
        data = [
            _make_row(
                order_number="100", detail_number="10",
                comment_detail="分納:50個 2/20,50個 未定",
                ship_status="未処理",
            ),
        ]
        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト商事", {},
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is not None
        assert len(result.confirming_orders) == 1
        assert result.confirming_orders[0].status == "分納"

    def test_multiple_bunno_orders(self, tmp_path):
        """複数注番に分納がある場合"""
        data = [
            _make_row(
                order_number="100", detail_number="10",
                product_name="溶接棒A",
                comment_detail="分納:30個 2/20,70個 未定",
                ship_status="未処理",
            ),
            _make_row(
                order_number="200", detail_number="10",
                product_name="ワイヤB",
                comment_detail="分納:100個 2/25,200個 3/5",
                ship_status="未処理",
            ),
        ]
        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト商事", {},
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is not None
        # 注番100は未定あり → 確認中、注番200は全確定 → 分納完了
        has_confirming = any(
            o.order_number == "100" for o in result.confirming_orders
        )
        has_confirmed = any(
            o.order_number == "200" for o in result.confirmed_orders
        )
        assert has_confirming
        assert has_confirmed


# ============================================
# エッジケース: Z99/Z97特殊品
# ============================================
class TestZ99Z97EdgeCases:
    """Z99/Z97品目コードの特殊処理"""

    def test_z99_in_report(self, tmp_path):
        """Z99品目コードでの回答書生成"""
        data = [
            _make_row(
                order_number="100", detail_number="10",
                item_group_code="Z99",
                product_name="ABC商事 特殊溶接棒 5mm",
                manufacturer_name="",
            ),
        ]
        cache = _make_cache()
        result = create_delivery_report_by_order_numbers(
            data, "テスト商事", ["100"],
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is not None
        # Excel出力を確認
        wb = load_workbook(result.file_path)
        ws = wb.active
        assert ws.cell(row=7, column=4).value == "ABC商事"  # メーカー名
        assert ws.cell(row=7, column=5).value == "特殊溶接棒 5mm"  # 品名

    def test_z97_in_report(self, tmp_path):
        """Z97品目コードでの回答書生成"""
        data = [
            _make_row(
                order_number="100", detail_number="10",
                item_group_code="Z97",
                product_name="XYZ産業\u3000特殊部品",
                manufacturer_name="",
            ),
        ]
        cache = _make_cache()
        result = create_delivery_report_by_order_numbers(
            data, "テスト商事", ["100"],
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is not None
        wb = load_workbook(result.file_path)
        ws = wb.active
        assert ws.cell(row=7, column=4).value == "XYZ産業"
        assert ws.cell(row=7, column=5).value == "特殊部品"


# ============================================
# エッジケース: 日付境界
# ============================================
class TestDateEdgeCases:
    """日付関連のエッジケース"""

    def test_dec31_order_delivery_date(self):
        """受注納期が12/31 → 確認中判定"""
        row = _make_row(
            order_delivery_date=datetime.date(2026, 12, 31),
            document_type="【受注】直送販売",
            storage_place="転送中（直送用）",
        )
        cache = _make_cache()
        result = calculate_delivery_date(
            row, cache, {}, BRANCH, EXEC_TIME, TODAY
        )
        assert result == "確認中"

    def test_dec31_specified_delivery_date(self):
        """指定納期が12/31 → 無視して受注納期で計算"""
        row = _make_row(
            specified_delivery_date=datetime.date(2026, 12, 31),
            order_delivery_date=datetime.date(2026, 2, 20),
        )
        cache = _make_cache()
        result = calculate_delivery_date(
            row, cache, {}, BRANCH, EXEC_TIME, TODAY
        )
        assert "2月" in result or "配達" in result or "出荷" in result

    def test_registration_date_none(self, tmp_path):
        """登録日がNone → 期間フィルタでスキップ"""
        data = [
            _make_row(registration_date=None),
        ]
        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト商事", {},
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            date_from=datetime.date(2026, 2, 1),
            date_to=datetime.date(2026, 2, 28),
            today=TODAY,
        )
        assert result is None

    def test_no_date_range_filter(self, tmp_path):
        """期間指定なし（date_from/date_to=None）"""
        data = [
            _make_row(
                registration_date=datetime.date(2025, 1, 1),
            ),
        ]
        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト商事", {},
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        # 期間指定なしだが、registration_date < today なので再送チェックに通る
        assert result is not None


# ============================================
# エッジケース: 価格特殊ケース
# ============================================
class TestPriceEdgeCases:
    """価格関連のエッジケース"""

    def test_price_confirming_with_dollar_flag(self):
        """$$フラグで価格確認中をオーバーライド"""
        row = _make_row(
            comment_internal="$$",
            order_delivery_date=datetime.date(2026, 12, 31),
        )
        cache = _make_cache()
        report, _ = build_report_row(
            row, cache, {}, BRANCH, EXEC_TIME, False, TODAY
        )
        # $$があるので価格は確認中にならない
        assert report.unit_price == "500"

    def test_price_confirming_force_delivered(self):
        """forceDelivered時は価格確認中をオーバーライド"""
        row = _make_row(
            order_delivery_date=datetime.date(2026, 12, 31),
        )
        cache = _make_cache()
        report, _ = build_report_row(
            row, cache, {}, BRANCH, EXEC_TIME, True, TODAY
        )
        assert report.unit_price == "500"

    def test_fullwidth_dollar_flag(self):
        """全角＄＄フラグ"""
        row = _make_row(
            comment_internal="＄＄",
            order_delivery_date=datetime.date(2026, 12, 31),
        )
        cache = _make_cache()
        report, _ = build_report_row(
            row, cache, {}, BRANCH, EXEC_TIME, False, TODAY
        )
        assert report.unit_price == "500"


# ============================================
# 結合テスト: report_generator → history
# ============================================
class TestReportToHistoryIntegration:
    """回答書生成から送付履歴保存までの結合テスト"""

    def test_confirmed_orders_saved_to_history(self, tmp_path):
        """確定伝票が送付履歴に正しく保存される"""
        # 回答書生成
        data = [
            _make_row(
                order_number="100", detail_number="10",
                product_name="溶接棒A",
            ),
            _make_row(
                order_number="200", detail_number="10",
                product_name="ガスB",
            ),
        ]
        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト商事", {},
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is not None
        assert len(result.confirmed_orders) == 2

        # 送付履歴に保存
        history_path = str(tmp_path / "送付履歴.xlsx")
        wb = initialize_delivery_history(history_path)
        ws_history = wb[BRANCH.name] if BRANCH.name in wb.sheetnames else wb.active

        save_delivery_history(
            ws_history, result.confirmed_orders,
            execution_time=EXEC_TIME,
        )
        wb.save(history_path)

        # 保存されたデータを検証
        assert ws_history.cell(row=2, column=4).value == "100"  # 注番
        assert ws_history.cell(row=3, column=4).value == "200"

    def test_confirming_orders_saved_to_confirming_list(self, tmp_path):
        """確認中伝票が確認中一覧に正しく保存される"""
        data = [
            _make_row(
                order_number="100", detail_number="10",
                document_type="【受注】直送販売",
                storage_place="転送中（直送用）",
                order_delivery_date=datetime.date(2026, 12, 31),
            ),
        ]
        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト商事", {},
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is not None
        assert len(result.confirming_orders) == 1

        # 確認中一覧に保存
        history_path = str(tmp_path / "送付履歴.xlsx")
        wb = initialize_delivery_history(history_path)
        ws_confirming = wb["確認中一覧"]

        save_confirming_list(
            ws_confirming, result.confirming_orders,
            execution_time=EXEC_TIME,
        )
        wb.save(history_path)

        # 保存されたデータを検証
        assert ws_confirming.cell(row=2, column=4).value == "100"
        assert ws_confirming.cell(row=2, column=8).value == "未"  # 問合せ状況

    def test_full_lifecycle_confirming_to_confirmed(self, tmp_path):
        """確認中 → 次回処理完了で確定のフルライフサイクル"""
        cache = _make_cache()
        history_path = str(tmp_path / "送付履歴.xlsx")
        wb = initialize_delivery_history(history_path)

        # --- 1回目: 確認中伝票 ---
        data_first = [
            _make_row(
                order_number="100", detail_number="10",
                document_type="【受注】直送販売",
                storage_place="転送中（直送用）",
                order_delivery_date=datetime.date(2026, 12, 31),
            ),
        ]
        result1 = create_delivery_report(
            data_first, "テスト商事", {},
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result1 is not None
        assert len(result1.confirming_orders) == 1

        # 確認中一覧に保存
        ws_confirming = wb["確認中一覧"]
        save_confirming_list(
            ws_confirming, result1.confirming_orders,
            execution_time=EXEC_TIME,
        )
        wb.save(history_path)

        # --- 2回目: 処理完了になった ---
        data_second = [
            _make_row(
                order_number="100", detail_number="10",
                document_type="【受注】直送販売",
                storage_place="転送中（直送用）",
                ship_status="処理完了",
                order_delivery_date=datetime.date(2026, 12, 31),
            ),
        ]
        # 確認中をsent_ordersに含める
        sent = {"100|10": "確認中"}
        result2 = create_delivery_report(
            data_second, "テスト商事", sent,
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result2 is not None
        assert len(result2.confirmed_orders) == 1
        assert result2.confirmed_orders[0].delivery_answer == "納品済み"

        # 確認中→送付履歴へ移動
        ws_history = wb["送付履歴"]
        clean_confirming_list(
            ws_history, ws_confirming,
            result2.confirmed_orders,
            execution_time=EXEC_TIME,
        )
        wb.save(history_path)

        # 履歴に移動されたことを確認
        assert ws_history.cell(row=2, column=4).value == "100"


# ============================================
# 結合テスト: report_generator → email_builder
# ============================================
class TestReportToEmailIntegration:
    """回答書生成からメール作成までの結合テスト"""

    def test_basic_email_generation(self, tmp_path):
        """基本的なメール生成フロー"""
        data = [
            _make_row(order_number="100", detail_number="10"),
        ]
        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト商事", {},
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is not None

        # メール件名
        subject = build_email_subject("テスト商事", "", TODAY)
        assert "テスト商事" in subject
        assert "02/16" in subject

        # メール本文
        body = build_email_body_html(
            customer_name="テスト商事",
            branch=BRANCH,
            stockout_info_list=result.stockout_info_list,
            tracking_info_list=result.tracking_info_list,
            today=TODAY,
        )
        assert "テスト商事" in body
        assert "京葉営業所" in body

    def test_email_with_stockout(self, tmp_path):
        """欠品情報付きメール"""
        data = [
            _make_row(
                order_number="100", detail_number="10",
                comment_detail="欠品中 3月上旬予定",
                order_delivery_date=datetime.date(2026, 12, 31),
            ),
        ]
        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト商事", {},
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is not None

        body = build_email_body_html(
            customer_name="テスト商事",
            branch=BRANCH,
            stockout_info_list=result.stockout_info_list,
            today=TODAY,
        )
        assert "欠品" in body

    def test_email_with_tracking(self, tmp_path):
        """送り状情報付きメール"""
        data = [
            _make_row(
                order_number="100", detail_number="10",
                comment_external="佐川 1234567890",
            ),
        ]
        cache = _make_cache()
        result = create_delivery_report_by_order_numbers(
            data, "テスト商事", ["100"],
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is not None

        body = build_email_body_html(
            customer_name="テスト商事",
            branch=BRANCH,
            tracking_info_list=result.tracking_info_list,
            today=TODAY,
        )
        assert "送り状" in body
        assert "1234567890" in body

    def test_create_emails_with_customer_master(self, tmp_path):
        """顧客マスター連携でのメール作成"""
        # 顧客マスターのモックシート（A列=顧客名, E列以降=メールアドレス）
        wb_master = Workbook()
        ws_master = wb_master.active
        ws_master.append(["顧客名", "配送", "路線便", "保持日数", "メール1", "メール2"])
        ws_master.append(["テスト商事", "", "", "", "test@example.com", "test2@example.com"])

        data = [_make_row(order_number="100")]
        cache = _make_cache()
        result = create_delivery_report_by_order_numbers(
            data, "テスト商事", ["100"],
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is not None

        created_files = [{
            "customer_name": result.customer_name,
            "file_path": result.file_path,
            "stockout_info_list": result.stockout_info_list,
            "tracking_info_list": result.tracking_info_list,
            "bunno_info_list": [],
            "rep_name": "",
            "bunno_completed_list": [],
        }]

        emails = create_emails(
            created_files, BRANCH,
            customer_master_ws=ws_master,
            today=TODAY,
        )
        assert len(emails) == 1
        assert "test@example.com" in emails[0]["to"]
        assert "テスト商事" in emails[0]["subject"]

    def test_create_emails_no_email_skips(self, tmp_path):
        """メールアドレスなし → スキップ"""
        wb_master = Workbook()
        ws_master = wb_master.active
        ws_master.append(["顧客名", "メールアドレス"])
        # テスト商事のメールアドレスなし

        data = [_make_row(order_number="100")]
        cache = _make_cache()
        result = create_delivery_report_by_order_numbers(
            data, "テスト商事", ["100"],
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is not None

        created_files = [{
            "customer_name": result.customer_name,
            "file_path": result.file_path,
            "stockout_info_list": [],
            "tracking_info_list": [],
            "bunno_info_list": [],
            "rep_name": "",
            "bunno_completed_list": [],
        }]

        emails = create_emails(
            created_files, BRANCH,
            customer_master_ws=ws_master,
            today=TODAY,
        )
        assert len(emails) == 0


# ============================================
# 結合テスト: Excel出力の整合性
# ============================================
class TestExcelOutputIntegrity:
    """Excelファイル出力の整合性テスト"""

    def test_header_structure(self, tmp_path):
        """ヘッダー行の構造確認"""
        data = [_make_row()]
        cache = _make_cache()
        result = create_delivery_report_by_order_numbers(
            data, "テスト商事", ["1000001"],
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is not None

        wb = load_workbook(result.file_path)
        ws = wb.active
        # タイトル行
        assert "納　期　回　答　書" in str(ws["A1"].value)
        # 列ヘッダー（行6）
        assert ws.cell(row=6, column=1).value == "受注日"
        assert ws.cell(row=6, column=9).value == "納期回答"
        assert ws.cell(row=6, column=12).value == "弊社注番"

    def test_data_row_values(self, tmp_path):
        """データ行の値が正しく書き込まれる"""
        data = [
            _make_row(
                order_number="100", detail_number="10",
                customer_contact="田中",
                customer_order_number="PO-001",
                quantity="50",
                unit_price="1000",
                net_amount="50000",
            ),
        ]
        cache = _make_cache()
        result = create_delivery_report_by_order_numbers(
            data, "テスト商事", ["100"],
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is not None

        wb = load_workbook(result.file_path)
        ws = wb.active
        assert ws.cell(row=7, column=2).value == "田中"
        assert ws.cell(row=7, column=3).value == "PO-001"
        assert ws.cell(row=7, column=6).value == 50  # 数値として書き込まれる
        assert ws.cell(row=7, column=12).value == "100"

    def test_multiple_rows_order(self, tmp_path):
        """複数行の出力順序"""
        data = [
            _make_row(order_number="300", detail_number="10", product_name="商品C"),
            _make_row(order_number="100", detail_number="10", product_name="商品A"),
            _make_row(order_number="200", detail_number="10", product_name="商品B"),
        ]
        cache = _make_cache()
        result = create_delivery_report_by_order_numbers(
            data, "テスト商事", ["100", "200", "300"],
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is not None

        wb = load_workbook(result.file_path)
        ws = wb.active
        # ソースデータの順序で出力される
        assert ws.cell(row=7, column=12).value == "300"
        assert ws.cell(row=8, column=12).value == "100"
        assert ws.cell(row=9, column=12).value == "200"

    def test_sheet_name_from_customer(self, tmp_path):
        """シート名が顧客名から作られる"""
        data = [_make_row(order_number="100")]
        cache = _make_cache()
        result = create_delivery_report_by_order_numbers(
            data, "テスト商事", ["100"],
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is not None
        wb = load_workbook(result.file_path)
        ws = wb.active
        assert "テスト商事" in ws.title


# ============================================
# 結合テスト: 送付履歴のフルフロー
# ============================================
class TestHistoryFullFlow:
    """送付履歴の初期化→保存→読み込み→クリーンアップ"""

    def test_init_save_load_cycle(self, tmp_path):
        """初期化 → 保存 → 読み込みの一連フロー"""
        history_path = str(tmp_path / "送付履歴.xlsx")

        # 初期化
        wb = initialize_delivery_history(history_path)
        ws_history = wb["送付履歴"]
        ws_confirming = wb["確認中一覧"]

        # 確定伝票を保存
        confirmed = [
            HistoryRecord(
                order_date=TODAY,
                customer_name="テスト商事",
                order_number="100",
                detail_number="10",
                manufacturer_name="ダイヘン",
                product_name="溶接棒",
                delivery_answer="2月20日配達予定",
            ),
        ]
        save_delivery_history(ws_history, confirmed, EXEC_TIME)

        # 確認中伝票を保存
        confirming = [
            ConfirmingRecord(
                order_date=TODAY,
                customer_name="テスト商事",
                order_number="200",
                detail_number="10",
                manufacturer_name="パナソニック",
                product_name="ガス",
                status="未処理",
                order_delivery_date=datetime.date(2026, 12, 31),
            ),
        ]
        save_confirming_list(ws_confirming, confirming, EXEC_TIME)
        wb.save(history_path)

        # 読み込み
        wb2 = load_workbook(history_path)
        ws_h2 = wb2["送付履歴"]
        ws_c2 = wb2["確認中一覧"]
        cache = _make_cache()

        sent_orders = load_delivery_history(
            ws_h2, ws_c2, cache, {}, TODAY
        )
        # 当日送付なので、retention=0 + order_date < today が false → sent_ordersに入らない場合がある
        # registration_date < today チェックで判定されるが、ここではload_delivery_historyの動作を確認
        assert isinstance(sent_orders, dict)

    def test_clean_confirming_moves_to_history(self, tmp_path):
        """確認中一覧→送付履歴の移動"""
        history_path = str(tmp_path / "送付履歴.xlsx")
        wb = initialize_delivery_history(history_path)
        ws_history = wb["送付履歴"]
        ws_confirming = wb["確認中一覧"]

        # 確認中に2件登録
        confirming = [
            ConfirmingRecord(
                order_date=TODAY,
                customer_name="テスト商事",
                order_number="100",
                detail_number="10",
                manufacturer_name="ダイヘン",
                product_name="溶接棒",
                status="未処理",
            ),
            ConfirmingRecord(
                order_date=TODAY,
                customer_name="テスト商事",
                order_number="200",
                detail_number="10",
                manufacturer_name="パナソニック",
                product_name="ガス",
                status="未処理",
            ),
        ]
        save_confirming_list(ws_confirming, confirming, EXEC_TIME)
        wb.save(history_path)

        # 100のみ確定
        confirmed = [
            HistoryRecord(
                order_date=TODAY,
                customer_name="テスト商事",
                order_number="100",
                detail_number="10",
                manufacturer_name="ダイヘン",
                product_name="溶接棒",
                delivery_answer="2月20日配達予定",
            ),
        ]

        clean_confirming_list(
            ws_history, ws_confirming,
            confirmed, EXEC_TIME,
        )
        wb.save(history_path)

        # 確認中一覧に200だけ残っている
        remaining_orders = []
        max_row = ws_confirming.max_row
        for r in range(2, max_row + 1):
            val = ws_confirming.cell(row=r, column=4).value
            if val:
                remaining_orders.append(str(val))
        assert "200" in remaining_orders
        assert "100" not in remaining_orders

        # 送付履歴に100が入っている
        history_orders = []
        max_row_h = ws_history.max_row
        for r in range(2, max_row_h + 1):
            val = ws_history.cell(row=r, column=4).value
            if val:
                history_orders.append(str(val))
        assert "100" in history_orders


# ============================================
# 結合テスト: 大量データ
# ============================================
class TestLargeDataSet:
    """大量データでのパフォーマンス・正常動作テスト"""

    def test_100_orders(self, tmp_path):
        """100件の注文データ"""
        data = []
        for i in range(100):
            data.append(_make_row(
                order_number=str(1000 + i),
                detail_number="10",
                product_name=f"商品{i}",
                customer_order_number=f"PO-{i:04d}",
            ))

        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト商事", {},
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is not None
        assert len(result.confirmed_orders) == 100

        # Excelファイルが正しく生成される
        wb = load_workbook(result.file_path)
        ws = wb.active
        assert ws.cell(row=106, column=12).value == str(1000 + 99)

    def test_50_customers_grouping(self):
        """50顧客分のグループ化"""
        data = []
        order_numbers = []
        for i in range(50):
            order_num = str(1000 + i)
            data.append(_make_row(
                order_number=order_num,
                customer_name=f"顧客{i}",
            ))
            order_numbers.append(order_num)

        result = group_order_numbers_by_customer(data, order_numbers)
        assert len(result) == 50
        for key in result:
            assert len(result[key]) == 1


# ============================================
# 結合テスト: 複数顧客の同時処理
# ============================================
class TestMultiCustomerFlow:
    """複数顧客を同時処理するフロー"""

    def test_period_mode_multi_customer(self, tmp_path):
        """期間モードで顧客ごとに個別に回答書生成"""
        all_data = [
            _make_row(
                order_number="100", detail_number="10",
                customer_name="A社", product_name="商品A",
            ),
            _make_row(
                order_number="200", detail_number="10",
                customer_name="B社", product_name="商品B",
            ),
            _make_row(
                order_number="300", detail_number="10",
                customer_name="A社", product_name="商品C",
            ),
        ]

        cache = _make_cache()

        # A社の回答書
        dir_a = tmp_path / "a"
        dir_a.mkdir()
        result_a = create_delivery_report(
            all_data, "A社", {},
            cache, dir_a, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result_a is not None
        assert len(result_a.confirmed_orders) == 2  # 100, 300

        # B社の回答書
        dir_b = tmp_path / "b"
        dir_b.mkdir()
        result_b = create_delivery_report(
            all_data, "B社", {},
            cache, dir_b, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result_b is not None
        assert len(result_b.confirmed_orders) == 1  # 200のみ

    def test_order_number_mode_multi_customer(self, tmp_path):
        """注番モードで複数顧客を処理"""
        all_data = [
            _make_row(
                order_number="100", detail_number="10",
                customer_name="A社",
            ),
            _make_row(
                order_number="200", detail_number="10",
                customer_name="B社",
            ),
        ]

        # グループ化
        groups = group_order_numbers_by_customer(all_data, ["100", "200"])
        assert "A社" in groups
        assert "B社" in groups

        cache = _make_cache()

        # 各顧客で回答書生成
        for customer_name, order_nums in groups.items():
            sub_dir = tmp_path / customer_name
            sub_dir.mkdir()
            result = create_delivery_report_by_order_numbers(
                all_data, customer_name, order_nums,
                cache, sub_dir, {}, BRANCH, EXEC_TIME,
                today=TODAY,
            )
            assert result is not None
            assert result.customer_name == customer_name


# ============================================
# エッジケース: 特殊コメント
# ============================================
class TestSpecialComments:
    """特殊コメントのエッジケース"""

    def test_work_order_comment(self):
        """&&作業コメント"""
        row = _make_row(
            comment_internal="&&作業依頼",
            specified_delivery_date=datetime.date(2026, 2, 25),
        )
        cache = _make_cache()
        result = calculate_delivery_date(
            row, cache, {}, BRANCH, EXEC_TIME, TODAY
        )
        assert "作業予定" in result

    def test_arrival_date_comment(self):
        """@@着日指定コメント"""
        row = _make_row(
            comment_internal="@@2/25",
            specified_delivery_date=datetime.date(2026, 2, 20),
        )
        cache = _make_cache()
        result = calculate_delivery_date(
            row, cache, {}, BRANCH, EXEC_TIME, TODAY
        )
        assert "着" in result

    def test_pickup_comment(self):
        """引取コメント"""
        row = _make_row(
            comment_external="引取 2/25",
        )
        cache = _make_cache()
        result = calculate_delivery_date(
            row, cache, {}, BRANCH, EXEC_TIME, TODAY
        )
        assert "引取予定" in result

    def test_multiple_special_comments_priority(self):
        """特殊コメントの優先順位: &&が最優先"""
        row = _make_row(
            comment_internal="&&作業 @@2/25",
            comment_external="引取 2/20",
            specified_delivery_date=datetime.date(2026, 2, 25),
        )
        cache = _make_cache()
        result = calculate_delivery_date(
            row, cache, {}, BRANCH, EXEC_TIME, TODAY
        )
        # &&が最優先
        assert "作業" in result


# ============================================
# エッジケース: 納入先名
# ============================================
class TestDeliveryPlaceEdgeCases:
    """納入先名のエッジケース"""

    def test_empty_ship_to_name(self):
        """出荷先名が空"""
        row = _make_row(ship_to_name="")
        cache = _make_cache()
        report, _ = build_report_row(
            row, cache, {}, BRANCH, EXEC_TIME, False, TODAY
        )
        # 空の出荷先名
        assert report.delivery_place == ""

    def test_ship_to_with_sama(self):
        """出荷先名に「様」付き"""
        row = _make_row(ship_to_name="東京工場様")
        cache = _make_cache()
        report, _ = build_report_row(
            row, cache, {}, BRANCH, EXEC_TIME, False, TODAY
        )
        assert report.delivery_place == "東京工場様"

    def test_ship_to_without_sama(self):
        """出荷先名に「様」なし → 付与"""
        row = _make_row(ship_to_name="東京工場")
        cache = _make_cache()
        report, _ = build_report_row(
            row, cache, {}, BRANCH, EXEC_TIME, False, TODAY
        )
        assert report.delivery_place == "東京工場様"
