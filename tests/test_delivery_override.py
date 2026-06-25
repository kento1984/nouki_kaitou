"""納期上書き（納期上書きシート＋逐語文字列）のテスト

SAP施錠等でツール計算の納期を直せない明細を、運用者が手書きした逐語文字列で
最優先に差し替える機能。期間限定運用。機能削除時はこのファイルごと削除してよい。

検証の柱:
- 上書きシート空 → 全行が従来どおり（完全no-op）
- 上書き1行 → その明細だけ逐語文字列で出る（指定納期もTWF納品済み上書きも越える）
- 同じ注番の他明細・他注番・他客は一切影響を受けない
- 納期上書きシートのラウンドトリップ保全（保存で消えない）
"""

import datetime

from openpyxl import load_workbook

from nouki_kaitou.delivery_calc import (
    calculate_delivery_date,
    get_delivery_override,
)
from nouki_kaitou.history import (
    OVERRIDE_SHEET_NAME,
    extract_override_rows,
    initialize_delivery_history,
    load_delivery_overrides,
    save_history_batch,
)
from nouki_kaitou.models import BranchSettings, CacheStore, OrderRow
from nouki_kaitou.report_generator import build_report_row, create_delivery_report
from nouki_kaitou.utils import normalize_order_detail_key

# ============================================
# テスト用ヘルパー（test_twf.py と同パターン）
# ============================================
TODAY = datetime.date(2026, 6, 26)
EXEC = datetime.datetime(2026, 6, 26, 10, 0, 0)
WRONG = datetime.date(2026, 6, 20)  # 計上済み行に焼き付いた間違った指定納期（過去日）
BRANCH = BranchSettings(
    name="京葉営業所", default_cutoff=15, base_center="関東商品センター"
)


def _make_cache(**kwargs) -> CacheStore:
    return CacheStore(
        mfg_name=kwargs.get("mfg_name", {"D01": "ダイヘン"}),
        mfg_days=kwargs.get("mfg_days", {"D01": 2}),
        cust_days=kwargs.get("cust_days", {}),
        cust_retention=kwargs.get("cust_retention", {}),
        cust_route=kwargs.get("cust_route", {}),
        confirm=kwargs.get("confirm", {}),
        storage=kwargs.get("storage", {}),
        delivery_overrides=kwargs.get("delivery_overrides", {}),
    )


def _make_row(**kwargs) -> OrderRow:
    defaults = {
        "order_number": "9999001",
        "detail_number": "10",
        "document_type": "【受注】在庫販売",
        "customer_name": "テスト商事",
        "product_name": "溶接棒 3.2mm",
        "ship_status": "未処理",
        "quantity": "100",
        "unit_price": "500",
        "net_amount": "50000",
        "manufacturer_name": "ダイヘン",
        "storage_place": "関東商品センター",
        "customer_order_number": "PO-001",
        "customer_contact": "田中",
        "comment_detail": "",
        "comment_external": "",
        "comment_internal": "",
        "rejection_reason": "",
        "ship_to_name": "テスト商事",
        "registration_date": TODAY,
        "time_value": "10:00:00",
        "order_delivery_date": datetime.date(2026, 6, 20),
        "specified_delivery_date": None,
        "item_group_code": "D01",
    }
    defaults.update(kwargs)
    return OrderRow(**defaults)


def _twf_answers(data, cache, sent, tmp_path) -> list[str]:
    """TWF専用回答書を生成し、データ行のI列(納期回答)を順に返す。"""
    result = create_delivery_report(
        data, "テスト商事", sent,
        cache, tmp_path, {}, BRANCH, EXEC,
        today=TODAY,
        include_only_orders={r.order_number.strip() for r in data},
        filter_already_sent=False,
        twf_mode=True,
    )
    ws = load_workbook(result.file_path).active
    answers = []
    for r in range(7, ws.max_row + 1):
        if ws.cell(r, 12).value:  # L列=弊社注番 がある＝データ行
            answers.append(ws.cell(r, 9).value)
    return answers


# ============================================
# get_delivery_override（単体）
# ============================================
class TestGetDeliveryOverride:
    def test_empty_cache_returns_none(self):
        """上書き辞書が空なら常にNone（no-op）"""
        cache = _make_cache()
        assert get_delivery_override("9999001", "10", cache) is None

    def test_hit_returns_verbatim(self):
        """ヒットすれば逐語文字列を返す"""
        cache = _make_cache(delivery_overrides={"9999001|10": "6月30日納品予定"})
        assert get_delivery_override("9999001", "10", cache) == "6月30日納品予定"

    def test_key_is_stripped(self):
        """注番・明細の前後空白を除去してキー照合する"""
        cache = _make_cache(delivery_overrides={"9999001|10": "X"})
        assert get_delivery_override(" 9999001 ", " 10 ", cache) == "X"

    def test_other_key_unaffected(self):
        """別の明細・別の注番はヒットしない"""
        cache = _make_cache(delivery_overrides={"9999001|10": "X"})
        assert get_delivery_override("9999001", "20", cache) is None
        assert get_delivery_override("9999002", "10", cache) is None

    def test_blank_value_treated_as_none(self):
        """空白だけの上書き値はNone扱い"""
        cache = _make_cache(delivery_overrides={"9999001|10": "   "})
        assert get_delivery_override("9999001", "10", cache) is None


# ============================================
# calculate_delivery_date（優先0で最優先）
# ============================================
class TestCalculatePriority:
    def test_override_beats_specified_date_on_completed_row(self):
        """計上済み（処理完了）＋指定納期入りでも、上書きが最優先で勝つ"""
        row = _make_row(
            ship_status="処理完了", specified_delivery_date=WRONG
        )
        # 上書きなし → 指定納期パスで間違った過去日
        base = _make_cache()
        assert calculate_delivery_date(row, base, today=TODAY, execution_time=EXEC) \
            == "6月20日配達済み"
        # 上書きあり → 逐語文字列が最優先
        over = _make_cache(delivery_overrides={"9999001|10": "6月30日納品予定"})
        assert calculate_delivery_date(row, over, today=TODAY, execution_time=EXEC) \
            == "6月30日納品予定"

    def test_no_override_is_noop_for_normal_row(self):
        """上書き辞書が空なら通常行の計算結果は一切変わらない"""
        row = _make_row(specified_delivery_date=WRONG)
        empty = _make_cache()
        # delivery_overrides={} は CacheStore のデフォルト。明示有無で差が出ないこと
        assert calculate_delivery_date(row, empty, today=TODAY, execution_time=EXEC) \
            == calculate_delivery_date(row, _make_cache(delivery_overrides={}),
                                       today=TODAY, execution_time=EXEC)


# ============================================
# TWF回答書（report_generator 統合）
# ============================================
class TestTwfIntegration:
    def _problem_set(self):
        """計上済み・指定納期入りの問題行(明細10) ＋ 正常な同注番3行"""
        problem = _make_row(
            detail_number="10", ship_status="処理完了",
            specified_delivery_date=WRONG, comment_detail="TWFNo.001 テスト商事様",
        )
        others = [
            _make_row(
                detail_number=str(n), ship_status="処理完了",
                specified_delivery_date=WRONG, product_name=f"部材{n}",
                comment_detail="TWFNo.001 テスト商事様",
            )
            for n in (20, 30, 40)
        ]
        return problem, others

    def test_override_empty_is_noop(self, tmp_path):
        """上書き空 → 全4行が従来どおり（処理完了+履歴あり→納品済み）"""
        problem, others = self._problem_set()
        data = [problem] + others
        sent = {f"9999001|{n}": "6月20日配達済み" for n in (10, 20, 30, 40)}
        cache = _make_cache()  # delivery_overrides 空
        answers = _twf_answers(data, cache, sent, tmp_path)
        assert answers == ["納品済み"] * 4

    def test_override_wins_over_nouhinzumi(self, tmp_path):
        """上書き1行 → その行だけ逐語、他3行は従来どおり納品済み"""
        problem, others = self._problem_set()
        data = [problem] + others
        sent = {f"9999001|{n}": "6月20日配達済み" for n in (10, 20, 30, 40)}
        cache = _make_cache(
            delivery_overrides={"9999001|10": "6月30日納品予定"}
        )
        answers = _twf_answers(data, cache, sent, tmp_path)
        # 明細10だけ上書き、残りは納品済み
        assert answers.count("6月30日納品予定") == 1
        assert answers.count("納品済み") == 3

    def test_other_customer_unaffected(self, tmp_path):
        """別客の同形の行は、上書きキーに無いので影響なし"""
        # 別客の注番8888001（上書き辞書には9999001しか無い）
        row = _make_row(
            order_number="8888001", detail_number="10", ship_status="処理完了",
            specified_delivery_date=WRONG, customer_name="別商事",
            ship_to_name="別商事", comment_detail="TWFNo.099 別商事様",
        )
        sent = {"8888001|10": "6月20日配達済み"}
        cache = _make_cache(delivery_overrides={"9999001|10": "6月30日納品予定"})
        result = create_delivery_report(
            [row], "別商事", sent,
            cache, tmp_path, {}, BRANCH, EXEC,
            today=TODAY, include_only_orders={"8888001"},
            filter_already_sent=False, twf_mode=True,
        )
        ws = load_workbook(result.file_path).active
        answers = [ws.cell(r, 9).value for r in range(7, ws.max_row + 1)
                   if ws.cell(r, 12).value]
        assert answers == ["納品済み"]  # 上書きされない


# ============================================
# 納期上書きシートのI/O・ラウンドトリップ保全
# ============================================
class TestSheetIO:
    def test_initialize_creates_override_sheet(self, tmp_path):
        """新規送付履歴.xlsxに納期上書きシートが作られる"""
        path = str(tmp_path / "送付履歴.xlsx")
        wb = initialize_delivery_history(path)
        assert OVERRIDE_SHEET_NAME in wb.sheetnames
        ws = wb[OVERRIDE_SHEET_NAME]
        assert ws.cell(1, 1).value == "受発注伝票"
        assert ws.cell(1, 3).value == "納期回答(上書き)"

    def test_load_empty_sheet_is_empty_dict(self, tmp_path):
        """空シート → 空dict（完全no-op）"""
        path = str(tmp_path / "送付履歴.xlsx")
        wb = initialize_delivery_history(path)
        assert load_delivery_overrides(wb[OVERRIDE_SHEET_NAME]) == {}

    def test_load_none_sheet_is_empty_dict(self):
        """シートNone（既存ファイルに無い）→ 空dict"""
        assert load_delivery_overrides(None) == {}

    def test_load_skips_blank_answer(self, tmp_path):
        """納期回答(上書き)が空の行は無視（注番だけ書いても上書きしない）"""
        path = str(tmp_path / "送付履歴.xlsx")
        wb = initialize_delivery_history(path)
        ws = wb[OVERRIDE_SHEET_NAME]
        ws.cell(2, 1).value = "9999001"
        ws.cell(2, 2).value = 10
        ws.cell(2, 3).value = "6月30日納品予定"
        ws.cell(3, 1).value = "9999001"     # 注番だけ・回答空
        ws.cell(3, 2).value = 20
        overrides = load_delivery_overrides(ws)
        assert overrides == {"9999001|10": "6月30日納品予定"}

    def test_int_detail_normalized(self, tmp_path):
        """明細が数値セルでも文字列キーに正規化される"""
        path = str(tmp_path / "送付履歴.xlsx")
        wb = initialize_delivery_history(path)
        ws = wb[OVERRIDE_SHEET_NAME]
        ws.cell(2, 1).value = "9999001"
        ws.cell(2, 2).value = 10           # int
        ws.cell(2, 3).value = "X"
        assert load_delivery_overrides(ws) == {"9999001|10": "X"}

    def test_roundtrip_survives_save(self, tmp_path):
        """手入力した上書き行が save_history_batch（WB新規作成）後も残る"""
        path = str(tmp_path / "送付履歴.xlsx")
        wb = initialize_delivery_history(path)
        ws = wb[OVERRIDE_SHEET_NAME]
        ws.cell(2, 1).value = "9999001"
        ws.cell(2, 2).value = 10
        ws.cell(2, 3).value = "6月30日納品予定"
        ws.cell(2, 4).value = "計上済みSAP施錠"
        wb.save(path)

        # 読み直し → 保全用に行を抽出 → 保存（履歴更新なし）
        wb_ro = load_workbook(path, read_only=True)
        override_rows = extract_override_rows(wb_ro[OVERRIDE_SHEET_NAME])
        wb_ro.close()
        assert len(override_rows) == 1

        save_history_batch(
            path, [], [], [], [], EXEC, "tester",
            today=TODAY, override_rows=override_rows,
        )

        # 保存後も上書きが生きている
        wb2 = load_workbook(path)
        assert OVERRIDE_SHEET_NAME in wb2.sheetnames
        assert load_delivery_overrides(wb2[OVERRIDE_SHEET_NAME]) == {
            "9999001|10": "6月30日納品予定"
        }

    def test_save_without_override_rows_makes_empty_sheet(self, tmp_path):
        """override_rows未指定でも空シートが作られクラッシュしない"""
        path = str(tmp_path / "送付履歴.xlsx")
        initialize_delivery_history(path)
        save_history_batch(path, [], [], [], [], EXEC, "tester", today=TODAY)
        wb = load_workbook(path)
        assert OVERRIDE_SHEET_NAME in wb.sheetnames
        assert load_delivery_overrides(wb[OVERRIDE_SHEET_NAME]) == {}


# ============================================
# Codexレビュー指摘の回帰テスト
# ============================================
class TestOverrideBeatsBunnoKeppin:
    """指摘①: build_report_row の分納/欠品再上書きが、上書きを壊さないこと。

    非処理完了（未処理）行で分納・欠品コメントがあっても、上書きが効いて
    いれば逐語文字列が最優先で残る（計上済み=処理完了は元々ガード済み）。
    """

    def _row(self, **kw):
        return _make_row(order_number="7000001", detail_number="10", **kw)

    def test_keppin_does_not_append_to_override(self):
        """欠品中コメント＋未処理＋上書き → 「（欠品）」が付かない"""
        cache = _make_cache(delivery_overrides={"7000001|10": "6月30日納品予定"})
        _, ans = build_report_row(
            self._row(ship_status="未処理", comment_detail="欠品中"),
            cache, today=TODAY, execution_time=EXEC,
        )
        assert ans == "6月30日納品予定"

    def test_bunno_does_not_clobber_override(self):
        """分納コメント＋未処理＋上書き → 「分納」で潰れない"""
        cache = _make_cache(delivery_overrides={"7000001|10": "6月30日納品予定"})
        _, ans = build_report_row(
            self._row(ship_status="未処理",
                      comment_detail="分納:50個 1/10,30個 未定"),
            cache, today=TODAY, execution_time=EXEC,
        )
        assert ans == "6月30日納品予定"

    def test_no_override_keppin_unchanged(self):
        """上書きが無ければ欠品の従来挙動（（欠品）付与）は維持＝no-op"""
        cache = _make_cache()
        _, ans = build_report_row(
            self._row(ship_status="未処理", comment_detail="欠品中",
                      specified_delivery_date=datetime.date(2026, 6, 30)),
            cache, today=TODAY, execution_time=EXEC,
        )
        assert "（欠品）" in ans  # 従来どおり付く


class TestKeyNormalization:
    """指摘③: 手入力の表記ゆれでマッチ漏れ/誤マッチしない共通キー正規化。"""

    def test_variants_map_to_same_key(self):
        base = normalize_order_detail_key("7000001", "10")
        assert base == "7000001|10"
        assert normalize_order_detail_key("7000001", "010") == base   # ゼロ詰め
        assert normalize_order_detail_key("7000001", "10.0") == base  # Excel数値
        assert normalize_order_detail_key("７０００００１", "１０") == base  # 全角
        assert normalize_order_detail_key(7000001, 10) == base        # 数値型セル

    def test_non_numeric_detail_kept_as_string(self):
        """非数値の明細は文字列のまま（誤って数値化しない）"""
        assert normalize_order_detail_key("A1", "10A") == "A1|10A"

    def test_lookup_matches_despite_excel_float(self):
        """シートに 10.0 で入っていても、SAP明細 '10' で照合できる"""
        from nouki_kaitou.delivery_calc import get_delivery_override
        # シート読込時に正規化されるので、登録キーは "7000001|10"
        cache = _make_cache(
            delivery_overrides={normalize_order_detail_key("7000001", 10.0): "X"}
        )
        assert get_delivery_override("7000001", "10", cache) == "X"


# ============================================
# B: 指定納期あり紐付き＋処理完了を配達予定/済み日で出す（TWF限定）
# ============================================
class TestHimozukiDeliveryDateInTwf:
    """TWFで「指定納期あり紐付き＋処理完了」を納品済みでなく具体的な
    配達予定/配達済み日で出す。指定納期なし紐付き・直送は納品済み維持。
    """

    def _himo(self, **kw):
        """紐付き(直送販売＋非転送中)＋処理完了 の TWF 明細。"""
        d = dict(
            order_number="8000001", detail_number="10",
            document_type="【受注】直送販売", ship_status="処理完了",
            storage_place="関東商品センター",  # 非転送中＝紐付き
            # 顧客名/出荷先は _make_row 既定（テスト商事）＝自社便。
            # create_delivery_report の顧客名フィルタと _twf_answers の
            # customer_name("テスト商事") に一致させる
            comment_detail="TWFNo.001 テスト商事様",
            registration_date=datetime.date(2026, 6, 20),
            specified_delivery_date=None,
        )
        d.update(kw)
        return _make_row(**d)

    def _twf(self, data, tmp_path):
        return _twf_answers(data, _make_cache(), {"8000001|10": "x"}, tmp_path)

    def test_future_spec_shows_haitatsu_yotei(self, tmp_path):
        """指定納期が未来 → ○月○日配達予定（納品済みにしない）"""
        data = [self._himo(specified_delivery_date=datetime.date(2026, 7, 1))]
        ans = self._twf(data, tmp_path)
        assert ans[0] == "7月3日配達予定"

    def test_past_spec_shows_haitatsu_zumi_date(self, tmp_path):
        """指定納期が過去 → ○月○日配達済み（具体日。納品済みにしない）"""
        data = [self._himo(specified_delivery_date=datetime.date(2026, 6, 20))]
        ans = self._twf(data, tmp_path)
        assert ans[0].endswith("配達済み") and ans[0] != "納品済み"

    def test_no_spec_stays_nouhinzumi(self, tmp_path):
        """指定納期なし紐付き（today基準で毎日スライド）→ 納品済み維持"""
        data = [self._himo(specified_delivery_date=None)]
        assert self._twf(data, tmp_path) == ["納品済み"]

    def test_dec31_spec_stays_nouhinzumi(self, tmp_path):
        """指定納期=12/31（未確定）→ 指定納期なし扱い → 納品済み維持"""
        data = [self._himo(specified_delivery_date=datetime.date(2026, 12, 31))]
        assert self._twf(data, tmp_path) == ["納品済み"]

    def test_chokusou_stays_nouhinzumi(self, tmp_path):
        """直送（転送中）＋処理完了は指定納期があっても納品済み維持"""
        data = [self._himo(storage_place="転送中（直送用）",
                           specified_delivery_date=datetime.date(2026, 7, 1))]
        assert self._twf(data, tmp_path) == ["納品済み"]

    def test_normal_mode_unaffected(self, tmp_path):
        """通常モード（非TWF）は変更前と同じく計算日を出す（B はTWF限定）"""
        row = self._himo(specified_delivery_date=datetime.date(2026, 7, 1))
        result = create_delivery_report(
            [row], "テスト商事", {},  # sentなし＝スキップされない
            _make_cache(), tmp_path, {}, BRANCH, EXEC, today=TODAY,
        )
        ws = load_workbook(result.file_path).active
        col = next(c for c in range(1, 15) if ws.cell(6, c).value == "納期回答")
        answers = [ws.cell(r, col).value for r in range(7, ws.max_row + 1)
                   if ws.cell(r, 12).value]
        assert answers == ["7月3日配達予定"]  # 元から納品済みにしていない
