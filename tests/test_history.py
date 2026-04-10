"""history.py のユニットテスト

送付履歴管理の全7関数を網羅的にテスト。
"""

import datetime
import os
import tempfile

import pytest
from openpyxl import Workbook, load_workbook

from nouki_kaitou.history import (
    CONFIRMING_SHEET_NAME,
    HISTORY_SHEET_NAME,
    clean_confirming_list,
    clean_old_confirming_list,
    clean_old_history,
    extract_sheet_rows,
    initialize_delivery_history,
    load_delivery_history,
    save_confirming_list,
    save_delivery_history,
    save_history_batch,
)
from nouki_kaitou.models import (
    CacheStore,
    ConfirmingRecord,
    HistoryRecord,
)


# ============================================
# ヘルパー
# ============================================
def _make_history_ws(wb=None):
    """テスト用の送付履歴シートを作成"""
    if wb is None:
        wb = Workbook()
    ws = wb.active
    ws.title = HISTORY_SHEET_NAME
    headers = [
        "送付日時", "受注日", "顧客名", "受発注伝票", "明細",
        "メーカー名", "品名", "納期回答", "送付者",
    ]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx).value = h
    return wb, ws


def _make_confirming_ws(wb=None):
    """テスト用の確認中一覧シートを作成"""
    if wb is None:
        wb = Workbook()
        ws = wb.active
    else:
        ws = wb.create_sheet(CONFIRMING_SHEET_NAME)
    ws.title = CONFIRMING_SHEET_NAME
    headers = [
        "送付日時", "受注日", "顧客名", "受発注伝票", "明細",
        "メーカー名", "品名", "問合せ状況", "ステータス", "受注納期", "送付者",
    ]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx).value = h
    return wb, ws


def _add_history_row(ws, row_num, sent_dt, order_date, customer, order_num,
                     detail, mfg, product, delivery, sender="test"):
    """送付履歴にデータ行を追加"""
    ws.cell(row=row_num, column=1).value = sent_dt
    ws.cell(row=row_num, column=2).value = order_date
    ws.cell(row=row_num, column=3).value = customer
    ws.cell(row=row_num, column=4).value = order_num
    ws.cell(row=row_num, column=5).value = detail
    ws.cell(row=row_num, column=6).value = mfg
    ws.cell(row=row_num, column=7).value = product
    ws.cell(row=row_num, column=8).value = delivery
    ws.cell(row=row_num, column=9).value = sender


def _add_confirming_row(ws, row_num, sent_dt, order_date, customer, order_num,
                        detail, mfg, product, inquiry, status,
                        order_delivery=None, sender="test"):
    """確認中一覧にデータ行を追加"""
    ws.cell(row=row_num, column=1).value = sent_dt
    ws.cell(row=row_num, column=2).value = order_date
    ws.cell(row=row_num, column=3).value = customer
    ws.cell(row=row_num, column=4).value = order_num
    ws.cell(row=row_num, column=5).value = detail
    ws.cell(row=row_num, column=6).value = mfg
    ws.cell(row=row_num, column=7).value = product
    ws.cell(row=row_num, column=8).value = inquiry
    ws.cell(row=row_num, column=9).value = status
    ws.cell(row=row_num, column=10).value = order_delivery
    ws.cell(row=row_num, column=11).value = sender


# ============================================
# InitializeDeliveryHistory
# ============================================
class TestInitializeDeliveryHistory:
    def test_creates_file(self):
        """ファイルが作成される"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            wb = initialize_delivery_history(path)
            wb.close()
            assert os.path.exists(path)

            wb2 = load_workbook(path)
            assert HISTORY_SHEET_NAME in wb2.sheetnames
            assert CONFIRMING_SHEET_NAME in wb2.sheetnames
            wb2.close()
        finally:
            os.unlink(path)

    def test_history_headers(self):
        """送付履歴のヘッダーが設定される"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            wb = initialize_delivery_history(path)
            ws = wb[HISTORY_SHEET_NAME]
            assert ws.cell(row=1, column=1).value == "送付日時"
            assert ws.cell(row=1, column=9).value == "送付者"
            wb.close()
        finally:
            os.unlink(path)

    def test_confirming_headers(self):
        """確認中一覧のヘッダーが設定される"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            wb = initialize_delivery_history(path)
            ws = wb[CONFIRMING_SHEET_NAME]
            assert ws.cell(row=1, column=1).value == "送付日時"
            assert ws.cell(row=1, column=8).value == "問合せ状況"
            assert ws.cell(row=1, column=11).value == "送付者"
            wb.close()
        finally:
            os.unlink(path)

    def test_tables_created(self):
        """テーブルが作成される"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            wb = initialize_delivery_history(path)
            ws_hist = wb[HISTORY_SHEET_NAME]
            ws_conf = wb[CONFIRMING_SHEET_NAME]
            assert len(ws_hist.tables) == 1
            assert len(ws_conf.tables) == 1
            wb.close()
        finally:
            os.unlink(path)


# ============================================
# LoadDeliveryHistory
# ============================================
class TestLoadDeliveryHistory:
    def test_empty_sheets(self):
        """空のシートでは空辞書"""
        wb, ws_hist = _make_history_ws()
        _, ws_conf = _make_confirming_ws(wb)
        cache = CacheStore()
        result = load_delivery_history(ws_hist, ws_conf, cache)
        assert result == {}

    def test_basic_loading(self):
        """基本的な読み込み"""
        wb, ws_hist = _make_history_ws()
        _, ws_conf = _make_confirming_ws(wb)
        cache = CacheStore()

        today = datetime.date(2026, 2, 16)
        _add_history_row(
            ws_hist, 2,
            sent_dt=datetime.datetime(2026, 2, 15, 10, 0),
            order_date=datetime.date(2026, 2, 10),
            customer="テスト顧客",
            order_num="12345",
            detail="10",
            mfg="メーカーA",
            product="製品A",
            delivery="2/17出荷予定",
        )

        result = load_delivery_history(ws_hist, ws_conf, cache, today=today)
        assert "12345|10" in result
        assert result["12345|10"] == "2/17出荷予定"

    def test_confirming_skip(self):
        """確認中ステータスはスキップ対象外"""
        wb, ws_hist = _make_history_ws()
        _, ws_conf = _make_confirming_ws(wb)
        cache = CacheStore()

        today = datetime.date(2026, 2, 16)
        _add_history_row(
            ws_hist, 2,
            sent_dt=datetime.datetime(2026, 2, 15, 10, 0),
            order_date=datetime.date(2026, 2, 10),
            customer="テスト顧客",
            order_num="12345",
            detail="10",
            mfg="メーカーA",
            product="製品A",
            delivery="確認中",
        )

        result = load_delivery_history(ws_hist, ws_conf, cache, today=today)
        assert "12345|10" not in result

    def test_bunno_kanryou_always_skip(self):
        """分納完了は常にスキップ対象"""
        wb, ws_hist = _make_history_ws()
        _, ws_conf = _make_confirming_ws(wb)
        cache = CacheStore()

        today = datetime.date(2026, 2, 16)
        # 受注日が今日（通常なら除外されない）
        _add_history_row(
            ws_hist, 2,
            sent_dt=datetime.datetime(2026, 2, 16, 10, 0),
            order_date=datetime.date(2026, 2, 16),
            customer="テスト顧客",
            order_num="12345",
            detail="10",
            mfg="メーカーA",
            product="製品A",
            delivery="分納完了",
        )

        result = load_delivery_history(ws_hist, ws_conf, cache, today=today)
        assert "12345|10" in result

    def test_order_date_future_not_skip(self):
        """受注日が今日以降→スキップ対象外（保持日数=0）"""
        wb, ws_hist = _make_history_ws()
        _, ws_conf = _make_confirming_ws(wb)
        cache = CacheStore()

        today = datetime.date(2026, 2, 16)
        _add_history_row(
            ws_hist, 2,
            sent_dt=datetime.datetime(2026, 2, 16, 10, 0),
            order_date=datetime.date(2026, 2, 16),  # 今日
            customer="テスト顧客",
            order_num="12345",
            detail="10",
            mfg="メーカーA",
            product="製品A",
            delivery="2/17出荷予定",
        )

        result = load_delivery_history(ws_hist, ws_conf, cache, today=today)
        assert "12345|10" not in result

    def test_retention_days(self):
        """保持日数による判定"""
        wb, ws_hist = _make_history_ws()
        _, ws_conf = _make_confirming_ws(wb)
        cache = CacheStore()
        cache.cust_retention = {"テスト顧客": 3}

        today = datetime.date(2026, 2, 16)
        # 送付日時が5営業日前 → 保持日数3を超えている
        _add_history_row(
            ws_hist, 2,
            sent_dt=datetime.datetime(2026, 2, 9, 10, 0),
            order_date=datetime.date(2026, 2, 5),
            customer="テスト顧客",
            order_num="12345",
            detail="10",
            mfg="メーカーA",
            product="製品A",
            delivery="2/10出荷予定",
        )

        result = load_delivery_history(ws_hist, ws_conf, cache, today=today)
        assert "12345|10" in result

    def test_excluded_from_confirming(self):
        """確認中一覧の「除外」はスキップ対象"""
        wb, ws_hist = _make_history_ws()
        _, ws_conf = _make_confirming_ws(wb)
        cache = CacheStore()

        today = datetime.date(2026, 2, 16)
        _add_confirming_row(
            ws_conf, 2,
            sent_dt=datetime.datetime(2026, 2, 15, 10, 0),
            order_date=datetime.date(2026, 2, 10),
            customer="テスト顧客",
            order_num="99999",
            detail="20",
            mfg="メーカーB",
            product="製品B",
            inquiry="除外",
            status="確認中",
        )

        result = load_delivery_history(ws_hist, ws_conf, cache, today=today)
        assert "99999|20" in result
        assert result["99999|20"] == "除外"

    def test_non_excluded_confirming_ignored(self):
        """確認中一覧の「未」はスキップ対象外"""
        wb, ws_hist = _make_history_ws()
        _, ws_conf = _make_confirming_ws(wb)
        cache = CacheStore()

        today = datetime.date(2026, 2, 16)
        _add_confirming_row(
            ws_conf, 2,
            sent_dt=datetime.datetime(2026, 2, 15, 10, 0),
            order_date=datetime.date(2026, 2, 10),
            customer="テスト顧客",
            order_num="99999",
            detail="20",
            mfg="メーカーB",
            product="製品B",
            inquiry="未",
            status="確認中",
        )

        result = load_delivery_history(ws_hist, ws_conf, cache, today=today)
        assert "99999|20" not in result


# ============================================
# SaveDeliveryHistory
# ============================================
class TestSaveDeliveryHistory:
    def test_basic_save(self):
        """基本的な書き込み"""
        wb, ws = _make_history_ws()
        exec_time = datetime.datetime(2026, 2, 16, 10, 0)

        orders = [
            HistoryRecord(
                order_date=datetime.date(2026, 2, 10),
                customer_name="テスト顧客",
                order_number="12345",
                detail_number="10",
                manufacturer_name="メーカーA",
                product_name="製品A",
                delivery_answer="2/17出荷予定",
                sender="テスト送付者",
            ),
        ]

        save_delivery_history(ws, orders, execution_time=exec_time)
        assert ws.cell(row=2, column=4).value == "12345"
        assert ws.cell(row=2, column=8).value == "2/17出荷予定"
        assert ws.cell(row=2, column=9).value == "テスト送付者"

    def test_empty_orders(self):
        """空リストでは何もしない"""
        wb, ws = _make_history_ws()
        save_delivery_history(ws, [])
        assert ws.cell(row=2, column=1).value is None

    def test_dedup(self):
        """重複チェック — 同じ注番・明細は追加しない"""
        wb, ws = _make_history_ws()
        exec_time = datetime.datetime(2026, 2, 16, 10, 0)

        # 既存データ
        _add_history_row(
            ws, 2,
            sent_dt=datetime.datetime(2026, 2, 15, 10, 0),
            order_date=datetime.date(2026, 2, 10),
            customer="テスト顧客",
            order_num="12345",
            detail="10",
            mfg="メーカーA",
            product="製品A",
            delivery="2/17出荷予定",
        )

        # 同じキーで新規追加を試みる
        orders = [
            HistoryRecord(
                order_number="12345",
                detail_number="10",
                delivery_answer="2/18出荷予定",
            ),
        ]

        save_delivery_history(ws, orders, execution_time=exec_time)
        # 2行目は既存のまま、3行目はない
        assert ws.cell(row=2, column=8).value == "2/17出荷予定"
        assert ws.cell(row=3, column=4).value is None

    def test_delivered_update(self):
        """重複+納品済み → 既存行の納期回答を更新"""
        wb, ws = _make_history_ws()
        exec_time = datetime.datetime(2026, 2, 16, 10, 0)

        _add_history_row(
            ws, 2,
            sent_dt=datetime.datetime(2026, 2, 15, 10, 0),
            order_date=datetime.date(2026, 2, 10),
            customer="テスト顧客",
            order_num="12345",
            detail="10",
            mfg="メーカーA",
            product="製品A",
            delivery="2/17出荷予定",
        )

        orders = [
            HistoryRecord(
                order_number="12345",
                detail_number="10",
                delivery_answer="納品済み",
            ),
        ]

        save_delivery_history(ws, orders, execution_time=exec_time)
        assert ws.cell(row=2, column=8).value == "納品済み"

    def test_sort_descending(self):
        """送付日時の降順でソートされる"""
        wb, ws = _make_history_ws()

        _add_history_row(
            ws, 2,
            sent_dt=datetime.datetime(2026, 2, 14, 10, 0),
            order_date=datetime.date(2026, 2, 10),
            customer="顧客A",
            order_num="11111",
            detail="10",
            mfg="メーカー",
            product="製品",
            delivery="古い",
        )

        orders = [
            HistoryRecord(
                order_date=datetime.date(2026, 2, 15),
                customer_name="顧客B",
                order_number="22222",
                detail_number="10",
                delivery_answer="新しい",
            ),
        ]

        save_delivery_history(
            ws, orders,
            execution_time=datetime.datetime(2026, 2, 16, 10, 0),
        )

        # 新しいレコードが先頭
        assert ws.cell(row=2, column=4).value == "22222"
        assert ws.cell(row=3, column=4).value == "11111"

    def test_multiple_orders(self):
        """複数レコードの書き込み"""
        wb, ws = _make_history_ws()
        exec_time = datetime.datetime(2026, 2, 16, 10, 0)

        orders = [
            HistoryRecord(
                order_number="11111", detail_number="10",
                delivery_answer="回答A",
            ),
            HistoryRecord(
                order_number="22222", detail_number="20",
                delivery_answer="回答B",
            ),
        ]

        save_delivery_history(ws, orders, execution_time=exec_time)
        # 2行書き込まれる
        assert ws.cell(row=2, column=4).value is not None
        assert ws.cell(row=3, column=4).value is not None

    def test_sender_override(self):
        """sender引数で送付者名を上書き"""
        wb, ws = _make_history_ws()

        orders = [
            HistoryRecord(
                order_number="12345", detail_number="10",
                delivery_answer="回答", sender="元の送付者",
            ),
        ]

        save_delivery_history(ws, orders, sender="上書き送付者")
        assert ws.cell(row=2, column=9).value == "上書き送付者"


# ============================================
# SaveConfirmingList
# ============================================
class TestSaveConfirmingList:
    def test_basic_save(self):
        """基本的な書き込み"""
        wb, ws = _make_confirming_ws()
        exec_time = datetime.datetime(2026, 2, 16, 10, 0)

        orders = [
            ConfirmingRecord(
                order_date=datetime.date(2026, 2, 10),
                customer_name="テスト顧客",
                order_number="12345",
                detail_number="10",
                manufacturer_name="メーカーA",
                product_name="製品A",
                inquiry_status="未",
                status="確認中",
                sender="送付者",
            ),
        ]

        save_confirming_list(ws, orders, execution_time=exec_time)
        assert ws.cell(row=2, column=4).value == "12345"
        assert ws.cell(row=2, column=8).value == "未"
        assert ws.cell(row=2, column=9).value == "確認中"

    def test_empty_orders(self):
        """空リストでは何もしない"""
        wb, ws = _make_confirming_ws()
        save_confirming_list(ws, [])
        assert ws.cell(row=2, column=1).value is None

    def test_dedup_status_update(self):
        """重複 → ステータスのみ更新"""
        wb, ws = _make_confirming_ws()
        exec_time = datetime.datetime(2026, 2, 16, 10, 0)

        _add_confirming_row(
            ws, 2,
            sent_dt=datetime.datetime(2026, 2, 15, 10, 0),
            order_date=datetime.date(2026, 2, 10),
            customer="テスト顧客",
            order_num="12345",
            detail="10",
            mfg="メーカーA",
            product="製品A",
            inquiry="済",
            status="旧ステータス",
        )

        orders = [
            ConfirmingRecord(
                order_number="12345",
                detail_number="10",
                status="新ステータス",
            ),
        ]

        save_confirming_list(ws, orders, execution_time=exec_time)
        # ステータスが更新される
        assert ws.cell(row=2, column=9).value == "新ステータス"
        # 問合せ状況は元のまま
        assert ws.cell(row=2, column=8).value == "済"

    def test_inquiry_validation(self):
        """入力規則が設定される"""
        wb, ws = _make_confirming_ws()
        exec_time = datetime.datetime(2026, 2, 16, 10, 0)

        orders = [
            ConfirmingRecord(
                order_number="12345",
                detail_number="10",
                status="確認中",
            ),
        ]

        save_confirming_list(ws, orders, execution_time=exec_time)
        # DataValidationが存在する
        assert len(ws.data_validations.dataValidation) > 0


# ============================================
# CleanConfirmingList
# ============================================
class TestCleanConfirmingList:
    def test_basic_clean(self):
        """確定伝票が確認中一覧から削除される"""
        wb, ws_hist = _make_history_ws()
        _, ws_conf = _make_confirming_ws(wb)

        _add_confirming_row(
            ws_conf, 2,
            sent_dt=datetime.datetime(2026, 2, 15, 10, 0),
            order_date=datetime.date(2026, 2, 10),
            customer="テスト顧客",
            order_num="12345",
            detail="10",
            mfg="メーカーA",
            product="製品A",
            inquiry="未",
            status="確認中",
        )
        _add_confirming_row(
            ws_conf, 3,
            sent_dt=datetime.datetime(2026, 2, 15, 10, 0),
            order_date=datetime.date(2026, 2, 10),
            customer="テスト顧客",
            order_num="99999",
            detail="20",
            mfg="メーカーB",
            product="製品B",
            inquiry="未",
            status="欠品中",
        )

        confirmed = [
            HistoryRecord(
                order_number="12345",
                detail_number="10",
                delivery_answer="2/17出荷予定",
                customer_name="テスト顧客",
            ),
        ]

        clean_confirming_list(ws_hist, ws_conf, confirmed)

        # 確認中一覧には99999だけ残る
        assert ws_conf.cell(row=2, column=4).value == "99999"
        assert ws_conf.cell(row=3, column=4).value is None

        # 送付履歴に移動されている
        assert ws_hist.cell(row=2, column=4).value == "12345"

    def test_all_confirmed(self):
        """全て確定 → 確認中一覧が空に"""
        wb, ws_hist = _make_history_ws()
        _, ws_conf = _make_confirming_ws(wb)

        _add_confirming_row(
            ws_conf, 2,
            sent_dt=datetime.datetime(2026, 2, 15, 10, 0),
            order_date=datetime.date(2026, 2, 10),
            customer="テスト顧客",
            order_num="12345",
            detail="10",
            mfg="メーカーA",
            product="製品A",
            inquiry="未",
            status="確認中",
        )

        confirmed = [
            HistoryRecord(
                order_number="12345",
                detail_number="10",
                delivery_answer="2/17出荷予定",
            ),
        ]

        clean_confirming_list(ws_hist, ws_conf, confirmed)
        assert ws_conf.cell(row=2, column=4).value is None

    def test_no_match(self):
        """マッチしない場合は何も変わらない"""
        wb, ws_hist = _make_history_ws()
        _, ws_conf = _make_confirming_ws(wb)

        _add_confirming_row(
            ws_conf, 2,
            sent_dt=datetime.datetime(2026, 2, 15, 10, 0),
            order_date=datetime.date(2026, 2, 10),
            customer="テスト顧客",
            order_num="12345",
            detail="10",
            mfg="メーカーA",
            product="製品A",
            inquiry="未",
            status="確認中",
        )

        confirmed = [
            HistoryRecord(
                order_number="99999",
                detail_number="99",
                delivery_answer="該当なし",
            ),
        ]

        clean_confirming_list(ws_hist, ws_conf, confirmed)
        # 確認中一覧はそのまま
        assert ws_conf.cell(row=2, column=4).value == "12345"

    def test_empty_confirmed(self):
        """確定リストが空なら何もしない"""
        wb, ws_hist = _make_history_ws()
        _, ws_conf = _make_confirming_ws(wb)
        clean_confirming_list(ws_hist, ws_conf, [])


# ============================================
# CleanOldHistory
# ============================================
class TestCleanOldHistory:
    def test_basic_cleanup(self):
        """古いレコードが削除される"""
        wb, ws = _make_history_ws()
        today = datetime.date(2026, 2, 16)

        # 400日前のレコード（365日超 → 削除対象）
        _add_history_row(
            ws, 2,
            sent_dt=datetime.datetime(2025, 1, 12, 10, 0),
            order_date=datetime.date(2025, 1, 5),
            customer="顧客A",
            order_num="11111",
            detail="10",
            mfg="メーカー",
            product="製品",
            delivery="回答",
        )
        # 昨日のレコード（残る）
        _add_history_row(
            ws, 3,
            sent_dt=datetime.datetime(2026, 2, 15, 10, 0),
            order_date=datetime.date(2026, 2, 10),
            customer="顧客B",
            order_num="22222",
            detail="20",
            mfg="メーカー",
            product="製品",
            delivery="回答",
        )

        deleted = clean_old_history(ws, days_to_keep=365, today=today)
        assert deleted == 1
        assert ws.cell(row=2, column=4).value == "22222"
        assert ws.cell(row=3, column=4).value is None

    def test_all_recent(self):
        """全て新しい場合は削除なし"""
        wb, ws = _make_history_ws()
        today = datetime.date(2026, 2, 16)

        _add_history_row(
            ws, 2,
            sent_dt=datetime.datetime(2026, 2, 15, 10, 0),
            order_date=datetime.date(2026, 2, 10),
            customer="顧客A",
            order_num="11111",
            detail="10",
            mfg="メーカー",
            product="製品",
            delivery="回答",
        )

        deleted = clean_old_history(ws, days_to_keep=365, today=today)
        assert deleted == 0

    def test_all_old(self):
        """全て古い場合は全削除"""
        wb, ws = _make_history_ws()
        today = datetime.date(2026, 2, 16)

        _add_history_row(
            ws, 2,
            sent_dt=datetime.datetime(2024, 12, 1, 10, 0),
            order_date=datetime.date(2024, 11, 25),
            customer="顧客A",
            order_num="11111",
            detail="10",
            mfg="メーカー",
            product="製品",
            delivery="回答",
        )

        deleted = clean_old_history(ws, days_to_keep=365, today=today)
        assert deleted == 1
        assert ws.cell(row=2, column=4).value is None

    def test_empty_sheet(self):
        """空シートでも正常動作"""
        wb, ws = _make_history_ws()
        deleted = clean_old_history(ws, today=datetime.date(2026, 2, 16))
        assert deleted == 0

    def test_no_date_kept(self):
        """日付が不明な行は残す"""
        wb, ws = _make_history_ws()
        today = datetime.date(2026, 2, 16)

        _add_history_row(
            ws, 2,
            sent_dt=None,  # 日付なし
            order_date=datetime.date(2025, 5, 25),
            customer="顧客A",
            order_num="11111",
            detail="10",
            mfg="メーカー",
            product="製品",
            delivery="回答",
        )

        deleted = clean_old_history(ws, days_to_keep=365, today=today)
        assert deleted == 0

    def test_cutoff_boundary(self):
        """ちょうどcutoff日のレコードは残る"""
        wb, ws = _make_history_ws()
        today = datetime.date(2026, 2, 16)
        cutoff = today - datetime.timedelta(days=365)

        _add_history_row(
            ws, 2,
            sent_dt=datetime.datetime(cutoff.year, cutoff.month, cutoff.day, 10, 0),
            order_date=datetime.date(2025, 2, 10),
            customer="顧客A",
            order_num="11111",
            detail="10",
            mfg="メーカー",
            product="製品",
            delivery="回答",
        )

        deleted = clean_old_history(ws, days_to_keep=365, today=today)
        assert deleted == 0


# ============================================
# CleanOldConfirmingList
# ============================================
class TestCleanOldConfirmingList:
    def test_basic_cleanup(self):
        """古いレコードが削除される"""
        wb, ws = _make_confirming_ws()
        today = datetime.date(2026, 2, 16)

        # 400日前のレコード（365日超 → 削除対象）
        _add_confirming_row(
            ws, 2,
            sent_dt=datetime.datetime(2025, 1, 12, 10, 0),
            order_date=datetime.date(2025, 1, 5),
            customer="顧客A",
            order_num="11111",
            detail="10",
            mfg="メーカー",
            product="製品",
            inquiry="未",
            status="確認中",
        )
        # 昨日のレコード
        _add_confirming_row(
            ws, 3,
            sent_dt=datetime.datetime(2026, 2, 15, 10, 0),
            order_date=datetime.date(2026, 2, 10),
            customer="顧客B",
            order_num="22222",
            detail="20",
            mfg="メーカー",
            product="製品",
            inquiry="未",
            status="欠品中",
        )

        deleted = clean_old_confirming_list(ws, days_to_keep=365, today=today)
        assert deleted == 1
        assert ws.cell(row=2, column=4).value == "22222"

    def test_empty_sheet(self):
        """空シートでもエラーなし"""
        wb, ws = _make_confirming_ws()
        deleted = clean_old_confirming_list(ws, today=datetime.date(2026, 2, 16))
        assert deleted == 0


# ============================================
# 統合テスト
# ============================================
class TestIntegration:
    def test_full_lifecycle(self):
        """ファイル初期化 → 保存 → 読み込み → クリーンアップの統合テスト"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name

        try:
            today = datetime.date(2026, 2, 16)
            exec_time = datetime.datetime(2026, 2, 16, 10, 0)

            # 1. 初期化
            wb = initialize_delivery_history(path)
            ws_hist = wb[HISTORY_SHEET_NAME]
            ws_conf = wb[CONFIRMING_SHEET_NAME]

            # 2. 確定伝票を保存
            confirmed = [
                HistoryRecord(
                    order_date=datetime.date(2026, 2, 10),
                    customer_name="テスト顧客",
                    order_number="12345",
                    detail_number="10",
                    manufacturer_name="メーカーA",
                    product_name="製品A",
                    delivery_answer="2/17出荷予定",
                ),
            ]
            save_delivery_history(ws_hist, confirmed, execution_time=exec_time)
            assert ws_hist.cell(row=2, column=4).value == "12345"

            # 3. 未確定伝票を保存
            unconfirmed = [
                ConfirmingRecord(
                    order_date=datetime.date(2026, 2, 12),
                    customer_name="テスト顧客",
                    order_number="99999",
                    detail_number="20",
                    manufacturer_name="メーカーB",
                    product_name="製品B",
                    status="確認中",
                ),
            ]
            save_confirming_list(ws_conf, unconfirmed, execution_time=exec_time)
            assert ws_conf.cell(row=2, column=4).value == "99999"

            # 4. 読み込み
            cache = CacheStore()
            sent = load_delivery_history(ws_hist, ws_conf, cache, today=today)
            assert "12345|10" in sent

            # 5. 確認中→送付履歴へ移動
            newly_confirmed = [
                HistoryRecord(
                    order_number="99999",
                    detail_number="20",
                    delivery_answer="2/18配達予定",
                    customer_name="テスト顧客",
                ),
            ]
            clean_confirming_list(ws_hist, ws_conf, newly_confirmed, execution_time=exec_time)
            assert ws_conf.cell(row=2, column=4).value is None  # 確認中から消えた

            # 送付履歴に2件
            history_keys = set()
            for r in range(2, ws_hist.max_row + 1):
                on = ws_hist.cell(row=r, column=4).value
                dn = ws_hist.cell(row=r, column=5).value
                if on:
                    history_keys.add(f"{on}|{dn}")
            assert "12345|10" in history_keys
            assert "99999|20" in history_keys

            wb.close()
        finally:
            os.unlink(path)


# ============================================
# ExtractSheetRows
# ============================================
class TestExtractSheetRows:
    def test_basic_read(self):
        """正常なデータ行を読み出す"""
        wb, ws = _make_history_ws()
        _add_history_row(
            ws, 2,
            sent_dt=datetime.datetime(2026, 2, 15, 10, 0),
            order_date=datetime.date(2026, 2, 10),
            customer="テスト顧客",
            order_num="12345",
            detail="10",
            mfg="メーカーA",
            product="製品A",
            delivery="2/17出荷予定",
        )
        _add_history_row(
            ws, 3,
            sent_dt=datetime.datetime(2026, 2, 14, 10, 0),
            order_date=datetime.date(2026, 2, 9),
            customer="顧客B",
            order_num="67890",
            detail="20",
            mfg="メーカーB",
            product="製品B",
            delivery="納品済み",
        )

        rows = extract_sheet_rows(ws, 9)
        assert len(rows) == 2
        assert rows[0][3] == "12345"
        assert rows[1][3] == "67890"

    def test_empty_sheet(self):
        """空シートでは空リスト"""
        wb, ws = _make_history_ws()
        rows = extract_sheet_rows(ws, 9)
        assert rows == []

    def test_skip_empty_order_number(self):
        """受発注伝票が空の行はスキップ"""
        wb, ws = _make_history_ws()
        # 正常な行
        _add_history_row(
            ws, 2,
            sent_dt=datetime.datetime(2026, 2, 15, 10, 0),
            order_date=datetime.date(2026, 2, 10),
            customer="テスト顧客",
            order_num="12345",
            detail="10",
            mfg="メーカーA",
            product="製品A",
            delivery="回答",
        )
        # 受発注伝票が空の行
        ws.cell(row=3, column=1).value = datetime.datetime(2026, 2, 14, 10, 0)
        ws.cell(row=3, column=3).value = "顧客X"
        # column 4 (受発注伝票) は空

        rows = extract_sheet_rows(ws, 9)
        assert len(rows) == 1

    def test_read_only_workbook(self):
        """read_only=True のワークブックで正常動作"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            wb = initialize_delivery_history(path)
            ws = wb[HISTORY_SHEET_NAME]
            _add_history_row(
                ws, 2,
                sent_dt=datetime.datetime(2026, 2, 15, 10, 0),
                order_date=datetime.date(2026, 2, 10),
                customer="テスト顧客",
                order_num="12345",
                detail="10",
                mfg="メーカーA",
                product="製品A",
                delivery="回答",
            )
            wb.save(path)
            wb.close()

            # read_only=True で読み込み
            wb_ro = load_workbook(path, read_only=True)
            ws_ro = wb_ro[HISTORY_SHEET_NAME]
            rows = extract_sheet_rows(ws_ro, 9)
            assert len(rows) == 1
            assert str(rows[0][3]).strip() == "12345"
            wb_ro.close()
        finally:
            os.unlink(path)


# ============================================
# SaveHistoryBatch
# ============================================
class TestSaveHistoryBatch:
    def test_basic_save(self):
        """基本的なバッチ保存 → テーブル・ヘッダー・データが正しい"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            exec_time = datetime.datetime(2026, 2, 16, 10, 0)
            today = datetime.date(2026, 2, 16)

            confirmed = [
                HistoryRecord(
                    order_date=datetime.date(2026, 2, 10),
                    customer_name="テスト顧客",
                    order_number="12345",
                    detail_number="10",
                    manufacturer_name="メーカーA",
                    product_name="製品A",
                    delivery_answer="2/17出荷予定",
                    sender="テスト送付者",
                ),
            ]
            confirming = [
                ConfirmingRecord(
                    order_date=datetime.date(2026, 2, 12),
                    customer_name="テスト顧客",
                    order_number="99999",
                    detail_number="20",
                    manufacturer_name="メーカーB",
                    product_name="製品B",
                    status="確認中",
                ),
            ]

            save_history_batch(
                path, [], [], confirmed, confirming,
                exec_time, "送付者", today=today,
            )

            # 検証
            wb = load_workbook(path)

            # 送付履歴シート
            ws_h = wb[HISTORY_SHEET_NAME]
            assert ws_h.cell(row=1, column=1).value == "送付日時"
            assert ws_h.cell(row=2, column=4).value == "12345"
            assert ws_h.cell(row=2, column=8).value == "2/17出荷予定"
            assert ws_h.cell(row=2, column=9).value == "送付者"
            assert len(ws_h.tables) == 1

            # 確認中一覧シート
            ws_c = wb[CONFIRMING_SHEET_NAME]
            assert ws_c.cell(row=1, column=1).value == "送付日時"
            assert ws_c.cell(row=2, column=4).value == "99999"
            assert ws_c.cell(row=2, column=8).value == "未"  # デフォルト問合せ状況
            assert ws_c.cell(row=2, column=9).value == "確認中"
            assert len(ws_c.tables) == 1

            # 入力規則
            assert len(ws_c.data_validations.dataValidation) > 0

            # 表示形式
            assert ws_h.cell(row=2, column=1).number_format == "mm/dd hh:mm"
            assert ws_h.cell(row=2, column=2).number_format == "mm/dd"

            wb.close()
        finally:
            os.unlink(path)

    def test_dedup_and_delivered_update(self):
        """重複チェック: 既存データに同じキーがあれば納品済み更新のみ"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            exec_time = datetime.datetime(2026, 2, 16, 10, 0)
            today = datetime.date(2026, 2, 16)

            # 既存の送付履歴データ
            existing_history = [
                [
                    datetime.datetime(2026, 2, 15, 10, 0),
                    datetime.date(2026, 2, 10),
                    "テスト顧客", "12345", 10,
                    "メーカーA", "製品A", "2/17出荷予定", "test",
                ],
            ]

            # 同じキーで「納品済み」→ 既存行を更新
            confirmed = [
                HistoryRecord(
                    order_number="12345",
                    detail_number="10",
                    delivery_answer="納品済み",
                ),
            ]

            save_history_batch(
                path, existing_history, [], confirmed, [],
                exec_time, today=today,
            )

            wb = load_workbook(path)
            ws_h = wb[HISTORY_SHEET_NAME]
            assert ws_h.cell(row=2, column=4).value == "12345"
            assert ws_h.cell(row=2, column=8).value == "納品済み"
            # 重複なので2行にはならない
            assert ws_h.cell(row=3, column=4).value is None
            wb.close()
        finally:
            os.unlink(path)

    def test_confirming_to_history_move(self):
        """確認中一覧から送付履歴への移動"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            exec_time = datetime.datetime(2026, 2, 16, 10, 0)
            today = datetime.date(2026, 2, 16)

            # 既存の確認中一覧
            existing_confirming = [
                [
                    datetime.datetime(2026, 2, 15, 10, 0),
                    datetime.date(2026, 2, 10),
                    "テスト顧客", "12345", 10,
                    "メーカーA", "製品A", "未", "確認中", None, "test",
                ],
                [
                    datetime.datetime(2026, 2, 15, 10, 0),
                    datetime.date(2026, 2, 10),
                    "テスト顧客", "99999", 20,
                    "メーカーB", "製品B", "未", "欠品中", None, "test",
                ],
            ]

            # 12345が確定
            confirmed = [
                HistoryRecord(
                    order_number="12345",
                    detail_number="10",
                    delivery_answer="2/17出荷予定",
                    customer_name="テスト顧客",
                ),
            ]

            save_history_batch(
                path, [], existing_confirming, confirmed, [],
                exec_time, today=today,
            )

            wb = load_workbook(path)

            # 送付履歴に12345が移動
            ws_h = wb[HISTORY_SHEET_NAME]
            assert ws_h.cell(row=2, column=4).value == "12345"

            # 確認中一覧には99999だけ残る
            ws_c = wb[CONFIRMING_SHEET_NAME]
            assert ws_c.cell(row=2, column=4).value == "99999"
            assert ws_c.cell(row=3, column=4).value is None

            wb.close()
        finally:
            os.unlink(path)

    def test_old_records_removed(self):
        """古いレコードが除去される"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            exec_time = datetime.datetime(2026, 2, 16, 10, 0)
            today = datetime.date(2026, 2, 16)

            # 400日前のレコード（365日超 → 削除対象）
            old_history = [
                [
                    datetime.datetime(2025, 1, 12, 10, 0),
                    datetime.date(2025, 1, 5),
                    "顧客A", "11111", 10,
                    "メーカー", "製品", "回答", "test",
                ],
            ]
            # 昨日のレコード（残る）
            new_history = [
                [
                    datetime.datetime(2026, 2, 15, 10, 0),
                    datetime.date(2026, 2, 10),
                    "顧客B", "22222", 20,
                    "メーカー", "製品", "回答", "test",
                ],
            ]

            save_history_batch(
                path, old_history + new_history, [], [], [],
                exec_time, today=today,
            )

            wb = load_workbook(path)
            ws_h = wb[HISTORY_SHEET_NAME]
            # 古いレコードは削除され、新しいレコードだけ残る
            assert ws_h.cell(row=2, column=4).value == "22222"
            assert ws_h.cell(row=3, column=4).value is None
            wb.close()
        finally:
            os.unlink(path)

    def test_sort_descending(self):
        """送付日時の降順でソートされる"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            exec_time = datetime.datetime(2026, 2, 16, 10, 0)
            today = datetime.date(2026, 2, 16)

            existing_history = [
                [
                    datetime.datetime(2026, 2, 14, 10, 0),
                    datetime.date(2026, 2, 10),
                    "顧客A", "11111", 10,
                    "メーカー", "製品", "古い", "test",
                ],
            ]

            confirmed = [
                HistoryRecord(
                    order_date=datetime.date(2026, 2, 15),
                    customer_name="顧客B",
                    order_number="22222",
                    detail_number="10",
                    delivery_answer="新しい",
                ),
            ]

            save_history_batch(
                path, existing_history, [], confirmed, [],
                exec_time, today=today,
            )

            wb = load_workbook(path)
            ws_h = wb[HISTORY_SHEET_NAME]
            # 新しいレコード（exec_time=2/16）が先頭
            assert ws_h.cell(row=2, column=4).value == "22222"
            assert ws_h.cell(row=3, column=4).value == "11111"
            wb.close()
        finally:
            os.unlink(path)

    def test_confirming_status_update(self):
        """確認中一覧の重複時はステータスのみ更新"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            exec_time = datetime.datetime(2026, 2, 16, 10, 0)
            today = datetime.date(2026, 2, 16)

            existing_confirming = [
                [
                    datetime.datetime(2026, 2, 15, 10, 0),
                    datetime.date(2026, 2, 10),
                    "テスト顧客", "12345", 10,
                    "メーカーA", "製品A", "済", "旧ステータス", None, "test",
                ],
            ]

            new_confirming = [
                ConfirmingRecord(
                    order_number="12345",
                    detail_number="10",
                    status="新ステータス",
                ),
            ]

            save_history_batch(
                path, [], existing_confirming, [], new_confirming,
                exec_time, today=today,
            )

            wb = load_workbook(path)
            ws_c = wb[CONFIRMING_SHEET_NAME]
            assert ws_c.cell(row=2, column=9).value == "新ステータス"
            # 問合せ状況は元のまま
            assert ws_c.cell(row=2, column=8).value == "済"
            wb.close()
        finally:
            os.unlink(path)

    def test_empty_input(self):
        """全て空でもファイルが正常に生成される"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            exec_time = datetime.datetime(2026, 2, 16, 10, 0)
            today = datetime.date(2026, 2, 16)

            save_history_batch(
                path, [], [], [], [],
                exec_time, today=today,
            )

            wb = load_workbook(path)
            assert HISTORY_SHEET_NAME in wb.sheetnames
            assert CONFIRMING_SHEET_NAME in wb.sheetnames
            ws_h = wb[HISTORY_SHEET_NAME]
            assert ws_h.cell(row=1, column=1).value == "送付日時"
            assert ws_h.cell(row=2, column=1).value is None
            wb.close()
        finally:
            os.unlink(path)


# ============================================
# save_history_batch: deleted_keys（明細削除の自動除去）
# ============================================
class TestSaveHistoryBatchDeletedKeys:
    """SAPで明細削除された伝票を確認中一覧から自動除去するテスト"""

    def _make_confirming_row(
        self, order_num, detail_num, inquiry="未", status="確認中",
        order_delivery=None,
    ):
        """確認中一覧の1行データを作成するヘルパー"""
        return [
            datetime.datetime(2026, 3, 1, 10, 0),
            datetime.date(2026, 2, 25),
            "テスト顧客", order_num, detail_num,
            "メーカーA", "製品A", inquiry, status, order_delivery, "test",
        ]

    def test_deleted_key_removed_from_confirming(self):
        """パターン1: 明細削除 + 確認中一覧にキーあり → 除去される"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            exec_time = datetime.datetime(2026, 3, 6, 10, 0)
            today = datetime.date(2026, 3, 6)

            existing_confirming = [
                self._make_confirming_row("12345", 10),
                self._make_confirming_row("99999", 20),
            ]

            # 12345|10 が明細削除
            deleted = {"12345|10"}

            save_history_batch(
                path, [], existing_confirming, [], [],
                exec_time, today=today, deleted_keys=deleted,
            )

            wb = load_workbook(path)
            ws_c = wb[CONFIRMING_SHEET_NAME]
            # 99999|20 だけ残る
            assert ws_c.cell(row=2, column=4).value == "99999"
            assert ws_c.cell(row=3, column=4).value is None

            # 送付履歴には移動していない
            ws_h = wb[HISTORY_SHEET_NAME]
            assert ws_h.cell(row=2, column=4).value is None
            wb.close()
        finally:
            os.unlink(path)

    def test_deleted_key_not_in_confirming(self):
        """パターン2: 明細削除 + 確認中一覧にキーなし → 何も起きない"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            exec_time = datetime.datetime(2026, 3, 6, 10, 0)
            today = datetime.date(2026, 3, 6)

            existing_confirming = [
                self._make_confirming_row("99999", 20),
            ]

            # 存在しないキーを削除
            deleted = {"77777|10"}

            save_history_batch(
                path, [], existing_confirming, [], [],
                exec_time, today=today, deleted_keys=deleted,
            )

            wb = load_workbook(path)
            ws_c = wb[CONFIRMING_SHEET_NAME]
            # 既存行はそのまま残る
            assert ws_c.cell(row=2, column=4).value == "99999"
            wb.close()
        finally:
            os.unlink(path)

    def test_deleted_keys_only_no_confirmed_no_confirming(self):
        """パターン3: deleted_keysのみ（confirmed/confirming空） → 除去される"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            exec_time = datetime.datetime(2026, 3, 6, 10, 0)
            today = datetime.date(2026, 3, 6)

            existing_confirming = [
                self._make_confirming_row("12345", 10),
                self._make_confirming_row("99999", 20),
            ]

            # confirmed/confirming は空、deleted_keys のみ
            deleted = {"12345|10"}

            save_history_batch(
                path, [], existing_confirming, [], [],
                exec_time, today=today, deleted_keys=deleted,
            )

            wb = load_workbook(path)
            ws_c = wb[CONFIRMING_SHEET_NAME]
            assert ws_c.cell(row=2, column=4).value == "99999"
            assert ws_c.cell(row=3, column=4).value is None
            wb.close()
        finally:
            os.unlink(path)

    def test_key_in_both_confirmed_and_deleted(self):
        """パターン4: 同じキーがconfirmed_keysとdeleted_keysの両方 → ステップ1で処理、二重処理なし"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            exec_time = datetime.datetime(2026, 3, 6, 10, 0)
            today = datetime.date(2026, 3, 6)

            existing_confirming = [
                self._make_confirming_row("12345", 10),
                self._make_confirming_row("99999", 20),
            ]

            # 12345|10 が confirmed（送付履歴へ移動）かつ deleted にも含まれる
            confirmed = [
                HistoryRecord(
                    order_number="12345",
                    detail_number="10",
                    delivery_answer="2/17出荷予定",
                    customer_name="テスト顧客",
                ),
            ]
            deleted = {"12345|10"}

            save_history_batch(
                path, [], existing_confirming, confirmed, [],
                exec_time, today=today, deleted_keys=deleted,
            )

            wb = load_workbook(path)

            # 送付履歴に移動されている（ステップ1で処理）
            ws_h = wb[HISTORY_SHEET_NAME]
            assert ws_h.cell(row=2, column=4).value == "12345"
            assert ws_h.cell(row=2, column=8).value == "2/17出荷予定"

            # 確認中一覧には99999だけ残る
            ws_c = wb[CONFIRMING_SHEET_NAME]
            assert ws_c.cell(row=2, column=4).value == "99999"
            assert ws_c.cell(row=3, column=4).value is None
            wb.close()
        finally:
            os.unlink(path)

    def test_non_deleted_rejection_reason_not_collected(self):
        """パターン5: deleted_keysが空set → 確認中一覧は変更されない"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            exec_time = datetime.datetime(2026, 3, 6, 10, 0)
            today = datetime.date(2026, 3, 6)

            existing_confirming = [
                self._make_confirming_row("12345", 10),
            ]

            # 空set（明細削除以外は収集されない）
            save_history_batch(
                path, [], existing_confirming, [], [],
                exec_time, today=today, deleted_keys=set(),
            )

            wb = load_workbook(path)
            ws_c = wb[CONFIRMING_SHEET_NAME]
            assert ws_c.cell(row=2, column=4).value == "12345"
            wb.close()
        finally:
            os.unlink(path)

    def test_unexpected_document_type_still_removed(self):
        """パターン6: 伝票タイプに関係なくキーが一致すれば除去される"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            exec_time = datetime.datetime(2026, 3, 6, 10, 0)
            today = datetime.date(2026, 3, 6)

            # 確認中一覧にはキーだけで伝票タイプ情報はない
            existing_confirming = [
                self._make_confirming_row("12345", 10),
            ]

            # 伝票タイプが何であれ、キーが一致すれば除去される
            deleted = {"12345|10"}

            save_history_batch(
                path, [], existing_confirming, [], [],
                exec_time, today=today, deleted_keys=deleted,
            )

            wb = load_workbook(path)
            ws_c = wb[CONFIRMING_SHEET_NAME]
            assert ws_c.cell(row=2, column=4).value is None
            wb.close()
        finally:
            os.unlink(path)

    def test_excluded_inquiry_status_still_removed(self):
        """パターン7: 問合せ状況が「除外」の確認中エントリも明細削除で除去される"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            exec_time = datetime.datetime(2026, 3, 6, 10, 0)
            today = datetime.date(2026, 3, 6)

            existing_confirming = [
                self._make_confirming_row("12345", 10, inquiry="除外"),
                self._make_confirming_row("99999", 20),
            ]

            deleted = {"12345|10"}

            save_history_batch(
                path, [], existing_confirming, [], [],
                exec_time, today=today, deleted_keys=deleted,
            )

            wb = load_workbook(path)
            ws_c = wb[CONFIRMING_SHEET_NAME]
            assert ws_c.cell(row=2, column=4).value == "99999"
            assert ws_c.cell(row=3, column=4).value is None
            wb.close()
        finally:
            os.unlink(path)

    def test_user_entered_delivery_date_still_removed(self):
        """パターン8: ユーザー手入力の受注納期がある確認中エントリも明細削除で除去される"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            exec_time = datetime.datetime(2026, 3, 6, 10, 0)
            today = datetime.date(2026, 3, 6)

            existing_confirming = [
                self._make_confirming_row(
                    "12345", 10,
                    order_delivery=datetime.date(2026, 3, 15),
                ),
                self._make_confirming_row("99999", 20),
            ]

            deleted = {"12345|10"}

            save_history_batch(
                path, [], existing_confirming, [], [],
                exec_time, today=today, deleted_keys=deleted,
            )

            wb = load_workbook(path)
            ws_c = wb[CONFIRMING_SHEET_NAME]
            assert ws_c.cell(row=2, column=4).value == "99999"
            assert ws_c.cell(row=3, column=4).value is None
            wb.close()
        finally:
            os.unlink(path)
