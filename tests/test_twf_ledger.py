"""twf_ledger（TWF2026 社内手配管理表）のテスト

期間限定機能。TWF運用終了時はこのテストファイルごと削除してよい。
"""

from pathlib import Path

import pytest
from openpyxl import load_workbook

from nouki_kaitou.models import BranchSettings, CacheStore, OrderRow
from nouki_kaitou import twf_ledger
from nouki_kaitou.twf_ledger import (
    DELIVERY_FAIL_ABORT_RATIO,
    MANUFACTURER_UNKNOWN,
    STATUS_CHOICES,
    STATUS_DEFAULT,
    LedgerRow,
    apply_carryover,
    build_delivery_context,
    build_ledger_rows,
    classify_tehai,
    format_twf_no,
    parse_quantity,
    read_existing_status,
    resolve_manufacturer,
    write_ledger,
)


def _order(
    order_number="GL001",
    detail_number="10",
    document_type="【受注】直送販売",
    storage_place="転送中（直送用）",
    comment_detail="TWFNo.003243　新成（株）様",
    rep_name="柏原　賢人",
    customer_name="京葉帝酸（株）",
    manufacturer_name="ベルテクノ（株）",
    product_name="溶接機",
    quantity="2",
    item_group_code="0001",
):
    return OrderRow(
        order_number=order_number,
        detail_number=detail_number,
        document_type=document_type,
        storage_place=storage_place,
        comment_detail=comment_detail,
        rep_name=rep_name,
        customer_name=customer_name,
        manufacturer_name=manufacturer_name,
        product_name=product_name,
        quantity=quantity,
        item_group_code=item_group_code,
    )


# ============================================
# 手配区分の判定
# ============================================
def test_classify_chokusou():
    o = _order(document_type="【受注】直送販売", storage_place="転送中（直送用）")
    assert classify_tehai(o) == "直送"


def test_classify_himozuki():
    o = _order(document_type="【受注】直送販売", storage_place="GL0")
    assert classify_tehai(o) == "紐付き"


def test_classify_zaiko():
    o = _order(document_type="【受注】在庫販売", storage_place="")
    assert classify_tehai(o) == "在庫販売"


# ============================================
# メーカー名の補完（在庫販売で「名称」が空のとき品目Groupで補完）
# ============================================
def test_resolve_keeps_existing_name():
    """名称があればマスターに関係なくそのまま（直送・紐付き）。"""
    o = _order(manufacturer_name="日鉄溶接工業（株）", item_group_code="0316")
    assert resolve_manufacturer(o, {"0316": "別メーカー"}) == "日鉄溶接工業（株）"


def test_resolve_backfills_blank_from_master():
    """名称が空（在庫販売）→ 品目Groupマスターで補完。"""
    o = _order(document_type="【受注】在庫販売", storage_place="関東商品センター",
               manufacturer_name="", item_group_code="0075")
    assert resolve_manufacturer(o, {"0075": "ダイヘン"}) == "ダイヘン"


def test_resolve_unknown_when_not_in_master():
    """名称が空でマスターにも無ければ「（要確認）」。"""
    o = _order(document_type="【受注】在庫販売", storage_place="関東商品センター",
               manufacturer_name="", item_group_code="1774")
    assert resolve_manufacturer(o, {"0075": "ダイヘン"}) == MANUFACTURER_UNKNOWN


def test_resolve_no_master_leaves_blank():
    """マスター未提供（None）なら補完せず空欄のまま。"""
    o = _order(document_type="【受注】在庫販売", manufacturer_name="",
               item_group_code="0075")
    assert resolve_manufacturer(o, None) == ""


def test_build_applies_backfill():
    orders = [
        _order(order_number="GLZ1", document_type="【受注】在庫販売",
               storage_place="関東商品センター", manufacturer_name="",
               item_group_code="0075",
               comment_detail="TWFNo.001　甲社様"),
        _order(order_number="GLZ2", document_type="【受注】在庫販売",
               storage_place="関東商品センター", manufacturer_name="",
               item_group_code="1774",
               comment_detail="TWFNo.002　乙社様"),
    ]
    rows = build_ledger_rows(orders, {"0075": "ダイヘン"})
    by_order = {r.order_number: r for r in rows}
    assert by_order["GLZ1"].manufacturer == "ダイヘン"
    assert by_order["GLZ2"].manufacturer == MANUFACTURER_UNKNOWN


# ============================================
# 台帳行の構築
# ============================================
def test_build_filters_non_twf():
    orders = [
        _order(order_number="GL001", comment_detail="TWFNo.003243　新成（株）様"),
        _order(order_number="GL999", comment_detail="通常受注（TWFなし）"),
    ]
    rows = build_ledger_rows(orders)
    assert len(rows) == 1
    assert rows[0].order_number == "GL001"
    assert rows[0].twf_no == "003243"
    assert rows[0].customer == "新成（株）様"
    assert rows[0].status == STATUS_DEFAULT


def test_build_propagates_within_order():
    """同一注番の入れ忘れ明細にも TWF No.・お客様名が引き継がれる。"""
    orders = [
        _order(order_number="GL010", detail_number="10",
               comment_detail="TWFNo.004041　金安工業様"),
        _order(order_number="GL010", detail_number="20",
               comment_detail="（TWF記載なし・入れ忘れ）"),
    ]
    rows = build_ledger_rows(orders)
    assert len(rows) == 2
    by_detail = {r.detail_number: r for r in rows}
    assert by_detail["20"].twf_no == "004041"
    assert by_detail["20"].customer == "金安工業様"


def test_build_question_mark_number_goes_to_number_column():
    """？はTWF№側へ。お客様名には？が付かない（4行の改善）。"""
    orders = [
        _order(order_number="GLQ1", detail_number="10",
               comment_detail="TWFNo.0014？？　㈱ハイプラン様"),
    ]
    rows = build_ledger_rows(orders)
    assert rows[0].twf_no == "0014？？"   # 数字は半角・？は全角保持
    assert rows[0].customer == "㈱ハイプラン様"


# ============================================
# 桁揃え（台帳のみ・6桁ゼロ埋め）
# ============================================
def test_format_twf_no_pads_short_pure_digits():
    assert format_twf_no("0014") == "000014"
    assert format_twf_no("2") == "000002"


def test_format_twf_no_keeps_six_or_more():
    assert format_twf_no("000022") == "000022"
    assert format_twf_no("0005017") == "0005017"  # 7桁はそのまま


def test_format_twf_no_skips_question_unknown_blank():
    assert format_twf_no("0014？？") == "0014？？"   # 全角？含みは埋めない
    assert format_twf_no("002？？？") == "002？？？"
    assert format_twf_no("不明") == "不明"
    assert format_twf_no("") == ""


def test_build_applies_zero_pad_display():
    orders = [
        _order(order_number="GLP1", comment_detail="TWFNo.0014　甲社様"),
    ]
    rows = build_ledger_rows(orders)
    assert rows[0].twf_no == "000014"


def test_build_sort_order():
    """TWF No.昇順 → 番号なし・不明は末尾。"""
    orders = [
        _order(order_number="A", comment_detail="TWFNo.005　甲社様"),
        _order(order_number="B", comment_detail="TWFNo.001　乙社様"),
        _order(order_number="C", comment_detail="TWFNo.不明　丙社様"),
    ]
    rows = build_ledger_rows(orders)
    # 台帳表示は6桁ゼロ埋め。不明は末尾。
    assert [r.twf_no for r in rows] == ["000001", "000005", "不明"]


# ============================================
# 引き継ぎ
# ============================================
def test_carryover_round_trip(tmp_path):
    orders = [
        _order(order_number="GL100", detail_number="10",
               comment_detail="TWFNo.003243　新成（株）様"),
    ]
    rows = build_ledger_rows(orders)
    # 手入力を模してステータス・備考を設定 → 書き出し
    rows[0].status = "メーカー発注済"
    rows[0].note = "6/20入荷予定"
    first = write_ledger(rows, tmp_path / "first.xlsx")

    # 再生成（手入力はデフォルトに戻る）→ 引き継ぎ適用
    rows2 = build_ledger_rows(orders)
    assert rows2[0].status == STATUS_DEFAULT
    existing = read_existing_status(first)
    applied = apply_carryover(rows2, existing)
    assert applied == 1
    assert rows2[0].status == "メーカー発注済"
    assert rows2[0].note == "6/20入荷予定"


def test_read_existing_missing_file_returns_empty(tmp_path):
    assert read_existing_status(tmp_path / "nope.xlsx") == {}


# ============================================
# 数量の数値化（Excelの文字列保存エラーマーク対策）
# ============================================
def test_parse_quantity_integer():
    assert parse_quantity("10") == 10
    assert isinstance(parse_quantity("10"), int)


def test_parse_quantity_whole_decimal_becomes_int():
    assert parse_quantity("1.00") == 1
    assert isinstance(parse_quantity("1.00"), int)
    assert parse_quantity("800.00") == 800
    assert isinstance(parse_quantity("800.00"), int)


def test_parse_quantity_real_decimal():
    assert parse_quantity("1.5") == 1.5
    assert isinstance(parse_quantity("1.5"), float)


def test_parse_quantity_comma_thousands():
    assert parse_quantity("1,000") == 1000


def test_parse_quantity_blank_is_none():
    assert parse_quantity("") is None
    assert parse_quantity("   ") is None


def test_parse_quantity_non_numeric_stays_string():
    assert parse_quantity("未定") == "未定"


def test_write_quantity_is_numeric_cell(tmp_path):
    """数量セルが数値型で書き込まれ、非数値・空でも落ちない。"""
    orders = [
        _order(order_number="GLN1", quantity="1.00",
               comment_detail="TWFNo.001　甲社様"),
        _order(order_number="GLN2", quantity="10",
               comment_detail="TWFNo.002　乙社様"),
        _order(order_number="GLN3", quantity="",
               comment_detail="TWFNo.003　丙社様"),
        _order(order_number="GLN4", quantity="未定",
               comment_detail="TWFNo.004　丁社様"),
    ]
    rows = build_ledger_rows(orders)
    out = write_ledger(rows, tmp_path / "qty.xlsx")
    wb = load_workbook(out)
    ws = wb.active
    hdr = {ws.cell(8, c).value: c for c in range(2, 13)}
    qcol = hdr["数量"]
    # 行はTWF No.順（001..004）
    vals = {ws.cell(r, hdr["注番"]).value: ws.cell(r, qcol).value
            for r in range(9, ws.max_row + 1)}
    assert vals["GLN1"] == 1 and isinstance(vals["GLN1"], int)
    assert vals["GLN2"] == 10 and isinstance(vals["GLN2"], int)
    assert vals["GLN3"] is None          # 空欄
    assert vals["GLN4"] == "未定"        # 非数値は文字列


# ============================================
# Excel 出力の構造
# ============================================
def test_write_creates_table_and_dropdown(tmp_path):
    orders = [
        _order(order_number="GL200", comment_detail="TWFNo.001　甲社様"),
        _order(order_number="GL201", comment_detail="TWFNo.002　乙社様"),
    ]
    rows = build_ledger_rows(orders)
    out = write_ledger(rows, tmp_path / "ledger.xlsx")
    assert out.exists()

    wb = load_workbook(out)
    ws = wb.active
    # テーブルが1つ存在し、B8 始まり
    assert len(ws.tables) == 1
    tbl = list(ws.tables.values())[0]
    assert tbl.ref.startswith("B8")
    # ステータスのドロップダウンが存在
    dvs = list(ws.data_validations.dataValidation)
    assert any(dv.type == "list" for dv in dvs)
    # ヘッダーはB8（左A列・上の余白を確保している）
    assert ws["B8"].value == "TWF No."
    assert ws["A8"].value is None  # 左余白
    # ヘッダー固定
    assert ws.freeze_panes == "A9"


def test_column_layout_dealer_user_rep(tmp_path):
    """受注先（販売店）追加・ユーザー名・担当者は最右端の列構成。"""
    orders = [
        _order(order_number="GLD1", customer_name="京葉帝酸（株）",
               rep_name="首藤　佑哉",
               comment_detail="TWFNo.001　藤原溶接様"),
    ]
    rows = build_ledger_rows(orders)
    out = write_ledger(rows, tmp_path / "layout.xlsx")
    wb = load_workbook(out)
    ws = wb.active
    headers = [ws.cell(8, c).value for c in range(2, 14)]
    assert headers == [
        "TWF No.", "注番", "明細", "受注先（販売店）", "ユーザー名",
        "メーカー名", "品名", "数量", "手配区分", "ステータス", "備考", "担当者",
    ]
    hdr = {ws.cell(8, c).value: c for c in range(2, 14)}
    # 受注先＝販売店、ユーザー名＝エンドユーザーが別列に入る
    assert ws.cell(9, hdr["受注先（販売店）"]).value == "京葉帝酸（株）"
    assert ws.cell(9, hdr["ユーザー名"]).value == "藤原溶接様"
    # 担当者は最右端（備考の後）
    assert hdr["担当者"] > hdr["備考"]
    assert ws.cell(9, hdr["担当者"]).value == "首藤　佑哉"


def test_status_choices_count():
    # 5項目に統一（メーカー手配済み/在庫計上済み/保留/その他 ＋ 初期値の未着手）
    assert STATUS_CHOICES == [
        "未着手", "メーカー手配済み", "在庫計上済み", "保留", "その他",
    ]
    assert STATUS_DEFAULT == "未着手"
    # 廃止した選択肢が残っていないこと
    for removed in ("一部手配", "対象外・キャンセル", "メーカー発注済", "完了"):
        assert removed not in STATUS_CHOICES


# ============================================
# 回答納期（delivery_ctx）— Codex review 対応の回帰防止
# ============================================
def _twf_orders(n):
    """TWF明細を n 件作る（注番・TWF No. を一意に）。"""
    return [
        _order(order_number=f"O{i:03d}", detail_number="10",
               comment_detail=f"TWFNo.{i + 1:06d}　甲社様")
        for i in range(n)
    ]


def test_delivery_ctx_populates_answer(monkeypatch):
    """delivery_ctx を渡すと回答納期が入る。"""
    monkeypatch.setattr(
        twf_ledger, "build_report_row",
        lambda *a, **k: (None, "7月2日配達予定"),
    )
    orders = _twf_orders(1)
    ctx = (object(), {}, None)  # cache は非Noneであればよい（build_report_row は差替）
    rows = build_ledger_rows(orders, None, delivery_ctx=ctx)
    assert rows[0].delivery_answer == "7月2日配達予定"


def test_delivery_ctx_none_keeps_blank_legacy():
    """delivery_ctx を渡さない（None）と従来どおり全件空欄（既存挙動非破壊）。"""
    orders = _twf_orders(2)
    rows = build_ledger_rows(orders)            # 第3引数省略
    assert all(r.delivery_answer == "" for r in rows)
    rows2 = build_ledger_rows(orders, None, delivery_ctx=None)
    assert all(r.delivery_answer == "" for r in rows2)


def _fail_first(n_fail):
    """最初の n_fail 回だけ例外、それ以降は正常値を返す build_report_row スタブ。"""
    state = {"n": 0}

    def stub(*a, **k):
        state["n"] += 1
        if state["n"] <= n_fail:
            raise ValueError("計算失敗（テスト）")
        return (None, "7月2日配達予定")

    return stub


def test_delivery_fail_below_threshold_continues(monkeypatch, capsys):
    """49%失敗（<50%）→ 停止せず、該当明細のみ空欄で続行。"""
    monkeypatch.setattr(twf_ledger, "build_report_row", _fail_first(49))
    orders = _twf_orders(100)
    rows = build_ledger_rows(orders, None, delivery_ctx=(object(), {}, None))
    assert len(rows) == 100
    assert sum(1 for r in rows if r.delivery_answer == "") == 49
    assert "失敗" in capsys.readouterr().out  # 警告は出る


def test_delivery_fail_at_threshold_aborts(monkeypatch):
    """50%失敗（=閾値）→ SystemExit で取込中止。"""
    monkeypatch.setattr(twf_ledger, "build_report_row", _fail_first(50))
    orders = _twf_orders(100)
    with pytest.raises(SystemExit):
        build_ledger_rows(orders, None, delivery_ctx=(object(), {}, None))


def test_delivery_fail_all_aborts(monkeypatch):
    """100%失敗 → SystemExit で取込中止。"""
    monkeypatch.setattr(twf_ledger, "build_report_row", _fail_first(10))
    orders = _twf_orders(10)
    with pytest.raises(SystemExit):
        build_ledger_rows(orders, None, delivery_ctx=(object(), {}, None))


def test_abort_ratio_constant():
    """閾値は 0.5（50%以上で停止）。"""
    assert DELIVERY_FAIL_ABORT_RATIO == 0.5


def test_delivery_context_missing_file_returns_none(tmp_path, capsys):
    """必須マスターのファイルが無い → None（全件空欄）＋警告。"""
    ctx = build_delivery_context([], {}, masters_dir=tmp_path)  # 空ディレクトリ
    assert ctx is None
    assert "回答納期を計算できません" in capsys.readouterr().out


def test_delivery_context_empty_master_returns_none(monkeypatch, capsys):
    """ファイルは開けるが必須キャッシュが空（シート欠落・空ファイル）→ None。"""
    def fake_load(path, **kwargs):
        # 送付履歴は optional（None）、メーカー/顧客は「開ける」体で非Noneを返す
        if Path(path).name == twf_ledger.HISTORY_NAME:
            return None
        return object()

    monkeypatch.setattr(twf_ledger, "_safe_load_workbook", fake_load)
    monkeypatch.setattr(twf_ledger, "build_all_caches", lambda *a, **k: CacheStore())
    monkeypatch.setattr(twf_ledger, "load_holidays", lambda wb: {})
    monkeypatch.setattr(twf_ledger, "load_branch_settings", lambda *a, **k: BranchSettings())

    ctx = build_delivery_context([], {}, masters_dir="dummy")
    assert ctx is None
    assert "回答納期を計算できません" in capsys.readouterr().out


def test_execution_time_fixed_across_rows(monkeypatch):
    """P1-1: execution_time / today は全行で同一（行ごとの datetime.now() でない）。"""
    seen = []

    def stub(row, cache, holidays, branch, execution_time, force_delivered, today):
        seen.append((execution_time, force_delivered, today))
        return (None, "7月2日配達予定")

    monkeypatch.setattr(twf_ledger, "build_report_row", stub)
    orders = _twf_orders(5)
    build_ledger_rows(orders, None, delivery_ctx=(object(), {}, None))

    assert len(seen) == 5
    exec_times = {s[0] for s in seen}
    todays = {s[2] for s in seen}
    assert len(exec_times) == 1          # 全行で同一の実行時刻
    assert len(todays) == 1              # 全行で同一の基準日
    assert next(iter(exec_times)) is not None
    assert next(iter(todays)) is not None
    assert all(s[1] is False for s in seen)  # force_delivered は False 固定
