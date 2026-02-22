"""excel_writer.py のユニットテスト

Excel出力・書式設定の全5関数 + ヘルパー関数を網羅的にテスト。
"""

import datetime

import pytest
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from nouki_kaitou.excel_writer import (
    _classify_delivery_color,
    _COLOR_BUNNO,
    _COLOR_CONFIRMING,
    _COLOR_DEFAULT,
    _COLOR_DELIVER_LATER,
    _COLOR_DELIVER_SOON,
    _COLOR_DELIVERED,
    _COLOR_DONE,
    _COLOR_OTHER_BRANCH,
    _COLOR_OTHER_PLAN,
    _COLOR_PICKED_UP,
    _COLOR_PICKUP_PLAN,
    _COLOR_SCHEDULING,
    _COLOR_SHIP_LATER,
    _COLOR_SHIP_SOON,
    _COLOR_STOCKOUT,
    _COLOR_STOCKOUT_PARTIAL,
    _COLOR_WORK,
    _HEADER_LABELS,
    check_same_date_in_bunno,
    color_confirming_list,
    copy_data_row,
    create_header,
    format_report,
)
from nouki_kaitou.models import (
    BranchSettings,
    BunnoEntry,
    ReportRow,
    StockoutEntry,
    TrackingEntry,
)


# ============================================
# CreateHeader
# ============================================
class TestCreateHeader:
    def test_title_row(self):
        """行1にタイトルが設定される"""
        wb = Workbook()
        ws = wb.active
        create_header(ws, "テスト株式会社")
        assert ws["A1"].value == "納　期　回　答　書"

    def test_customer_name(self):
        """行4に顧客名が設定される"""
        wb = Workbook()
        ws = wb.active
        create_header(ws, "テスト株式会社")
        assert "テスト株式会社 御中" in ws["B4"].value

    def test_customer_with_rep(self):
        """担当者名がある場合"""
        wb = Workbook()
        ws = wb.active
        create_header(ws, "テスト株式会社", rep_name="田中")
        assert "ご担当：田中 様" in ws["B4"].value

    def test_customer_without_rep(self):
        """担当者名がない場合"""
        wb = Workbook()
        ws = wb.active
        create_header(ws, "テスト株式会社", rep_name="")
        assert ws["B4"].value == "テスト株式会社 御中"

    def test_customer_other_rep(self):
        """__OTHER__の場合は担当者名なし"""
        wb = Workbook()
        ws = wb.active
        create_header(ws, "テスト株式会社", rep_name="__OTHER__")
        assert ws["B4"].value == "テスト株式会社 御中"

    def test_issue_date(self):
        """発行日が指定される"""
        wb = Workbook()
        ws = wb.active
        d = datetime.date(2026, 2, 16)
        create_header(ws, "テスト株式会社", issue_date=d)
        assert "2026年2月16日(月)" in ws.cell(row=4, column=12).value

    def test_weekday_names(self):
        """曜日名の正確性"""
        wb = Workbook()
        ws = wb.active
        # 2026-02-16 は月曜日
        d = datetime.date(2026, 2, 16)
        create_header(ws, "テスト", issue_date=d)
        assert "(月)" in ws.cell(row=4, column=12).value

    def test_header_labels(self):
        """行6にヘッダーラベルが設定される"""
        wb = Workbook()
        ws = wb.active
        create_header(ws, "テスト")
        for col_idx, label in enumerate(_HEADER_LABELS, start=1):
            assert ws.cell(row=6, column=col_idx).value == label

    def test_row_heights(self):
        """行の高さ設定"""
        wb = Workbook()
        ws = wb.active
        create_header(ws, "テスト")
        assert ws.row_dimensions[1].height == 55
        assert ws.row_dimensions[2].height == 4
        assert ws.row_dimensions[6].height == 28

    def test_title_fill(self):
        """タイトル行の背景色"""
        wb = Workbook()
        ws = wb.active
        create_header(ws, "テスト")
        assert ws["A1"].fill.start_color.rgb == "00142846"

    def test_accent_line(self):
        """行2のアクセントライン色"""
        wb = Workbook()
        ws = wb.active
        create_header(ws, "テスト")
        assert ws.cell(row=2, column=1).fill.start_color.rgb == "00B49646"


# ============================================
# CopyDataRow
# ============================================
class TestCopyDataRow:
    def _make_row(self, **kwargs) -> ReportRow:
        return ReportRow(**kwargs)

    def test_basic_write(self):
        """基本的なデータ書き込み"""
        wb = Workbook()
        ws = wb.active
        row = self._make_row(
            registration_date=datetime.date(2026, 1, 15),
            customer_contact="田中",
            customer_order_number="PO-001",
            manufacturer_name="テストメーカー",
            product_name="テスト製品A",
            quantity="100",
            unit_price="500",
            net_amount="50000",
            delivery_answer="2/20出荷予定",
            delivery_place="東京支店",
            remarks="備考テスト",
            order_number="12345-10",
        )
        copy_data_row(ws, 7, row)

        assert ws.cell(row=7, column=1).value == datetime.date(2026, 1, 15)
        assert ws.cell(row=7, column=2).value == "田中"
        assert ws.cell(row=7, column=5).value == "テスト製品A"
        assert ws.cell(row=7, column=9).value == "2/20出荷予定"
        assert ws.cell(row=7, column=12).value == "12345-10"

    def test_even_row_fill(self):
        """偶数行に背景色が設定される"""
        wb = Workbook()
        ws = wb.active
        row = self._make_row()
        copy_data_row(ws, 8, row)  # 偶数行
        assert ws.cell(row=8, column=1).fill.start_color.rgb == "00E1EBF8"

    def test_odd_row_no_fill(self):
        """奇数行には背景色が設定されない"""
        wb = Workbook()
        ws = wb.active
        row = self._make_row()
        copy_data_row(ws, 7, row)  # 奇数行
        # デフォルトのfill（なし）
        assert ws.cell(row=7, column=1).fill.patternType is None

    def test_customer_order_number_numeric(self):
        """数字のみの貴社注番はint型で書き込まれる"""
        wb = Workbook()
        ws = wb.active
        row = self._make_row(customer_order_number="12345")
        copy_data_row(ws, 7, row)
        val = ws.cell(row=7, column=3).value
        assert val == 12345
        assert isinstance(val, int)

    def test_customer_order_number_alphanumeric(self):
        """文字混じりの貴社注番はそのまま文字列"""
        wb = Workbook()
        ws = wb.active
        row = self._make_row(customer_order_number="PO-001")
        copy_data_row(ws, 7, row)
        val = ws.cell(row=7, column=3).value
        assert val == "PO-001"
        assert isinstance(val, str)

    def test_customer_order_number_empty(self):
        """空の貴社注番はそのまま"""
        wb = Workbook()
        ws = wb.active
        row = self._make_row(customer_order_number="")
        copy_data_row(ws, 7, row)
        assert ws.cell(row=7, column=3).value == ""

    def test_all_columns_written(self):
        """全12列が書き込まれる"""
        wb = Workbook()
        ws = wb.active
        row = self._make_row(
            registration_date=datetime.date(2026, 1, 1),
            customer_contact="CC",
            customer_order_number="CON",
            manufacturer_name="MN",
            product_name="PN",
            quantity="Q",
            unit_price="UP",
            net_amount="NA",
            delivery_answer="DA",
            delivery_place="DP",
            remarks="RM",
            order_number="ON",
        )
        copy_data_row(ws, 7, row)
        values = [ws.cell(row=7, column=c).value for c in range(1, 13)]
        assert values == [
            datetime.date(2026, 1, 1), "CC", "CON", "MN", "PN",
            "Q", "UP", "NA", "DA", "DP", "RM", "ON",
        ]


# ============================================
# CheckSameDateInBunno
# ============================================
class TestCheckSameDateInBunno:
    def test_no_entries(self):
        assert check_same_date_in_bunno([]) is False

    def test_single_entry(self):
        entries = [BunnoEntry(quantity="100", date_str="3/10")]
        assert check_same_date_in_bunno(entries) is False

    def test_different_dates(self):
        entries = [
            BunnoEntry(quantity="100", date_str="3/10"),
            BunnoEntry(quantity="200", date_str="3/15"),
        ]
        assert check_same_date_in_bunno(entries) is False

    def test_same_dates(self):
        entries = [
            BunnoEntry(quantity="100", date_str="3/10", location="東京"),
            BunnoEntry(quantity="200", date_str="3/10", location="大阪"),
        ]
        assert check_same_date_in_bunno(entries) is True

    def test_mitei_excluded(self):
        """「未定」は比較対象外"""
        entries = [
            BunnoEntry(quantity="100", date_str="未定"),
            BunnoEntry(quantity="200", date_str="未定"),
        ]
        assert check_same_date_in_bunno(entries) is False

    def test_yotei_excluded(self):
        """「○旬予定」は比較対象外"""
        entries = [
            BunnoEntry(quantity="100", date_str="3月上旬予定"),
            BunnoEntry(quantity="200", date_str="3月上旬予定"),
        ]
        assert check_same_date_in_bunno(entries) is False

    def test_mixed_same_date(self):
        """未定と確定日が混在し、確定日が重複"""
        entries = [
            BunnoEntry(quantity="100", date_str="未定"),
            BunnoEntry(quantity="200", date_str="3/10"),
            BunnoEntry(quantity="300", date_str="3/10"),
        ]
        assert check_same_date_in_bunno(entries) is True


# ============================================
# _classify_delivery_color
# ============================================
class TestClassifyDeliveryColor:
    """納期回答列の色分類テスト"""

    TODAY = datetime.date(2026, 2, 16)
    TOMORROW = datetime.date(2026, 2, 17)
    DAY_AFTER = datetime.date(2026, 2, 18)

    @staticmethod
    def _dummy_extract_date(value: str):
        """テスト用の日付抽出関数"""
        import re
        m = re.search(r"(\d{1,2})/(\d{1,2})", value)
        if m:
            month, day = int(m.group(1)), int(m.group(2))
            return datetime.date(2026, month, day)
        return None

    def _classify(self, value):
        return _classify_delivery_color(
            value, self.TODAY, self.TOMORROW, self.DAY_AFTER,
            self._dummy_extract_date,
        )

    def test_delivered(self):
        assert self._classify("納品済み") == _COLOR_DELIVERED

    def test_stockout(self):
        assert self._classify("欠品中") == _COLOR_STOCKOUT

    def test_stockout_partial(self):
        assert self._classify("2/20出荷予定（欠品）") == _COLOR_STOCKOUT_PARTIAL

    def test_bunno(self):
        assert self._classify("分納") == _COLOR_BUNNO

    def test_confirming(self):
        assert self._classify("確認中") == _COLOR_CONFIRMING

    def test_scheduling(self):
        assert self._classify("日程調整中") == _COLOR_SCHEDULING

    def test_work(self):
        assert self._classify("作業") == _COLOR_WORK
        assert self._classify("作業中") == _COLOR_WORK

    def test_other_branch(self):
        assert self._classify("他拠点より出荷") == _COLOR_OTHER_BRANCH

    def test_picked_up(self):
        assert self._classify("引取済み") == _COLOR_PICKED_UP

    def test_done_generic(self):
        """「○○済み」（引取済み以外）"""
        assert self._classify("手配済み") == _COLOR_DONE
        assert self._classify("出荷済") == _COLOR_DONE

    def test_ship_today(self):
        """本日出荷予定"""
        assert self._classify("2/16出荷予定") == _COLOR_SHIP_SOON

    def test_ship_tomorrow(self):
        """明日出荷予定"""
        assert self._classify("2/17出荷予定") == _COLOR_SHIP_SOON

    def test_ship_later(self):
        """それ以降の出荷予定"""
        assert self._classify("2/20出荷予定") == _COLOR_SHIP_LATER

    def test_deliver_tomorrow(self):
        """明日配達予定"""
        assert self._classify("2/17配達予定") == _COLOR_DELIVER_SOON

    def test_deliver_day_after(self):
        """明後日配達予定"""
        assert self._classify("2/18配達予定") == _COLOR_DELIVER_SOON

    def test_deliver_later(self):
        """それ以降の配達予定"""
        assert self._classify("2/25配達予定") == _COLOR_DELIVER_LATER

    def test_pickup_plan(self):
        """引取予定"""
        assert self._classify("2/20引取予定") == _COLOR_PICKUP_PLAN

    def test_other_plan_with_date(self):
        """日付あり・その他の予定"""
        assert self._classify("2/20入荷予定") == _COLOR_OTHER_PLAN

    def test_other_plan_no_date(self):
        """日付なし・予定"""
        assert self._classify("3月上旬入荷予定") == _COLOR_OTHER_PLAN

    def test_default(self):
        """いずれにも当てはまらない"""
        assert self._classify("不明な値") == _COLOR_DEFAULT

    def test_empty(self):
        assert self._classify("") == _COLOR_DEFAULT


# ============================================
# FormatReport
# ============================================
class TestFormatReport:
    def _setup_ws_with_data(self, num_rows=3):
        """テスト用WSにヘッダー+データ行をセットアップ"""
        wb = Workbook()
        ws = wb.active
        create_header(ws, "テスト顧客")

        for i in range(num_rows):
            row_num = 7 + i
            row = ReportRow(
                registration_date=datetime.date(2026, 1, 15),
                customer_contact="田中",
                manufacturer_name="メーカーA",
                product_name="製品A",
                quantity="100",
                unit_price=500 if i != 1 else "確認中",
                net_amount=50000 if i != 1 else "確認中",
                delivery_answer=["2/17出荷予定", "確認中", "納品済み"][i],
                order_number=f"12345-{i + 1}0",
            )
            copy_data_row(ws, row_num, row)

        return wb, ws

    def test_basic_format(self):
        """基本書式適用が成功する"""
        wb, ws = self._setup_ws_with_data()
        # エラーなく完了するか
        format_report(ws, 9, today=datetime.date(2026, 2, 16))

    def test_delivery_color_applied(self):
        """納期回答列に色が設定される"""
        wb, ws = self._setup_ws_with_data()
        format_report(ws, 9, today=datetime.date(2026, 2, 16))

        # I列（9列目）に色が設定されている
        for row in range(7, 10):
            cell = ws.cell(row=row, column=9)
            assert cell.fill.patternType == "solid"

    def test_price_confirming_red(self):
        """「確認中」の単価・金額が赤字になる"""
        wb, ws = self._setup_ws_with_data()
        format_report(ws, 9, today=datetime.date(2026, 2, 16))

        # 行8（2番目）はunit_price="確認中"
        cell_g = ws.cell(row=8, column=7)
        assert cell_g.font.color.rgb == "00B41E1E"

    def test_tax_note(self):
        """税抜き注記が設定される"""
        wb, ws = self._setup_ws_with_data()
        format_report(ws, 9, today=datetime.date(2026, 2, 16))
        assert ws.cell(row=10, column=7).value == "※表示金額は税抜きです"

    def test_borders_applied(self):
        """罫線が設定される"""
        wb, ws = self._setup_ws_with_data()
        format_report(ws, 9, today=datetime.date(2026, 2, 16))

        # ヘッダー行（行6）とデータ行（行7-9）に罫線
        for row in range(6, 10):
            for col in range(1, 13):
                cell = ws.cell(row=row, column=col)
                assert cell.border.left.style is not None

    def test_number_format_quantity(self):
        """F列に数値書式が設定される"""
        wb, ws = self._setup_ws_with_data()
        format_report(ws, 9, today=datetime.date(2026, 2, 16))
        assert ws.cell(row=7, column=6).number_format == "#,##0"

    def test_print_settings(self):
        """印刷設定"""
        wb, ws = self._setup_ws_with_data()
        format_report(ws, 9, today=datetime.date(2026, 2, 16))
        assert ws.page_setup.orientation == "landscape"
        assert ws.print_title_rows == "$1:$6"

    def test_row_height_data(self):
        """データ行の高さ"""
        wb, ws = self._setup_ws_with_data()
        format_report(ws, 9, today=datetime.date(2026, 2, 16))
        for row in range(7, 10):
            assert ws.row_dimensions[row].height == 22

    def test_info_section_header(self):
        """ご連絡事項ヘッダーと署名"""
        wb, ws = self._setup_ws_with_data()
        branch = BranchSettings(name="京葉営業所")
        format_report(
            ws, 9, branch=branch, today=datetime.date(2026, 2, 16)
        )
        # 署名が存在する
        found_sign = False
        for row in range(10, ws.max_row + 1):
            val = ws.cell(row=row, column=9).value
            if val and "マツモト産業" in str(val):
                found_sign = True
                break
        assert found_sign, "署名が見つからない"

    def test_info_section_branch_name(self):
        """営業所名が署名セクションに表示される"""
        wb, ws = self._setup_ws_with_data()
        branch = BranchSettings(name="京葉営業所")
        format_report(
            ws, 9, branch=branch, today=datetime.date(2026, 2, 16)
        )
        found_branch = False
        for row in range(10, ws.max_row + 1):
            val = ws.cell(row=row, column=9).value
            if val and "京葉営業所" in str(val):
                found_branch = True
                break
        assert found_branch, "営業所名が見つからない"


# ============================================
# FormatReport — 送り状セクション
# ============================================
class TestFormatReportTracking:
    def test_tracking_section(self):
        """送り状情報セクションが書き込まれる"""
        wb = Workbook()
        ws = wb.active
        create_header(ws, "テスト")
        row = ReportRow(delivery_answer="2/17出荷予定")
        copy_data_row(ws, 7, row)

        tracking_info = [
            ("メーカーA", "製品A", "100",
             TrackingEntry(carrier_name="ヤマト運輸", tracking_number="1234567890123")),
        ]
        format_report(
            ws, 7,
            tracking_info_list=tracking_info,
            today=datetime.date(2026, 2, 16),
        )

        # 送り状番号がどこかに表示される
        found = False
        for row_num in range(1, ws.max_row + 1):
            val = ws.cell(row=row_num, column=1).value
            if val and "1234567890123" in str(val):
                found = True
                break
        assert found, "送り状番号が見つからない"

    def test_tracking_hyperlink(self):
        """ヤマト運輸はハイパーリンク付き"""
        wb = Workbook()
        ws = wb.active
        create_header(ws, "テスト")
        row = ReportRow(delivery_answer="2/17出荷予定")
        copy_data_row(ws, 7, row)

        tracking_info = [
            ("メーカーA", "製品A", "100",
             TrackingEntry(carrier_name="ヤマト運輸", tracking_number="1234567890123")),
        ]
        format_report(
            ws, 7,
            tracking_info_list=tracking_info,
            today=datetime.date(2026, 2, 16),
        )

        # HYPERLINK関数でリンクが設定されていることを確認
        found_link = False
        for row_num in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_num, column=1)
            val = str(cell.value or "")
            if "HYPERLINK" in val and "kuronekoyamato" in val:
                found_link = True
                break
        assert found_link, "ヤマトの追跡リンクが見つからない"

    def test_indirect_tracking(self):
        """間接追跡（トナミ等）は追跡ページリンクが別行"""
        wb = Workbook()
        ws = wb.active
        create_header(ws, "テスト")
        row = ReportRow(delivery_answer="2/17出荷予定")
        copy_data_row(ws, 7, row)

        tracking_info = [
            ("メーカーA", "製品A", "100",
             TrackingEntry(carrier_name="トナミ運輸", tracking_number="1234567890")),
        ]
        format_report(
            ws, 7,
            tracking_info_list=tracking_info,
            today=datetime.date(2026, 2, 16),
        )

        # HYPERLINK関数で間接追跡リンクが設定されていることを確認
        found = False
        for row_num in range(1, ws.max_row + 1):
            val = str(ws.cell(row=row_num, column=1).value or "")
            if "追跡ページ" in val and "HYPERLINK" in val:
                found = True
                break
        assert found, "間接追跡リンクの案内が見つからない"


# ============================================
# FormatReport — 欠品セクション
# ============================================
class TestFormatReportStockout:
    def test_stockout_section(self):
        """欠品情報セクションが書き込まれる"""
        wb = Workbook()
        ws = wb.active
        create_header(ws, "テスト")
        row = ReportRow(delivery_answer="欠品中")
        copy_data_row(ws, 7, row)

        stockout_list = [
            StockoutEntry(
                manufacturer_name="メーカーX",
                product_name="製品X",
                quantity="50",
                approx_delivery="3月上旬入荷予定",
            ),
        ]
        format_report(
            ws, 7,
            stockout_info_list=stockout_list,
            today=datetime.date(2026, 2, 16),
        )

        found = False
        for row_num in range(1, ws.max_row + 1):
            val = ws.cell(row=row_num, column=1).value
            if val and "欠品中" in str(val):
                found = True
                break
        assert found, "欠品情報が見つからない"

    def test_stockout_approx_delivery(self):
        """概算入荷日の表示"""
        wb = Workbook()
        ws = wb.active
        create_header(ws, "テスト")
        row = ReportRow(delivery_answer="欠品中")
        copy_data_row(ws, 7, row)

        stockout_list = [
            StockoutEntry(
                manufacturer_name="メーカーX",
                product_name="製品X",
                quantity="50",
                approx_delivery="3月上旬入荷予定",
            ),
        ]
        format_report(
            ws, 7,
            stockout_info_list=stockout_list,
            today=datetime.date(2026, 2, 16),
        )

        found = False
        for row_num in range(1, ws.max_row + 1):
            val = ws.cell(row=row_num, column=1).value
            if val and "3月上旬入荷予定" in str(val):
                found = True
                break
        assert found, "概算入荷日が見つからない"

    def test_stockout_no_approx(self):
        """概算入荷日がない場合"""
        wb = Workbook()
        ws = wb.active
        create_header(ws, "テスト")
        row = ReportRow(delivery_answer="欠品中")
        copy_data_row(ws, 7, row)

        stockout_list = [
            StockoutEntry(
                manufacturer_name="メーカーX",
                product_name="製品X",
                quantity="50",
            ),
        ]
        format_report(
            ws, 7,
            stockout_info_list=stockout_list,
            today=datetime.date(2026, 2, 16),
        )

        found = False
        for row_num in range(1, ws.max_row + 1):
            val = ws.cell(row=row_num, column=1).value
            if val and "入荷次第ご連絡" in str(val):
                found = True
                break
        assert found, "入荷次第の文言が見つからない"


# ============================================
# FormatReport — 分納セクション
# ============================================
class TestFormatReportBunno:
    def test_bunno_section(self):
        """分納情報セクションが書き込まれる"""
        wb = Workbook()
        ws = wb.active
        create_header(ws, "テスト")
        row = ReportRow(delivery_answer="分納")
        copy_data_row(ws, 7, row)

        bunno_info = [
            {
                "manufacturer": "メーカーA",
                "product": "製品A",
                "quantity": "300",
                "entries": [
                    BunnoEntry(quantity="200m", date_str="3/10"),
                    BunnoEntry(quantity="100m", date_str="3/15"),
                ],
                "calc_details": [
                    ("200m", "3/10", "", "3/10出荷予定"),
                    ("100m", "3/15", "", "3/15出荷予定"),
                ],
                "is_ship_rule": True,
                "days_to_add": 0,
                "order_number": "12345",
                "detail_number": "10",
                "is_rosenbin": False,
            },
        ]
        format_report(
            ws, 7,
            bunno_info_list=bunno_info,
            today=datetime.date(2026, 2, 16),
        )

        found = False
        for row_num in range(1, ws.max_row + 1):
            val = ws.cell(row=row_num, column=1).value
            if val and "分納" in str(val):
                found = True
                break
        assert found, "分納セクションが見つからない"

    def test_bunno_circled_numbers(self):
        """丸数字が使われる"""
        wb = Workbook()
        ws = wb.active
        create_header(ws, "テスト")
        row = ReportRow(delivery_answer="分納")
        copy_data_row(ws, 7, row)

        bunno_info = [
            {
                "manufacturer": "メーカーA",
                "product": "製品A",
                "quantity": "300",
                "entries": [
                    BunnoEntry(quantity="200m", date_str="3/10"),
                    BunnoEntry(quantity="100m", date_str="3/15"),
                ],
                "calc_details": [
                    ("200m", "3/10", "", "3/10出荷予定"),
                    ("100m", "3/15", "", "3/15出荷予定"),
                ],
                "is_ship_rule": True,
                "days_to_add": 0,
                "order_number": "12345",
                "detail_number": "10",
                "is_rosenbin": False,
            },
        ]
        format_report(
            ws, 7,
            bunno_info_list=bunno_info,
            today=datetime.date(2026, 2, 16),
        )

        # ①と②が見つかる
        found_1 = False
        found_2 = False
        for row_num in range(1, ws.max_row + 1):
            val = str(ws.cell(row=row_num, column=1).value or "")
            if "①" in val:
                found_1 = True
            if "②" in val:
                found_2 = True
        assert found_1, "①が見つからない"
        assert found_2, "②が見つからない"

    def test_bunno_same_date_note(self):
        """同じ日付の場合の注釈"""
        wb = Workbook()
        ws = wb.active
        create_header(ws, "テスト")
        row = ReportRow(delivery_answer="分納")
        copy_data_row(ws, 7, row)

        bunno_info = [
            {
                "manufacturer": "メーカーA",
                "product": "製品A",
                "quantity": "300",
                "entries": [
                    BunnoEntry(quantity="200m", date_str="3/10", location="東京"),
                    BunnoEntry(quantity="100m", date_str="3/10", location="大阪"),
                ],
                "calc_details": [
                    ("200m", "3/10", "東京", "3/10出荷予定"),
                    ("100m", "3/10", "大阪", "3/10出荷予定"),
                ],
                "is_ship_rule": True,
                "days_to_add": 0,
                "order_number": "12345",
                "detail_number": "10",
                "is_rosenbin": False,
            },
        ]
        format_report(
            ws, 7,
            bunno_info_list=bunno_info,
            today=datetime.date(2026, 2, 16),
        )

        found = False
        for row_num in range(1, ws.max_row + 1):
            val = ws.cell(row=row_num, column=1).value
            if val and "別々の場所" in str(val):
                found = True
                break
        assert found, "同日注釈が見つからない"


# ============================================
# FormatReport — 分納完了セクション
# ============================================
class TestFormatReportBunnoCompleted:
    def test_bunno_completed_section(self):
        """分納完了通知が書き込まれる"""
        wb = Workbook()
        ws = wb.active
        create_header(ws, "テスト")
        row = ReportRow(delivery_answer="納品済み")
        copy_data_row(ws, 7, row)

        completed = [("メーカーA", "製品A", "300")]
        format_report(
            ws, 7,
            bunno_completed_list=completed,
            today=datetime.date(2026, 2, 16),
        )

        found = False
        for row_num in range(1, ws.max_row + 1):
            val = ws.cell(row=row_num, column=1).value
            if val and "全て出荷が完了" in str(val):
                found = True
                break
        assert found, "分納完了通知が見つからない"


# ============================================
# ColorConfirmingList
# ============================================
class TestColorConfirmingList:
    def _make_ws(self, rows_data):
        """テスト用の確認中一覧シートを作成

        rows_data: list of (sent_date, col2..col8, status)
        """
        wb = Workbook()
        ws = wb.active
        # ヘッダー行
        headers = ["送付日時", "受注日", "顧客名", "注番", "明細",
                    "メーカー", "品名", "問合せ", "ステータス"]
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx).value = h

        for row_idx, data in enumerate(rows_data, 2):
            ws.cell(row=row_idx, column=1).value = data[0]  # 送付日時
            ws.cell(row=row_idx, column=9).value = data[1]  # ステータス

        return wb, ws

    def test_ship_done(self):
        """出荷完了 → 赤系背景"""
        today = datetime.date(2026, 2, 16)
        wb, ws = self._make_ws([
            (datetime.date(2026, 2, 15), "出荷完了"),
        ])
        color_confirming_list(ws, today=today)
        assert ws.cell(row=2, column=1).fill.start_color.rgb == "00FFB4B4"

    def test_stockout(self):
        """欠品中 → 薄紫"""
        today = datetime.date(2026, 2, 16)
        wb, ws = self._make_ws([
            (datetime.date(2026, 2, 15), "欠品中"),
        ])
        color_confirming_list(ws, today=today)
        assert ws.cell(row=2, column=1).fill.start_color.rgb == "00DCC8FF"

    def test_bunno(self):
        """分納 → 薄青"""
        today = datetime.date(2026, 2, 16)
        wb, ws = self._make_ws([
            (datetime.date(2026, 2, 15), "分納"),
        ])
        color_confirming_list(ws, today=today)
        assert ws.cell(row=2, column=1).fill.start_color.rgb == "00C8DCFF"

    def test_partial(self):
        """一部処理済み → 薄緑"""
        today = datetime.date(2026, 2, 16)
        wb, ws = self._make_ws([
            (datetime.date(2026, 2, 15), "一部処理済み"),
        ])
        color_confirming_list(ws, today=today)
        assert ws.cell(row=2, column=1).fill.start_color.rgb == "00C8FFD4"

    def test_week_old(self):
        """1週間以上経過 → オレンジ"""
        today = datetime.date(2026, 2, 16)
        wb, ws = self._make_ws([
            (datetime.date(2026, 2, 9), ""),  # 7日前
        ])
        color_confirming_list(ws, today=today)
        assert ws.cell(row=2, column=1).fill.start_color.rgb == "00FFC896"

    def test_three_days_old(self):
        """3日以上経過 → 黄色"""
        today = datetime.date(2026, 2, 16)
        wb, ws = self._make_ws([
            (datetime.date(2026, 2, 13), ""),  # 3日前
        ])
        color_confirming_list(ws, today=today)
        assert ws.cell(row=2, column=1).fill.start_color.rgb == "00FFFFC8"

    def test_recent_no_color(self):
        """2日前 → 色なし"""
        today = datetime.date(2026, 2, 16)
        wb, ws = self._make_ws([
            (datetime.date(2026, 2, 14), ""),  # 2日前
        ])
        color_confirming_list(ws, today=today)
        assert ws.cell(row=2, column=1).fill.patternType is None

    def test_status_priority_over_age(self):
        """ステータスが日数より優先"""
        today = datetime.date(2026, 2, 16)
        wb, ws = self._make_ws([
            (datetime.date(2026, 2, 1), "出荷完了"),  # 15日前だが出荷完了
        ])
        color_confirming_list(ws, today=today)
        # 出荷完了の色になる（1週間以上のオレンジではなく）
        assert ws.cell(row=2, column=1).fill.start_color.rgb == "00FFB4B4"

    def test_all_columns_colored(self):
        """色は全列に適用される"""
        today = datetime.date(2026, 2, 16)
        wb, ws = self._make_ws([
            (datetime.date(2026, 2, 15), "出荷完了"),
        ])
        color_confirming_list(ws, today=today)
        for col in range(1, 10):
            assert ws.cell(row=2, column=col).fill.start_color.rgb == "00FFB4B4"

    def test_empty_sheet(self):
        """データなしでもエラーにならない"""
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1).value = "送付日時"
        color_confirming_list(ws, today=datetime.date(2026, 2, 16))

    def test_datetime_value(self):
        """datetimeオブジェクトでも正しく処理"""
        today = datetime.date(2026, 2, 16)
        wb, ws = self._make_ws([
            (datetime.datetime(2026, 2, 9, 10, 30), ""),  # 7日前
        ])
        color_confirming_list(ws, today=today)
        assert ws.cell(row=2, column=1).fill.start_color.rgb == "00FFC896"

    def test_multiple_rows(self):
        """複数行の色分け"""
        today = datetime.date(2026, 2, 16)
        wb, ws = self._make_ws([
            (datetime.date(2026, 2, 15), "出荷完了"),
            (datetime.date(2026, 2, 15), "欠品中"),
            (datetime.date(2026, 2, 9), ""),
            (datetime.date(2026, 2, 14), ""),
        ])
        color_confirming_list(ws, today=today)
        assert ws.cell(row=2, column=1).fill.start_color.rgb == "00FFB4B4"  # 出荷完了
        assert ws.cell(row=3, column=1).fill.start_color.rgb == "00DCC8FF"  # 欠品中
        assert ws.cell(row=4, column=1).fill.start_color.rgb == "00FFC896"  # 1週間以上
        assert ws.cell(row=5, column=1).fill.patternType is None             # 2日前


# ============================================
# 統合テスト
# ============================================
class TestIntegration:
    def test_full_report_generation(self):
        """ヘッダー + データ + 書式設定の統合テスト"""
        wb = Workbook()
        ws = wb.active
        branch = BranchSettings(name="京葉営業所")
        today = datetime.date(2026, 2, 16)

        # ヘッダー
        create_header(ws, "テスト顧客", issue_date=today)

        # データ行
        rows = [
            ReportRow(
                registration_date=datetime.date(2026, 1, 10),
                customer_contact="田中",
                manufacturer_name="メーカーA",
                product_name="製品A",
                quantity="100",
                unit_price=500,
                net_amount=50000,
                delivery_answer="2/17出荷予定",
                order_number="12345-10",
            ),
            ReportRow(
                registration_date=datetime.date(2026, 1, 12),
                customer_contact="佐藤",
                manufacturer_name="メーカーB",
                product_name="製品B",
                quantity="200",
                unit_price="確認中",
                net_amount="確認中",
                delivery_answer="確認中",
                order_number="12346-10",
            ),
        ]
        for i, r in enumerate(rows):
            copy_data_row(ws, 7 + i, r)

        # 書式設定
        tracking_info = [
            ("メーカーA", "製品A", "100",
             TrackingEntry(carrier_name="佐川急便", tracking_number="9876543210")),
        ]
        stockout_list = [
            StockoutEntry(
                manufacturer_name="メーカーC",
                product_name="製品C",
                quantity="30",
                approx_delivery="3月中旬",
            ),
        ]

        format_report(
            ws, 8,
            branch=branch,
            tracking_info_list=tracking_info,
            stockout_info_list=stockout_list,
            today=today,
        )

        # 検証
        assert ws["A1"].value == "納　期　回　答　書"
        assert ws.cell(row=7, column=9).value == "2/17出荷予定"
        assert ws.cell(row=8, column=9).value == "確認中"

    def test_empty_report(self):
        """データなしでもエラーにならない"""
        wb = Workbook()
        ws = wb.active
        create_header(ws, "テスト")
        format_report(ws, 6, today=datetime.date(2026, 2, 16))
