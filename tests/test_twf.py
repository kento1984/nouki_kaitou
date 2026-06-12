"""twf モジュール（TWF展示会受注の専用回答書）のテスト

期間限定機能。機能削除時はこのテストファイルごと削除してよい。
"""

import datetime
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from nouki_kaitou.models import (
    BranchSettings,
    CacheStore,
    OrderRow,
    ReportResult,
    StockoutEntry,
)
from nouki_kaitou.report_generator import create_delivery_report
from nouki_kaitou.twf import (
    TWF_END_DATE,
    TWF_FILENAME_TAG,
    TWF_NOTICE_EMAIL,
    TWF_NOTICE_EXCEL,
    TWF_REPORT_TITLE,
    TWF_SHEET_PREFIX,
    TWF_THANKS_EXCEL,
    TwfDetailInfo,
    build_twf_info_map,
    collect_twf_orders,
    is_twf_active,
    is_twf_comment,
    merge_email_input,
    normalize_twf_text,
    parse_twf_comment,
    remove_twf_text,
    twf_sort_key,
)


# ============================================
# テスト用ヘルパー（test_report_generator.py と同パターン）
# ============================================
TODAY = datetime.date(2026, 2, 16)
EXEC_TIME = datetime.datetime(2026, 2, 16, 10, 0, 0)
BRANCH = BranchSettings(
    name="京葉営業所",
    default_cutoff=15,
    base_center="関東商品センター",
)


def _make_cache(**kwargs) -> CacheStore:
    return CacheStore(
        mfg_name=kwargs.get("mfg_name", {"D01": "ダイヘン"}),
        mfg_days=kwargs.get("mfg_days", {"D01": 2}),
        cust_days=kwargs.get("cust_days", {}),
        cust_retention=kwargs.get("cust_retention", {}),
        cust_route=kwargs.get("cust_route", {}),
        confirm=kwargs.get("confirm", {}),
        storage=kwargs.get("storage", {}),
    )


def _make_row(**kwargs) -> OrderRow:
    defaults = {
        "order_number": "1000001",
        "detail_number": "10",
        "document_type": "【受注】在庫販売",
        "customer_name": "テスト商事",
        "product_name": "溶接棒 3.2mm",
        "ship_status": "未処理",
        "quantity": "100",
        "unit_price": "500",
        "net_amount": "50000",
        "manufacturer_name": "ダイヘン",
        "storage_place": "関東商品センター",
        "customer_order_number": "PO-001",
        "customer_contact": "田中",
        "comment_detail": "",
        "comment_external": "",
        "comment_internal": "",
        "rejection_reason": "",
        "ship_to_name": "テスト商事",
        "registration_date": TODAY,
        "time_value": "10:00:00",
        "order_delivery_date": datetime.date(2026, 2, 20),
        "specified_delivery_date": None,
        "item_group_code": "D01",
    }
    defaults.update(kwargs)
    return OrderRow(**defaults)


# ============================================
# normalize_twf_text
# ============================================
class TestNormalizeTwfText:
    def test_fullwidth_to_halfwidth(self):
        assert normalize_twf_text("ＴＷＦＮｏ．００３２４３") == "TWFNO.003243"

    def test_lowercase_to_upper(self):
        assert normalize_twf_text("twfno.003243") == "TWFNO.003243"

    def test_remove_spaces(self):
        assert normalize_twf_text("TWF No. 003243") == "TWFNO.003243"

    def test_remove_fullwidth_spaces(self):
        assert normalize_twf_text("TWF　No．003243　新成（株）") == "TWFNO.003243新成(株)"

    def test_numero_sign(self):
        # NFKC正規化で № → No に展開される
        assert normalize_twf_text("TWF№003243") == "TWFNO003243"

    def test_empty(self):
        assert normalize_twf_text("") == ""


# ============================================
# is_twf_comment
# ============================================
class TestIsTwfComment:
    """正規化後に「TWFNO」を含むか（数字の有無は問わない）"""

    @pytest.mark.parametrize("comment", [
        "TWFNo.003243　新成（株）",
        "ＴＷＦＮｏ．００３２４３　新成（株）",
        "TWF No. 003243",
        "TWF　No　3243",
        "twf no.3243",
        "TWF№003243",
        "ＴＷＦ№３２４３",
        "TWFNo.",          # 番号の書き忘れ・省略も検知
        "TWFNO3243",
        "分納:50個 2/20、TWFNo.003243",  # 他コメントとの併記
    ])
    def test_positive(self, comment):
        assert is_twf_comment(comment) is True

    @pytest.mark.parametrize("comment", [
        "ＴＷＦ特価",          # 展示会前の特価先行受注（対象外）
        "TWF特価",
        "ダイヘン期間限定ＷＦキャンペーン",
        "TWF",                # Noなし単独は対象外
        "欠品中 3月上旬入荷予定",
        "",
    ])
    def test_negative(self, comment):
        assert is_twf_comment(comment) is False


# ============================================
# collect_twf_orders（注番単位の伝播）
# ============================================
class TestCollectTwfOrders:
    def test_propagates_to_all_details_of_order(self):
        """同一注番の1明細にTWF記載 → 注番全体が対象"""
        orders = [
            _make_row(order_number="100", detail_number="10",
                      comment_detail="TWFNo.003243　新成（株）"),
            _make_row(order_number="100", detail_number="20",
                      comment_detail=""),  # 入れ忘れ
            _make_row(order_number="200", detail_number="10",
                      comment_detail=""),
        ]
        twf_orders, detected = collect_twf_orders(orders)
        assert twf_orders == {"100"}
        assert len(detected) == 1
        assert detected[0] == ("100", "10", "TWFNo.003243　新成（株）")

    def test_multiple_orders(self):
        orders = [
            _make_row(order_number="100", comment_detail="TWFNo.001"),
            _make_row(order_number="200", comment_detail="ＴＷＦ№002"),
            _make_row(order_number="300", comment_detail="ＴＷＦ特価"),  # 対象外
        ]
        twf_orders, detected = collect_twf_orders(orders)
        assert twf_orders == {"100", "200"}
        assert len(detected) == 2

    def test_empty(self):
        twf_orders, detected = collect_twf_orders([])
        assert twf_orders == set()
        assert detected == []


# ============================================
# remove_twf_text
# ============================================
class TestRemoveTwfText:
    def test_remove_to_end_of_line(self):
        """TWF記載から行末（得意先名含む）まで除去"""
        assert remove_twf_text("TWFNo.003243　新成（株）") == ""

    def test_keep_preceding_text(self):
        result = remove_twf_text("至急対応 TWFNo.003243　新成（株）")
        assert result.strip() == "至急対応"

    def test_fullwidth(self):
        assert remove_twf_text("ＴＷＦＮｏ．００３２４３　新成（株）") == ""

    def test_numero_sign(self):
        assert remove_twf_text("TWF№003243 新成") == ""

    def test_no_twf_text_unchanged(self):
        assert remove_twf_text("欠品中 3月上旬入荷予定") == "欠品中 3月上旬入荷予定"

    def test_twf_tokka_unchanged(self):
        """ＴＷＦ特価（Noなし）は除去しない（判定対象外のため）"""
        assert remove_twf_text("ＴＷＦ特価") == "ＴＷＦ特価"

    def test_empty(self):
        assert remove_twf_text("") == ""

    def test_multiline_keeps_other_lines(self):
        result = remove_twf_text("分納:50個 2/20\nTWFNo.003243　新成（株）")
        assert "分納:50個 2/20" in result
        assert "TWF" not in result


# ============================================
# parse_twf_comment（TWF記載の構造化）
# ============================================
class TestParseTwfComment:
    def test_number_and_customer(self):
        info = parse_twf_comment("TWFNo.003243　新成（株）様")
        assert info == TwfDetailInfo(number="003243", customer="新成（株）様", memo="")

    def test_number_only(self):
        info = parse_twf_comment("TWFNo.005210")
        assert info == TwfDetailInfo(number="005210", customer="", memo="")

    def test_fumei(self):
        info = parse_twf_comment("TWFNo.不明　新成㈱第三工場様")
        assert info == TwfDetailInfo(number="不明", customer="新成㈱第三工場様", memo="")

    def test_fullwidth_number_normalized(self):
        """全角数字は半角に正規化される"""
        info = parse_twf_comment("ＴＷＦＮｏ．００３２４３　新成（株）様")
        assert info.number == "003243"
        assert info.customer == "新成（株）様"

    def test_numero_sign(self):
        info = parse_twf_comment("TWF№003243 新成様")
        assert info == TwfDetailInfo(number="003243", customer="新成様", memo="")

    def test_no_space_after_number(self):
        info = parse_twf_comment("TWFNo.004662先進機設（株）様")
        assert info == TwfDetailInfo(number="004662", customer="先進機設（株）様", memo="")

    def test_customer_and_memo(self):
        """「様」の後にメモが続く → 分割"""
        info = parse_twf_comment("TWFNo.004982　(有)狩野溶接工業様お持ち帰り")
        assert info == TwfDetailInfo(
            number="004982", customer="(有)狩野溶接工業様", memo="お持ち帰り")

    def test_sama_muke_exception(self):
        """「様向け」特例: 向けまでお客様名に含める"""
        info = parse_twf_comment("TWFNo.005409　住友建機㈱横須賀工場様向け")
        assert info.customer == "住友建機㈱横須賀工場様向け"
        assert info.memo == ""

    def test_no_sama_all_memo(self):
        """「様」なし → 全文を備考メモへ（社名でも安全側）"""
        info = parse_twf_comment("TWFNo.003281　6/15着")
        assert info == TwfDetailInfo(number="003281", customer="", memo="6/15着")

    def test_no_sama_company_name(self):
        info = parse_twf_comment("TWFNo.004761　立山工業所")
        assert info == TwfDetailInfo(number="004761", customer="", memo="立山工業所")

    def test_memo_contains_sama_after_customer(self):
        """お客様名の後のメモ中の「様」は分割に影響しない"""
        info = parse_twf_comment("TWFNo.000882　（株）ハイプラン様　カミマル様が主催店様へ配達")
        assert info.customer == "（株）ハイプラン様"
        assert info.memo == "カミマル様が主催店様へ配達"

    def test_no_number(self):
        """番号の書き忘れ → number空欄、後続はそのまま振り分け"""
        info = parse_twf_comment("TWFNo.　新成（株）様")
        assert info == TwfDetailInfo(number="", customer="新成（株）様", memo="")

    def test_bare(self):
        info = parse_twf_comment("TWFNo.")
        assert info == TwfDetailInfo(number="", customer="", memo="")

    def test_no_twf_returns_none(self):
        assert parse_twf_comment("欠品中 3月上旬入荷予定") is None

    def test_twf_tokka_returns_none(self):
        assert parse_twf_comment("ＴＷＦ特価") is None

    def test_empty_returns_none(self):
        assert parse_twf_comment("") is None


# ============================================
# build_twf_info_map（注番→TWF情報の引き継ぎマップ）
# ============================================
class TestBuildTwfInfoMap:
    def test_first_occurrence_wins(self):
        orders = [
            _make_row(order_number="100", detail_number="10",
                      comment_detail="TWFNo.001　新成様"),
            _make_row(order_number="100", detail_number="20",
                      comment_detail="TWFNo.001　別会社様"),
            _make_row(order_number="200", detail_number="10",
                      comment_detail=""),
        ]
        m = build_twf_info_map(orders)
        assert set(m.keys()) == {"100"}
        assert m["100"].number == "001"
        assert m["100"].customer == "新成様"

    def test_empty(self):
        assert build_twf_info_map([]) == {}


# ============================================
# 持ち帰り判定（is_twf_pickup_memo / is_twf_pickup_only_memo）
# ============================================
class TestTwfPickupMemo:
    @pytest.mark.parametrize("memo", [
        "お持ち帰り",
        "お持ち帰り済み",
        "お持ち帰り済",
        "持ち帰り",
        "お持帰り",          # 表記ゆれ（実データ未出だが許容）
        "持帰",
        "6/15or6/16引取りとなります。",
        "お引き取り希望",
        "今回のみ旧値 お渡し済み",
    ])
    def test_pickup_positive(self, memo):
        from nouki_kaitou.twf import is_twf_pickup_memo
        assert is_twf_pickup_memo(memo) is True

    @pytest.mark.parametrize("memo", [
        "",
        "三脚Ｂ2ケｻｰﾋﾞｽ",
        "6/15着",
        "後日納品日ご連絡",
        "後日お渡し予定",    # 「渡し済」でないため対象外（未来の受け渡し）
    ])
    def test_pickup_negative(self, memo):
        from nouki_kaitou.twf import is_twf_pickup_memo
        assert is_twf_pickup_memo(memo) is False

    @pytest.mark.parametrize("memo,expected", [
        ("お持ち帰り", True),
        ("持ち帰り", True),
        ("お持帰り", True),
        ("お持ち帰り済み", False),   # 状態情報あり→備考に残す
        ("お持ち帰り済", False),
        ("6/15or6/16引取りとなります。", False),
        ("今回のみ旧値 お渡し済み", False),
        ("", False),
    ])
    def test_pickup_only(self, memo, expected):
        from nouki_kaitou.twf import is_twf_pickup_only_memo
        assert is_twf_pickup_only_memo(memo) is expected


# ============================================
# twf_sort_key（TWF No.昇順ソート）
# ============================================
class TestTwfSortKey:
    def test_numeric_ascending(self):
        rows = [
            ("005409", "300", "10"),
            ("000123", "100", "10"),
            ("003281", "200", "10"),
        ]
        ordered = sorted(rows, key=lambda r: twf_sort_key(*r))
        assert [r[0] for r in ordered] == ["000123", "003281", "005409"]

    def test_no_number_last(self):
        """番号なし・「不明」は末尾"""
        rows = [
            ("不明", "300", "10"),
            ("", "400", "10"),
            ("005409", "100", "10"),
        ]
        ordered = sorted(rows, key=lambda r: twf_sort_key(*r))
        assert [r[0] for r in ordered][0] == "005409"
        assert {r[0] for r in ordered[1:]} == {"不明", ""}

    def test_same_number_by_order_and_detail(self):
        """同一No.内は注番→明細順"""
        rows = [
            ("003341", "200", "10"),
            ("003341", "100", "20"),
            ("003341", "100", "10"),
        ]
        ordered = sorted(rows, key=lambda r: twf_sort_key(*r))
        assert [(r[1], r[2]) for r in ordered] == [("100", "10"), ("100", "20"), ("200", "10")]


# ============================================
# is_twf_active（期間ゲート）
# ============================================
class TestIsTwfActive:
    def test_before_end_date(self):
        assert is_twf_active(datetime.date(2026, 6, 12)) is True

    def test_on_end_date(self):
        assert is_twf_active(TWF_END_DATE) is True

    def test_after_end_date(self):
        assert is_twf_active(TWF_END_DATE + datetime.timedelta(days=1)) is False


# ============================================
# merge_email_input
# ============================================
class TestMergeEmailInput:
    def _result(self, file_path="r.xlsx", is_twf=False, **kwargs) -> ReportResult:
        return ReportResult(
            file_path=file_path,
            customer_name=kwargs.get("customer_name", "テスト商事"),
            rep_name=kwargs.get("rep_name", ""),
            stockout_info_list=kwargs.get("stockout_info_list", []),
            tracking_info_list=kwargs.get("tracking_info_list", []),
            bunno_info_list=kwargs.get("bunno_info_list", []),
            bunno_completed_list=kwargs.get("bunno_completed_list", []),
            is_twf=is_twf,
        )

    def test_normal_only(self):
        merged = merge_email_input(self._result("normal.xlsx"), None)
        assert merged["attachments"] == ["normal.xlsx"]
        assert merged["twf_notice"] is None
        assert merged["customer_name"] == "テスト商事"

    def test_normal_and_twf(self):
        merged = merge_email_input(
            self._result("normal.xlsx"),
            self._result("twf.xlsx", is_twf=True),
        )
        assert merged["attachments"] == ["normal.xlsx", "twf.xlsx"]
        assert merged["twf_notice"] == TWF_NOTICE_EMAIL

    def test_twf_only(self):
        """TWF受注のみの顧客（通常回答書なし）"""
        merged = merge_email_input(None, self._result("twf.xlsx", is_twf=True))
        assert merged["attachments"] == ["twf.xlsx"]
        assert merged["twf_notice"] == TWF_NOTICE_EMAIL
        assert merged["customer_name"] == "テスト商事"

    def test_info_lists_merged(self):
        normal = self._result(
            "normal.xlsx",
            stockout_info_list=[StockoutEntry(product_name="商品A")],
        )
        twf = self._result(
            "twf.xlsx", is_twf=True,
            stockout_info_list=[StockoutEntry(product_name="商品B")],
        )
        merged = merge_email_input(normal, twf)
        names = [s.product_name for s in merged["stockout_info_list"]]
        assert names == ["商品A", "商品B"]

    def test_both_none_raises(self):
        with pytest.raises(ValueError):
            merge_email_input(None, None)


# ============================================
# create_delivery_report のTWF関連パラメータ
# ============================================
class TestCreateDeliveryReportTwf:
    def test_exclude_orders(self, tmp_path):
        """exclude_orders指定の注番は通常回答書から除外される"""
        data = [
            _make_row(order_number="100", comment_detail="TWFNo.001"),
            _make_row(order_number="200"),
        ]
        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト商事", {},
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            date_from=datetime.date(2026, 2, 1),
            date_to=datetime.date(2026, 2, 28),
            today=TODAY,
            exclude_orders={"100"},
        )
        assert result is not None
        wb = load_workbook(result.file_path)
        ws = wb.active
        order_nums = [ws.cell(row=r, column=12).value for r in range(7, 9)]
        assert "200" in order_nums
        assert "100" not in order_nums

    def test_exclude_all_returns_none(self, tmp_path):
        """全注番が除外されたら None（回答書なし）"""
        data = [_make_row(order_number="100")]
        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト商事", {},
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
            exclude_orders={"100"},
        )
        assert result is None

    def test_include_only_orders(self, tmp_path):
        """include_only_orders指定の注番のみが出力される"""
        data = [
            _make_row(order_number="100", comment_detail="TWFNo.001"),
            _make_row(order_number="200"),
        ]
        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト商事", {},
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
            include_only_orders={"100"},
        )
        assert result is not None
        wb = load_workbook(result.file_path)
        ws = wb.active
        assert ws.cell(row=7, column=12).value == "100"
        assert ws.cell(row=8, column=12).value is None

    def test_filter_already_sent_false_shows_sent(self, tmp_path):
        """filter_already_sent=False なら送付済みでも表示される"""
        data = [
            _make_row(registration_date=datetime.date(2026, 2, 10)),
        ]
        sent = {"1000001|10": "2月15日配達予定"}
        cache = _make_cache()

        # 従来動作（True）: スキップされ None
        result_default = create_delivery_report(
            data, "テスト商事", sent,
            cache, tmp_path / "a", {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result_default is None

        # False: 表示される
        (tmp_path / "b").mkdir()
        result = create_delivery_report(
            data, "テスト商事", sent,
            cache, tmp_path / "b", {}, BRANCH, EXEC_TIME,
            today=TODAY,
            filter_already_sent=False,
        )
        assert result is not None

    def test_filter_already_sent_false_excluded_still_skipped(self, tmp_path):
        """filter_already_sent=False でも「除外」ステータスは常にスキップ"""
        data = [_make_row()]
        sent = {"1000001|10": "除外"}
        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト商事", sent,
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
            filter_already_sent=False,
        )
        assert result is None

    def test_filter_already_sent_false_hash_marker_still_skipped(self, tmp_path):
        """filter_already_sent=False でも##除外マーカーは常にスキップ"""
        data = [_make_row(comment_internal="##テスト除外")]
        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト商事", {},
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
            filter_already_sent=False,
        )
        assert result is None

    def test_twf_mode_title_and_sheet_name(self, tmp_path):
        """twf_mode: タイトル・シート名・ファイル名がTWF用になる"""
        data = [_make_row(order_number="100", comment_detail="TWFNo.001")]
        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト商事", {},
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
            include_only_orders={"100"},
            filter_already_sent=False,
            twf_mode=True,
        )
        assert result is not None
        assert result.is_twf is True
        assert TWF_FILENAME_TAG in Path(result.file_path).name

        wb = load_workbook(result.file_path)
        ws = wb.active
        assert ws.title.startswith(TWF_SHEET_PREFIX)
        assert ws["A1"].value == TWF_REPORT_TITLE

    def test_twf_mode_notice_in_info_section(self, tmp_path):
        """twf_mode: ご連絡事項にTWF注記が出る"""
        data = [_make_row(order_number="100", comment_detail="TWFNo.001")]
        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト商事", {},
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
            include_only_orders={"100"},
            filter_already_sent=False,
            twf_mode=True,
        )
        wb = load_workbook(result.file_path)
        ws = wb.active
        values = [
            ws.cell(row=r, column=1).value
            for r in range(7, ws.max_row + 1)
            if ws.cell(row=r, column=1).value
        ]
        assert TWF_NOTICE_EXCEL in values

    def _twf_report(self, data, tmp_path, sent=None):
        """TWF専用回答書を生成して(result, ws)を返すヘルパー。"""
        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト商事", sent or {},
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
            include_only_orders={r.order_number.strip() for r in data},
            filter_already_sent=False,
            twf_mode=True,
        )
        wb = load_workbook(result.file_path)
        return result, wb.active

    def test_twf_layout_columns(self, tmp_path):
        """TWF専用レイアウト: A=TWF No.（先頭ゼロ保持）, C=お客様名"""
        data = [_make_row(order_number="100",
                          comment_detail="TWFNo.003243　新成（株）様")]
        result, ws = self._twf_report(data, tmp_path)
        assert ws.cell(row=7, column=1).value == "003243"   # 文字列で先頭ゼロ保持
        assert ws.cell(row=7, column=3).value == "新成（株）様"
        assert ws.cell(row=7, column=11).value is None      # メモなし→備考空欄

    def test_twf_layout_header_labels(self, tmp_path):
        """TWF版ヘッダー: A=TWF No., C=お客様名、他は通常版と同じ"""
        data = [_make_row(order_number="100", comment_detail="TWFNo.001")]
        result, ws = self._twf_report(data, tmp_path)
        assert ws.cell(row=6, column=1).value == "TWF No."
        assert ws.cell(row=6, column=3).value == "お客様名"
        assert ws.cell(row=6, column=9).value == "納期回答"
        assert ws.cell(row=6, column=12).value == "弊社注番"

    def test_twf_memo_in_remarks(self, tmp_path):
        """メモ部分はK列備考に表示"""
        data = [_make_row(order_number="100",
                          comment_detail="TWFNo.002041　先進機設㈱様　三脚Ｂ2ケｻｰﾋﾞｽ")]
        result, ws = self._twf_report(data, tmp_path)
        assert ws.cell(row=7, column=3).value == "先進機設㈱様"
        assert ws.cell(row=7, column=11).value == "三脚Ｂ2ケｻｰﾋﾞｽ"

    def test_twf_memo_coexists_with_existing_remark(self, tmp_path):
        """メモと既存備考は「メモ ／ 既存備考」で共存"""
        data = [_make_row(
            order_number="100",
            comment_detail="至急対応\nTWFNo.003243　新成様　三脚サービス",
        )]
        result, ws = self._twf_report(data, tmp_path)
        assert ws.cell(row=7, column=11).value == "三脚サービス ／ 至急対応"

    def test_twf_pickup_pure_memo(self, tmp_path):
        """純粋な「お持ち帰り」→ 納入先「お引き取り」・備考から省略"""
        data = [_make_row(order_number="100",
                          comment_detail="TWFNo.004982　(有)狩野溶接工業様お持ち帰り")]
        result, ws = self._twf_report(data, tmp_path)
        assert ws.cell(row=7, column=10).value == "お引き取り"
        assert ws.cell(row=7, column=11).value is None  # 重複のため省略

    def test_twf_pickup_with_extra_info_keeps_memo(self, tmp_path):
        """「お持ち帰り済み」→ お引き取り＋備考に残す（状態情報あり）"""
        data = [_make_row(order_number="100",
                          comment_detail="TWFNo.004987　東進産業様　お持ち帰り済み")]
        result, ws = self._twf_report(data, tmp_path)
        assert ws.cell(row=7, column=10).value == "お引き取り"
        assert ws.cell(row=7, column=11).value == "お持ち帰り済み"

    def test_twf_pickup_hikitori_with_date(self, tmp_path):
        """「6/15or6/16引取りとなります。」→ お引き取り＋日程は備考に残す"""
        data = [_make_row(order_number="100",
                          comment_detail="TWFNo.003322　6/15or6/16引取りとなります。")]
        result, ws = self._twf_report(data, tmp_path)
        assert ws.cell(row=7, column=10).value == "お引き取り"
        assert ws.cell(row=7, column=11).value == "6/15or6/16引取りとなります。"

    def test_twf_pickup_watashizumi(self, tmp_path):
        """「お渡し済み」→ お引き取り＋備考に残す"""
        data = [_make_row(order_number="100",
                          comment_detail="TWFNo.000887　今回のみ旧値　お渡し済み")]
        result, ws = self._twf_report(data, tmp_path)
        assert ws.cell(row=7, column=10).value == "お引き取り"
        assert ws.cell(row=7, column=11).value == "今回のみ旧値 お渡し済み"

    def test_twf_pickup_not_inherited(self, tmp_path):
        """memoは引き継がないため、入れ忘れ明細の納入先は上書きされない"""
        data = [
            _make_row(order_number="100", detail_number="10",
                      comment_detail="TWFNo.001　新成様　お持ち帰り"),
            _make_row(order_number="100", detail_number="20",
                      comment_detail=""),  # 入れ忘れ
        ]
        result, ws = self._twf_report(data, tmp_path)
        assert ws.cell(row=7, column=10).value == "お引き取り"  # 明細10
        assert ws.cell(row=8, column=10).value == "貴社"        # 明細20は従来表示

    def test_twf_thanks_in_info_section(self, tmp_path):
        """ご連絡事項欄に感謝文＋赤字注記の両方が出る（感謝文が先）"""
        data = [_make_row(order_number="100", comment_detail="TWFNo.001　新成様")]
        result, ws = self._twf_report(data, tmp_path)
        values = [
            ws.cell(row=r, column=1).value
            for r in range(7, ws.max_row + 1)
            if ws.cell(row=r, column=1).value
        ]
        assert TWF_THANKS_EXCEL in values
        assert TWF_NOTICE_EXCEL in values
        assert values.index(TWF_THANKS_EXCEL) < values.index(TWF_NOTICE_EXCEL)

    def test_twf_total_row_mixed(self, tmp_path):
        """合計行: 数値行のみ合計、未確定行があれば注記を併記"""
        data = [
            _make_row(order_number="100", detail_number="10",
                      comment_detail="TWFNo.001　新成様",
                      net_amount="50000"),
            _make_row(order_number="100", detail_number="20",
                      comment_detail="TWFNo.001　新成様",
                      net_amount="30000"),
            # 仮単価（1円）→ 金額「確認中」になる行
            _make_row(order_number="100", detail_number="30",
                      comment_detail="TWFNo.001　新成様",
                      unit_price="1", net_amount="1"),
        ]
        result, ws = self._twf_report(data, tmp_path)
        total_row = 10  # データ7〜9行 → 合計は10行目
        assert ws.cell(row=total_row, column=6).value == "展示会ご成約合計（税抜）："
        assert ws.cell(row=total_row, column=8).value == 80000
        assert ws.cell(row=total_row, column=8).number_format == '"￥"#,##0'
        assert ws.cell(row=total_row, column=9).value == "※金額確定分の合計です"
        # 税抜注記は1行下がる
        assert ws.cell(row=total_row + 1, column=7).value == "※表示金額は税抜きです"

    def test_twf_total_row_all_confirmed_no_note(self, tmp_path):
        """全行確定済みなら注記は出ない"""
        data = [
            _make_row(order_number="100", detail_number="10",
                      comment_detail="TWFNo.001　新成様", net_amount="50000"),
            _make_row(order_number="100", detail_number="20",
                      comment_detail="TWFNo.001　新成様", net_amount="30000"),
        ]
        result, ws = self._twf_report(data, tmp_path)
        total_row = 9
        assert ws.cell(row=total_row, column=8).value == 80000
        assert ws.cell(row=total_row, column=9).value is None  # 注記なし

    def test_twf_total_excluded_from_autofilter(self, tmp_path):
        """合計行はオートフィルタ範囲外"""
        data = [
            _make_row(order_number="100", detail_number="10",
                      comment_detail="TWFNo.001", net_amount="50000"),
        ]
        result, ws = self._twf_report(data, tmp_path)
        assert ws.auto_filter.ref == "A6:L7"  # 合計行（8行目）は含まない

    def test_normal_mode_no_total_row(self, tmp_path):
        """通常版には合計行なし（税抜注記がデータ直下のまま）"""
        data = [_make_row(net_amount="50000")]
        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト商事", {},
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        wb = load_workbook(result.file_path)
        ws = wb.active
        assert ws.cell(row=8, column=7).value == "※表示金額は税抜きです"
        assert ws.cell(row=8, column=6).value is None  # 合計ラベルなし

    def test_normal_mode_no_thanks(self, tmp_path):
        """通常版には感謝文は出ない"""
        data = [_make_row()]
        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト商事", {},
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        wb = load_workbook(result.file_path)
        ws = wb.active
        values = [
            ws.cell(row=r, column=1).value
            for r in range(7, ws.max_row + 1)
            if ws.cell(row=r, column=1).value
        ]
        assert TWF_THANKS_EXCEL not in values

    def test_twf_inherited_number_and_customer(self, tmp_path):
        """入れ忘れ明細は同注番から番号とお客様名を引き継ぐ（memoは引き継がない）"""
        data = [
            _make_row(order_number="100", detail_number="10",
                      comment_detail="TWFNo.003243　新成様　お持ち帰り"),
            _make_row(order_number="100", detail_number="20",
                      comment_detail=""),  # 入れ忘れ
        ]
        result, ws = self._twf_report(data, tmp_path)
        # 明細20の行（同一No.内は明細順なので2行目）
        assert ws.cell(row=8, column=1).value == "003243"
        assert ws.cell(row=8, column=3).value == "新成様"
        assert ws.cell(row=8, column=11).value is None  # memoは引き継がない

    def test_twf_sorted_by_number(self, tmp_path):
        """TWF No.昇順ソート（番号なしは末尾）"""
        data = [
            _make_row(order_number="300", comment_detail="TWFNo.不明　Ｃ社様"),
            _make_row(order_number="100", comment_detail="TWFNo.005409　Ａ社様"),
            _make_row(order_number="200", comment_detail="TWFNo.000123　Ｂ社様"),
        ]
        result, ws = self._twf_report(data, tmp_path)
        assert [ws.cell(row=r, column=1).value for r in (7, 8, 9)] == [
            "000123", "005409", "不明"]
        assert [ws.cell(row=r, column=3).value for r in (7, 8, 9)] == [
            "Ｂ社様", "Ａ社様", "Ｃ社様"]

    def test_twf_delivery_place_chokusou_unchanged(self, tmp_path):
        """直送でも納入先名は従来表示のまま（付記なし）"""
        data = [_make_row(
            order_number="100",
            document_type="【受注】直送販売",
            storage_place="転送中（直送用）",
            comment_detail="TWFNo.001　新成様",
        )]
        result, ws = self._twf_report(data, tmp_path)
        assert ws.cell(row=7, column=10).value == "貴社"

    def test_twf_delivery_place_onetime_replaced(self, tmp_path):
        """「ワンタイム出荷先」→「ご指定先」置換"""
        data = [_make_row(
            order_number="100",
            document_type="【受注】直送販売",
            storage_place="転送中（直送用）",
            ship_to_name="ワンタイム出荷先",
            comment_detail="TWFNo.001　新成様",
        )]
        result, ws = self._twf_report(data, tmp_path)
        assert ws.cell(row=7, column=10).value == "ご指定先"

    def test_twf_delivery_place_stock_unchanged(self, tmp_path):
        """在庫販売・紐付き（弊社出荷）も従来表示のまま"""
        data = [_make_row(order_number="100", comment_detail="TWFNo.001　新成様")]
        result, ws = self._twf_report(data, tmp_path)
        assert ws.cell(row=7, column=10).value == "貴社"

    def test_twf_title_and_notice_have_2026(self, tmp_path):
        """タイトル・赤字注記とも「東京ウェルディングフェスタ2026」表記"""
        assert "東京ウェルディングフェスタ2026" in TWF_REPORT_TITLE
        assert "東京ウェルディングフェスタ2026" in TWF_NOTICE_EXCEL
        assert "東京ウェルディングフェスタ2026" in TWF_NOTICE_EMAIL

    def test_normal_mode_layout_unchanged(self, tmp_path):
        """通常版はヘッダー・A列・C列・納入先とも従来どおり"""
        data = [_make_row(ship_to_name="ワンタイム出荷先",
                          document_type="【受注】直送販売",
                          storage_place="転送中（直送用）")]
        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト商事", {},
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        wb = load_workbook(result.file_path)
        ws = wb.active
        assert ws.cell(row=6, column=1).value == "受注日"
        assert ws.cell(row=6, column=3).value == "貴社注番"
        assert ws.cell(row=7, column=1).value.date() == TODAY  # 受注日のまま
        assert ws.cell(row=7, column=10).value == "ワンタイム出荷先様"  # 置換なし

    def test_normal_mode_remark_still_removed(self, tmp_path):
        """通常モードでは従来どおりTWF記載は備考から除去される"""
        data = [_make_row(order_number="100",
                          comment_detail="至急対応\nTWFNo.003243　新成（株）様")]
        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト商事", {},
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        wb = load_workbook(result.file_path)
        ws = wb.active
        assert ws.cell(row=7, column=11).value == "至急対応"

    def test_twf_mode_auto_filter_set(self, tmp_path):
        """TWF専用回答書にはオートフィルタが設定される"""
        data = [
            _make_row(order_number="100", detail_number="10",
                      comment_detail="TWFNo.001"),
            _make_row(order_number="100", detail_number="20",
                      comment_detail="TWFNo.001"),
        ]
        result, ws = self._twf_report(data, tmp_path)
        assert ws.auto_filter.ref == "A6:L8"

    def test_normal_mode_no_auto_filter(self, tmp_path):
        """通常回答書にはオートフィルタを設定しない（従来どおり）"""
        data = [_make_row()]
        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト商事", {},
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        wb = load_workbook(result.file_path)
        ws = wb.active
        assert ws.auto_filter.ref is None

    def test_twf_mode_delivered_override(self, tmp_path):
        """twf_mode: 処理完了+回答済み（履歴に確定記録）→「納品済み」表示"""
        data = [
            _make_row(
                order_number="100",
                document_type="【受注】直送販売",
                storage_place="転送中（直送用）",
                ship_status="処理完了",
                registration_date=datetime.date(2026, 2, 10),
                order_delivery_date=datetime.date(2026, 12, 31),
            ),
        ]
        # 既に「2月13日出荷予定」で回答済み
        sent = {"100|10": "2月13日出荷予定"}
        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト商事", sent,
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
            include_only_orders={"100"},
            filter_already_sent=False,
            twf_mode=True,
        )
        assert result is not None
        wb = load_workbook(result.file_path)
        ws = wb.active
        assert ws.cell(row=7, column=9).value == "納品済み"

    def test_twf_mode_no_override_when_confirming(self, tmp_path):
        """twf_mode: 前回「確認中」+処理完了は既存ルール（force_delivered）で処理"""
        data = [
            _make_row(
                order_number="100",
                document_type="【受注】直送販売",
                storage_place="転送中（直送用）",
                ship_status="処理完了",
                registration_date=datetime.date(2026, 2, 10),
                order_delivery_date=datetime.date(2026, 12, 31),
            ),
        ]
        sent = {"100|10": "確認中"}
        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト商事", sent,
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
            include_only_orders={"100"},
            filter_already_sent=False,
            twf_mode=True,
        )
        assert result is not None
        wb = load_workbook(result.file_path)
        ws = wb.active
        # 既存のforce_deliveredルールで納品済みになる
        assert ws.cell(row=7, column=9).value == "納品済み"

    def test_twf_mode_classification_still_recorded(self, tmp_path):
        """案Y: TWF回答書でも確定/確認中の分類（履歴記録用）は行われる"""
        data = [
            _make_row(order_number="100", comment_detail="TWFNo.001"),
            _make_row(
                order_number="100", detail_number="20",
                order_delivery_date=datetime.date(2026, 12, 31),  # 未確定
                specified_delivery_date=datetime.date(2026, 12, 31),
            ),
        ]
        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト商事", {},
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
            include_only_orders={"100"},
            filter_already_sent=False,
            twf_mode=True,
        )
        assert result is not None
        assert len(result.confirmed_orders) == 1   # 明細10: 確定
        assert len(result.confirming_orders) == 1  # 明細20: 日程調整中

    def test_default_behavior_unchanged(self, tmp_path):
        """新パラメータ未指定なら従来と同一の出力"""
        data = [_make_row()]
        cache = _make_cache()
        result = create_delivery_report(
            data, "テスト商事", {},
            cache, tmp_path, {}, BRANCH, EXEC_TIME,
            today=TODAY,
        )
        assert result is not None
        assert result.is_twf is False
        assert TWF_FILENAME_TAG not in Path(result.file_path).name
        wb = load_workbook(result.file_path)
        ws = wb.active
        assert ws["A1"].value == "納　期　回　答　書"
        assert not ws.title.startswith(TWF_SHEET_PREFIX)


# ============================================
# email_builder のTWF対応
# ============================================
class TestEmailBuilderTwf:
    def _make_customer_ws(self, rows):
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1).value = "顧客名"
        ws.cell(row=1, column=2).value = "住所"
        ws.cell(row=1, column=3).value = "電話"
        ws.cell(row=1, column=4).value = "担当"
        ws.cell(row=1, column=5).value = "メール1"
        for row_idx, data in enumerate(rows, 2):
            for col_idx, val in enumerate(data, 1):
                ws.cell(row=row_idx, column=col_idx).value = val
        return ws

    def test_body_contains_twf_notice(self):
        from nouki_kaitou.email_builder import build_email_body_html

        body = build_email_body_html(
            "テスト商事", BRANCH, today=TODAY,
            twf_notice=TWF_NOTICE_EMAIL,
        )
        assert "東京ウェルディングフェスタ2026" in body
        assert "多大なるご尽力を賜り" in body       # 感謝文
        assert "ご成約いただきました商品" in body
        assert "<br>" in body                       # 改行が変換されている

    def test_body_without_twf_notice(self):
        from nouki_kaitou.email_builder import build_email_body_html

        body = build_email_body_html("テスト商事", BRANCH, today=TODAY)
        assert "東京ウェルディングフェスタ" not in body

    def test_create_emails_multiple_attachments(self):
        from nouki_kaitou.email_builder import create_emails

        cust_ws = self._make_customer_ws([
            ("テスト商事", "", "", "", "test@example.com"),
        ])
        files = [{
            "customer_name": "テスト商事",
            "file_path": "normal.xlsx",
            "attachments": ["normal.xlsx", "twf.xlsx"],
            "twf_notice": TWF_NOTICE_EMAIL,
        }]
        results, skipped = create_emails(files, BRANCH, cust_ws, today=TODAY)
        assert len(results) == 1
        assert results[0]["attachments"] == ["normal.xlsx", "twf.xlsx"]
        assert "東京ウェルディングフェスタ" in results[0]["html_body"]

    def test_create_emails_fallback_single_attachment(self):
        """attachments未指定なら従来どおりfile_path 1件添付"""
        from nouki_kaitou.email_builder import create_emails

        cust_ws = self._make_customer_ws([
            ("テスト商事", "", "", "", "test@example.com"),
        ])
        files = [{
            "customer_name": "テスト商事",
            "file_path": "normal.xlsx",
        }]
        results, skipped = create_emails(files, BRANCH, cust_ws, today=TODAY)
        assert len(results) == 1
        assert results[0]["attachments"] == ["normal.xlsx"]
        assert "東京ウェルディングフェスタ" not in results[0]["html_body"]


# ============================================
# utils のTWF対応（ファイル名タグ・シート名プレフィックス）
# ============================================
class TestUtilsTwf:
    def test_filename_with_tag(self):
        from nouki_kaitou.utils import build_report_filename

        name = build_report_filename(
            "テスト商事", EXEC_TIME, filename_tag=TWF_FILENAME_TAG,
        )
        assert name == "納期回答書【TWF2026】_テスト商事様_20260216.xlsx"

    def test_filename_with_tag_and_rep(self):
        from nouki_kaitou.utils import build_report_filename

        name = build_report_filename(
            "テスト商事", EXEC_TIME, rep_name="田中",
            filename_tag=TWF_FILENAME_TAG,
        )
        assert name == "納期回答書【TWF2026】_テスト商事様_田中様_20260216.xlsx"

    def test_filename_without_tag_unchanged(self):
        from nouki_kaitou.utils import build_report_filename

        name = build_report_filename("テスト商事", EXEC_TIME)
        assert name == "納期回答書_テスト商事様_20260216.xlsx"

    def test_sheet_name_with_prefix(self):
        from nouki_kaitou.utils import build_sheet_name

        assert build_sheet_name("テスト商事", prefix=TWF_SHEET_PREFIX) == "TWF2026_テスト商事様"

    def test_sheet_name_prefix_31_chars_limit(self):
        from nouki_kaitou.utils import build_sheet_name

        name = build_sheet_name("あ" * 40, prefix=TWF_SHEET_PREFIX)
        assert len(name) <= 31
        assert name.startswith(TWF_SHEET_PREFIX)

    def test_sheet_name_without_prefix_unchanged(self):
        from nouki_kaitou.utils import build_sheet_name

        assert build_sheet_name("テスト商事") == "テスト商事様"
