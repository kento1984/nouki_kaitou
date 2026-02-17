"""出荷ステータス整合性WARNING 276件のパターン別集計"""

import sys
from pathlib import Path
import re
from collections import defaultdict

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import load_workbook

def main():
    validation_path = Path(__file__).parent / "validation_result.xlsx"

    print("validation_result.xlsxを読み込み中...")
    wb = load_workbook(validation_path, read_only=True, data_only=True)

    # 出荷ステータス整合性シートを探す
    target_sheet = None
    for sheet_name in wb.sheetnames:
        if "出荷ステータス" in sheet_name:
            target_sheet = sheet_name
            break

    if not target_sheet:
        print(f"シートが見つかりません。利用可能なシート: {wb.sheetnames}")
        return

    print(f"対象シート: {target_sheet}")
    ws = wb[target_sheet]

    # ヘッダーを確認
    headers = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    print(f"ヘッダー: {headers}")

    # 問題列のインデックスを特定
    problem_col = None
    for idx, h in enumerate(headers):
        if h and "問題" in str(h):
            problem_col = idx
            break

    if problem_col is None:
        # ヘッダーから推測
        problem_col = 10  # デフォルト

    print(f"問題列: {problem_col}")

    # データを収集
    warnings = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if row and len(row) > 3 and row[2]:  # 注番は3列目（インデックス2）
            order_num = str(row[2]).strip()
            detail_num = str(row[3] or "").strip() if len(row) > 3 else ""
            customer = str(row[4] or "").strip()[:20] if len(row) > 4 else ""
            ship_status = str(row[8] or "").strip() if len(row) > 8 else ""
            problem = str(row[problem_col] or "").strip() if len(row) > problem_col else ""

            warnings.append({
                "order": order_num,
                "detail": detail_num,
                "customer": customer,
                "ship_status": ship_status,
                "problem": problem,
                "raw_row": row,
            })

    wb.close()
    print(f"読み込み完了: {len(warnings)}件\n")

    # パターン別に集計
    pattern_counts = defaultdict(list)

    for w in warnings:
        problem = w["problem"]

        # パターンを抽出（数値部分を正規化）
        # "処理完了だが結果日付が123日後" → "処理完了だが結果日付がN日後"
        normalized = re.sub(r"\d+日後", "N日後", problem)
        normalized = re.sub(r"\d+日前", "N日前", normalized)

        pattern_counts[normalized].append(w)

    # 結果表示
    print("=" * 100)
    print("出荷ステータス整合性 WARNING パターン別集計")
    print("=" * 100)

    # 件数順にソート
    sorted_patterns = sorted(pattern_counts.items(), key=lambda x: -len(x[1]))

    for pattern, items in sorted_patterns:
        print(f"\n【{pattern}】({len(items)}件)")
        print("-" * 80)

        # 代表例を3件まで表示
        for i, item in enumerate(items[:3]):
            print(f"  例{i+1}: {item['order']}|{item['detail']}")
            print(f"       出荷ステータス: {item['ship_status']}")
            print(f"       顧客: {item['customer']}")
            if item['problem'] != pattern:
                print(f"       問題: {item['problem']}")

        if len(items) > 3:
            print(f"  ... 他 {len(items) - 3}件")

    # サマリー
    print("\n" + "=" * 100)
    print("サマリー")
    print("=" * 100)
    print(f"総件数: {len(warnings)}件")
    print(f"パターン数: {len(pattern_counts)}種類")
    print()
    print("パターン別件数:")
    for pattern, items in sorted_patterns:
        print(f"  {len(items):3d}件: {pattern[:70]}")

if __name__ == "__main__":
    main()
