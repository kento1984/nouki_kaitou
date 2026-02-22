"""送付履歴管理モジュール

VBAの以下の関数を移植:
- InitializeDeliveryHistory (L1396): 送付履歴ファイル初期化
- LoadDeliveryHistory (L1495): 送付履歴読み込み → スキップ対象dict
- SaveDeliveryHistory (L1614): 確定伝票を送付履歴に書き込み
- SaveConfirmingList (L1775): 未確定伝票を確認中一覧に書き込み
- CleanConfirmingList (L2057): 確認中→送付履歴への移動
- CleanOldHistory (L2150): 古い送付履歴レコード削除
- CleanOldConfirmingList (L6506): 古い確認中一覧削除
"""

from __future__ import annotations

import datetime
import os
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from nouki_kaitou.models import (
    CacheStore,
    ConfirmingRecord,
    HolidayMap,
    HistoryRecord,
)

# テーブル名
_HISTORY_TABLE_NAME = "送付履歴テーブル"
_CONFIRMING_TABLE_NAME = "確認中テーブル"

# シート名
HISTORY_SHEET_NAME = "送付履歴"
CONFIRMING_SHEET_NAME = "確認中一覧"

# 送付履歴ヘッダー（9列）
_HISTORY_HEADERS = [
    "送付日時", "受注日", "顧客名", "受発注伝票", "明細",
    "メーカー名", "品名", "納期回答", "送付者",
]

# 確認中一覧ヘッダー（11列）
_CONFIRMING_HEADERS = [
    "送付日時", "受注日", "顧客名", "受発注伝票", "明細",
    "メーカー名", "品名", "問合せ状況", "ステータス", "受注納期", "送付者",
]

# 列幅
_HISTORY_WIDTHS = [17, 12, 25, 15, 8, 20, 47, 22, 18]
_CONFIRMING_WIDTHS = [17, 12, 25, 15, 8, 20, 47, 13, 12, 18, 18]


def _to_int_detail(detail: object) -> int | str:
    """明細番号をint型に変換する。VBAとの互換性のため数値で書き込む。"""
    try:
        return int(detail)
    except (ValueError, TypeError):
        return str(detail) if detail else ""


def _get_default_sender() -> str:
    """Windowsのログインユーザー名を取得する。VBAと同じ形式。"""
    return os.environ.get("USERNAME", "")


# ============================================
# VBA: InitializeDeliveryHistory (L1396-1491)
# 送付履歴ファイルの初期作成
# ============================================
def initialize_delivery_history(file_path: str) -> Workbook:
    """送付履歴ファイル(.xlsx)を初期化する。

    送付履歴シートと確認中一覧シートを作成し、
    それぞれにテーブルを設定する。

    Args:
        file_path: 保存先ファイルパス

    Returns:
        作成したWorkbook
    """
    wb = Workbook()

    # === 送付履歴シート ===
    ws_history = wb.active
    ws_history.title = HISTORY_SHEET_NAME

    for col_idx, header in enumerate(_HISTORY_HEADERS, 1):
        ws_history.cell(row=1, column=col_idx).value = header

    for col_idx, width in enumerate(_HISTORY_WIDTHS, 1):
        ws_history.column_dimensions[get_column_letter(col_idx)].width = width

    # テーブル設定
    ref = f"A1:I1"
    tbl_history = Table(displayName=_HISTORY_TABLE_NAME, ref=ref)
    tbl_history.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True
    )
    ws_history.add_table(tbl_history)

    # === 確認中一覧シート ===
    ws_confirming = wb.create_sheet(CONFIRMING_SHEET_NAME)

    for col_idx, header in enumerate(_CONFIRMING_HEADERS, 1):
        ws_confirming.cell(row=1, column=col_idx).value = header

    for col_idx, width in enumerate(_CONFIRMING_WIDTHS, 1):
        ws_confirming.column_dimensions[get_column_letter(col_idx)].width = width

    # テーブル設定
    ref = f"A1:K1"
    tbl_confirming = Table(displayName=_CONFIRMING_TABLE_NAME, ref=ref)
    tbl_confirming.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium1", showRowStripes=True
    )
    ws_confirming.add_table(tbl_confirming)

    wb.save(file_path)
    return wb


# ============================================
# VBA: LoadDeliveryHistory (L1495-1609)
# 送付履歴の読み込み → スキップ対象dict
# ============================================
def load_delivery_history(
    ws_history: Worksheet,
    ws_confirming: Worksheet,
    cache: CacheStore,
    holidays: HolidayMap | None = None,
    today: Optional[datetime.date] = None,
) -> dict[str, str]:
    """送付履歴と確認中一覧から、スキップ対象の伝票を読み込む。

    送付履歴テーブルの確定伝票 + 確認中一覧の「除外」伝票を
    辞書に格納して返す。

    Args:
        ws_history: 送付履歴シート
        ws_confirming: 確認中一覧シート
        cache: キャッシュストア（保持日数判定用）
        holidays: 祝日辞書（営業日計算用）
        today: 基準日（テスト用）

    Returns:
        dict[キー(受発注伝票|明細), 納期回答ステータス]
    """
    from nouki_kaitou.business_days import count_business_days_between

    if today is None:
        today = datetime.date.today()

    sent_orders: dict[str, str] = {}

    # === 送付履歴テーブルを読み込み（iter_rowsで高速化） ===
    for row_data in ws_history.iter_rows(min_row=2, values_only=True):
        if not row_data or len(row_data) < 8:
            continue

        sent_datetime_val = row_data[0]
        order_date_val = row_data[1]
        customer_name = str(row_data[2] or "").strip()
        order_number = str(row_data[3] or "").strip()
        detail_number = str(row_data[4] or "").strip()
        delivery_status = str(row_data[7] or "").strip()

        if not order_number or not delivery_status or delivery_status == "確認中":
            continue

        history_key = f"{order_number}|{detail_number}"

        # 送付日時パース
        sent_date = _parse_date_value(sent_datetime_val)

        # 受注日パース
        order_date = _parse_date_value(order_date_val)

        # 保持日数判定
        retention_days = cache.cust_retention.get(customer_name, 0)

        # 分納完了は常にスキップ対象
        if delivery_status == "分納完了":
            if history_key not in sent_orders:
                sent_orders[history_key] = delivery_status
        elif retention_days == 0:
            # 従来通り：受注日が今日より前なら除外
            if order_date and order_date < today:
                if history_key not in sent_orders:
                    sent_orders[history_key] = delivery_status
        else:
            # 保持日数設定あり：営業日経過で判定
            if sent_date:
                business_days_passed = count_business_days_between(
                    sent_date, today, holidays
                )
                if business_days_passed > retention_days:
                    if history_key not in sent_orders:
                        sent_orders[history_key] = delivery_status

    # === 確認中テーブルから「除外」の伝票も読み込み ===
    for row_data in ws_confirming.iter_rows(min_row=2, values_only=True):
        if not row_data or len(row_data) < 8:
            continue

        inquiry_status = str(row_data[7] or "").strip()
        if inquiry_status != "除外":
            continue

        order_number = str(row_data[3] or "").strip()
        detail_number = str(row_data[4] or "").strip()

        if order_number:
            history_key = f"{order_number}|{detail_number}"
            if history_key not in sent_orders:
                sent_orders[history_key] = "除外"

    return sent_orders


# ============================================
# VBA: SaveDeliveryHistory (L1614-1770)
# 確定伝票を送付履歴に書き込み
# ============================================
def save_delivery_history(
    ws: Worksheet,
    new_orders: list[HistoryRecord],
    execution_time: Optional[datetime.datetime] = None,
    sender: str = "",
) -> None:
    """確定した伝票を送付履歴テーブルに書き込む。

    重複チェックを行い、既存の伝票は更新（納品済みのみ）。
    新規伝票は先頭に追加し、送付日時の降順でソート。

    Args:
        ws: 送付履歴シート
        new_orders: 書き込むHistoryRecordのリスト
        execution_time: 実行時刻（Noneなら現在時刻）
        sender: 送付者名（空ならHistoryRecord.senderを使用）
    """
    if not new_orders:
        return

    if execution_time is None:
        execution_time = datetime.datetime.now()

    # === 既存データ読み取り ===
    existing_rows: list[list] = []
    existing_keys: dict[str, int] = {}  # キー → 既存行index

    max_row = ws.max_row
    if max_row >= 2:
        for row_idx in range(2, max_row + 1):
            row_data = []
            for col in range(1, 10):
                row_data.append(ws.cell(row=row_idx, column=col).value)

            order_num = str(row_data[3] or "").strip()
            detail_num = str(row_data[4] or "").strip()

            if order_num:
                existing_rows.append(row_data)
                key = f"{order_num}|{detail_num}"
                if key not in existing_keys:
                    existing_keys[key] = len(existing_rows) - 1

    # === 新規データ処理 ===
    new_items: list[list] = []

    for record in new_orders:
        key = f"{record.order_number}|{record.detail_number}"
        record_sender = sender or record.sender or _get_default_sender()

        if key in existing_keys:
            # 重複：納品済みの場合は既存行の納期回答を更新
            if record.delivery_answer == "納品済み":
                idx = existing_keys[key]
                existing_rows[idx][7] = "納品済み"
        else:
            # 新規
            new_row = [
                execution_time,
                record.order_date,
                record.customer_name,
                record.order_number,
                _to_int_detail(record.detail_number),
                record.manufacturer_name,
                record.product_name,
                record.delivery_answer,
                record_sender,
            ]
            new_items.append(new_row)
            existing_keys[key] = -1  # 後続重複チェック用

    # === 結果をシートに書き戻し ===
    all_rows = new_items + existing_rows
    if not all_rows:
        return

    # 送付日時の降順でソート
    all_rows.sort(
        key=lambda r: r[0] if isinstance(r[0], datetime.datetime) else datetime.datetime.min,
        reverse=True,
    )

    _write_rows_to_sheet(ws, all_rows, len(_HISTORY_HEADERS))

    # 表示形式
    for row_idx in range(2, len(all_rows) + 2):
        ws.cell(row=row_idx, column=1).number_format = "mm/dd hh:mm"
        ws.cell(row=row_idx, column=2).number_format = "mm/dd"


# ============================================
# VBA: SaveConfirmingList (L1775-1950)
# 未確定伝票を確認中一覧に書き込み
# ============================================
def save_confirming_list(
    ws: Worksheet,
    new_orders: list[ConfirmingRecord],
    execution_time: Optional[datetime.datetime] = None,
    sender: str = "",
) -> None:
    """未確定伝票を確認中一覧テーブルに書き込む。

    重複チェックを行い、既存の伝票はステータスのみ更新。
    新規伝票は問合せ状況「未」で追加。
    問合せ状況列に入力規則（未/済/回答待ち/除外）を設定。

    Args:
        ws: 確認中一覧シート
        new_orders: 書き込むConfirmingRecordのリスト
        execution_time: 実行時刻（Noneなら現在時刻）
        sender: 送付者名（空ならConfirmingRecord.senderを使用）
    """
    from nouki_kaitou.excel_writer import color_confirming_list

    if not new_orders:
        return

    if execution_time is None:
        execution_time = datetime.datetime.now()

    # === 既存データ読み取り ===
    existing_rows: list[list] = []
    existing_keys: dict[str, int] = {}

    max_row = ws.max_row
    if max_row >= 2:
        for row_idx in range(2, max_row + 1):
            row_data = []
            for col in range(1, 12):
                row_data.append(ws.cell(row=row_idx, column=col).value)

            order_num = str(row_data[3] or "").strip()
            detail_num = str(row_data[4] or "").strip()

            if order_num:
                existing_rows.append(row_data)
                key = f"{order_num}|{detail_num}"
                if key not in existing_keys:
                    existing_keys[key] = len(existing_rows) - 1

    # === 新規データ処理 ===
    new_items: list[list] = []

    for record in new_orders:
        key = f"{record.order_number}|{record.detail_number}"
        record_sender = sender or record.sender or _get_default_sender()

        if key in existing_keys:
            # 重複：ステータス（9列目=index 8）を更新
            idx = existing_keys[key]
            if idx >= 0:
                existing_rows[idx][8] = record.status
        else:
            # 新規
            new_row = [
                execution_time,
                record.order_date,
                record.customer_name,
                record.order_number,
                _to_int_detail(record.detail_number),
                record.manufacturer_name,
                record.product_name,
                record.inquiry_status,  # 問合せ状況（デフォルト「未」）
                record.status,           # ステータス
                None,                    # 受注納期（ユーザー手入力欄。常に空で書き込む）
                record_sender,
            ]
            new_items.append(new_row)
            existing_keys[key] = -1

    # === 結果をシートに書き戻し ===
    all_rows = new_items + existing_rows
    if not all_rows:
        return

    # 送付日時の降順でソート
    all_rows.sort(
        key=lambda r: r[0] if isinstance(r[0], datetime.datetime) else datetime.datetime.min,
        reverse=True,
    )

    _write_rows_to_sheet(ws, all_rows, len(_CONFIRMING_HEADERS))

    # 問合せ状況列（H列=8列目）に入力規則を設定
    _set_inquiry_validation(ws, len(all_rows))

    # 表示形式
    for row_idx in range(2, len(all_rows) + 2):
        ws.cell(row=row_idx, column=1).number_format = "mm/dd hh:mm"
        ws.cell(row=row_idx, column=2).number_format = "mm/dd"

    # 色分け
    color_confirming_list(ws)


# ============================================
# VBA: CleanConfirmingList (L2057-2145)
# 確認中→送付履歴への移動
# ============================================
def clean_confirming_list(
    ws_history: Worksheet,
    ws_confirming: Worksheet,
    confirmed_orders: list[HistoryRecord],
    execution_time: Optional[datetime.datetime] = None,
    sender: str = "",
) -> None:
    """確定した伝票を確認中一覧から削除し、送付履歴に移動する。

    Args:
        ws_history: 送付履歴シート
        ws_confirming: 確認中一覧シート
        confirmed_orders: 確定した伝票のリスト
        execution_time: 実行時刻
        sender: 送付者名
    """
    if not confirmed_orders:
        return

    # 確定伝票のキーセット
    confirmed_keys: dict[str, HistoryRecord] = {}
    for record in confirmed_orders:
        key = f"{record.order_number}|{record.detail_number}"
        confirmed_keys[key] = record

    # === 確認中テーブルから残す行と移動する行を分類 ===
    keep_rows: list[list] = []
    moved_orders: list[HistoryRecord] = []

    max_row = ws_confirming.max_row
    if max_row < 2:
        return

    for row_idx in range(2, max_row + 1):
        order_num = str(
            ws_confirming.cell(row=row_idx, column=4).value or ""
        ).strip()
        detail_num = str(
            ws_confirming.cell(row=row_idx, column=5).value or ""
        ).strip()

        conf_key = f"{order_num}|{detail_num}"

        if conf_key in confirmed_keys:
            moved_orders.append(confirmed_keys[conf_key])
        else:
            row_data = []
            for col in range(1, 12):
                row_data.append(ws_confirming.cell(row=row_idx, column=col).value)
            if order_num:  # 空行除外
                keep_rows.append(row_data)

    # === 確認中テーブルを残す行だけで再構築 ===
    _write_rows_to_sheet(ws_confirming, keep_rows, len(_CONFIRMING_HEADERS))

    # 入力規則を再設定
    if keep_rows:
        _set_inquiry_validation(ws_confirming, len(keep_rows))

    # === 移動した伝票を送付履歴に追加 ===
    if moved_orders:
        save_delivery_history(
            ws_history, moved_orders,
            execution_time=execution_time,
            sender=sender,
        )


# ============================================
# VBA: CleanOldHistory (L2150-2217)
# 古い送付履歴レコードの削除
# ============================================
def clean_old_history(
    ws: Worksheet,
    days_to_keep: int = 180,
    today: Optional[datetime.date] = None,
) -> int:
    """送付履歴テーブルから古いレコードを削除する。

    Args:
        ws: 送付履歴シート
        days_to_keep: 保持日数（デフォルト180日）
        today: 基準日（テスト用）

    Returns:
        削除した行数
    """
    return _clean_old_records(ws, days_to_keep, len(_HISTORY_HEADERS), today)


# ============================================
# VBA: CleanOldConfirmingList (L6506-6573)
# 古い確認中一覧レコードの削除
# ============================================
def clean_old_confirming_list(
    ws: Worksheet,
    days_to_keep: int = 180,
    today: Optional[datetime.date] = None,
) -> int:
    """確認中一覧テーブルから古いレコードを削除する。

    Args:
        ws: 確認中一覧シート
        days_to_keep: 保持日数（デフォルト180日）
        today: 基準日（テスト用）

    Returns:
        削除した行数
    """
    return _clean_old_records(ws, days_to_keep, len(_CONFIRMING_HEADERS), today)


# ============================================
# ヘルパー関数
# ============================================
def _parse_date_value(val) -> Optional[datetime.date]:
    """セル値をdatetime.dateに変換する。"""
    if isinstance(val, datetime.datetime):
        return val.date()
    if isinstance(val, datetime.date):
        return val
    return None


def _write_rows_to_sheet(
    ws: Worksheet,
    rows: list[list],
    col_count: int,
) -> None:
    """シートのデータ部分（行2以降）をrowsで置き換える。"""
    # 既存データをクリア
    max_row = ws.max_row
    if max_row >= 2:
        for row_idx in range(2, max_row + 1):
            for col in range(1, col_count + 1):
                ws.cell(row=row_idx, column=col).value = None

    # 新しいデータを書き込み
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            if col_idx <= col_count:
                ws.cell(row=row_idx, column=col_idx).value = value

    # テーブル参照範囲の更新
    _update_table_ref(ws, len(rows), col_count)


def _update_table_ref(ws: Worksheet, data_row_count: int, col_count: int) -> None:
    """テーブルの参照範囲を更新する。"""
    if not ws.tables:
        return

    last_col = get_column_letter(col_count)
    # データが0行でもヘッダーは残す（最小2行）
    last_row = max(data_row_count + 1, 1)

    for table_name in list(ws.tables):
        tbl = ws.tables[table_name]
        tbl.ref = f"A1:{last_col}{last_row}"


def _set_inquiry_validation(ws: Worksheet, data_row_count: int) -> None:
    """問合せ状況列（H列）に入力規則を設定する。"""
    if data_row_count <= 0:
        return

    # 既存のバリデーションをクリア
    ws.data_validations = ws.data_validations.__class__()

    dv = DataValidation(
        type="list",
        formula1='"未,済,回答待ち,除外"',
        allow_blank=True,
    )
    dv.error = "未/済/回答待ち/除外 から選択してください"
    dv.errorTitle = "入力エラー"

    dv.add(f"H2:H{data_row_count + 1}")
    ws.add_data_validation(dv)


def _clean_old_records(
    ws: Worksheet,
    days_to_keep: int,
    col_count: int,
    today: Optional[datetime.date] = None,
) -> int:
    """古いレコードを削除する共通処理。

    Returns:
        削除した行数
    """
    if today is None:
        today = datetime.date.today()

    cutoff_date = today - datetime.timedelta(days=days_to_keep)

    max_row = ws.max_row
    if max_row < 2:
        return 0

    keep_rows: list[list] = []
    total_rows = 0

    for row_idx in range(2, max_row + 1):
        row_data = []
        for col in range(1, col_count + 1):
            row_data.append(ws.cell(row=row_idx, column=col).value)

        # 空行チェック
        if not any(v is not None and str(v).strip() for v in row_data):
            continue

        total_rows += 1

        # 送付日時（1列目）をパース
        sent_date = _parse_date_value(row_data[0])

        # 日付不明 or cutoff以降 → 残す
        if sent_date is None or sent_date >= cutoff_date:
            keep_rows.append(row_data)

    deleted_count = total_rows - len(keep_rows)

    if deleted_count > 0:
        _write_rows_to_sheet(ws, keep_rows, col_count)

    return deleted_count


# ============================================
# 高速バッチ処理（read_only=True → インメモリ → 新規Workbook書き出し）
# ============================================
def extract_sheet_rows(ws, col_count: int) -> list[list]:
    """read_only=True のワークシートからデータ行をメモリに読み出す。

    Args:
        ws: read_only=True で開いたワークシート
        col_count: 読み取る列数

    Returns:
        各行のリスト（ヘッダー行は除く、受発注伝票が空の行はスキップ）
    """
    rows: list[list] = []
    for row_data in ws.iter_rows(min_row=2, max_col=col_count, values_only=True):
        if not row_data:
            continue
        # 受発注伝票（4列目=index 3）が空の行はスキップ
        order_num = row_data[3] if len(row_data) > 3 else None
        if order_num is None or str(order_num).strip() == "":
            continue
        rows.append(list(row_data))
    return rows


def save_history_batch(
    file_path: str,
    history_rows: list[list],
    confirming_rows: list[list],
    new_confirmed: list[HistoryRecord],
    new_confirming: list[ConfirmingRecord],
    execution_time: datetime.datetime,
    sender: str = "",
    days_to_keep: int = 180,
    today: datetime.date | None = None,
) -> None:
    """送付履歴ファイルの全更新を1回のバッチで実行する。

    read_only=True で読み込んだ行データをインメモリで処理し、
    新規Workbookに書き出す。load_workbook(read_only=False) を使わないため高速。

    Args:
        file_path: 送付履歴.xlsx のパス
        history_rows: 既存の送付履歴データ（extract_sheet_rowsで読み込み済み）
        confirming_rows: 既存の確認中一覧データ（同上）
        new_confirmed: 新規確定伝票
        new_confirming: 新規確認中伝票
        execution_time: 実行時刻
        sender: 送付者名
        days_to_keep: 保持日数（デフォルト180日）
        today: 基準日（テスト用）
    """
    if today is None:
        today = datetime.date.today()

    default_sender = sender or _get_default_sender()

    # 作業用コピー（元データを破壊しない）
    conf_rows = [row[:] for row in confirming_rows]
    hist_rows = [row[:] for row in history_rows]

    # --- ステップ1: 確認中一覧から確定キーを除去 → moved_orders作成 ---
    confirmed_keys: dict[str, HistoryRecord] = {}
    for record in new_confirmed:
        key = f"{record.order_number}|{record.detail_number}"
        confirmed_keys[key] = record

    moved_orders: list[HistoryRecord] = []
    if confirmed_keys:
        keep_conf: list[list] = []
        for row in conf_rows:
            order_num = str(row[3] or "").strip()
            detail_num = str(row[4] or "").strip()
            conf_key = f"{order_num}|{detail_num}"
            if conf_key in confirmed_keys:
                moved_orders.append(confirmed_keys[conf_key])
            else:
                keep_conf.append(row)
        conf_rows = keep_conf

    # --- ステップ2: 確認中一覧に new_confirming を追加（重複時はステータス更新） ---
    if new_confirming:
        existing_conf_keys: dict[str, int] = {}
        for i, row in enumerate(conf_rows):
            order_num = str(row[3] or "").strip()
            detail_num = str(row[4] or "").strip()
            if order_num:
                key = f"{order_num}|{detail_num}"
                if key not in existing_conf_keys:
                    existing_conf_keys[key] = i

        new_conf_items: list[list] = []
        for record in new_confirming:
            key = f"{record.order_number}|{record.detail_number}"
            record_sender = sender or record.sender or default_sender

            if key in existing_conf_keys:
                # 重複：ステータス（9列目=index 8）を更新
                idx = existing_conf_keys[key]
                if idx >= 0:
                    conf_rows[idx][8] = record.status
            else:
                new_row = [
                    execution_time,
                    record.order_date,
                    record.customer_name,
                    record.order_number,
                    _to_int_detail(record.detail_number),
                    record.manufacturer_name,
                    record.product_name,
                    record.inquiry_status,  # 問合せ状況（デフォルト「未」）
                    record.status,           # ステータス
                    None,                    # 受注納期（ユーザー手入力欄。常に空で書き込む）
                    record_sender,
                ]
                new_conf_items.append(new_row)
                existing_conf_keys[key] = -1

        conf_rows = new_conf_items + conf_rows

    # --- ステップ3: 確認中一覧の古いレコードを除去 ---
    cutoff_date = today - datetime.timedelta(days=days_to_keep)
    conf_rows = _filter_old_rows(conf_rows, cutoff_date)

    # --- ステップ4: 確認中一覧をソート（送付日時降順） ---
    conf_rows.sort(
        key=lambda r: r[0] if isinstance(r[0], datetime.datetime) else datetime.datetime.min,
        reverse=True,
    )

    # --- ステップ5: 送付履歴に new_confirmed + moved_orders を追加 ---
    all_new_history = list(new_confirmed) + moved_orders
    if all_new_history:
        existing_hist_keys: dict[str, int] = {}
        for i, row in enumerate(hist_rows):
            order_num = str(row[3] or "").strip()
            detail_num = str(row[4] or "").strip()
            if order_num:
                key = f"{order_num}|{detail_num}"
                if key not in existing_hist_keys:
                    existing_hist_keys[key] = i

        new_hist_items: list[list] = []
        for record in all_new_history:
            key = f"{record.order_number}|{record.detail_number}"
            record_sender = sender or record.sender or default_sender

            if key in existing_hist_keys:
                # 重複：納品済みの場合は既存行の納期回答を更新
                if record.delivery_answer == "納品済み":
                    idx = existing_hist_keys[key]
                    hist_rows[idx][7] = "納品済み"
            else:
                new_row = [
                    execution_time,
                    record.order_date,
                    record.customer_name,
                    record.order_number,
                    _to_int_detail(record.detail_number),
                    record.manufacturer_name,
                    record.product_name,
                    record.delivery_answer,
                    record_sender,
                ]
                new_hist_items.append(new_row)
                existing_hist_keys[key] = -1

        hist_rows = new_hist_items + hist_rows

    # --- ステップ6: 送付履歴の古いレコードを除去 ---
    hist_rows = _filter_old_rows(hist_rows, cutoff_date)

    # --- ステップ7: 送付履歴をソート（送付日時降順） ---
    hist_rows.sort(
        key=lambda r: r[0] if isinstance(r[0], datetime.datetime) else datetime.datetime.min,
        reverse=True,
    )

    # --- ステップ8: 新規Workbook書き出し ---
    _write_history_workbook(file_path, hist_rows, conf_rows, today)


def _filter_old_rows(rows: list[list], cutoff_date: datetime.date) -> list[list]:
    """cutoff_date より古い行を除外する。日付不明は残す。"""
    keep: list[list] = []
    for row in rows:
        sent_date = _parse_date_value(row[0]) if row else None
        if sent_date is None or sent_date >= cutoff_date:
            keep.append(row)
    return keep


def _write_history_workbook(
    file_path: str,
    history_rows: list[list],
    confirming_rows: list[list],
    today: datetime.date | None = None,
) -> None:
    """送付履歴Workbookを新規作成し、データを書き出す。

    initialize_delivery_history と同じ構造（テーブル・列幅・表示形式・入力規則）で
    新規Workbookを作成し、インメモリの行データを書き込んで保存する。

    Args:
        file_path: 保存先ファイルパス
        history_rows: 送付履歴のデータ行
        confirming_rows: 確認中一覧のデータ行
        today: 基準日（色分け用）
    """
    from nouki_kaitou.excel_writer import color_confirming_list

    wb = Workbook()

    # === 送付履歴シート ===
    ws_history = wb.active
    ws_history.title = HISTORY_SHEET_NAME

    # ヘッダー
    for col_idx, header in enumerate(_HISTORY_HEADERS, 1):
        ws_history.cell(row=1, column=col_idx).value = header

    # 列幅
    for col_idx, width in enumerate(_HISTORY_WIDTHS, 1):
        ws_history.column_dimensions[get_column_letter(col_idx)].width = width

    # データ書き込み
    for row_idx, row_data in enumerate(history_rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            if col_idx <= len(_HISTORY_HEADERS):
                ws_history.cell(row=row_idx, column=col_idx).value = value

    # テーブル設定
    hist_last_row = max(len(history_rows) + 1, 1)
    tbl_history = Table(
        displayName=_HISTORY_TABLE_NAME,
        ref=f"A1:I{hist_last_row}",
    )
    tbl_history.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True
    )
    ws_history.add_table(tbl_history)

    # 表示形式
    for row_idx in range(2, len(history_rows) + 2):
        ws_history.cell(row=row_idx, column=1).number_format = "mm/dd hh:mm"
        ws_history.cell(row=row_idx, column=2).number_format = "mm/dd"

    # === 確認中一覧シート ===
    ws_confirming = wb.create_sheet(CONFIRMING_SHEET_NAME)

    # ヘッダー
    for col_idx, header in enumerate(_CONFIRMING_HEADERS, 1):
        ws_confirming.cell(row=1, column=col_idx).value = header

    # 列幅
    for col_idx, width in enumerate(_CONFIRMING_WIDTHS, 1):
        ws_confirming.column_dimensions[get_column_letter(col_idx)].width = width

    # データ書き込み
    for row_idx, row_data in enumerate(confirming_rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            if col_idx <= len(_CONFIRMING_HEADERS):
                ws_confirming.cell(row=row_idx, column=col_idx).value = value

    # テーブル設定
    conf_last_row = max(len(confirming_rows) + 1, 1)
    tbl_confirming = Table(
        displayName=_CONFIRMING_TABLE_NAME,
        ref=f"A1:K{conf_last_row}",
    )
    tbl_confirming.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium1", showRowStripes=True
    )
    ws_confirming.add_table(tbl_confirming)

    # 入力規則
    if confirming_rows:
        _set_inquiry_validation(ws_confirming, len(confirming_rows))

    # 表示形式
    for row_idx in range(2, len(confirming_rows) + 2):
        ws_confirming.cell(row=row_idx, column=1).number_format = "mm/dd hh:mm"
        ws_confirming.cell(row=row_idx, column=2).number_format = "mm/dd"

    # 色分け
    color_confirming_list(ws_confirming, today)

    wb.save(file_path)
