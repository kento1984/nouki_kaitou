"""配達/出荷の使い分けWARNING 17件の保管場所確認

validation_result.xlsxの「配達/出荷の使い分け」カテゴリの注番について、
保管場所が「他拠点より出荷」や別拠点からの出荷になっているか確認する。
"""

import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import load_workbook
from nouki_kaitou.data_loader import load_source_file, get_column_positions

def main():
    # validation_result.xlsxから配達/出荷の使い分けの注番を取得
    validation_path = Path(__file__).parent / "validation_result.xlsx"

    print("validation_result.xlsxを読み込み中...")
    wb = load_workbook(validation_path, read_only=True, data_only=True)

    # 配達_出荷の使い分けシートを探す
    target_sheet = None
    for sheet_name in wb.sheetnames:
        if "配達" in sheet_name and "出荷" in sheet_name:
            target_sheet = sheet_name
            break

    if not target_sheet:
        print(f"シートが見つかりません。利用可能なシート: {wb.sheetnames}")
        return

    print(f"シート: {target_sheet}")
    ws = wb[target_sheet]

    # 注番リストを取得（ヘッダー行をスキップ）
    order_details = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if row and row[0]:  # 注番列
            order_num = str(row[0]).strip()
            detail_num = str(row[1] or "").strip() if len(row) > 1 else ""
            problem = str(row[8] or "").strip() if len(row) > 8 else ""  # 問題列
            order_details.append((order_num, detail_num, problem))

    wb.close()
    print(f"対象件数: {len(order_details)}件\n")

    # 16AM.xlsから保管場所を取得
    print("16AM.xlsを読み込み中...")
    xls_path = Path(r"\\flsv04\316京葉\納期回答書ツールフォルダ\受注一覧\16AM.xls")
    source_data = load_source_file(xls_path)
    cols = get_column_positions(source_data)

    if cols is None:
        print("列位置を取得できませんでした")
        return

    # 保管場所の列を取得
    col_order = cols.get("受発注伝票", 0)
    col_detail = cols.get("明細", 1)
    col_storage = cols.get("保管場所")
    col_customer = cols.get("受注先")
    col_ship_to = cols.get("出荷先名")
    col_product = cols.get("品名")

    print(f"保管場所列: {col_storage}")

    # 注番→保管場所のマッピングを作成
    storage_map = {}
    for row_idx in range(5, len(source_data)):
        row = source_data[row_idx]
        if len(row) <= col_order:
            continue

        order_num = str(row[col_order] or "").strip()
        if not order_num:
            continue

        detail_num = str(row[col_detail] if col_detail < len(row) else "").strip()
        storage = str(row[col_storage] if col_storage and col_storage < len(row) else "").strip()
        customer = str(row[col_customer] if col_customer and col_customer < len(row) else "").strip()
        ship_to = str(row[col_ship_to] if col_ship_to and col_ship_to < len(row) else "").strip()
        product = str(row[col_product] if col_product and col_product < len(row) else "").strip()

        key = f"{order_num}|{detail_num}"
        storage_map[key] = {
            "storage": storage,
            "customer": customer[:20],
            "ship_to": ship_to[:20],
            "product": product[:30],
        }

    print(f"読み込み完了\n")

    # 結果表示
    print("=" * 100)
    print("配達/出荷の使い分け WARNING 17件の保管場所確認")
    print("=" * 100)

    other_branch_count = 0
    same_branch_count = 0
    not_found_count = 0

    for order_num, detail_num, problem in order_details:
        key = f"{order_num}|{detail_num}"

        if key in storage_map:
            info = storage_map[key]
            storage = info["storage"]

            # 他拠点かどうか判定
            is_other_branch = (
                "他拠点" in storage or
                storage not in ("", "1600", "16京葉", "京葉") and storage.strip() != ""
            )

            marker = "○他拠点" if is_other_branch else "×同一拠点"

            if is_other_branch:
                other_branch_count += 1
            else:
                same_branch_count += 1

            print(f"\n{order_num}|{detail_num}")
            print(f"  保管場所: {storage} → {marker}")
            print(f"  受注先: {info['customer']}")
            print(f"  出荷先: {info['ship_to']}")
            print(f"  品名: {info['product']}")
            print(f"  問題: {problem[:50]}")
        else:
            not_found_count += 1
            print(f"\n{order_num}|{detail_num} - データなし")

    # 集計
    print("\n" + "=" * 100)
    print("集計結果")
    print("=" * 100)
    print(f"他拠点からの出荷（問題なし）: {other_branch_count}件")
    print(f"同一拠点（要確認）: {same_branch_count}件")
    print(f"データなし: {not_found_count}件")

    if same_branch_count > 0:
        print("\n【要確認】同一拠点なのに「出荷予定」になっているケース:")
        for order_num, detail_num, problem in order_details:
            key = f"{order_num}|{detail_num}"
            if key in storage_map:
                info = storage_map[key]
                storage = info["storage"]
                is_other_branch = (
                    "他拠点" in storage or
                    storage not in ("", "1600", "16京葉", "京葉") and storage.strip() != ""
                )
                if not is_other_branch:
                    print(f"  {order_num}|{detail_num} - 保管場所: {storage}")

if __name__ == "__main__":
    main()
