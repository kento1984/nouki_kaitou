"""GL2Z446911の明細10と20の両方を確認 + VBA回答書の本多酸素ファイルの全行を出力"""

import sys
import io
import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from openpyxl import load_workbook

from nouki_kaitou.data_loader import (
    get_column_positions,
    get_data_rows_range,
    is_data_row,
    load_source_file,
    parse_order_row,
)


def main():
    tool_folder = Path(r"\\flsv04\316京葉\納期回答書ツールフォルダ")
    source_path = tool_folder / "受注一覧" / "17PM.xls"
    vba_folder = tool_folder / "納期回答書" / "2月17日(火)_②回目"

    # SAPデータからGL2Z446911の全明細を取得
    print("=" * 80)
    print("GL2Z446911 の全明細（SAPデータ）")
    print("=" * 80)

    source_data_raw = load_source_file(str(source_path))
    cols = get_column_positions(source_data_raw)

    for i in get_data_rows_range(source_data_raw, cols):
        if is_data_row(source_data_raw, i, cols):
            row = parse_order_row(source_data_raw, i, cols)
            if row.order_number == "GL2Z446911":
                print(f"\n  明細: {row.detail_number}")
                print(f"    顧客名:       [{row.customer_name}]")
                print(f"    出荷先:       [{row.ship_to_name}]")
                print(f"    品名:         [{row.product_name}]")
                print(f"    伝票タイプ:   [{row.document_type}]")
                print(f"    保管場所:     [{row.storage_place}]")
                print(f"    出荷ステータス: [{row.ship_status}]")
                print(f"    指定納期:     [{row.specified_delivery_date}]")
                print(f"    受注納期:     [{row.order_delivery_date}]")
                print(f"    登録日時:     [{row.registration_date}] [{row.time_value}]")
                print(f"    数量:         [{row.quantity}]")
                print(f"    単価:         [{row.unit_price}]")
                print(f"    金額:         [{row.net_amount}]")
                print(f"    コメント社内: [{row.comment_internal}]")
                print(f"    メーカー名:   [{row.manufacturer_name}]")

    # VBA回答書の本多酸素（株）八潮営業所ファイルの全行を出力
    print("\n" + "=" * 80)
    print("VBA回答書: 本多酸素（株）　八潮営業所　の全行")
    print("=" * 80)

    vba_file = vba_folder / "納期回答書_本多酸素（株）　八潮営業所様_20260217.xlsx"
    wb = load_workbook(str(vba_file), data_only=True)
    ws = wb.active

    # ヘッダー行を探す
    header_row = None
    for row_idx in range(1, min(20, ws.max_row + 1)):
        cell_val = ws.cell(row=row_idx, column=1).value
        if cell_val and str(cell_val).strip() == "受注日":
            header_row = row_idx
            break

    print(f"ヘッダー行: {header_row}")
    print(f"ヘッダー: ", end="")
    for col_idx in range(1, 12):
        print(f"[{ws.cell(row=header_row, column=col_idx).value}]", end=" ")
    print()

    print(f"\n全データ行:")
    for row_idx in range(header_row + 1, ws.max_row + 1):
        order_date_val = ws.cell(row=row_idx, column=1).value
        if not order_date_val:
            continue

        unit_price_val = ws.cell(row=row_idx, column=7).value
        if unit_price_val and "※" in str(unit_price_val):
            print(f"\n  行{row_idx}: (フッター行)")
            break

        vals = {}
        for col_idx in range(1, 12):
            vals[col_idx] = ws.cell(row=row_idx, column=col_idx).value

        print(f"\n  行{row_idx}:")
        print(f"    受注日={vals[1]}")
        print(f"    担当者={vals[2]}")
        print(f"    注番={vals[3]}")
        print(f"    メーカー={vals[4]}")
        print(f"    品名={vals[5]}")
        print(f"    数量={vals[6]}")
        print(f"    単価={vals[7]}")
        print(f"    金額={vals[8]}")
        print(f"    納期回答={vals[9]}")
        print(f"    納入先={vals[10]}")
        print(f"    備考={vals[11]}")

    wb.close()


if __name__ == "__main__":
    main()
