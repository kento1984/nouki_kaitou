"""パターン1（VBA=納品済み / PY=確認中）の7件を調査"""

import sys
import io
import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from openpyxl import load_workbook

from nouki_kaitou.cache import build_all_caches
from nouki_kaitou.config import load_branch_settings, load_holidays
from nouki_kaitou.data_loader import (
    get_column_positions,
    get_data_rows_range,
    is_data_row,
    load_source_file,
    parse_order_row,
)


def main():
    tool_folder = Path(r"\\flsv04\316京葉\納期回答書ツールフォルダ")

    # 比較結果からVBA=納品済み/PY=確認中の注番を抽出
    result_path = Path(__file__).parent / "comparison_result.xlsx"
    wb_result = load_workbook(str(result_path), data_only=True)
    ws_result = wb_result["比較結果"]

    pattern1_orders = []
    for row in ws_result.iter_rows(min_row=2, values_only=True):
        if row[4] == "納品済み" and row[5] == "確認中":
            order_key = row[0]  # 注番|明細
            if order_key and order_key not in pattern1_orders:
                pattern1_orders.append(order_key)

    wb_result.close()

    print("=" * 80)
    print("パターン1: VBA=納品済み / PY=確認中 の注番リスト")
    print("=" * 80)
    print(f"件数: {len(pattern1_orders)}")
    for o in pattern1_orders:
        print(f"  {o}")

    # 送付履歴を読み込んで登録日を確認
    print()
    print("=" * 80)
    print("送付履歴の確認")
    print("=" * 80)

    history_path = tool_folder / "送付履歴.xlsx"
    wb_history = load_workbook(str(history_path), data_only=True, read_only=True)
    ws_history = wb_history["送付履歴"]

    # 送付履歴の構造を確認（ヘッダー行）
    header_row = list(ws_history.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    print(f"\n送付履歴ヘッダー: {header_row[:10]}")

    # 各注番の送付履歴を検索
    print()
    print("-" * 60)
    today = datetime.date.today()
    print(f"今日: {today}")
    print("-" * 60)

    # 列の位置: 送付日時(0), 受注日(1), 顧客名(2), 受発注伝票(3), 明細(4), メーカー名(5), 品名(6), 納期回答(7)
    for order_key in pattern1_orders:
        order_no, detail_no = order_key.split("|")
        print(f"\n【{order_key}】")

        found = False
        for row in ws_history.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 8:
                continue

            history_order = str(row[3] or "").strip()
            history_detail = str(row[4] or "").strip()

            if history_order == order_no and history_detail == detail_no:
                sent_datetime = row[0]
                order_date = row[1]
                answer = row[7]

                if isinstance(sent_datetime, datetime.datetime):
                    sent_date = sent_datetime.date()
                elif isinstance(sent_datetime, datetime.date):
                    sent_date = sent_datetime
                else:
                    sent_date = None

                print(f"  送付履歴に存在")
                print(f"    送付日時: {sent_datetime}")
                print(f"    受注日: {order_date}")
                print(f"    納期回答: {answer}")
                print(f"    今日送付？: {sent_date == today if sent_date else 'N/A'}")
                found = True
                break

        if not found:
            print(f"  送付履歴になし")

    wb_history.close()

    # SAPデータで登録日も確認
    print()
    print("=" * 80)
    print("SAPデータの登録日（受注日）確認")
    print("=" * 80)

    source_path = tool_folder / "受注一覧" / "17PM.xls"
    source_data_raw = load_source_file(str(source_path))
    cols = get_column_positions(source_data_raw)

    for order_key in pattern1_orders:
        order_no, detail_no = order_key.split("|")
        print(f"\n【{order_key}】")

        for i in get_data_rows_range(source_data_raw, cols):
            if is_data_row(source_data_raw, i, cols):
                row = parse_order_row(source_data_raw, i, cols)
                if row.order_number == order_no and row.detail_number == detail_no:
                    print(f"  登録日(registration_date): {row.registration_date}")
                    print(f"  今日より前？: {row.registration_date < today if row.registration_date else 'N/A'}")
                    print(f"  出荷ステータス: {row.ship_status}")
                    print(f"  保管場所: {row.storage_place}")
                    break


if __name__ == "__main__":
    main()
