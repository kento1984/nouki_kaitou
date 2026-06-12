"""Excel出力・書式設定モジュール

VBAの以下の関数を移植:
- CreateHeader (L2933): ヘッダー行生成（行1-6）
- CopyDataRow (L3784): データ行のコピー・変換
- FormatReport (L4037): 書式設定全般（色分け・分納行挿入・署名・連絡事項）
- ColorConfirmingList (L1954): 確認中一覧の色分け
- CheckSameDateInBunno (L4943): 分納の同日判定
"""

from __future__ import annotations

import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from nouki_kaitou.models import (
    BranchSettings,
    BunnoEntry,
    CacheStore,
    HolidayMap,
    ReportRow,
    StockoutEntry,
    TrackingEntry,
    is_stockout_confirmed,
)

# ============================================
# 定数：色・フォント
# ============================================
_FONT_NAME = "游ゴシック"

# タイトル行（行1）
_TITLE_BG = "142846"        # RGB(20, 40, 70)
_TITLE_FG = "FFFFFF"
_ACCENT_BG = "B49646"       # RGB(180, 150, 70)

# ヘッダー行（行6）
_HEADER_BG = "23375A"       # RGB(35, 55, 90)
_HEADER_FG = "FFFFFF"

# 偶数行
_EVEN_ROW_BG = "E1EBF8"     # RGB(225, 235, 248)

# 納期回答列の色定義
_COLOR_DELIVERED = ("DCDCDC", "505050")         # 納品済み
_COLOR_STOCKOUT = ("DC5014", "FFFFFF")          # 欠品中
_COLOR_STOCKOUT_PARTIAL = ("FF9678", "8C2814")  # （欠品）
_COLOR_BUNNO = ("C8DCFF", "00468C")             # 分納
_COLOR_CONFIRMING = ("FFC8C8", "B41E1E")        # 確認中
_COLOR_SCHEDULING = ("FAF5DC", "8C6428")        # 日程調整中
_COLOR_WORK = ("E1D2F5", "503278")              # 作業
_COLOR_OTHER_BRANCH = ("FFDCB4", "B45000")      # 他拠点より
_COLOR_PICKED_UP = ("DCCDF0", "46286E")         # 引取済み
_COLOR_DONE = ("C8F0D2", "146432")              # ○○済み
_COLOR_SHIP_SOON = ("C8E1FF", "14468C")         # 今日/明日出荷予定
_COLOR_SHIP_LATER = ("FFEBB4", "8C5A0A")        # それ以降出荷予定
_COLOR_DELIVER_SOON = ("C8E1FF", "14468C")       # 明日/明後日配達予定
_COLOR_DELIVER_LATER = ("FFEBB4", "8C5A0A")      # それ以降配達予定
_COLOR_PICKUP_PLAN = ("E6D7FA", "5A3282")        # 引取予定
_COLOR_OTHER_PLAN = ("FFEBB4", "8C5A0A")         # その他予定
_COLOR_DEFAULT = ("FAF5DC", "505050")            # デフォルト

# 確認中一覧の色
_CONFIRM_SHIP_DONE = ("FFB4B4", None)    # 出荷完了
_CONFIRM_PARTIAL = ("C8FFD4", None)      # 一部処理済み
_CONFIRM_STOCKOUT = ("DCC8FF", None)     # 欠品中
_CONFIRM_BUNNO = ("C8DCFF", None)        # 分納
_CONFIRM_PRICE_PENDING = ("FFD2A0", None)  # 価格確認中（オレンジ）
_CONFIRM_WEEK_OLD = ("FFC896", None)     # 1週間以上
_CONFIRM_THREE_DAYS = ("FFFFC8", None)   # 3日以上

# 列幅
# VBAのColumnWidthはxlsx保存時に約0.7のパディングが加算される。
# openpyxlではそのまま保存されるため、VBA値+0.7で補正する。
_VBA_WIDTH_OFFSET = 0.7
_COLUMN_WIDTHS = [
    9 + _VBA_WIDTH_OFFSET,    # A: 受注日
    16 + _VBA_WIDTH_OFFSET,   # B: 担当者様
    14 + _VBA_WIDTH_OFFSET,   # C: 貴社注番
    20 + _VBA_WIDTH_OFFSET,   # D: メーカー名
    50 + _VBA_WIDTH_OFFSET,   # E: 品名
    7 + _VBA_WIDTH_OFFSET,    # F: 数量
    9 + _VBA_WIDTH_OFFSET,    # G: 単価
    11 + _VBA_WIDTH_OFFSET,   # H: 金額
    18 + _VBA_WIDTH_OFFSET,   # I: 納期回答
    20 + _VBA_WIDTH_OFFSET,   # J: 納入先名
    26 + _VBA_WIDTH_OFFSET,   # K: 備考
    14 + _VBA_WIDTH_OFFSET,   # L: 弊社注番
]

# ヘッダー列名
_HEADER_LABELS = [
    "受注日", "担当者様", "貴社注番", "メーカー名", "品名",
    "数量", "単価", "金額", "納期回答", "納入先名", "備考", "弊社注番",
]

# TWF専用回答書のレイアウト（期間限定。twf.py参照）
# A=TWF No., C=お客様名 に差し替え。D〜L列は通常版と完全一致のため
# 書式エンジン（色分け・数値書式・罫線）はそのまま流用できる
_TWF_HEADER_LABELS = [
    "TWF No.", "担当者様", "お客様名", "メーカー名", "品名",
    "数量", "単価", "金額", "納期回答", "納入先名", "備考", "弊社注番",
]
_TWF_COLUMN_WIDTHS = list(_COLUMN_WIDTHS)
_TWF_COLUMN_WIDTHS[0] = 10 + _VBA_WIDTH_OFFSET   # A: TWF No.
_TWF_COLUMN_WIDTHS[2] = 22 + _VBA_WIDTH_OFFSET   # C: お客様名


def _make_fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _make_font(
    size: int = 10,
    bold: bool = False,
    color: str = "000000",
    italic: bool = False,
    underline: str | None = None,
) -> Font:
    return Font(
        name=_FONT_NAME, size=size, bold=bold, color=color,
        italic=italic, underline=underline,
    )


# --- 頻出スタイルのキャッシュ（パフォーマンス最適化） ---
_FONT_10 = Font(name=_FONT_NAME, size=10)
_FONT_10_BOLD = Font(name=_FONT_NAME, size=10, bold=True)
_FONT_PRICE_CONFIRM = Font(name=_FONT_NAME, size=10, bold=True, color="B41E1E")
_ALIGN_SHRINK = Alignment(shrink_to_fit=True)
_ALIGN_CENTER_SHRINK = Alignment(horizontal="center", shrink_to_fit=True)
_ALIGN_CENTER = Alignment(horizontal="center")
_FILL_EVEN = PatternFill(
    start_color=_EVEN_ROW_BG, end_color=_EVEN_ROW_BG, fill_type="solid"
)
_THIN_SIDE = Side(style="thin", color="A0A0A0")
_THIN_BORDER = Border(
    left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE
)

# 納期回答色のFillキャッシュ
_DELIVERY_FILL_CACHE: dict[str, PatternFill] = {}
_DELIVERY_FONT_CACHE: dict[str, Font] = {}


def _get_delivery_fill(hex_color: str) -> PatternFill:
    """納期回答色のFillをキャッシュ付きで取得する。"""
    if hex_color not in _DELIVERY_FILL_CACHE:
        _DELIVERY_FILL_CACHE[hex_color] = _make_fill(hex_color)
    return _DELIVERY_FILL_CACHE[hex_color]


def _get_delivery_font(hex_color: str) -> Font:
    """納期回答色のFontをキャッシュ付きで取得する。"""
    if hex_color not in _DELIVERY_FONT_CACHE:
        _DELIVERY_FONT_CACHE[hex_color] = Font(
            name=_FONT_NAME, size=10, bold=True, color=hex_color
        )
    return _DELIVERY_FONT_CACHE[hex_color]


# ============================================
# VBA: CreateHeader (L2933-2999)
# ヘッダー行生成（行1-6）
# ============================================
def create_header(
    ws: Worksheet,
    customer_name: str,
    rep_name: str = "",
    issue_date: Optional[datetime.date] = None,
    branch: Optional[BranchSettings] = None,
    title: Optional[str] = None,
    twf_layout: bool = False,
) -> None:
    """回答書のヘッダー（行1〜6）を作成する。

    Args:
        ws: 対象ワークシート
        customer_name: 顧客名
        rep_name: 担当者名（担当者分割時）
        issue_date: 発行日（Noneなら今日）
        branch: 営業所設定（L列ヘッダー切替用）
        title: タイトル文字列（Noneなら「納　期　回　答　書」。
            長いタイトルはフォントを縮小して表示する）
        twf_layout: TrueならTWF専用レイアウト（A=TWF No., C=お客様名）の
            ヘッダーラベル・列幅を使う。external修飾は適用しない
    """
    if issue_date is None:
        issue_date = datetime.date.today()

    # フォントを全体に設定
    ws.sheet_properties.defaultRowHeight = 15

    # --- 行1: タイトル ---
    ws.merge_cells("A1:L1")
    cell_title = ws["A1"]
    if title is None:
        title = "納　期　回　答　書"
    # 長いタイトル（TWF専用回答書等）はセル幅に収まるよう縮小
    title_size = 26 if len(title) <= 15 else 18
    cell_title.value = title
    cell_title.font = _make_font(size=title_size, bold=True, color=_TITLE_FG)
    cell_title.fill = _make_fill(_TITLE_BG)
    cell_title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 55

    # --- 行2: アクセントライン ---
    for col in range(1, 13):
        ws.cell(row=2, column=col).fill = _make_fill(_ACCENT_BG)
    ws.row_dimensions[2].height = 4

    # --- 行3: 空白 ---
    ws.row_dimensions[3].height = 10

    # --- 行4: 顧客名・発行日 ---
    ws.cell(row=4, column=1).value = "お客様："
    ws.cell(row=4, column=1).font = _make_font(size=12)

    ws.merge_cells("B4:E4")
    if rep_name and rep_name != "__OTHER__":
        customer_text = f"{customer_name} 御中（ご担当：{rep_name} 様）"
    else:
        customer_text = f"{customer_name} 御中"
    cell_cust = ws["B4"]
    cell_cust.value = customer_text
    cell_cust.font = _make_font(size=16, bold=True, color=_TITLE_BG)

    weekday_names = ["月", "火", "水", "木", "金", "土", "日"]
    weekday_str = weekday_names[issue_date.weekday()]
    date_text = f"発行日： {issue_date.year}年{issue_date.month}月{issue_date.day}日({weekday_str})"
    cell_date = ws.cell(row=4, column=12)
    cell_date.value = date_text
    cell_date.font = _make_font(size=12, color="323232")
    cell_date.alignment = Alignment(horizontal="right")
    ws.row_dimensions[4].height = 35

    # --- 行5: 空白 ---
    ws.row_dimensions[5].height = 8

    # --- 行6: 列ヘッダー ---
    header_font = _make_font(size=11, bold=True, color=_HEADER_FG)
    header_fill = _make_fill(_HEADER_BG)
    header_align = Alignment(horizontal="center", vertical="center")

    # remarks_mode=external のときL列ヘッダーを「連絡事項」に変更
    # （TWFレイアウトはL列=弊社注番固定なので適用しない）
    if twf_layout:
        labels = list(_TWF_HEADER_LABELS)
        widths = _TWF_COLUMN_WIDTHS
    else:
        labels = list(_HEADER_LABELS)
        widths = _COLUMN_WIDTHS
        if branch and branch.remarks_mode == "external":
            labels[11] = "連絡事項"

    for col_idx, label in enumerate(labels, start=1):
        cell = ws.cell(row=6, column=col_idx)
        cell.value = label
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    ws.row_dimensions[6].height = 28

    # 列幅設定
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# ============================================
# VBA: CopyDataRow (L3784-4001)
# データ行の値変換
# ============================================
def _to_numeric(value: object) -> object:
    """文字列を数値に変換する。変換できなければそのまま返す。

    カンマ区切りの数値文字列（例: "5,865.00"）にも対応。
    "確認中"等のテキストはそのまま返す。
    """
    if isinstance(value, (int, float)):
        return value
    if not isinstance(value, str):
        return value
    s = value.strip()
    if not s:
        return value
    # カンマを除去して数値変換を試みる
    try:
        num = float(s.replace(",", ""))
        # 整数化できるなら整数で返す
        if num == int(num):
            return int(num)
        return num
    except (ValueError, OverflowError):
        return value


def copy_data_row(
    ws: Worksheet,
    target_row: int,
    report_row: ReportRow,
    external_comment: str | None = None,
) -> None:
    """レポート行をExcelシートに書き込む。

    CopyDataRowのExcel書き込み部分のみ。
    データ変換（納期計算・メーカー名解決等）はreport_generator側で実施済み。

    Args:
        ws: 対象ワークシート
        target_row: 書き込み先行番号（7〜）
        report_row: 書き込みデータ
        external_comment: L列に表示する社外コメント。
            None=注番を表示（デフォルト）、str=社外コメントを表示（空文字含む）
    """
    ws.cell(row=target_row, column=1).value = report_row.registration_date
    ws.cell(row=target_row, column=2).value = report_row.customer_contact
    cust_order = report_row.customer_order_number
    if isinstance(cust_order, str) and cust_order.strip().isdigit():
        ws.cell(row=target_row, column=3).value = int(cust_order.strip())
    else:
        ws.cell(row=target_row, column=3).value = cust_order
    ws.cell(row=target_row, column=3).alignment = Alignment(horizontal="left")
    ws.cell(row=target_row, column=4).value = report_row.manufacturer_name
    ws.cell(row=target_row, column=5).value = report_row.product_name
    ws.cell(row=target_row, column=6).value = _to_numeric(report_row.quantity)
    ws.cell(row=target_row, column=7).value = _to_numeric(report_row.unit_price)
    ws.cell(row=target_row, column=8).value = _to_numeric(report_row.net_amount)
    ws.cell(row=target_row, column=9).value = report_row.delivery_answer
    ws.cell(row=target_row, column=10).value = report_row.delivery_place
    ws.cell(row=target_row, column=11).value = report_row.remarks
    # L列: external_commentがNone以外なら社外コメント、Noneなら注番
    ws.cell(row=target_row, column=12).value = (
        external_comment if external_comment is not None else report_row.order_number
    )

    # 偶数行は薄いブルー
    if target_row % 2 == 0:
        for col in range(1, 13):
            ws.cell(row=target_row, column=col).fill = _FILL_EVEN


def copy_twf_data_row(
    ws: Worksheet,
    target_row: int,
    report_row: ReportRow,
    twf_number: str,
    twf_customer: str,
) -> None:
    """TWF専用レイアウトのデータ行を書き込む（期間限定。twf.py参照）。

    copy_data_rowで通常レイアウトを書いた後、A列をTWF No.、
    C列をお客様名に差し替える。D〜L列は通常版と同一のため、
    書式エンジン（色分け・数値書式・罫線）はそのまま機能する。

    Args:
        ws: 対象ワークシート
        target_row: 書き込み先行番号
        report_row: 書き込みデータ
        twf_number: TWF No.（例: "003281"。先頭ゼロ保持のため文字列で書き込む）
        twf_customer: お客様名（例: "三友工業様"。なければ空欄）
    """
    copy_data_row(ws, target_row, report_row, external_comment=None)

    # A列: 受注日 → TWF No.（番号なしは空欄）
    cell_a = ws.cell(row=target_row, column=1)
    cell_a.value = twf_number or None
    cell_a.alignment = Alignment(horizontal="center")

    # C列: 貴社注番 → お客様名
    cell_c = ws.cell(row=target_row, column=3)
    cell_c.value = twf_customer or None
    cell_c.alignment = Alignment(horizontal="left")


# ============================================
# VBA: CheckSameDateInBunno (L4943-4963)
# 分納に同じ日付があるかチェック
# ============================================
def check_same_date_in_bunno(bunno_detail: list[BunnoEntry]) -> bool:
    """分納詳細に同じ日付が複数あるかチェックする。

    同じ日付＝別々の場所からの出荷 → 注釈表示用。
    「未定」や「○旬予定」は除外。

    Args:
        bunno_detail: 分納情報リスト

    Returns:
        True = 同じ日付がある
    """
    seen_dates: set[str] = set()

    for entry in bunno_detail:
        ds = entry.date_str
        if ds == "未定" or "予定" in ds:
            continue
        if ds in seen_dates:
            return True
        seen_dates.add(ds)

    return False


# ============================================
# VBA: FormatReport (L4037-4939) — 分割実装
# 書式設定全般
# ============================================
def format_report(
    ws: Worksheet,
    last_data_row: int,
    branch: Optional[BranchSettings] = None,
    tracking_info_list: Optional[list[tuple[str, str, str, TrackingEntry]]] = None,
    stockout_info_list: Optional[list[StockoutEntry]] = None,
    bunno_info_list: Optional[list[dict]] = None,
    bunno_completed_list: Optional[list[tuple[str, str, str]]] = None,
    holidays: HolidayMap | None = None,
    cache: Optional[CacheStore] = None,
    today: Optional[datetime.date] = None,
    twf_notice: Optional[str] = None,
    twf_thanks: Optional[str] = None,
    with_auto_filter: bool = False,
) -> None:
    """回答書の書式設定を行う。

    色分け・罫線・数値書式・ご連絡事項・署名を設定する。

    Args:
        ws: 対象ワークシート
        last_data_row: 最終データ行番号
        branch: 営業所設定
        tracking_info_list: 送り状情報 [(メーカー, 品名, 数量, TrackingEntry), ...]
        stockout_info_list: 欠品情報
        bunno_info_list: 分納情報（FormatReport用の辞書リスト）
        bunno_completed_list: 分納完了リスト [(メーカー, 品名, 数量), ...]
        holidays: 祝日辞書
        cache: キャッシュストア
        today: 基準日（テスト用）
        twf_notice: TWF展示会注記（指定時のみご連絡事項に赤字表示）
        twf_thanks: TWF感謝文（指定時のみ注記の上に通常色で表示）
        with_auto_filter: Trueならヘッダー行（行6）〜データ最終行に
            オートフィルタを設定する（TWF専用回答書用）
    """
    if today is None:
        today = datetime.date.today()

    # データ行の全書式を1パスで適用（5ループ→1ループに統合して高速化）
    _apply_all_data_formatting(ws, last_data_row, today, branch)

    # オートフィルタ（ヘッダー行6〜データ最終行。マージセルは行1-5なので干渉しない）
    if with_auto_filter:
        ws.auto_filter.ref = f"A6:L{last_data_row}"

    # 税抜き注記
    note_row = last_data_row + 1
    ws.merge_cells(f"G{note_row}:H{note_row}")
    cell_note = ws.cell(row=note_row, column=7)
    cell_note.value = "※表示金額は税抜きです"
    cell_note.font = _make_font(size=9, color="787878")
    cell_note.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[note_row].height = 16

    # ご連絡事項と署名
    info_row = last_data_row + 2
    info_row = _write_info_section(
        ws, info_row, branch,
        tracking_info_list, stockout_info_list,
        bunno_info_list, bunno_completed_list,
        holidays, cache, today,
        twf_notice=twf_notice,
        twf_thanks=twf_thanks,
    )

    # 印刷設定
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "1:6"

    # 余白設定（インチ単位: 0.5cm ≒ 0.197in）
    ws.page_margins.left = 0.2
    ws.page_margins.right = 0.2
    ws.page_margins.top = 0.2
    ws.page_margins.bottom = 0.2
    ws.page_margins.header = 0.0
    ws.page_margins.footer = 0.0


def _apply_delivery_colors(
    ws: Worksheet,
    last_data_row: int,
    today: datetime.date,
) -> None:
    """納期回答列（I列）の色分けを適用する。"""
    from nouki_kaitou.utils import extract_date_from_string

    tomorrow = today + datetime.timedelta(days=1)
    day_after = today + datetime.timedelta(days=2)
    col = 9  # I列

    # 共通書式（shrink_to_fitを維持しつつ中央揃え）
    bold_center_shrink = Alignment(horizontal="center", shrink_to_fit=True)

    for row in range(7, last_data_row + 1):
        cell = ws.cell(row=row, column=col)
        value = str(cell.value or "").strip()
        cell.font = _make_font(size=10, bold=True)
        cell.alignment = bold_center_shrink

        color_pair = _classify_delivery_color(
            value, today, tomorrow, day_after, extract_date_from_string
        )
        if color_pair:
            cell.fill = _make_fill(color_pair[0])
            cell.font = _make_font(size=10, bold=True, color=color_pair[1])


def _classify_delivery_color(
    value: str,
    today: datetime.date,
    tomorrow: datetime.date,
    day_after: datetime.date,
    extract_date_fn,
) -> tuple[str, str] | None:
    """納期回答の値から色ペア (背景, 文字) を判定する。"""
    if value == "納品済み":
        return _COLOR_DELIVERED
    if value == "欠品中":
        return _COLOR_STOCKOUT
    if "（欠品）" in value:
        return _COLOR_STOCKOUT_PARTIAL
    if "分納" in value:
        return _COLOR_BUNNO
    if value == "確認中":
        return _COLOR_CONFIRMING
    if value == "日程調整中":
        return _COLOR_SCHEDULING
    if "作業" in value:
        return _COLOR_WORK
    if "他拠点より" in value:
        return _COLOR_OTHER_BRANCH
    if "引取済み" in value:
        return _COLOR_PICKED_UP
    if "済み" in value or "済" in value:
        return _COLOR_DONE
    if "予定" in value:
        d = extract_date_fn(value)
        if d:
            if "出荷予定" in value:
                if d == today or d == tomorrow:
                    return _COLOR_SHIP_SOON
                return _COLOR_SHIP_LATER
            if "配達予定" in value:
                if d == tomorrow or d == day_after:
                    return _COLOR_DELIVER_SOON
                return _COLOR_DELIVER_LATER
            if "引取予定" in value:
                return _COLOR_PICKUP_PLAN
            return _COLOR_OTHER_PLAN
        return _COLOR_OTHER_PLAN
    return _COLOR_DEFAULT


def _apply_price_confirming_style(ws: Worksheet, last_data_row: int) -> None:
    """単価・金額の「確認中」を赤字太字にする。"""
    red_font = _make_font(size=10, bold=True, color="B41E1E")
    center = Alignment(horizontal="center")

    for row in range(7, last_data_row + 1):
        for col in [7, 8]:  # G列, H列
            cell = ws.cell(row=row, column=col)
            if str(cell.value or "").strip() == "確認中":
                cell.font = red_font
                cell.alignment = center
                cell.number_format = "@"


def _apply_number_formats(ws: Worksheet, last_data_row: int) -> None:
    """数値書式を設定する。"""
    for row in range(7, last_data_row + 1):
        # A列（受注日）
        ws.cell(row=row, column=1).number_format = "m/d(aaa)"

        # F列（数量）: カンマ区切り（小数があればそのまま表示）
        cell_f = ws.cell(row=row, column=6)
        if isinstance(cell_f.value, (int, float)):
            if isinstance(cell_f.value, float) and cell_f.value != int(cell_f.value):
                cell_f.number_format = "#,##0.###"
            else:
                cell_f.number_format = "#,##0"

        # G列・H列（単価・金額）: カンマ区切り
        for col in [7, 8]:
            cell = ws.cell(row=row, column=col)
            val = cell.value
            if isinstance(val, (int, float)):
                if isinstance(val, float) and val != int(val):
                    cell.number_format = "#,##0.00"
                else:
                    cell.number_format = "#,##0"


def _apply_borders(ws: Worksheet, last_data_row: int) -> None:
    """罫線を設定する。"""
    thin_side = Side(style="thin", color="A0A0A0")
    thin_border = Border(
        left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
    )

    for row in range(6, last_data_row + 1):
        for col in range(1, 13):
            ws.cell(row=row, column=col).border = thin_border

    # ヘッダー行下線（ゴールド）
    gold_bottom = Side(style="medium", color=_ACCENT_BG)
    for col in range(1, 13):
        cell = ws.cell(row=6, column=col)
        cell.border = Border(
            left=thin_side, right=thin_side, top=thin_side, bottom=gold_bottom
        )


def _apply_all_data_formatting(
    ws: Worksheet,
    last_data_row: int,
    today: datetime.date,
    branch: Optional[BranchSettings] = None,
) -> None:
    """データ行のフォント・色分け・数値書式・罫線を1パスで適用する。

    従来の5ループ（フォント設定 → 納期色分け → 確認中赤字 → 数値書式 → 罫線）
    を1ループに統合して、セルアクセス回数を大幅に削減する。
    """
    from nouki_kaitou.utils import extract_date_from_string

    tomorrow = today + datetime.timedelta(days=1)
    day_after = today + datetime.timedelta(days=2)

    # ヘッダー行（行6）の罫線
    gold_bottom = Side(style="medium", color=_ACCENT_BG)
    header_border = Border(
        left=_THIN_SIDE, right=_THIN_SIDE,
        top=_THIN_SIDE, bottom=gold_bottom,
    )
    for col in range(1, 13):
        ws.cell(row=6, column=col).border = header_border

    # データ行を1パスで処理
    for row in range(7, last_data_row + 1):
        ws.row_dimensions[row].height = 22

        for col in range(1, 13):
            cell = ws.cell(row=row, column=col)

            # 罫線（全セル共通）
            cell.border = _THIN_BORDER

            # 列別のフォーマット
            if col == 1:
                # A列: 受注日 — フォント + 日付書式
                cell.font = _FONT_10
                cell.number_format = "m/d(aaa)"

            elif col in (2, 3, 4, 5):
                # B,C,D,E列: テキスト列 — フォントのみ
                cell.font = _FONT_10

            elif col == 12:
                # L列: 注番 or 連絡事項
                cell.font = _FONT_10
                if branch and branch.remarks_mode == "external":
                    cell.alignment = _ALIGN_SHRINK

            elif col == 6:
                # F列: 数量 — フォント + カンマ区切り（小数があればそのまま表示）
                cell.font = _FONT_10
                if isinstance(cell.value, (int, float)):
                    if isinstance(cell.value, float) and cell.value != int(cell.value):
                        cell.number_format = "#,##0.###"
                    else:
                        cell.number_format = "#,##0"

            elif col in (7, 8):
                # G,H列: 単価・金額 — 確認中赤字 or 数値書式
                val = cell.value
                val_str = str(val or "").strip()
                if val_str == "確認中":
                    cell.font = _FONT_PRICE_CONFIRM
                    cell.alignment = _ALIGN_CENTER
                    cell.number_format = "@"
                else:
                    cell.font = _FONT_10
                    if isinstance(val, (int, float)):
                        if isinstance(val, float) and val != int(val):
                            cell.number_format = "#,##0.00"
                        else:
                            cell.number_format = "#,##0"

            elif col == 9:
                # I列: 納期回答 — 太字中央 + 色分け
                value_str = str(cell.value or "").strip()
                color_pair = _classify_delivery_color(
                    value_str, today, tomorrow, day_after,
                    extract_date_from_string,
                )
                if color_pair:
                    cell.fill = _get_delivery_fill(color_pair[0])
                    cell.font = _get_delivery_font(color_pair[1])
                else:
                    cell.font = _FONT_10_BOLD
                cell.alignment = _ALIGN_CENTER_SHRINK

            elif col in (10, 11):
                # J,K列: 納入先・備考 — フォント + 縮小表示
                cell.font = _FONT_10
                cell.alignment = _ALIGN_SHRINK


def _write_info_section(
    ws: Worksheet,
    start_row: int,
    branch: Optional[BranchSettings],
    tracking_info_list,
    stockout_info_list,
    bunno_info_list,
    bunno_completed_list,
    holidays,
    cache,
    today: datetime.date,
    twf_notice: Optional[str] = None,
    twf_thanks: Optional[str] = None,
) -> int:
    """ご連絡事項と署名セクションを書き込む。返り値は次の空き行。"""
    from nouki_kaitou.tracking import can_direct_track, get_tracking_url
    from nouki_kaitou.utils import to_circled_number

    has_tracking = bool(tracking_info_list)
    has_stockout = bool(stockout_info_list)
    has_bunno = bool(bunno_info_list)
    has_twf_notice = bool(twf_notice)
    has_twf_thanks = bool(twf_thanks)

    row = start_row

    # --- ご連絡事項ヘッダー ---
    ws.merge_cells(f"A{row}:G{row}")
    if has_tracking or has_stockout or has_bunno or has_twf_notice or has_twf_thanks:
        cell = ws.cell(row=row, column=1)
        cell.value = "【ご連絡事項】"
        cell.font = _make_font(size=14, bold=True, color=_HEADER_BG)

    # 下線（ダブル）
    double_side = Side(style="double", color=_ACCENT_BG)
    for col in range(1, 9):
        ws.cell(row=row, column=col).border = Border(bottom=double_side)

    # --- 署名 ---
    ws.merge_cells(f"I{row}:J{row}")
    cell_sign = ws.cell(row=row, column=9)
    cell_sign.value = "◆ マツモト産業株式会社 ◆"
    cell_sign.font = _make_font(size=14, bold=True, color=_HEADER_BG)
    cell_sign.alignment = Alignment(horizontal="center", vertical="bottom")

    for col in range(11, 13):
        ws.cell(row=row, column=col).border = Border(bottom=double_side)

    ws.row_dimensions[row].height = 28
    row += 1

    # 営業所名
    ws.merge_cells(f"I{row}:J{row}")
    branch_name = branch.name if branch else ""
    cell_branch = ws.cell(row=row, column=9)
    cell_branch.value = branch_name
    cell_branch.font = _make_font(size=14, bold=True, color=_HEADER_BG)
    cell_branch.alignment = Alignment(horizontal="center", vertical="top")

    # --- TWF感謝文・展示会注記（期間限定。twf.py参照） ---
    # 感謝文（通常色）→ 赤字注記の順。書いた分だけ行を進め、
    # 最後に1行進めて後続セクションの開始行を確保する
    if has_twf_thanks:
        row += 1
        ws.merge_cells(f"A{row}:L{row}")
        cell_thanks = ws.cell(row=row, column=1)
        cell_thanks.value = twf_thanks
        cell_thanks.font = _make_font(size=11, bold=True, color=_HEADER_BG)
        cell_thanks.alignment = Alignment(
            horizontal="left", vertical="center", wrap_text=True
        )
        ws.row_dimensions[row].height = 30

    if has_twf_notice:
        row += 1
        ws.merge_cells(f"A{row}:L{row}")
        cell_notice = ws.cell(row=row, column=1)
        cell_notice.value = twf_notice
        cell_notice.font = _make_font(size=11, bold=True, color="B40000")
        cell_notice.alignment = Alignment(
            horizontal="left", vertical="center", wrap_text=True
        )
        ws.row_dimensions[row].height = 55

    if has_twf_thanks or has_twf_notice:
        row += 1

    # --- 送り状情報 ---
    if has_tracking:
        row = _write_tracking_section(ws, row, tracking_info_list)

    # --- 欠品情報 ---
    if has_stockout:
        row = _write_stockout_section(ws, row, stockout_info_list)

    # --- 分納情報 ---
    if has_bunno:
        row = _write_bunno_section(
            ws, row, bunno_info_list, holidays, cache, today
        )

    # --- 分納完了通知 ---
    if bunno_completed_list:
        row = _write_bunno_completed_section(ws, row, bunno_completed_list)

    return row


def _make_hyperlink_formula(url: str, display_text: str) -> str:
    """HYPERLINK関数の数式文字列を生成する。

    マージセルでもクリック可能なハイパーリンクを確実に作成するため、
    cell.hyperlinkではなくHYPERLINK関数を使用する。
    """
    # 数式内のダブルクォートをエスケープ
    safe_url = url.replace('"', '""')
    safe_text = display_text.replace('"', '""')
    return f'=HYPERLINK("{safe_url}","{safe_text}")'


def _write_tracking_section(
    ws: Worksheet,
    start_row: int,
    tracking_info_list: list[tuple[str, str, str, TrackingEntry]],
) -> int:
    """送り状情報セクションを書き込む。"""
    from nouki_kaitou.tracking import can_direct_track, get_tracking_url

    row = start_row
    ws.merge_cells(f"A{row}:G{row}")
    cell = ws.cell(row=row, column=1)
    cell.value = "    下記商品の送り状番号をご連絡いたします。"
    cell.font = _make_font(size=11, color="282828")
    ws.row_dimensions[row].height = 24
    row += 1

    for mfg, product, qty, entry in tracking_info_list:
        # 送り状番号表示
        ws.merge_cells(f"A{row}:G{row}")
        cell = ws.cell(row=row, column=1)
        text = f"    ■ {entry.carrier_name}  {entry.tracking_number}"

        url = get_tracking_url(entry.carrier_name, entry.tracking_number)
        direct = can_direct_track(entry.carrier_name)

        if direct and url:
            cell.value = _make_hyperlink_formula(url, text)
            cell.font = _make_font(
                size=11, bold=True, color="0000FF", underline="single",
            )
        else:
            cell.value = text
            cell.font = _make_font(size=11, bold=True, color=_HEADER_BG)

        ws.row_dimensions[row].height = 24
        row += 1

        # 間接追跡リンク
        if not direct and url:
            ws.merge_cells(f"A{row}:G{row}")
            cell = ws.cell(row=row, column=1)
            link_text = "        → 追跡ページ（番号を入力してください）"
            cell.value = _make_hyperlink_formula(url, link_text)
            cell.font = _make_font(
                size=10, color="0000FF", underline="single",
            )
            ws.row_dimensions[row].height = 20
            row += 1

        # 商品表示
        ws.merge_cells(f"A{row}:G{row}")
        cell = ws.cell(row=row, column=1)
        from nouki_kaitou.utils import format_quantity
        cell.value = f"        - {mfg}  {product}  x{format_quantity(qty)}"
        cell.font = _make_font(size=10, bold=True, color="282828")
        ws.row_dimensions[row].height = 20
        row += 1

    return row


def _write_stockout_section(
    ws: Worksheet,
    start_row: int,
    stockout_info_list: list[StockoutEntry],
) -> int:
    """欠品情報セクションを書き込む（入荷確定 + 欠品継続の2グループ）。"""
    confirmed = [i for i in stockout_info_list if is_stockout_confirmed(i)]
    pending = [i for i in stockout_info_list if not is_stockout_confirmed(i)]

    row = start_row
    if confirmed:
        row = _write_stockout_confirmed(ws, row, confirmed)
    if pending:
        row = _write_stockout_pending(ws, row, pending)
    return row


def _write_stockout_confirmed(
    ws: Worksheet,
    start_row: int,
    items: list[StockoutEntry],
) -> int:
    """入荷確定グループを書き込む。"""
    from nouki_kaitou.utils import format_quantity

    row = start_row
    ws.merge_cells(f"A{row}:G{row}")
    cell = ws.cell(row=row, column=1)
    cell.value = "    欠品しておりました商品の入荷日が確定いたしました。"
    cell.font = _make_font(size=11, bold=True, color="338833")
    ws.row_dimensions[row].height = 24
    row += 1

    for item in items:
        text = (
            f"        - {item.manufacturer_name}  {item.product_name}"
            f"  x{format_quantity(item.quantity)}"
            f" → {item.delivery}"
        )
        ws.merge_cells(f"A{row}:G{row}")
        cell = ws.cell(row=row, column=1)
        cell.value = text
        cell.font = _make_font(size=10, bold=True, color="338833")
        ws.row_dimensions[row].height = 20
        row += 1

    return row


def _write_stockout_pending(
    ws: Worksheet,
    start_row: int,
    items: list[StockoutEntry],
) -> int:
    """欠品継続グループを書き込む。"""
    from nouki_kaitou.utils import format_quantity

    row = start_row
    ws.merge_cells(f"A{row}:G{row}")
    cell = ws.cell(row=row, column=1)
    cell.value = "    下記商品は現在欠品中です。ご迷惑をおかけし申し訳ございません。"
    cell.font = _make_font(size=11, bold=True, color="B40000")
    ws.row_dimensions[row].height = 24
    row += 1

    for item in items:
        text = (
            f"        - {item.manufacturer_name}  {item.product_name}"
            f"  x{format_quantity(item.quantity)}"
        )
        if item.approx_delivery:
            text += f" → {item.approx_delivery}"
        else:
            text += " → 入荷次第ご連絡"

        ws.merge_cells(f"A{row}:G{row}")
        cell = ws.cell(row=row, column=1)
        cell.value = text
        cell.font = _make_font(size=10, bold=True, color="8B0000")
        ws.row_dimensions[row].height = 20
        row += 1

    return row


def _write_bunno_section(
    ws: Worksheet,
    start_row: int,
    bunno_info_list: list[dict],
    holidays,
    cache,
    today: datetime.date,
) -> int:
    """分納情報セクションを書き込む。"""
    from nouki_kaitou.bunno import (
        calculate_bunno_date,
        has_bunno_kakuninchu,
    )
    from nouki_kaitou.utils import to_circled_number

    row = start_row

    # 確認中があるかチェック
    has_mitei_note = False
    for item in bunno_info_list:
        calc_details = item.get("calc_details", [])
        if has_bunno_kakuninchu(calc_details):
            has_mitei_note = True
            break

    ws.merge_cells(f"A{row}:G{row}")
    cell = ws.cell(row=row, column=1)
    cell.value = "    下記商品は分納にてお届けいたします。"
    cell.font = _make_font(size=11, color="00468C")
    ws.row_dimensions[row].height = 24
    row += 1

    if has_mitei_note:
        ws.merge_cells(f"A{row}:G{row}")
        cell = ws.cell(row=row, column=1)
        cell.value = "    ※一部納期未定のためご迷惑をおかけいたします。確定次第ご連絡いたします。"
        cell.font = _make_font(size=10, color="B40000")
        ws.row_dimensions[row].height = 20
        row += 1

    for item in bunno_info_list:
        mfg = item.get("manufacturer", "")
        product = item.get("product", "")
        qty = item.get("quantity", "")
        entries = item.get("entries", [])
        calc_details = item.get("calc_details", [])
        is_ship_rule = item.get("is_ship_rule", False)
        days_to_add = item.get("days_to_add", 0)
        order_num = item.get("order_number", "")
        detail_num = item.get("detail_number", "")
        is_rosenbin = item.get("is_rosenbin", False)

        # 商品ヘッダー
        ws.merge_cells(f"A{row}:G{row}")
        cell = ws.cell(row=row, column=1)
        from nouki_kaitou.utils import format_quantity
        cell.value = f"    ■ {mfg}  {product}  x{format_quantity(qty)}"
        cell.font = _make_font(size=11, bold=True, color="00468C")

        # 左罫線（アクセント）
        accent_side = Side(style="thick", color="6496C8")
        cell.border = Border(left=accent_side)

        ws.row_dimensions[row].height = 24
        row += 1

        # 分納詳細行
        same_date = check_same_date_in_bunno(entries)

        for idx, entry in enumerate(entries):
            counter = idx + 1

            # 計算済み納期を使用（あれば）、なければ動的計算
            calc_date = ""
            if idx < len(calc_details) and len(calc_details[idx]) >= 4:
                calc_date = str(calc_details[idx][3])

            if not calc_date:
                calc_date = calculate_bunno_date(
                    entry.date_str, is_ship_rule, days_to_add,
                    holidays, cache, order_num, detail_num,
                    is_rosenbin, today,
                )

            location_text = f"（{entry.location}）" if entry.location else ""
            circled = to_circled_number(counter)

            ws.merge_cells(f"A{row}:G{row}")
            cell = ws.cell(row=row, column=1)
            cell.value = f"        {circled}{entry.quantity} → {calc_date}{location_text}"

            # 色分け
            is_from_mitei = (
                entry.date_str == "未定"
                or "欠品" in entry.date_str
                or "確認中" in entry.date_str
            )
            still_unconfirmed = (
                calc_date == "確認中"
                or ("予定" in calc_date and "出荷" not in calc_date and "配達" not in calc_date)
            )

            if still_unconfirmed:
                cell.font = _make_font(size=10, bold=True, color="B41E1E")
            elif is_from_mitei:
                cell.font = _make_font(size=10, bold=True, color="C86400")
            else:
                cell.font = _make_font(size=10, color="282828")

            ws.row_dimensions[row].height = 20
            row += 1

        # 同じ日付の場合の注釈
        if same_date:
            ws.merge_cells(f"A{row}:G{row}")
            cell = ws.cell(row=row, column=1)
            cell.value = "        ※別々の場所からの出荷になります"
            cell.font = _make_font(size=9, color="646464", italic=True)
            ws.row_dimensions[row].height = 18
            row += 1

        row += 1

    return row


def _write_bunno_completed_section(
    ws: Worksheet,
    start_row: int,
    bunno_completed_list: list[tuple[str, str, str]],
) -> int:
    """分納完了通知セクションを書き込む。"""
    row = start_row

    ws.merge_cells(f"A{row}:G{row}")
    cell = ws.cell(row=row, column=1)
    cell.value = "    分納でご注文いただいた商品は全て出荷が完了しました。"
    cell.font = _make_font(size=11, bold=True, color="00783C")
    ws.row_dimensions[row].height = 24
    row += 1

    for mfg, product, qty in bunno_completed_list:
        ws.merge_cells(f"A{row}:G{row}")
        cell = ws.cell(row=row, column=1)
        from nouki_kaitou.utils import format_quantity
        cell.value = f"        ■ {mfg}  {product}  x{format_quantity(qty)}"
        cell.font = _make_font(size=10, color="006432")
        ws.row_dimensions[row].height = 20
        row += 1

    return row + 1


# ============================================
# VBA: ColorConfirmingList (L1954-2051)
# 確認中一覧の色分け
# ============================================
def color_confirming_list(
    ws: Worksheet,
    today: Optional[datetime.date] = None,
) -> None:
    """確認中一覧シートの行を色分けする。

    条件:
    - 出荷完了 → 赤系
    - 欠品中 → 薄紫
    - 分納 → 薄青
    - 1週間以上経過 → オレンジ
    - 3日以上経過 → 黄色
    - その他 → 色なし

    Args:
        ws: 確認中一覧のワークシート
        today: 基準日（テスト用）
    """
    if today is None:
        today = datetime.date.today()

    # テーブルのデータ範囲を探す（ヘッダー行=1, データ行=2〜）
    max_row = ws.max_row
    if max_row < 2:
        return

    for row in range(2, max_row + 1):
        sent_date_val = ws.cell(row=row, column=1).value
        status_val = str(ws.cell(row=row, column=9).value or "").strip()

        # 送付日時をパース
        sent_date: Optional[datetime.date] = None
        if isinstance(sent_date_val, datetime.datetime):
            sent_date = sent_date_val.date()
        elif isinstance(sent_date_val, datetime.date):
            sent_date = sent_date_val

        fill: Optional[PatternFill] = None

        if status_val == "出荷完了":
            fill = _make_fill(_CONFIRM_SHIP_DONE[0])
        elif status_val in ("一部処理済み", "一部処理済"):
            fill = _make_fill(_CONFIRM_PARTIAL[0])
        elif status_val == "欠品中":
            fill = _make_fill(_CONFIRM_STOCKOUT[0])
        elif status_val == "分納":
            fill = _make_fill(_CONFIRM_BUNNO[0])
        elif status_val == "価格確認中":
            fill = _make_fill(_CONFIRM_PRICE_PENDING[0])
        elif sent_date and (today - sent_date).days >= 7:
            fill = _make_fill(_CONFIRM_WEEK_OLD[0])
        elif sent_date and (today - sent_date).days >= 3:
            fill = _make_fill(_CONFIRM_THREE_DAYS[0])

        # VBA同様、全行に色を設定（通常行は色クリア）
        max_col = ws.max_column
        if fill:
            for col in range(1, max_col + 1):
                ws.cell(row=row, column=col).fill = fill
        else:
            no_fill = PatternFill(fill_type=None)
            for col in range(1, max_col + 1):
                ws.cell(row=row, column=col).fill = no_fill
