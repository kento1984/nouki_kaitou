"""納期回答計算デバッグスクリプト

指定した注番の納期回答計算過程を詳細に表示する。
"""

import datetime
import sys
from pathlib import Path

# パッケージパスを追加（親ディレクトリをパスに追加してnouki_kaitouをインポート可能に）
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import load_workbook

# nouki_kaitouパッケージをインポート
from nouki_kaitou.models import BranchSettings, CacheStore, OrderRow, HolidayMap
from nouki_kaitou.data_loader import load_source_file, get_column_positions, parse_order_row, is_data_row, get_data_rows_range
from nouki_kaitou.cache import build_all_caches
from nouki_kaitou.config import load_branch_settings, load_holidays, get_branch_settings
from nouki_kaitou.delivery_calc import (
    calculate_delivery_date,
    extract_pickup_date,
    extract_arrival_date_from_internal,
    _resolve_storage_place,
    _check_work_order,
    _check_arrival_date,
    _check_pickup,
    _check_specified_date,
    _check_stock_completed,
    _check_himozuki_completed,
    _check_dec31,
    _calc_normal,
)
from nouki_kaitou.business_days import add_business_days, get_next_business_day, get_next_delivery_day, get_previous_business_day
from nouki_kaitou.customer import get_customer_delivery_days, is_route_delivery
from nouki_kaitou.manufacturer import get_delivery_days_to_add
from nouki_kaitou.confirming import get_confirmed_delivery_date
from nouki_kaitou.utils import format_date_japanese, is_december_31, parse_time


def print_separator(title: str = ""):
    """区切り線を出力"""
    if title:
        print(f"\n{'='*60}")
        print(f"  {title}")
        print('='*60)
    else:
        print('-'*60)


def print_order_row(row: OrderRow):
    """OrderRowの全フィールドを表示"""
    print_separator("OrderRow 全フィールド")
    fields = [
        ("order_number", "受発注伝票(注番)"),
        ("detail_number", "明細"),
        ("document_type", "伝票タイプ"),
        ("customer_name", "受注先(顧客名)"),
        ("product_name", "テキスト(品名)"),
        ("ship_status", "出荷ステータス"),
        ("quantity", "受注数量"),
        ("unit_price", "受注単価"),
        ("net_amount", "正味額"),
        ("manufacturer_name", "名称(メーカー)"),
        ("storage_place", "保管場所"),
        ("customer_order_number", "得意先発注番号"),
        ("customer_contact", "得意先担当者"),
        ("comment_detail", "コメント(明細)"),
        ("comment_external", "コメント(社外)"),
        ("comment_internal", "コメント(社内)"),
        ("rejection_reason", "拒否理由"),
        ("ship_to_name", "出荷先名"),
        ("registration_date", "登録日"),
        ("time_value", "時刻"),
        ("order_delivery_date", "受注納期"),
        ("specified_delivery_date", "指定納期"),
        ("item_group_code", "品目Group"),
        ("source_row", "元データ行番号"),
    ]

    for attr, label in fields:
        value = getattr(row, attr)
        if value is None:
            display = "(None)"
        elif value == "":
            display = "(空文字)"
        else:
            display = str(value)
        print(f"  {label:20s}: {display}")


def debug_calculate_delivery_date(
    row: OrderRow,
    cache: CacheStore,
    holidays: HolidayMap,
    branch: BranchSettings,
    execution_time: datetime.datetime,
    today: datetime.date,
) -> str:
    """納期計算の判定ステップを詳細表示"""

    print_separator("calculate_delivery_date 判定ステップ")

    # ============================================
    # Step 1: &&作業チェック
    # ============================================
    print("\n[Step 1] &&作業チェック")
    internal = row.comment_internal.strip()
    has_work_marker = "&&" in internal or "＆＆" in internal
    print(f"  コメント(社内): '{internal}'")
    print(f"  &&マーカー検出: {has_work_marker}")

    result = _check_work_order(row, today)
    if result is not None:
        print(f"  --> 判定結果: '{result}'")
        print(f"  --> このステップで終了")
        return result
    print(f"  --> 該当なし、次へ")

    # ============================================
    # Step 2: @@着日指定チェック
    # ============================================
    print("\n[Step 2] @@着日指定チェック")
    arrival_date = extract_arrival_date_from_internal(internal, today)
    print(f"  コメント(社内): '{internal}'")
    print(f"  @@着日抽出結果: {arrival_date}")

    result = _check_arrival_date(row, today)
    if result is not None:
        print(f"  --> 判定結果: '{result}'")
        print(f"  --> このステップで終了")
        return result
    print(f"  --> 該当なし、次へ")

    # ============================================
    # Step 3: 引取チェック
    # ============================================
    print("\n[Step 3] 引取チェック")
    comment = (row.comment_external.strip() + " " + row.comment_internal.strip()).strip()
    pickup_date = extract_pickup_date(comment, today)
    print(f"  コメント(社外+社内): '{comment}'")
    print(f"  引取日抽出結果: {pickup_date}")

    result = _check_pickup(row, today)
    if result is not None:
        print(f"  --> 判定結果: '{result}'")
        print(f"  --> このステップで終了")
        return result
    print(f"  --> 該当なし、次へ")

    # ============================================
    # フラグ設定
    # ============================================
    print("\n[フラグ設定]")
    storage_place = _resolve_storage_place(row, cache)
    print(f"  保管場所(解決後): '{storage_place}'")

    use_ship_rule = False
    if row.document_type == "【受注】在庫販売":
        if row.customer_name != row.ship_to_name:
            use_ship_rule = True
    print(f"  伝票タイプ: '{row.document_type}'")
    print(f"  受注先: '{row.customer_name}'")
    print(f"  出荷先名: '{row.ship_to_name}'")
    print(f"  受注先≠出荷先: {row.customer_name != row.ship_to_name}")
    print(f"  use_ship_rule(初期): {use_ship_rule}")

    is_rosenbin = is_route_delivery(row.customer_name, cache)
    print(f"  路線便フラグ: {is_rosenbin}")

    original_use_ship_rule = use_ship_rule
    if not use_ship_rule and is_rosenbin:
        use_ship_rule = True
    print(f"  use_ship_rule(路線便考慮後): {use_ship_rule}")
    print(f"  original_use_ship_rule: {original_use_ship_rule}")

    # ============================================
    # Step 4: 指定納期チェック
    # ============================================
    print("\n[Step 4] 指定納期チェック")
    spec_date = row.specified_delivery_date
    print(f"  指定納期: {spec_date}")
    if spec_date is not None:
        print(f"  12/31判定: {is_december_31(spec_date)}")

    result = _check_specified_date(row, cache, holidays, today, storage_place, use_ship_rule, is_rosenbin)
    if result is not None:
        # 詳細な計算過程を表示
        print(f"\n  [Step 4 詳細計算]")
        print(f"    伝票タイプ: '{row.document_type}'")

        if row.document_type == "【受注】在庫販売":
            print(f"    --> 在庫販売パス")
            if storage_place == "転送中（直送用）":
                print(f"    --> 転送中 → 指定納期そのまま出荷予定")
            elif use_ship_rule:
                ship_date = get_previous_business_day(spec_date, holidays)
                print(f"    --> use_ship_rule=True → 1営業日前を出荷日")
                print(f"    --> 指定納期({spec_date}) の前営業日 = {ship_date}")
            else:
                print(f"    --> 自社便配達 → 指定納期そのまま配達予定")
        else:
            print(f"    --> 直送販売（紐付き）パス")
            days_to_add = get_delivery_days_to_add(row.item_group_code, cache)
            print(f"    品目Group: '{row.item_group_code}'")
            print(f"    配送加算日数: {days_to_add}")

            if storage_place == "転送中（直送用）":
                print(f"    --> 転送中 → 指定納期そのまま出荷予定")
            else:
                delivery_days = get_customer_delivery_days(row.customer_name, cache)
                print(f"    顧客配送曜日: {delivery_days}")
                adjusted = add_business_days(spec_date, days_to_add, holidays)
                print(f"    指定納期 + {days_to_add}営業日 = {adjusted}")

                if delivery_days:
                    next_del = get_next_delivery_day(adjusted, delivery_days, holidays)
                    print(f"    次の配送曜日に調整: {adjusted} → {next_del}")
                    print(f"    --> 曜日制限あり → 出荷予定")
                elif is_rosenbin:
                    rosenbin_date = add_business_days(spec_date, max(days_to_add - 1, 0), holidays)
                    print(f"    路線便: 指定納期 + {max(days_to_add - 1, 0)}営業日 = {rosenbin_date}")
                    print(f"    --> 路線便 → 出荷予定")
                elif row.customer_name != row.ship_to_name:
                    ship_date = get_previous_business_day(adjusted, holidays)
                    print(f"    受注先≠出荷先: {adjusted}の前営業日 = {ship_date}")
                    print(f"    --> 受注先≠出荷先 → 出荷予定")
                else:
                    print(f"    --> 曜日制限なし、路線便なし、受注先=出荷先 → 配達予定")

        print(f"\n  --> 判定結果: '{result}'")
        print(f"  --> このステップで終了")
        return result
    print(f"  --> 該当なし、次へ")

    # ============================================
    # Step 5: 在庫販売 + 処理完了
    # ============================================
    print("\n[Step 5] 在庫販売 + 処理完了チェック")
    print(f"  伝票タイプ: '{row.document_type}'")
    print(f"  出荷ステータス: '{row.ship_status}'")
    print(f"  条件: 伝票タイプ='【受注】在庫販売' AND 出荷ステータス='処理完了'")

    result = _check_stock_completed(row, cache, holidays, branch, today, storage_place, use_ship_rule)
    if result is not None:
        print(f"  --> 判定結果: '{result}'")
        print(f"  --> このステップで終了")
        return result
    print(f"  --> 該当なし、次へ")

    # ============================================
    # Step 6: 紐付き + 処理完了
    # ============================================
    print("\n[Step 6] 紐付き + 処理完了チェック")
    print(f"  伝票タイプ: '{row.document_type}'")
    print(f"  出荷ステータス: '{row.ship_status}'")
    print(f"  保管場所: '{storage_place}'")
    print(f"  条件: 伝票タイプ='【受注】直送販売' AND 出荷ステータス='処理完了' AND 保管場所≠'転送中（直送用）'")

    result = _check_himozuki_completed(row, cache, holidays, branch, execution_time, today, storage_place, is_rosenbin)
    if result is not None:
        print(f"  --> 判定結果: '{result}'")
        print(f"  --> このステップで終了")
        return result
    print(f"  --> 該当なし、次へ")

    # ============================================
    # Step 7: 受注納期なし
    # ============================================
    print("\n[Step 7] 受注納期チェック")
    delivery_date = row.order_delivery_date
    print(f"  受注納期: {delivery_date}")

    if delivery_date is None:
        print(f"  --> 受注納期なし → '日程調整中'")
        return "日程調整中"
    print(f"  --> 受注納期あり、次へ")

    # ============================================
    # Step 8: 受注納期 = 12/31
    # ============================================
    print("\n[Step 8] 受注納期=12/31チェック")
    is_dec31 = is_december_31(delivery_date)
    print(f"  受注納期: {delivery_date}")
    print(f"  12/31判定: {is_dec31}")

    if is_dec31:
        # 確認中一覧から確定納期を取得
        confirmed_date = get_confirmed_delivery_date(row.order_number, row.detail_number, cache)
        print(f"  確認中一覧から確定納期: {confirmed_date}")

        result = _check_dec31(row, cache, holidays, today, storage_place, original_use_ship_rule, is_rosenbin)
        print(f"  --> 判定結果: '{result}'")
        return result
    print(f"  --> 12/31ではない、次へ")

    # ============================================
    # Step 9: 通常の納期計算
    # ============================================
    print("\n[Step 9] 通常の納期計算")
    days_to_add = get_delivery_days_to_add(row.item_group_code, cache)
    print(f"  品目Group: '{row.item_group_code}'")
    print(f"  配送加算日数: {days_to_add}")

    delivery_days = get_customer_delivery_days(row.customer_name, cache)
    print(f"  顧客配送曜日: {delivery_days} (VBA Weekday: 日=1,月=2,火=3,水=4,木=5,金=6,土=7)")

    print(f"\n  計算パラメータ:")
    print(f"    受注納期(ベース): {delivery_date}")
    print(f"    use_ship_rule: {use_ship_rule}")
    print(f"    original_use_ship_rule: {original_use_ship_rule}")
    print(f"    is_rosenbin: {is_rosenbin}")
    print(f"    storage_place: '{storage_place}'")

    # 転送中チェック
    if storage_place == "転送中（直送用）":
        print(f"\n  --> 転送中（直送用） → 受注納期そのまま出荷予定")

    # +営業日計算
    adjusted = add_business_days(delivery_date, days_to_add, holidays)
    print(f"\n  計算過程:")
    print(f"    受注納期 + {days_to_add}営業日 = {adjusted}")

    if delivery_days:
        next_delivery = get_next_delivery_day(adjusted, delivery_days, holidays)
        print(f"    次の配送曜日に調整: {adjusted} → {next_delivery}")

    result = _calc_normal(row, cache, holidays, today, delivery_date, storage_place, use_ship_rule, original_use_ship_rule, is_rosenbin)
    print(f"\n  --> 最終判定結果: '{result}'")

    return result


def main():
    # パラメータ
    source_file = r"\\flsv04\316京葉\納期回答書ツールフォルダ\受注一覧\16AM.xls"
    target_order_number = "GL2F446767"
    tool_folder = r"\\flsv04\316京葉\納期回答書ツールフォルダ"

    print_separator("納期回答デバッグ")
    print(f"対象ファイル: {source_file}")
    print(f"対象注番: {target_order_number}")
    print(f"ツールフォルダ: {tool_folder}")

    today = datetime.date.today()
    execution_time = datetime.datetime.now()
    print(f"今日の日付: {today}")
    print(f"実行時刻: {execution_time}")

    # ファイル読み込み
    print_separator("ファイル読み込み")

    print("受注一覧を読み込み中...")
    source_data = load_source_file(source_file)
    print(f"  読み込み行数: {len(source_data)}")

    cols = get_column_positions(source_data)
    if cols is None:
        print("ERROR: 列位置を取得できませんでした")
        sys.exit(1)
    print(f"  列位置取得: OK")

    # マスターファイル読み込み
    manufacturer_master_path = Path(tool_folder) / "メーカー一覧.xlsx"
    customer_master_path = Path(tool_folder) / "顧客マスター_v2.xlsm"
    history_path = Path(tool_folder) / "送付履歴.xlsx"

    print(f"\nメーカー一覧: {manufacturer_master_path}")
    manufacturer_wb = load_workbook(manufacturer_master_path, data_only=True)
    print(f"  読み込み: OK")

    print(f"\n顧客マスター: {customer_master_path}")
    customer_wb = load_workbook(customer_master_path, data_only=True)
    print(f"  読み込み: OK")

    print(f"\n送付履歴: {history_path}")
    try:
        history_wb = load_workbook(history_path, data_only=True)
        confirming_ws = history_wb["確認中一覧"] if "確認中一覧" in history_wb.sheetnames else None
        print(f"  読み込み: OK")
        print(f"  確認中一覧シート: {'あり' if confirming_ws else 'なし'}")
    except Exception as e:
        print(f"  読み込みエラー: {e}")
        confirming_ws = None

    # 営業所設定
    print("\n営業所設定を読み込み中...")
    branch = load_branch_settings(manufacturer_wb, source_data, cols)
    print(f"  営業所名: {branch.name}")
    print(f"  デフォルト締切時間: {branch.default_cutoff}時")
    print(f"  商品センター: {branch.base_center}")

    # 祝日読み込み
    print("\n祝日を読み込み中...")
    holidays = load_holidays(manufacturer_wb)
    print(f"  祝日件数: {len(holidays)}")

    # キャッシュ構築
    print("\nキャッシュを構築中...")
    cache = build_all_caches(manufacturer_wb, customer_wb, confirming_ws, source_data, cols)
    print(f"  mfg_name件数: {len(cache.mfg_name)}")
    print(f"  mfg_days件数: {len(cache.mfg_days)}")
    print(f"  cust_days件数: {len(cache.cust_days)}")
    print(f"  cust_route件数: {len(cache.cust_route)}")
    print(f"  confirm件数: {len(cache.confirm)}")
    print(f"  storage件数: {len(cache.storage)}")

    # 対象注番を検索
    print_separator("対象注番検索")
    target_row = None

    for row_idx in get_data_rows_range(source_data, cols):
        if not is_data_row(source_data, row_idx, cols):
            continue

        order_row = parse_order_row(source_data, row_idx, cols)
        if order_row.order_number == target_order_number:
            target_row = order_row
            print(f"注番 {target_order_number} を発見 (行番号: {row_idx})")
            break

    if target_row is None:
        print(f"ERROR: 注番 {target_order_number} が見つかりませんでした")
        sys.exit(1)

    # OrderRow全フィールド表示
    print_order_row(target_row)

    # 納期計算デバッグ
    result = debug_calculate_delivery_date(
        target_row, cache, holidays, branch, execution_time, today
    )

    print_separator("最終結果")
    print(f"納期回答: {result}")


if __name__ == "__main__":
    import io
    # 出力をファイルに保存
    output_file = Path(__file__).parent / "debug_output.txt"
    with open(output_file, "w", encoding="utf-8") as f:
        old_stdout = sys.stdout
        sys.stdout = f
        main()
        sys.stdout = old_stdout
    print(f"結果を {output_file} に保存しました")
