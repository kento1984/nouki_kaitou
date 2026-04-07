"""納期回答書作成 CLIエントリーポイント

コマンドラインまたは対話モードで納期回答書を生成する。

使用例:
    # 対話モード
    python -m nouki_kaitou.main

    # コマンドライン指定
    python -m nouki_kaitou.main --source path/to/10PM.XLS --customer "顧客名"

    # メールHTML出力付き
    python -m nouki_kaitou.main --source path/to/10PM.XLS --email-mode draft
"""

from __future__ import annotations

import argparse
import datetime
import os
import pickle
import sys
import time
import unicodedata
import warnings
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

from nouki_kaitou.cache import build_all_caches
from nouki_kaitou.config import load_branch_settings, load_holidays
from nouki_kaitou.customer import check_customer_master
from nouki_kaitou.data_loader import (
    get_column_positions,
    get_data_rows_range,
    is_data_row,
    load_source_file,
    parse_order_row,
)
from nouki_kaitou.email_builder import create_emails
from nouki_kaitou.history import (
    CONFIRMING_SHEET_NAME,
    HISTORY_SHEET_NAME,
    extract_sheet_rows,
    initialize_delivery_history,
    load_delivery_history,
    save_history_batch,
)
from nouki_kaitou.models import BranchSettings, OrderRow, ReportResult
from nouki_kaitou.representative import get_rep_list, is_split_by_rep
from nouki_kaitou.report_generator import (
    create_delivery_report,
    create_delivery_report_by_order_numbers,
)
from nouki_kaitou.utils import get_output_folder, is_file_open


def _normalize(s: str) -> str:
    """Unicode正規化（NFC）して比較用文字列を返す。"""
    return unicodedata.normalize("NFC", s)


def _find_file_in_dir(directory: Path, target_name: str) -> Path | None:
    """ディレクトリ内からファイルを探す（Unicode正規化で比較）。

    Windows + MINGW環境ではファイル名のUnicode正規化形式（NFC/NFD）が
    ずれることがあるため、正規化して比較する。
    """
    normalized_target = _normalize(target_name)
    # まず直接パスで試す（最速）
    direct = directory / target_name
    if direct.exists():
        return direct
    # 見つからない場合はディレクトリ走査でNFC正規化比較
    try:
        for p in directory.iterdir():
            if p.is_file() and _normalize(p.name) == normalized_target:
                return p
    except OSError:
        pass
    return None


def _resolve_tool_folder(source_path: Path) -> Path:
    """マスターファイルのあるツールフォルダを優先順で探索する。

    探索順:
        1. ソースファイルと同じフォルダ
        2. ソースファイルの親の親フォルダ（受注一覧/17PM.xls → ツールフォルダ）
        3. exe化時: exeが置かれたフォルダ（sys.executable の親）
           通常時: nouki_kaitouパッケージの親フォルダ（≒プロジェクトルート）
        4. カレントワーキングディレクトリ

    メーカー一覧.xlsxが最初に見つかったフォルダを返す。
    どこにもなければCWDを返す。
    """
    marker = "メーカー一覧.xlsx"

    # 候補3: PyInstaller exe化時は __file__ が一時展開フォルダ(_MEIPASS)を
    # 指すためマスターが見つからない。exe自体の場所を使う。
    if getattr(sys, "frozen", False):
        pkg_or_exe = Path(sys.executable).resolve().parent
    else:
        pkg_or_exe = Path(__file__).resolve().parent.parent

    candidates = [
        source_path.parent.resolve(),
        source_path.parent.parent.resolve(),
        pkg_or_exe,
        Path(os.getcwd()).resolve(),
    ]
    # 重複除去（順序維持）
    seen: set[Path] = set()
    for folder in candidates:
        if folder in seen:
            continue
        seen.add(folder)
        if _find_file_in_dir(folder, marker) is not None:
            return folder
    # どこにもなければCWD
    return Path(os.getcwd())


def find_source_file(directory: Path) -> Path | None:
    """ディレクトリから10PM.XLSを探す。

    大文字小文字を区別せず、10PM.XLS / 10PM.xls 等を検索。
    複数見つかった場合は更新日時が最新のものを返す。
    """
    candidates: list[Path] = []
    for p in directory.iterdir():
        if p.is_file() and p.name.upper() == "10PM.XLS":
            candidates.append(p)
    if not candidates:
        return None
    # 更新日時が最新のものを返す
    return max(candidates, key=lambda p: p.stat().st_mtime)


def get_unique_customers(orders: list[OrderRow]) -> list[str]:
    """OrderRowリストから重複なし顧客名リストを取得する。

    出現順を維持する。
    """
    seen: set[str] = set()
    result: list[str] = []
    for order in orders:
        name = order.customer_name
        if name and name not in seen:
            seen.add(name)
            result.append(name)
    return result


def result_to_email_input(result: ReportResult) -> dict:
    """ReportResultをcreate_emails用のdict形式に変換する。"""
    return {
        "customer_name": result.customer_name,
        "file_path": result.file_path,
        "stockout_info_list": result.stockout_info_list,
        "tracking_info_list": result.tracking_info_list,
        "bunno_info_list": result.bunno_info_list,
        "rep_name": result.rep_name,
        "bunno_completed_list": result.bunno_completed_list,
    }


def _load_sent_orders_cached(
    history_path: Path,
    ws_history: object,
    ws_confirming: object,
    cache: object,
    holidays: object,
    today: datetime.date | None = None,
) -> dict[str, str]:
    """送付履歴をキャッシュ付きで読み込む。

    pickleキャッシュが有効な場合（xlsx未変更かつ同日）はキャッシュから
    即座に読み込む。無効な場合は通常読み込みし、結果をキャッシュする。

    Args:
        history_path: 送付履歴.xlsxのパス
        ws_history: 送付履歴シート（read_only）
        ws_confirming: 確認中一覧シート（read_only）
        cache: キャッシュストア
        holidays: 祝日辞書
        today: 基準日

    Returns:
        送付済み伝票辞書 dict[キー(受発注伝票|明細), 納期回答ステータス]
    """
    if today is None:
        today = datetime.date.today()

    cache_file = Path(str(history_path) + ".cache")

    # キャッシュキー: xlsxの更新日時 + 本日日付
    try:
        xlsx_mtime = os.path.getmtime(str(history_path))
    except OSError:
        xlsx_mtime = 0
    cache_key = f"{xlsx_mtime}|{today.isoformat()}"

    # キャッシュ読み込み試行
    if cache_file.exists():
        try:
            with open(cache_file, "rb") as f:
                cached = pickle.load(f)
            if isinstance(cached, dict) and cached.get("_key") == cache_key:
                return cached["sent_orders"]
        except (pickle.UnpicklingError, KeyError, EOFError, OSError):
            pass

    # 通常読み込み
    sent_orders = load_delivery_history(
        ws_history, ws_confirming, cache, holidays, today
    )

    # キャッシュ保存
    try:
        with open(cache_file, "wb") as f:
            pickle.dump(
                {
                    "_key": cache_key,
                    "sent_orders": sent_orders,
                },
                f,
                protocol=pickle.HIGHEST_PROTOCOL,
            )
    except OSError:
        pass

    return sent_orders


def parse_args() -> argparse.Namespace:
    """コマンドライン引数をパースする。"""
    parser = argparse.ArgumentParser(
        description="納期回答書作成マクロ（Python版）",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "引数なしで起動すると対話モードになります。\n"
            "例:\n"
            "  python -m nouki_kaitou.main\n"
            '  python -m nouki_kaitou.main --source 10PM.XLS --customer "顧客名"\n'
        ),
    )
    parser.add_argument(
        "--source",
        type=str,
        default=None,
        help="10PM.XLSのパス（省略時: カレントディレクトリを探索）",
    )
    parser.add_argument(
        "--date-from",
        type=str,
        default=None,
        help="期間開始日 YYYY/MM/DD",
    )
    parser.add_argument(
        "--date-to",
        type=str,
        default=None,
        help="期間終了日 YYYY/MM/DD",
    )
    parser.add_argument(
        "--customer",
        type=str,
        default=None,
        help="顧客名（省略時: 全顧客 or 対話で選択）",
    )
    parser.add_argument(
        "--email-mode",
        type=str,
        choices=["send", "draft", "none"],
        default="none",
        help="メールモード: send=送信, draft=下書き, none=メール生成なし（デフォルト: none）",
    )
    parser.add_argument(
        "--output-dir",
        type=str,
        default=None,
        help="出力先ディレクトリ（省略時: 自動生成）",
    )
    parser.add_argument(
        "--sender",
        type=str,
        default="",
        help="送付者名（送付履歴に記録する名前）",
    )
    return parser.parse_args()


def _parse_date_arg(date_str: str) -> datetime.date:
    """YYYY/MM/DD形式の文字列をdatetime.dateに変換する。"""
    try:
        return datetime.datetime.strptime(date_str, "%Y/%m/%d").date()
    except ValueError:
        pass
    try:
        return datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        raise ValueError(
            f"日付の形式が不正です: '{date_str}'（YYYY/MM/DD または YYYY-MM-DD）"
        )


def interactive_mode(args: argparse.Namespace, orders: list[OrderRow] | None = None) -> argparse.Namespace:
    """対話モードで不足引数を補完する。

    Args:
        args: コマンドライン引数（一部が未指定の場合あり）
        orders: 受注データ（顧客選択に使用。Noneならソースファイル確定後に読み込む）
    """
    print("=" * 50)
    print("  納期回答書作成マクロ（対話モード）")
    print("=" * 50)
    print()

    # 1. ソースファイル
    if args.source is None:
        default_file = find_source_file(Path.cwd())
        if default_file:
            prompt = f"10PM.XLSのパス [Enter で {default_file.name}]: "
            user_input = input(prompt).strip()
            if not user_input:
                args.source = str(default_file)
            else:
                args.source = user_input
        else:
            args.source = input("10PM.XLSのパス: ").strip()

    if not args.source or not Path(args.source).exists():
        print(f"エラー: ファイルが見つかりません: {args.source}")
        sys.exit(1)

    # 2. 期間
    if args.date_from is None:
        user_input = input("期間開始日 (YYYY/MM/DD、Enter でスキップ): ").strip()
        if user_input:
            args.date_from = user_input

    if args.date_to is None:
        user_input = input("期間終了日 (YYYY/MM/DD、Enter でスキップ): ").strip()
        if user_input:
            args.date_to = user_input

    # 3. 顧客選択
    if args.customer is None and orders is not None:
        customer_names = get_unique_customers(orders)
        if customer_names:
            print()
            print("--- 顧客一覧 ---")
            for i, name in enumerate(customer_names, 1):
                print(f"  {i:3d}. {name}")
            print()
            user_input = input(
                '顧客番号を入力 (複数はカンマ区切り、"all" で全顧客、Enter で全顧客): '
            ).strip()
            if user_input and user_input.lower() != "all":
                try:
                    indices = [int(x.strip()) for x in user_input.split(",")]
                    selected = [customer_names[i - 1] for i in indices if 1 <= i <= len(customer_names)]
                    if selected:
                        # 複数顧客の場合はカンマ区切りで格納
                        args.customer = ",".join(selected) if len(selected) > 1 else selected[0]
                except (ValueError, IndexError):
                    print("入力が不正です。全顧客で処理します。")

    # 4. メールモード
    if args.email_mode == "none":
        user_input = input("メールモード (send/draft/none) [Enter で none]: ").strip().lower()
        if user_input in ("send", "draft", "none"):
            args.email_mode = user_input

    # 5. 送付者名
    if not args.sender:
        user_input = input("送付者名 (Enter でスキップ): ").strip()
        if user_input:
            args.sender = user_input

    print()
    return args


def _build_email_customers(
    cust_wb: object, email_start_col: int = 4
) -> set[str]:
    """顧客マスターからメールアドレス登録済み顧客のセットを構築する。"""
    email_customers: set[str] = set()
    try:
        cust_ws = cust_wb["顧客マスター"]
    except KeyError:
        return email_customers

    for row in cust_ws.iter_rows(min_row=2, values_only=True):
        name = str(row[0]).strip() if row[0] else ""
        if not name:
            continue
        for j in range(email_start_col, len(row)):
            if row[j] and str(row[j]).strip():
                email_customers.add(name)
                break

    return email_customers


def _show_gui(args: argparse.Namespace) -> tuple[dict | None, dict]:
    """GUI表示に必要なデータを読み込み、SelectionDialogを表示する。

    Returns:
        (ダイアログの結果dict or None, 読み込み済みデータdict)
        キャンセル時でも読み込み済みデータは返す（呼び出し側で判定）。
    """
    source_path = Path(args.source)
    tool_folder = _resolve_tool_folder(source_path)

    # ソースファイル読み込み
    source_data_raw = load_source_file(str(source_path))
    result = get_column_positions(source_data_raw)
    if result is None:
        print("エラー: ヘッダー行の列位置を検出できませんでした。")
        sys.exit(1)
    cols, header_row_idx = result

    orders: list[OrderRow] = []
    for i in get_data_rows_range(source_data_raw, cols, header_row_idx):
        if is_data_row(source_data_raw, i, cols):
            orders.append(parse_order_row(source_data_raw, i, cols))

    if not orders:
        print("処理対象の受注データがありません。")
        sys.exit(1)

    # マスターファイル読み込み（GUI表示用 → run()でも再利用）
    mfg_found = _find_file_in_dir(tool_folder, "メーカー一覧.xlsx")
    if mfg_found and is_file_open(str(mfg_found)):
        print("エラー: メーカー一覧.xlsxが使用中です。")
        print("  メーカー一覧.xlsxを閉じてから再実行してください。")
        sys.exit(1)
    mfg_wb = load_workbook(str(mfg_found), data_only=True) if mfg_found else None

    cust_found = _find_file_in_dir(tool_folder, "顧客マスター_v2.xlsm")
    if not cust_found:
        print(f"エラー: 顧客マスターが見つかりません: {tool_folder}")
        sys.exit(1)
    if is_file_open(str(cust_found)):
        print("エラー: 顧客マスター_v2.xlsmが使用中です。")
        print("  顧客マスター_v2.xlsmを閉じてから再実行してください。")
        sys.exit(1)
    warnings.filterwarnings('ignore', message='Data Validation extension', category=UserWarning)
    cust_wb = load_workbook(str(cust_found), data_only=True)
    warnings.resetwarnings()

    # キャッシュ構築（master_customers用。confirming_wsはGUIに不要なのでNone）
    cache = build_all_caches(mfg_wb, cust_wb, None, source_data_raw, cols)
    branch = (
        load_branch_settings(mfg_wb, source_data_raw, cols)
        if mfg_wb is not None
        else BranchSettings()
    )

    master_customers = set(cache.cust_days.keys())
    email_customers = _build_email_customers(cust_wb, cache.cust_email_start_col)

    from nouki_kaitou.gui import SelectionDialog

    dialog = SelectionDialog(orders, branch, master_customers, email_customers)
    gui_result = dialog.show()

    # run()で再利用するためのデータ
    preloaded = {
        "source_data_raw": source_data_raw,
        "cols": cols,
        "header_row_idx": header_row_idx,
        "orders": orders,
        "mfg_wb": mfg_wb,
        "cust_wb": cust_wb,
        "tool_folder": tool_folder,
        "branch": branch,
    }
    return gui_result, preloaded


def run(args: argparse.Namespace, preloaded: dict | None = None) -> None:
    """メイン処理を実行する。

    Args:
        args: コマンドライン引数
        preloaded: _show_guiで読み込み済みのデータ（GUI経由時に重複読み込みを回避）
    """
    execution_time = datetime.datetime.now()
    t_run_start = time.perf_counter()

    # --- 1-3. ソースファイル・マスター読み込み ---
    source_path = Path(args.source)
    print(f"ソースファイル: {source_path}")

    if preloaded is not None:
        # GUI経由: _show_guiで読み込み済みのデータを再利用
        source_data_raw = preloaded["source_data_raw"]
        cols = preloaded["cols"]
        orders = preloaded["orders"]
        mfg_wb = preloaded["mfg_wb"]
        cust_wb = preloaded["cust_wb"]
        tool_folder = preloaded["tool_folder"]
        branch = preloaded["branch"]
        print(f"受注データ: {len(orders)}件（GUI読み込み済み）")
    else:
        # CLI経由: 自前で読み込む
        source_data_raw = load_source_file(str(source_path))

        result = get_column_positions(source_data_raw)
        if result is None:
            print("エラー: ヘッダー行の列位置を検出できませんでした。")
            print("受注リストファイルの形式を確認してください。")
            sys.exit(1)
        cols, header_row_idx = result

        orders = []
        for i in get_data_rows_range(source_data_raw, cols, header_row_idx):
            if is_data_row(source_data_raw, i, cols):
                orders.append(parse_order_row(source_data_raw, i, cols))

        print(f"受注データ: {len(orders)}件")

        if not orders:
            print("処理対象の受注データがありません。")
            return

        tool_folder = _resolve_tool_folder(source_path)

        mfg_found = _find_file_in_dir(tool_folder, "メーカー一覧.xlsx")
        if mfg_found:
            if is_file_open(str(mfg_found)):
                print("エラー: メーカー一覧.xlsxが使用中です。")
                print("  メーカー一覧.xlsxを閉じてから再実行してください。")
                sys.exit(1)
            mfg_wb = load_workbook(str(mfg_found), data_only=True)
        else:
            print(f"警告: メーカー一覧が見つかりません: {tool_folder}")
            print("  祝日カレンダー・営業所設定・メーカーキャッシュなしで続行します。")
            mfg_wb = None

        cust_found = _find_file_in_dir(tool_folder, "顧客マスター_v2.xlsm")
        if not cust_found:
            print(f"エラー: 顧客マスターが見つかりません: {tool_folder}")
            sys.exit(1)
        if is_file_open(str(cust_found)):
            print("エラー: 顧客マスター_v2.xlsmが使用中です。")
            print("  顧客マスター_v2.xlsmを閉じてから再実行してください。")
            sys.exit(1)
        warnings.filterwarnings('ignore', message='Data Validation extension', category=UserWarning)
        cust_wb = load_workbook(str(cust_found), data_only=True)
        warnings.resetwarnings()

        branch = load_branch_settings(mfg_wb, source_data_raw, cols) if mfg_wb is not None else BranchSettings()

    print(f"ツールフォルダ: {tool_folder}")

    try:

        # 担当者マスターシート読み込み（シートが存在しない場合はNone → 担当者分割なし）
        try:
            rep_master_ws = cust_wb["担当者マスター"]
        except KeyError:
            rep_master_ws = None

        history_found = _find_file_in_dir(tool_folder, "送付履歴.xlsx")
        history_path = history_found if history_found else tool_folder / "送付履歴.xlsx"
        history_exists = history_found is not None

        # 送付履歴ファイルの排他チェック（他のプロセスで使用中なら中断）
        if history_exists and is_file_open(str(history_path)):
            print("エラー: 送付履歴ファイルが使用中です。")
            print("  他の人が作業中のため実行できません。")
            print(f"  ファイル: {history_path}")
            print("  使用中の人に声をかけて閉じてもらってから再実行してください。")
            sys.exit(1)

        t_history_load = time.perf_counter()
        if history_exists:
            # read_only=True で高速読み込み（書き込みは後で別途ロード）
            history_wb_ro = load_workbook(str(history_found), read_only=True)
        else:
            print("送付履歴ファイルを新規作成します。")
            history_wb_ro = initialize_delivery_history(str(history_path))

        # --- 4-5. キャッシュ構築・送付履歴読み込み ---
        try:
            ws_confirming_ro = history_wb_ro[CONFIRMING_SHEET_NAME]
            cache = build_all_caches(mfg_wb, cust_wb, ws_confirming_ro, source_data_raw, cols)
            holidays = load_holidays(mfg_wb) if mfg_wb is not None else {}

            t_data_ready = time.perf_counter()
            print(f"営業所: {branch.name}")
            print(f"データ読み込み: {t_data_ready - t_run_start:.2f}s（うち送付履歴: {t_data_ready - t_history_load:.2f}s）")

            # 送付履歴読み込み（pickleキャッシュ付き）
            ws_history_ro = history_wb_ro[HISTORY_SHEET_NAME]
            t0 = time.perf_counter()
            sent_orders = _load_sent_orders_cached(
                history_path, ws_history_ro, ws_confirming_ro, cache, holidays
            )
            history_elapsed = time.perf_counter() - t0
            print(f"送付済み伝票: {len(sent_orders)}件 ({history_elapsed:.2f}s)")

            # 送付履歴の行データをメモリに保持（後のバッチ書き込みで使用）
            history_rows = extract_sheet_rows(ws_history_ro, 9)
            confirming_rows = extract_sheet_rows(ws_confirming_ro, 11)
        finally:
            # read_onlyワークブックを閉じる（例外発生時も確実にクローズ）
            history_wb_ro.close()

        # --- 6. モード判定・期間パース ---
        order_numbers_mode = getattr(args, "order_numbers", None)

        date_from: Optional[datetime.date] = None
        date_to: Optional[datetime.date] = None
        if not order_numbers_mode:
            if args.date_from:
                date_from = _parse_date_arg(args.date_from)
            if args.date_to:
                date_to = _parse_date_arg(args.date_to)

            if date_from or date_to:
                period_str = f"{date_from or '---'} ～ {date_to or '---'}"
                print(f"期間: {period_str}")
        else:
            print(f"伝票番号指定モード: {len(order_numbers_mode)}件")

        # --- 7. 顧客別事前グルーピング ---
        orders_by_customer: dict[str, list[OrderRow]] = {}
        for order in orders:
            name = order.customer_name.strip()
            if name:
                if name not in orders_by_customer:
                    orders_by_customer[name] = []
                orders_by_customer[name].append(order)

        all_customer_names = list(orders_by_customer.keys())
        master_customers = set(cache.cust_days.keys())
        skipped_order_count = 0

        if order_numbers_mode:
            # 伝票番号モード: 指定された番号から対象顧客を自動決定
            order_number_set = set(order_numbers_mode)
            target_names: set[str] = set()
            for order in orders:
                if order.order_number in order_number_set:
                    name = order.customer_name.strip()
                    if name and name in master_customers:
                        target_names.add(name)
            customer_names = sorted(target_names)
            if not customer_names:
                print("エラー: 指定された伝票番号に対応する顧客が見つかりません。")
                return
        elif args.customer:
            # カンマ区切りの複数顧客に対応
            selected_names = [n.strip() for n in args.customer.split(",")]
            # 存在チェック
            for name in selected_names:
                if name not in all_customer_names:
                    print(f"警告: 顧客 '{name}' は受注データに存在しません。")
            customer_names = [n for n in selected_names if n in all_customer_names]
            if not customer_names:
                print("エラー: 指定された顧客が受注データに見つかりません。")
                return
            # 顧客マスターに登録されている顧客のみ対象（VBA版と同じ動作）
            skipped = [n for n in customer_names if n not in master_customers]
            customer_names = [n for n in customer_names if n in master_customers]
            skipped_order_count = sum(len(orders_by_customer.get(n, [])) for n in skipped)
            if skipped:
                print(f"顧客マスター未登録（スキップ）: {len(skipped)}件")
        else:
            customer_names = all_customer_names
            # 顧客マスターに登録されている顧客のみ対象（VBA版と同じ動作）
            skipped = [n for n in customer_names if n not in master_customers]
            customer_names = [n for n in customer_names if n in master_customers]
            skipped_order_count = sum(len(orders_by_customer.get(n, [])) for n in skipped)
            if skipped:
                print(f"顧客マスター未登録（スキップ）: {len(skipped)}件")

        print(f"対象顧客: {len(customer_names)}件")

        # 顧客マスターのメールアドレスチェック（警告のみ）
        if args.email_mode != "none":
            cust_ws = cust_wb["顧客マスター"]
            missing = check_customer_master(
                customer_names, cust_ws, cache.cust_email_start_col
            )
            if missing:
                print("警告: 以下の顧客はメールアドレスが未登録です:")
                print(missing)

        # --- 8. 出力フォルダ ---
        if args.output_dir:
            output_dir = Path(args.output_dir)
            output_dir.mkdir(parents=True, exist_ok=True)
        else:
            output_dir = get_output_folder(str(tool_folder), execution_time)

        print(f"出力先: {output_dir}")
        print()

        # --- 9. 回答書生成ループ ---
        all_results: list[ReportResult] = []
        gen_start = time.perf_counter()

        for cust in customer_names:
            # 担当者別分割判定
            if rep_master_ws is not None and is_split_by_rep(cust, rep_master_ws):
                rep_list = get_rep_list(cust, rep_master_ws)
                # 登録担当者ごと + __OTHER__（未登録担当者用）
                rep_names_to_process = rep_list + ["__OTHER__"]
            else:
                rep_names_to_process = [""]  # 分割なし

            for rep_name in rep_names_to_process:
                if order_numbers_mode:
                    # 伝票番号指定モード
                    result = create_delivery_report_by_order_numbers(
                        source_data=orders_by_customer.get(cust, []),
                        customer_name=cust,
                        order_numbers=order_numbers_mode,
                        cache=cache,
                        output_dir=output_dir,
                        holidays=holidays,
                        branch=branch,
                        execution_time=execution_time,
                        rep_name=rep_name,
                        rep_master_ws=rep_master_ws,
                    )
                else:
                    # 期間指定モード
                    result = create_delivery_report(
                        source_data=orders_by_customer.get(cust, []),
                        customer_name=cust,
                        sent_orders=sent_orders,
                        cache=cache,
                        output_dir=output_dir,
                        holidays=holidays,
                        branch=branch,
                        execution_time=execution_time,
                        date_from=date_from,
                        date_to=date_to,
                        rep_name=rep_name,
                        rep_master_ws=rep_master_ws,
                    )

                if result:
                    all_results.append(result)
                    confirmed_count = len(result.confirmed_orders)
                    confirming_count = len(result.confirming_orders)
                    print(f"  生成: {result.file_path}")
                    print(f"    確定: {confirmed_count}件 / 確認中: {confirming_count}件")
                elif rep_name == "" or rep_name == rep_names_to_process[0]:
                    # 分割なし or 最初の担当者のみスキップ表示（__OTHER__等は静かにスキップ）
                    print(f"  スキップ: {cust}（対象データなし）")

        gen_elapsed = time.perf_counter() - gen_start
        print(f"\n回答書生成: {gen_elapsed:.2f}s")

        print()
        print(f"生成件数: {len(all_results)}件 / {len(customer_names)}顧客")

        if not all_results:
            print("生成対象のデータがありませんでした。")
            return

        # --- 10. 送付履歴保存（バッチ化して1回の読み書きで完了） ---
        # 全顧客の結果をまとめてから1回だけ保存（毎回44K行の再読み書きを回避）
        all_confirmed = []
        all_confirming = []
        for result in all_results:
            all_confirmed.extend(result.confirmed_orders)
            all_confirming.extend(result.confirming_orders)

        # SAPで明細削除された伝票を確認中一覧から除去するためのキー収集
        deleted_keys: set[str] = set()
        for order in orders:
            if order.rejection_reason.strip() == "明細削除":
                key = f"{order.order_number}|{order.detail_number}"
                if key in cache.confirm:
                    deleted_keys.add(key)
        if deleted_keys:
            print(f"明細削除（確認中一覧から除去）: {len(deleted_keys)}件")

        has_updates = bool(all_confirmed or all_confirming or deleted_keys)
        if has_updates:
            t_save = time.perf_counter()
            save_history_batch(
                str(history_path), history_rows, confirming_rows,
                all_confirmed, all_confirming,
                execution_time, args.sender,
                deleted_keys=deleted_keys,
            )
            save_elapsed = time.perf_counter() - t_save
            print(f"送付履歴保存: {history_path} ({save_elapsed:.2f}s)")
        else:
            print("送付履歴: 更新なし")

        # --- 11. メール生成 ---
        if args.email_mode != "none" and all_results:
            from nouki_kaitou.email_builder import (
                create_outlook_drafts,
                create_outlook_sends,
            )

            created_files = [result_to_email_input(r) for r in all_results]
            cust_ws = cust_wb["顧客マスター"]
            emails, skipped_email_customers = create_emails(
                created_files=created_files,
                branch=branch,
                customer_master_ws=cust_ws,
                rep_master_ws=rep_master_ws,
                holidays=holidays,
                cache=cache,
                send_directly=(args.email_mode == "send"),
            )

            if emails:
                if args.email_mode == "send":
                    # 直接送信
                    sent = create_outlook_sends(emails)
                    print(f"メール送信: {len(sent)}件")
                elif args.email_mode == "draft":
                    # Outlook下書き作成
                    created = create_outlook_drafts(emails)
                    print(f"Outlook下書き作成: {len(created)}件")
            else:
                print("メール生成: 0件（宛先未登録等でスキップ）")

            # スキップした顧客名を表示（VBA版 CreateEmails L5655-5661 相当）
            if skipped_email_customers:
                print()
                print("※以下の顧客はメールアドレス未登録のためメール作成をスキップしました：")
                for name in skipped_email_customers:
                    print(f"  ・{name}")
                print("  手動で送付してください。")

        # --- 処理結果サマリー ---
        confirmed_total = len(all_confirmed)
        confirming_total = len(all_confirming)
        total = confirmed_total + confirming_total + skipped_order_count
        print("=" * 40)
        print("  処理結果サマリー")
        print("=" * 40)
        print(f"  確定:      {confirmed_total}件")
        print(f"  確認中:    {confirming_total}件")
        print(f"  スキップ:  {skipped_order_count}件")
        print(f"  合計:      {total}件")
        print("=" * 40)
        print()
        print("完了しました。")
    finally:
        # マスターワークブックを閉じる（例外発生時も確実にクローズ）
        if mfg_wb is not None:
            mfg_wb.close()
        cust_wb.close()


def main() -> None:
    """エントリーポイント。"""
    args = parse_args()

    # --- ソースファイル確定 ---
    if args.source is None:
        # 対話モードでソースファイルを選択
        default_file = find_source_file(Path.cwd())
        if default_file:
            prompt = f"10PM.XLSのパス [Enter で {default_file.name}]: "
            user_input = input(prompt).strip()
            args.source = user_input if user_input else str(default_file)
        else:
            args.source = input("10PM.XLSのパス: ").strip()

    if not args.source or not Path(args.source).exists():
        print(f"エラー: ファイルが見つかりません: {args.source}")
        sys.exit(1)

    # --- GUIモード判定: --customer未指定ならGUI表示 ---
    preloaded = None
    if args.customer is None:
        gui_result, preloaded = _show_gui(args)
        if gui_result is None:
            # キャンセル
            print("キャンセルされました。")
            return

        # メール送信モード（GUIで選択された3択）
        args.email_mode = gui_result.get("email_mode", "none")

        if gui_result["mode"] == "period":
            args.date_from = gui_result["date_from"].strftime("%Y/%m/%d")
            args.date_to = gui_result["date_to"].strftime("%Y/%m/%d")
            args.customer = ",".join(gui_result["customers"])
        elif gui_result["mode"] == "ordernumber":
            args.order_numbers = gui_result["order_numbers"]

    try:
        run(args, preloaded=preloaded)
    except KeyboardInterrupt:
        print("\n中断しました。")
        sys.exit(1)
    except Exception as e:
        print(f"\nエラーが発生しました: {e}")
        raise


if __name__ == "__main__":
    main()
