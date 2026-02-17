"""送付履歴を検索するスクリプト"""

import sys
import io
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from openpyxl import load_workbook

def main():
    order_no = sys.argv[1] if len(sys.argv) > 1 else "GL2V446963"

    tool_folder = Path(r"\\flsv04\316京葉\納期回答書ツールフォルダ")
    history_path = tool_folder / "送付履歴.xlsx"

    print(f"検索: {order_no}")
    print()

    wb = load_workbook(str(history_path), data_only=True, read_only=True)

    # 送付履歴シート
    print("=== 送付履歴 ===")
    ws = wb["送付履歴"]
    found = False
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0] and order_no in str(row[0]):
            print(f"  注番: {row[0]}")
            print(f"  確定: {row[2] if len(row) > 2 else 'N/A'}")
            print(f"  回答: {row[3] if len(row) > 3 else 'N/A'}")
            print()
            found = True
    if not found:
        print("  見つからず")

    # 確認中一覧シート
    print()
    print("=== 確認中一覧 ===")
    ws2 = wb["確認中一覧"]
    found = False
    for row in ws2.iter_rows(min_row=2, values_only=True):
        if row and row[0] and order_no in str(row[0]):
            print(f"  注番: {row[0]}")
            print(f"  確定日: {row[1] if len(row) > 1 else 'N/A'}")
            found = True
    if not found:
        print("  見つからず")

    wb.close()


if __name__ == "__main__":
    main()
