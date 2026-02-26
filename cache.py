"""キャッシュ構築モジュール

VBAの以下の関数を移植:
- BuildManufacturerCache (L796): メーカー一覧からメーカー名・加算日数キャッシュ
- BuildCustomerCache (L827): 顧客マスターから配送曜日・保持日数・路線便キャッシュ
- BuildConfirmingCache (L864): 確認中テーブルから問合せ状況キャッシュ
- BuildStorageCache (L906): 受注一覧から保管場所キャッシュ
"""

from __future__ import annotations

import datetime
from typing import TYPE_CHECKING, Optional

from nouki_kaitou.models import CacheStore, ColumnMap, DeliveryPattern
from nouki_kaitou.utils import parse_date

if TYPE_CHECKING:
    import openpyxl.worksheet.worksheet as ws_type


# ============================================
# 曜日文字列パーサー
# ============================================
# VBA Weekday番号: 日=1, 月=2, 火=3, 水=4, 木=5, 金=6, 土=7
_WEEKDAY_MAP: dict[str, int] = {
    "日": 1, "月": 2, "火": 3, "水": 4,
    "木": 5, "金": 6, "土": 7,
}


def _parse_weekday_string(text: str) -> list[int]:
    """出荷曜日文字列をVBA Weekday番号リストに変換する。

    例: "月水金" → [2, 4, 6]
    """
    if not text:
        return []
    result: list[int] = []
    for ch in text:
        if ch in _WEEKDAY_MAP:
            result.append(_WEEKDAY_MAP[ch])
    return result


# ============================================
# VBA: BuildManufacturerCache (L796-822)
# メーカー一覧シートからメーカー名・配送加算日数キャッシュ
# ============================================
def build_manufacturer_cache(
    manufacturer_master_wb: object,
) -> tuple[dict[str, str], dict[str, int]]:
    """メーカー一覧シートからメーカー名・配送加算日数キャッシュを構築する。

    メーカー一覧シートの構造:
    - A列: 品目GroupCode (キー)
    - B列: メーカー名
    - C列: 配送加算日数 (数値。空欄ならデフォルト2)

    Returns:
        (mfg_name辞書, mfg_days辞書)
    """
    mfg_name: dict[str, str] = {}
    mfg_days: dict[str, int] = {}

    try:
        ws = manufacturer_master_wb["メーカー一覧"]
    except KeyError:
        return mfg_name, mfg_days

    for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
        key = str(row[0]).strip() if row[0] else ""
        if not key or key in mfg_name:
            continue

        # B列: メーカー名
        mfg_name[key] = str(row[1]).strip() if row[1] else ""

        # C列: 配送加算日数
        days_val = row[2] if len(row) > 2 else None
        try:
            if days_val is not None and str(days_val).strip() != "":
                mfg_days[key] = int(days_val)
            else:
                mfg_days[key] = 2  # デフォルト値
        except (ValueError, TypeError):
            mfg_days[key] = 2

    return mfg_name, mfg_days


# ============================================
# 顧客マスターのフォーマット検出
# ============================================
def _detect_customer_master_format(ws: object) -> bool:
    """顧客マスターが新フォーマット（E列=配送パターン）かどうかを検出する。

    E列（0-indexed: 4）の最初の非空データセルを調べて判断する。
    - ``@`` を含む → 旧フォーマット（E列=メールアドレス）
    - ``@`` を含まない → 新フォーマット（E列=配送パターン名）
    - E列が全て空 → 旧フォーマット（安全デフォルト）

    Returns:
        True = 新フォーマット（E列が配送パターン列）
    """
    for row in ws.iter_rows(min_row=2, max_col=5, values_only=True):
        if len(row) < 5:
            continue
        val = str(row[4]).strip() if row[4] else ""
        if val:
            return "@" not in val
    return False


# ============================================
# VBA: BuildCustomerCache (L827-859)
# 顧客マスターから配送曜日・保持日数・路線便キャッシュ
# ============================================
def build_customer_cache(
    customer_master_wb: object,
) -> tuple[dict[str, list[int]], dict[str, int], dict[str, bool], dict[str, str]]:
    """顧客マスターシートから配送曜日・保持日数・路線便・配送パターンキャッシュを構築する。

    顧客マスターシートの構造（旧フォーマット: A-D列+E列以降メール）:
    - A列: 顧客名 (キー)
    - B列: 出荷曜日 ("月水金"等)
    - C列: 保持日数 (数値)
    - D列: 路線便 (空欄以外=True)

    新フォーマット（E列=配送パターン名、F列以降メール）:
    - E列: 配送パターン名（"近隣2便", "遠方午前" 等）

    Returns:
        (cust_days辞書, cust_retention辞書, cust_route辞書, cust_pattern辞書)
    """
    cust_days: dict[str, list[int]] = {}
    cust_retention: dict[str, int] = {}
    cust_route: dict[str, bool] = {}
    cust_pattern: dict[str, str] = {}

    try:
        ws = customer_master_wb["顧客マスター"]
    except KeyError:
        return cust_days, cust_retention, cust_route, cust_pattern

    has_pattern = _detect_customer_master_format(ws)
    max_col = 5 if has_pattern else 4

    for row in ws.iter_rows(min_row=2, max_col=max_col, values_only=True):
        key = str(row[0]).strip() if row[0] else ""
        if not key or key in cust_days:
            continue

        # B列: 出荷曜日
        days_str = str(row[1]).strip() if row[1] else ""
        cust_days[key] = _parse_weekday_string(days_str)

        # C列: 保持日数
        retention_val = row[2] if len(row) > 2 else None
        try:
            if (
                retention_val is not None
                and str(retention_val).strip() != ""
                and int(retention_val) > 0
            ):
                cust_retention[key] = int(retention_val)
            else:
                cust_retention[key] = 0
        except (ValueError, TypeError):
            cust_retention[key] = 0

        # D列: 路線便フラグ
        route_val = str(row[3]).strip() if (len(row) > 3 and row[3]) else ""
        cust_route[key] = route_val != ""

        # E列: 配送パターン（新フォーマットのみ）
        if has_pattern and len(row) > 4:
            pattern_val = str(row[4]).strip() if row[4] else ""
            if pattern_val:
                cust_pattern[key] = pattern_val

    return cust_days, cust_retention, cust_route, cust_pattern


# ============================================
# 配送パターンキャッシュ
# ============================================
_DAYS_LABEL_MAP: dict[str, int] = {
    "当日": 0,
    "翌日": 1,
    "翌々日": 2,
}


def _parse_time_str(text: str) -> tuple[int, int] | None:
    """'HH:MM'形式の文字列を(hour, minute)タプルに変換する。"""
    if not text:
        return None
    parts = str(text).strip().split(":")
    if len(parts) < 2:
        return None
    try:
        return int(parts[0]), int(parts[1])
    except (ValueError, TypeError):
        return None


def build_pattern_cache(
    manufacturer_master_wb: object,
) -> dict[str, DeliveryPattern]:
    """メーカー一覧.xlsx「配送パターン」シートからパターン定義を構築する。

    シート構造:
    - A列: パターン名 (例: "近隣2便")
    - B列: cutoff1 (例: "11:30")
    - C列: cutoff1前の日数ラベル (例: "当日")
    - D列: cutoff2 (例: "16:00", 空欄なら1段階)
    - E列: cutoff2前の日数ラベル (例: "翌日")

    Returns:
        {パターン名: DeliveryPattern}。シートなしなら空dict。
    """
    patterns: dict[str, DeliveryPattern] = {}

    try:
        ws = manufacturer_master_wb["配送パターン"]
    except KeyError:
        return patterns

    for row in ws.iter_rows(min_row=2, max_col=5, values_only=True):
        name = str(row[0]).strip() if row[0] else ""
        if not name or name in patterns:
            continue

        cutoff1 = _parse_time_str(row[1]) if len(row) > 1 else None
        if cutoff1 is None:
            continue

        days_before = _DAYS_LABEL_MAP.get(
            str(row[2]).strip() if (len(row) > 2 and row[2]) else "", 0
        )

        cutoff2 = None
        days_between = 1
        if len(row) > 3 and row[3]:
            cutoff2 = _parse_time_str(row[3])
        if cutoff2 is not None and len(row) > 4 and row[4]:
            days_between = _DAYS_LABEL_MAP.get(str(row[4]).strip(), 1)

        patterns[name] = DeliveryPattern(
            name=name,
            cutoff1=cutoff1,
            days_before_cutoff1=days_before,
            cutoff2=cutoff2,
            days_between_cutoffs=days_between,
        )

    return patterns


# ============================================
# VBA: BuildConfirmingCache (L864-901)
# 確認中テーブルから問合せ状況キャッシュ
# ============================================
def build_confirming_cache(
    confirming_ws: object,
) -> dict[str, tuple[str, str, Optional[datetime.date]]]:
    """確認中テーブルから問合せ状況キャッシュを構築する。

    確認中テーブルの構造 (テーブル名: 確認中テーブル):
    - 列1: 送付日時
    - 列2: 受注日
    - 列3: 顧客名
    - 列4: 受発注伝票
    - 列5: 明細
    - 列6: メーカー名
    - 列7: 品名
    - 列8: 問合せ状況
    - 列9: ステータス
    - 列10: 受注納期
    - 列11: 送付者

    キー: "注番|明細"
    値: (問合せ状況, ステータス, 受注納期date)

    Returns:
        {注番|明細: (問合せ状況, ステータス, 受注納期)}
    """
    cache: dict[str, tuple[str, str, Optional[datetime.date]]] = {}

    if confirming_ws is None:
        return cache

    # openpyxlではiter_rowsで2行目以降を読む（1行目はヘッダー）
    for row in confirming_ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 10:
            continue

        # 列4: 受発注伝票、列5: 明細（0-indexed: 3, 4）
        order_num = str(row[3]).strip() if row[3] else ""
        detail_num = str(row[4]).strip() if row[4] else ""

        key = f"{order_num}|{detail_num}"
        if key == "|" or key in cache:
            continue

        # 列8: 問合せ状況、列9: ステータス、列10: 受注納期（0-indexed: 7, 8, 9）
        inquiry_status = str(row[7]).strip() if row[7] else ""
        status = str(row[8]).strip() if row[8] else ""
        delivery_date = parse_date(row[9]) if len(row) > 9 else None

        cache[key] = (inquiry_status, status, delivery_date)

    return cache


# ============================================
# VBA: BuildStorageCache (L906-928)
# 受注一覧から注番→保管場所キャッシュ
# ============================================
def build_storage_cache(
    source_data: list[list[str]],
    cols: ColumnMap,
) -> dict[str, str]:
    """受注一覧から保管場所キャッシュを構築する。

    キー: 注番
    値: 保管場所

    Returns:
        {注番: 保管場所}
    """
    cache: dict[str, str] = {}

    order_col = cols.get("受発注伝票")
    storage_col = cols.get("保管場所")
    if order_col is None or storage_col is None:
        return cache

    # データ行は7行目（0-indexed: 6）から
    for row in source_data[6:]:
        if order_col >= len(row) or storage_col >= len(row):
            continue

        order_num = str(row[order_col]).strip()
        if not order_num or order_num in cache:
            continue

        storage_place = str(row[storage_col]).strip()
        if storage_place:
            cache[order_num] = storage_place

    return cache


# ============================================
# 全キャッシュ一括構築
# ============================================
def build_all_caches(
    manufacturer_master_wb: object | None,
    customer_master_wb: object | None,
    confirming_ws: object | None,
    source_data: list[list[str]],
    cols: ColumnMap,
) -> CacheStore:
    """全キャッシュを一括構築する。

    各build_xxx関数を呼び出してCacheStoreに集約する。

    Args:
        manufacturer_master_wb: メーカー一覧.xlsxのWorkbook
        customer_master_wb: 顧客マスター.xlsmのWorkbook
        confirming_ws: 送付履歴.xlsxの確認中一覧シート
        source_data: 受注一覧データ（load_source_fileの返り値）
        cols: 列位置マッピング（get_column_positionsの返り値）

    Returns:
        CacheStore
    """
    store = CacheStore()

    if manufacturer_master_wb is not None:
        store.mfg_name, store.mfg_days = build_manufacturer_cache(
            manufacturer_master_wb
        )
        store.delivery_patterns = build_pattern_cache(manufacturer_master_wb)

    if customer_master_wb is not None:
        (
            store.cust_days,
            store.cust_retention,
            store.cust_route,
            store.cust_pattern,
        ) = build_customer_cache(customer_master_wb)
        # フォーマット検出: cust_patternが1件でもあれば新フォーマット
        if store.cust_pattern:
            store.cust_email_start_col = 5
        else:
            store.cust_email_start_col = 4

    if confirming_ws is not None:
        store.confirm = build_confirming_cache(confirming_ws)

    store.storage = build_storage_cache(source_data, cols)

    return store
