"""VBA回答書とPython回答書の比較検証

17PMの受注一覧からPythonで納期計算を実行し、
VBAで生成済みの回答書と比較して不一致を抽出する。

修正版: 送付履歴を考慮した完全な処理フローで比較
"""

import sys
import io
import datetime
import re
from pathlib import Path
from typing import Optional

# パッケージの親ディレクトリをsys.pathに追加
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from openpyxl import Workbook, load_workbook

from nouki_kaitou.cache import build_all_caches
from nouki_kaitou.config import load_branch_settings, load_holidays
from nouki_kaitou.data_loader import (
    get_column_positions,
    get_data_rows_range,
    is_data_row,
    load_source_file,
    parse_order_row,
)
from nouki_kaitou.history import load_delivery_history
from nouki_kaitou.models import BranchSettings, OrderRow
from nouki_kaitou.report_generator import build_report_row, _determine_flags, _is_excluded, _pass_basic_filter
from nouki_kaitou.confirming import get_confirming_status


def normalize_delivery_answer(answer: str) -> tuple[str, str, str]:
    """納期回答を正規化して(日付部分, 配達/出荷, 予定/済み)に分解"""
    if not answer:
        return ("", "", "")

    answer = answer.strip()

    # 特殊ケース
    if answer in ("確認中", "欠品中", "日程調整中", "納品済み", "分納完了"):
        return (answer, "", "")

    # @@着日指定パターン
    match = re.match(r"(\d+/\d+)(出荷)(済み|済)?→(\d+/\d+)(着)(予定)?", answer)
    if match:
        ship_date = match.group(1)
        arrival_date = match.group(4)
        tense = "済み" if match.group(3) else "予定"
        return (f"{ship_date}→{arrival_date}着", "出荷", tense)

    # 通常パターン
    match = re.match(r"(.+?)(配達|出荷)(予定|済み|済)", answer)
    if match:
        date_part = match.group(1)
        delivery_type = match.group(2)
        tense = match.group(3)
        if tense == "済":
            tense = "済み"
        return (date_part, delivery_type, tense)

    # 引取パターン
    match = re.match(r"(.+?)(引取)(予定|済み|済)", answer)
    if match:
        return (match.group(1), "引取", match.group(3))

    # 作業パターン
    match = re.match(r"(.+?)(作業)(予定|済み|済)", answer)
    if match:
        return (match.group(1), "作業", match.group(3))

    return (answer, "", "")


def compare_answers(vba_answer: str, py_answer: str) -> tuple[bool, str]:
    """2つの納期回答を比較"""
    vba_norm = normalize_delivery_answer(vba_answer)
    py_norm = normalize_delivery_answer(py_answer)

    # 日付部分が異なる
    if vba_norm[0] != py_norm[0]:
        return (False, f"日付: VBA={vba_norm[0]} vs PY={py_norm[0]}")

    # 配達/出荷の区別が異なる
    if vba_norm[1] != py_norm[1]:
        return (False, f"区分: VBA={vba_norm[1]} vs PY={py_norm[1]}")

    return (True, "")


def normalize_price(value) -> str:
    """価格を正規化"""
    if value is None:
        return ""
    s = str(value).strip()
    # カンマを除去
    s = s.replace(",", "")
    if s in ("確認中", ""):
        return s
    try:
        return str(int(float(s)))
    except ValueError:
        return s


def normalize_quantity(value) -> str:
    """数量を正規化"""
    if value is None:
        return ""
    s = str(value).strip()
    # カンマを除去
    s = s.replace(",", "")
    try:
        # 整数化（1.0 -> 1）
        return str(int(float(s)))
    except ValueError:
        return s


def normalize_product_name(name: str) -> str:
    """品名を正規化（比較用）"""
    if not name:
        return ""
    # 全角スペースを半角に
    s = name.replace("\u3000", " ").strip()
    # 先頭25文字（VBA側で切られている場合がある）
    return s[:30]


def parse_date_cell(val) -> Optional[datetime.date]:
    """セルの日付値をパース"""
    if val is None:
        return None
    if isinstance(val, datetime.datetime):
        return val.date()
    if isinstance(val, datetime.date):
        return val
    s = str(val).strip()
    for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S"]:
        try:
            return datetime.datetime.strptime(s.split()[0], fmt.split()[0]).date()
        except ValueError:
            continue
    return None


def read_vba_excel(file_path: Path, customer_name: str) -> list[dict]:
    """VBA生成のExcelファイルを読み込み"""
    result = []

    try:
        wb = load_workbook(str(file_path), data_only=True)
    except Exception as e:
        print(f"  読み込みエラー: {file_path.name} - {e}")
        return result

    ws = wb.active

    # ヘッダー行を探す（A列が「受注日」の行）
    header_row = None
    for row_idx in range(1, min(20, ws.max_row + 1)):
        cell_val = ws.cell(row=row_idx, column=1).value
        if cell_val and str(cell_val).strip() == "受注日":
            header_row = row_idx
            break

    if header_row is None:
        print(f"  ヘッダー行が見つかりません: {file_path.name}")
        wb.close()
        return result

    # 列インデックス（A=1から）
    # 受注日(1), 担当者様(2), 貴社注番(3), メーカー名(4), 品名(5), 数量(6), 単価(7), 金額(8), 納期回答(9), 納入先名(10), 備考(11)
    col_order_date = 1
    col_manufacturer = 4
    col_product = 5
    col_quantity = 6
    col_unit_price = 7
    col_amount = 8
    col_delivery = 9

    # データ行を読み込み
    for row_idx in range(header_row + 1, ws.max_row + 1):
        order_date_val = ws.cell(row=row_idx, column=col_order_date).value
        if not order_date_val:
            continue

        # フッター行の判定（単価列に「※表示金額」等がある）
        unit_price_val = ws.cell(row=row_idx, column=col_unit_price).value
        if unit_price_val and "※" in str(unit_price_val):
            break

        order_date = parse_date_cell(order_date_val)
        manufacturer = str(ws.cell(row=row_idx, column=col_manufacturer).value or "").strip()
        product = str(ws.cell(row=row_idx, column=col_product).value or "").strip()
        quantity = str(ws.cell(row=row_idx, column=col_quantity).value or "").strip()
        delivery = str(ws.cell(row=row_idx, column=col_delivery).value or "").strip()
        unit_price = normalize_price(ws.cell(row=row_idx, column=col_unit_price).value)
        amount = normalize_price(ws.cell(row=row_idx, column=col_amount).value)

        result.append({
            "顧客": customer_name,
            "受注日": order_date,
            "メーカー": manufacturer,
            "品名": product,
            "品名_正規化": normalize_product_name(product),
            "数量": normalize_quantity(quantity),
            "納期回答": delivery,
            "単価": unit_price,
            "金額": amount,
            "ファイル": file_path.name,
        })

    wb.close()
    return result


def normalize_customer_name(name: str) -> str:
    """顧客名を正規化（比較用）"""
    if not name:
        return ""
    # 全角スペースを半角に
    s = name.replace("\u3000", " ")
    # 連続スペースを1つに
    s = re.sub(r"\s+", " ", s)
    # 前後のスペースを除去
    s = s.strip()
    # 株式会社の表記統一
    s = s.replace("（株）", "(株)")
    s = s.replace("（有）", "(有)")
    return s


def extract_customer_name_from_filename(filename: str) -> str:
    """ファイル名から顧客名を抽出"""
    # 例: "納期回答書_コイケ酸商（株）　白井営業所様_20260217.xlsx"
    name = filename.replace("納期回答書_", "").replace(".xlsx", "")
    # 日付部分を除去
    name = re.sub(r"_\d{8}$", "", name)
    # 末尾の「様」を除去
    if name.endswith("様"):
        name = name[:-1]
    return normalize_customer_name(name)


def load_history_raw(ws_history) -> dict[str, str]:
    """送付履歴を直接読み込んで注番|明細→納期回答のマップを作成"""
    history_map = {}
    # 列: 送付日時(0), 受注日(1), 顧客名(2), 受発注伝票(3), 明細(4), メーカー名(5), 品名(6), 納期回答(7)
    for row in ws_history.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 8:
            continue
        order_number = str(row[3] or "").strip()
        detail_number = str(row[4] or "").strip()
        delivery_answer = str(row[7] or "").strip()

        if not order_number:
            continue

        key = f"{order_number}|{detail_number}"
        # 最新の送付履歴を保持（後の行が新しい）
        history_map[key] = delivery_answer

    return history_map


def main():
    print("=" * 80)
    print("VBA vs Python 納期回答書比較検証（送付履歴考慮版）")
    print("=" * 80)

    # パス設定
    source_path = Path(r"\\flsv04\316京葉\納期回答書ツールフォルダ\受注一覧\17PM.xls")
    vba_folder = Path(r"\\flsv04\316京葉\納期回答書ツールフォルダ\納期回答書\2月17日(火)_②回目")
    tool_folder = Path(r"\\flsv04\316京葉\納期回答書ツールフォルダ")

    print(f"\nソースファイル: {source_path}")
    print(f"VBA回答書フォルダ: {vba_folder}")

    # ソースファイル読み込み
    print("\n17PM.xlsを読み込み中...")
    source_data_raw = load_source_file(str(source_path))
    cols = get_column_positions(source_data_raw)

    if cols is None:
        print("エラー: ヘッダー行の列位置を検出できませんでした。")
        return

    # OrderRow変換
    orders: list[OrderRow] = []
    for i in get_data_rows_range(source_data_raw, cols):
        if is_data_row(source_data_raw, i, cols):
            orders.append(parse_order_row(source_data_raw, i, cols))

    print(f"受注データ: {len(orders)}件")

    # マスターファイル読み込み
    print("\nマスターファイルを読み込み中...")
    mfg_path = tool_folder / "メーカー一覧.xlsx"
    cust_path = tool_folder / "顧客マスター_v2.xlsm"
    history_path = tool_folder / "送付履歴.xlsx"

    mfg_wb = load_workbook(str(mfg_path), data_only=True)
    cust_wb = load_workbook(str(cust_path), data_only=True)

    try:
        confirming_ws = cust_wb["確認中一覧"]
    except KeyError:
        confirming_ws = None

    cache = build_all_caches(mfg_wb, cust_wb, confirming_ws, source_data_raw, cols)
    branch = load_branch_settings(mfg_wb, source_data_raw, cols)
    holidays = load_holidays(mfg_wb)

    # 送付履歴読み込み（2種類）
    print("\n送付履歴を読み込み中...")
    history_wb = load_workbook(str(history_path), data_only=True, read_only=True)
    ws_history = history_wb["送付履歴"]
    ws_confirming = history_wb["確認中一覧"]

    today = datetime.date.today()

    # 1. load_delivery_history: スキップ判定用（確認中以外の確定伝票）
    sent_orders = load_delivery_history(ws_history, ws_confirming, cache, holidays, today)
    print(f"  スキップ判定用: {len(sent_orders)}件")

    # 2. load_history_raw: 全ての送付履歴（納期回答取得用）
    history_raw = load_history_raw(ws_history)
    print(f"  全送付履歴: {len(history_raw)}件")

    history_wb.close()

    # VBA回答書の読み込み
    print("\nVBA回答書を読み込み中...")
    vba_data: list[dict] = []

    vba_files = list(vba_folder.glob("*.xlsx"))
    print(f"VBA回答書ファイル: {len(vba_files)}件")

    for vba_file in vba_files:
        if vba_file.name.startswith("~$"):
            continue
        customer_name = extract_customer_name_from_filename(vba_file.name)
        file_data = read_vba_excel(vba_file, customer_name)
        vba_data.extend(file_data)

    print(f"VBA回答書の伝票数: {len(vba_data)}件")

    # Python側で納期計算（完全な処理フロー）
    print("\nPython側で納期計算中（送付履歴考慮）...")
    execution_time = datetime.datetime.now()

    # Pythonの計算結果を (顧客, 品名_正規化, 数量, 受注日) でインデックス
    py_data: dict[tuple, dict] = {}

    for row in orders:
        # 基本フィルタ（拒否理由・伝票タイプ・##除外）
        if not _pass_basic_filter(row):
            continue

        history_key = f"{row.order_number}|{row.detail_number}"
        previous_status = sent_orders.get(history_key, "")

        # 除外判定
        is_excluded_flag = _is_excluded(row, previous_status, cache)

        # スキップ判定（期間指定モードでスキップされるか）
        is_already_sent = False
        if is_excluded_flag:
            is_already_sent = True
        elif previous_status == "分納完了":
            is_already_sent = True
        elif row.ship_status == "処理完了" and previous_status == "確認中":
            is_already_sent = False  # 処理完了で確認中→今回出す
        elif row.ship_status == "処理完了" and previous_status:
            is_already_sent = True
        elif previous_status and row.registration_date and row.registration_date < today:
            is_already_sent = True

        # force_delivered判定
        force_delivered, is_himozuki, is_bunno_completed = _determine_flags(
            row, cache, sent_orders, history_key, previous_status
        )

        # 納期計算
        report_row, delivery_status = build_report_row(
            row, cache, holidays, branch, execution_time,
            force_delivered, today
        )

        # forceDelivered時の納品済み上書き
        if force_delivered and delivery_status in ("確認中", "欠品中", "日程調整中"):
            delivery_status = "納品済み"
        elif force_delivered and "（欠品）" in delivery_status:
            delivery_status = "納品済み"

        # 分納完了
        if is_bunno_completed:
            delivery_status = "納品済み"

        # スキップ対象の場合、送付履歴のステータスを使用
        unit_price = normalize_price(report_row.unit_price)
        net_amount = normalize_price(report_row.net_amount)

        if is_already_sent:
            # 送付履歴から取得（history_rawには全て入っている）
            hist_answer = history_raw.get(history_key, "")
            if hist_answer:
                delivery_status = hist_answer
                # 送付履歴で確定している場合、単価・金額も実際の値を使用
                if hist_answer not in ("確認中", "欠品中", "日程調整中"):
                    unit_price = normalize_price(row.unit_price)
                    net_amount = normalize_price(row.net_amount)

        key = (
            normalize_customer_name(row.customer_name),
            normalize_product_name(row.product_name),
            normalize_quantity(row.quantity),
            row.registration_date,
        )

        py_data[key] = {
            "注番|明細": f"{row.order_number}|{row.detail_number}",
            "納期回答": delivery_status,
            "単価": unit_price,
            "金額": net_amount,
            "品名": row.product_name,
            "is_already_sent": is_already_sent,
            "force_delivered": force_delivered,
        }

    print(f"Python計算の伝票数: {len(py_data)}件")

    # 比較
    print("\n比較実行中...")
    differences = []
    matched = 0
    not_found_in_python = 0
    skipped_both = 0  # 両方でスキップされた（Python側がis_already_sent=True）

    for vba_item in vba_data:
        key = (
            normalize_customer_name(vba_item["顧客"]),
            vba_item["品名_正規化"],
            vba_item["数量"],
            vba_item["受注日"],
        )

        if key not in py_data:
            not_found_in_python += 1
            # デバッグ用: 最初の数件だけ表示
            if not_found_in_python <= 5:
                print(f"  Python側なし: {vba_item['顧客']} / {vba_item['品名'][:20]} / {vba_item['数量']} / {vba_item['受注日']}")
            continue

        py = py_data[key]
        matched += 1

        # 納期回答の比較
        is_match, reason = compare_answers(vba_item["納期回答"], py["納期回答"])
        if not is_match:
            differences.append({
                "注番|明細": py["注番|明細"],
                "顧客": vba_item["顧客"],
                "品名": vba_item["品名"][:30],
                "項目": "納期回答",
                "VBA": vba_item["納期回答"],
                "Python": py["納期回答"],
                "詳細": reason,
                "送付済み": "○" if py["is_already_sent"] else "",
                "force_delivered": "○" if py["force_delivered"] else "",
            })

        # 単価の比較
        if vba_item["単価"] != py["単価"]:
            differences.append({
                "注番|明細": py["注番|明細"],
                "顧客": vba_item["顧客"],
                "品名": vba_item["品名"][:30],
                "項目": "単価",
                "VBA": vba_item["単価"],
                "Python": py["単価"],
                "詳細": "",
                "送付済み": "○" if py["is_already_sent"] else "",
                "force_delivered": "○" if py["force_delivered"] else "",
            })

        # 金額の比較
        if vba_item["金額"] != py["金額"]:
            differences.append({
                "注番|明細": py["注番|明細"],
                "顧客": vba_item["顧客"],
                "品名": vba_item["品名"][:30],
                "項目": "金額",
                "VBA": vba_item["金額"],
                "Python": py["金額"],
                "詳細": "",
                "送付済み": "○" if py["is_already_sent"] else "",
                "force_delivered": "○" if py["force_delivered"] else "",
            })

    print(f"\nマッチした伝票: {matched}件")
    print(f"Python側に見つからなかった伝票: {not_found_in_python}件")
    print(f"不一致件数: {len(differences)}件")

    # 結果をExcelに出力
    output_path = Path(__file__).parent / "comparison_result.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "比較結果"

    # ヘッダー
    headers = ["注番|明細", "顧客", "品名", "項目", "VBA", "Python", "詳細", "送付済み", "force_delivered"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # データ
    for row_idx, diff in enumerate(differences, 2):
        ws.cell(row=row_idx, column=1, value=diff["注番|明細"])
        ws.cell(row=row_idx, column=2, value=diff["顧客"])
        ws.cell(row=row_idx, column=3, value=diff["品名"])
        ws.cell(row=row_idx, column=4, value=diff["項目"])
        ws.cell(row=row_idx, column=5, value=diff["VBA"])
        ws.cell(row=row_idx, column=6, value=diff["Python"])
        ws.cell(row=row_idx, column=7, value=diff["詳細"])
        ws.cell(row=row_idx, column=8, value=diff.get("送付済み", ""))
        ws.cell(row=row_idx, column=9, value=diff.get("force_delivered", ""))

    # サマリーシート
    ws_summary = wb.create_sheet("サマリー")
    ws_summary.cell(row=1, column=1, value="項目")
    ws_summary.cell(row=1, column=2, value="件数")
    ws_summary.cell(row=2, column=1, value="VBA回答書の伝票数")
    ws_summary.cell(row=2, column=2, value=len(vba_data))
    ws_summary.cell(row=3, column=1, value="マッチした伝票数")
    ws_summary.cell(row=3, column=2, value=matched)
    ws_summary.cell(row=4, column=1, value="Python側に見つからなかった伝票")
    ws_summary.cell(row=4, column=2, value=not_found_in_python)
    ws_summary.cell(row=5, column=1, value="不一致件数")
    ws_summary.cell(row=5, column=2, value=len(differences))

    # 不一致の内訳
    item_counts = {}
    for diff in differences:
        item = diff["項目"]
        item_counts[item] = item_counts.get(item, 0) + 1

    row_idx = 7
    ws_summary.cell(row=row_idx, column=1, value="不一致内訳")
    row_idx += 1
    for item, count in sorted(item_counts.items()):
        ws_summary.cell(row=row_idx, column=1, value=f"  {item}")
        ws_summary.cell(row=row_idx, column=2, value=count)
        row_idx += 1

    # 送付済み伝票の不一致
    sent_diff_count = sum(1 for d in differences if d.get("送付済み") == "○")
    row_idx += 1
    ws_summary.cell(row=row_idx, column=1, value="送付済み伝票の不一致")
    ws_summary.cell(row=row_idx, column=2, value=sent_diff_count)

    wb.save(str(output_path))
    print(f"\n結果を保存しました: {output_path}")

    # 不一致の先頭10件を表示
    if differences:
        print("\n--- 不一致の先頭10件 ---")
        for diff in differences[:10]:
            print(f"  {diff['注番|明細']}: {diff['項目']}")
            print(f"    顧客: {diff['顧客']}")
            print(f"    品名: {diff['品名']}")
            print(f"    VBA: {diff['VBA']}")
            print(f"    Python: {diff['Python']}")
            if diff['詳細']:
                print(f"    詳細: {diff['詳細']}")
            if diff.get('送付済み'):
                print(f"    送付済み: {diff['送付済み']}")
            print()


if __name__ == "__main__":
    main()
