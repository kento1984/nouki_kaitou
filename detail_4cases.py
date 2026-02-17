"""4件の新規不一致の完全な詳細を出力"""

import sys
import io
import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from openpyxl import load_workbook

from nouki_kaitou.cache import build_all_caches
from nouki_kaitou.config import load_branch_settings, load_holidays
from nouki_kaitou.data_loader import (
    get_column_positions,
    get_data_rows_range,
    is_data_row,
    load_source_file,
    parse_order_row,
)
from nouki_kaitou.delivery_calc import calculate_delivery_date
from nouki_kaitou.customer import is_route_delivery


def main():
    tool_folder = Path(r"\\flsv04\316京葉\納期回答書ツールフォルダ")
    source_path = tool_folder / "受注一覧" / "17PM.xls"

    # 対象注番とVBA回答
    targets = [
        ("GL2C447202", "10", "2月17日出荷済み"),
        ("GL2Z446911", "20", "2月18日出荷予定"),
        ("GL2S447221", "10", "2月20日配達予定"),
        ("GL2V447251", "10", "2月17日出荷済み"),
    ]

    # ソースファイル読み込み
    source_data_raw = load_source_file(str(source_path))
    cols = get_column_positions(source_data_raw)

    # マスターファイル読み込み
    mfg_wb = load_workbook(str(tool_folder / "メーカー一覧.xlsx"), data_only=True)
    cust_wb = load_workbook(str(tool_folder / "顧客マスター_v2.xlsm"), data_only=True)

    try:
        confirming_ws = cust_wb["確認中一覧"]
    except KeyError:
        confirming_ws = None

    cache = build_all_caches(mfg_wb, cust_wb, confirming_ws, source_data_raw, cols)
    branch = load_branch_settings(mfg_wb, source_data_raw, cols)
    holidays = load_holidays(mfg_wb)

    today = datetime.date.today()
    execution_time = datetime.datetime.now()

    # 対象注番を検索してトレース
    for order_no, detail_no, vba_answer in targets:
        for i in get_data_rows_range(source_data_raw, cols):
            if is_data_row(source_data_raw, i, cols):
                row = parse_order_row(source_data_raw, i, cols)
                if row.order_number == order_no and row.detail_number == detail_no:
                    # Python計算
                    py_answer = calculate_delivery_date(
                        row, cache, holidays, branch, execution_time, today
                    )

                    # 差異の説明を生成
                    explanation = generate_explanation(row, vba_answer, py_answer, branch, cache)

                    print("=" * 100)
                    print(f"【{order_no}|{detail_no}】")
                    print("=" * 100)
                    print(f"顧客名:       {row.customer_name}")
                    print(f"出荷先名:     {row.ship_to_name}")
                    print(f"品名:         {row.product_name}")
                    print(f"数量:         {row.quantity}")
                    print(f"伝票タイプ:   {row.document_type}")
                    print(f"保管場所:     {row.storage_place}")
                    print(f"出荷ステータス: {row.ship_status}")
                    print(f"指定納期:     {row.specified_delivery_date}")
                    print(f"受注納期:     {row.order_delivery_date}")
                    print(f"コメント社内: {row.comment_internal if row.comment_internal else '(なし)'}")
                    print(f"登録日時:     {row.registration_date} {row.time_value}")
                    print()
                    print(f"VBAの回答:    {vba_answer}")
                    print(f"Pythonの回答: {py_answer}")
                    print()
                    print(f"【差異の説明】")
                    print(explanation)
                    print()
                    break


def generate_explanation(row, vba_answer, py_answer, branch, cache):
    """差異の説明を生成"""
    lines = []

    # 受注先と出荷先の比較
    customer = row.customer_name.strip()
    ship_to = row.ship_to_name.strip()
    is_same = (customer == ship_to)

    # 紐付き判定
    is_himozuki = (
        row.document_type == "【受注】直送販売"
        and row.storage_place != "転送中（直送用）"
    )

    # 他拠点判定
    is_other_center = (
        row.storage_place != ""
        and row.storage_place != branch.base_center
    )

    # 路線便判定
    is_rosenbin = is_route_delivery(row.customer_name, cache)

    # 転送中判定
    is_tensou = (row.storage_place == "転送中（直送用）")

    lines.append(f"条件:")
    lines.append(f"  - 受注先=出荷先: {is_same}")
    lines.append(f"  - 紐付き(直送+非転送中): {is_himozuki}")
    lines.append(f"  - 他拠点: {is_other_center} (base_center={branch.base_center})")
    lines.append(f"  - 路線便: {is_rosenbin}")
    lines.append(f"  - 転送中: {is_tensou}")
    lines.append("")

    # 差異の原因分析
    vba_has_ship = "出荷" in vba_answer
    py_has_ship = "出荷" in py_answer
    vba_has_deliver = "配達" in vba_answer
    py_has_deliver = "配達" in py_answer

    # 配達/出荷の区別
    if vba_has_ship and py_has_deliver:
        lines.append("【問題1: 配達/出荷の区別】")
        lines.append(f"  VBAは「出荷」、Pythonは「配達」と判定。")
        if is_same and not is_rosenbin:
            lines.append(f"  → 受注先=出荷先 かつ 路線便でない → 自社便配達")
            lines.append(f"  → Pythonの「配達」が正しい。VBAが誤り。")
        lines.append("")

    if vba_has_deliver and py_has_ship:
        lines.append("【問題1: 配達/出荷の区別】")
        lines.append(f"  VBAは「配達」、Pythonは「出荷」と判定。")
        if is_himozuki and not is_same:
            lines.append(f"  → 紐付き + 受注先≠出荷先 → 出荷予定")
            lines.append(f"  → Pythonの「出荷」が正しい。VBAが誤り。")
        elif is_tensou:
            lines.append(f"  → 転送中（直送用）→ 出荷予定")
            lines.append(f"  → Pythonの「出荷」が正しい。VBAが誤り。")
        lines.append("")

    # 日付の違い
    import re
    vba_date = re.search(r"(\d+)月(\d+)日", vba_answer)
    py_date = re.search(r"(\d+)月(\d+)日", py_answer)

    if vba_date and py_date:
        vba_day = int(vba_date.group(2))
        py_day = int(py_date.group(2))
        if vba_day != py_day:
            lines.append("【問題2: 日付のずれ】")
            lines.append(f"  VBA: {vba_date.group(0)}, Python: {py_date.group(0)}")

            # 指定納期がある場合
            if row.specified_delivery_date and str(row.specified_delivery_date) != "2026-12-31":
                lines.append(f"  → 指定納期 {row.specified_delivery_date} があるため、それを基準に計算")

            # 処理完了の場合
            if row.ship_status == "処理完了":
                lines.append(f"  → 処理完了 + 登録日時 {row.registration_date} {row.time_value}")
                lines.append(f"  → カットオフ時間との比較で日付が決まる")

            lines.append("")

    # 他拠点表記
    if "他拠点" in py_answer and "他拠点" not in vba_answer:
        lines.append("【問題3: 他拠点表記の有無】")
        lines.append(f"  Pythonは「他拠点より」を付与、VBAは付与していない。")
        lines.append(f"  → 保管場所={row.storage_place} ≠ base_center={branch.base_center}")
        lines.append(f"  → 意味は同じ。表記の違いのみ（許容範囲）。")
        lines.append("")

    return "\n".join(lines)


if __name__ == "__main__":
    main()
