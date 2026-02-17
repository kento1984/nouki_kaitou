"""配達/出荷の使い分けWARNING 17件の保管場所確認"""

import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import load_workbook
from nouki_kaitou.data_loader import load_source_file, get_column_positions

def main():
    # validation_result.xlsxから配達/出荷の使い分けの注番を取得
    validation_path = Path(__file__).parent / "validation_result.xlsx"

    print("validation_result.xlsxを読み込み中...")
    wb = load_workbook(validation_path, read_only=True, data_only=True)

    # シート一覧を表示
    print(f"シート一覧: {wb.sheetnames}")

    # 配達_出荷の使い分けシートを探す
    target_sheet = None
    for sheet_name in wb.sheetnames:
        if "配達" in sheet_name and "出荷" in sheet_name:
            target_sheet = sheet_name
            break

    if not target_sheet:
        print("シートが見つかりません")
        return

    print(f"対象シート: {target_sheet}")
    ws = wb[target_sheet]

    # ヘッダーを確認
    headers = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    print(f"ヘッダー: {headers[:10]}")

    # 最初の数行を表示
    print("\n最初の5行:")
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=6, values_only=True), 1):
        print(f"  {row_idx}: {row[:5]}")

    # 注番リストを取得（列: 重要度, カテゴリ, 注番, 明細, 受注先, ...）
    order_details = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if row and len(row) > 3 and row[2]:  # 注番は3列目（インデックス2）
            order_num = str(row[2]).strip()
            detail_num = str(row[3] or "").strip() if len(row) > 3 else ""
            problem = str(row[10] or "").strip() if len(row) > 10 else ""  # 問題列
            order_details.append((order_num, detail_num, problem, row))

    wb.close()
    print(f"\n対象件数: {len(order_details)}件")

    # 16AM.xlsから保管場所を取得
    print("\n16AM.xlsを読み込み中...")
    xls_path = Path(r"\\flsv04\316京葉\納期回答書ツールフォルダ\受注一覧\16AM.xls")
    source_data = load_source_file(xls_path)
    cols = get_column_positions(source_data)

    if cols is None:
        print("列位置を取得できませんでした")
        return

    col_order = cols.get("受発注伝票", 0)
    col_detail = cols.get("明細", 1)
    col_storage = cols.get("保管場所")
    col_customer = cols.get("受注先")
    col_ship_to = cols.get("出荷先名")
    col_product = cols.get("品名")

    # 注番→保管場所のマッピング
    storage_map = {}
    for row_idx in range(5, len(source_data)):
        row = source_data[row_idx]
        if len(row) <= col_order:
            continue
        order_num = str(row[col_order] or "").strip()
        if not order_num:
            continue
        detail_num = str(row[col_detail] if col_detail < len(row) else "").strip()
        storage = str(row[col_storage] if col_storage and col_storage < len(row) else "").strip()
        customer = str(row[col_customer] if col_customer and col_customer < len(row) else "").strip()
        ship_to = str(row[col_ship_to] if col_ship_to and col_ship_to < len(row) else "").strip()
        product = str(row[col_product] if col_product and col_product < len(row) else "").strip()

        key = f"{order_num}|{detail_num}"
        storage_map[key] = {
            "storage": storage,
            "customer": customer[:25],
            "ship_to": ship_to[:25],
            "product": product[:35],
        }

    print(f"読み込み完了: {len(storage_map)}件\n")

    # 結果表示
    print("=" * 110)
    print("配達/出荷の使い分け WARNING 17件の保管場所確認")
    print("=" * 110)

    other_branch_count = 0
    same_branch_count = 0
    not_found_count = 0
    same_branch_items = []

    for order_num, detail_num, problem, raw_row in order_details:
        key = f"{order_num}|{detail_num}"

        if key in storage_map:
            info = storage_map[key]
            storage = info["storage"]

            # 他拠点かどうか判定（京葉=16xx以外は他拠点）
            is_other_branch = False
            if storage:
                # 他拠点明示的な表記
                if "他拠点" in storage:
                    is_other_branch = True
                # 16xx以外の拠点コード
                elif not storage.startswith("16") and storage not in ("", "京葉"):
                    is_other_branch = True

            marker = "○他拠点" if is_other_branch else "×京葉拠点"

            if is_other_branch:
                other_branch_count += 1
            else:
                same_branch_count += 1
                same_branch_items.append((order_num, detail_num, storage, info))

            print(f"\n{order_num}|{detail_num}")
            print(f"  保管場所: [{storage}] → {marker}")
            print(f"  受注先: {info['customer']}")
            print(f"  出荷先: {info['ship_to']}")
            print(f"  品名: {info['product']}")
        else:
            not_found_count += 1
            print(f"\n{order_num}|{detail_num} - 16AM.xlsにデータなし")

    # 集計
    print("\n" + "=" * 110)
    print("集計結果")
    print("=" * 110)
    print(f"他拠点からの出荷（「出荷予定」で正しい）: {other_branch_count}件")
    print(f"京葉拠点（要確認）: {same_branch_count}件")
    print(f"データなし: {not_found_count}件")

    if same_branch_items:
        print("\n【要確認】京葉拠点なのに「出荷予定」になっているケース:")
        for order_num, detail_num, storage, info in same_branch_items:
            print(f"  {order_num}|{detail_num}")
            print(f"    保管場所: {storage}")
            print(f"    受注先≠出荷先: {info['customer']} ≠ {info['ship_to']}")

if __name__ == "__main__":
    main()
