"""リアルタイム比較スクリプト

同じ瞬間のVBA回答書とPython計算を全件比較する。
除外なし、送付履歴なし。弊社注番（L列）でマッチング。
"""

import sys
import io
import datetime
import re
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from nouki_kaitou.cache import build_all_caches
from nouki_kaitou.config import load_branch_settings, load_holidays
from nouki_kaitou.data_loader import (
    get_column_positions, get_data_rows_range, is_data_row,
    load_source_file, parse_order_row,
)
from nouki_kaitou.models import OrderRow
from nouki_kaitou.report_generator import (
    build_report_row, _determine_flags, _pass_basic_filter,
)


# ============================================
# 設定（ここを変更して実行）
# ============================================
TOOL_FOLDER = Path(r"\\flsv04\316京葉\納期回答書ツールフォルダ")
LOCAL_FOLDER = Path(__file__).resolve().parent.parent  # nouki_kaitouルート

SOURCE_FILE = TOOL_FOLDER / "受注一覧" / "18PM.XLS"
VBA_FOLDERS = [
    TOOL_FOLDER / "納期回答書" / "2月18日(水)_②回目",
]
EXECUTION_TIME = datetime.datetime(2026, 2, 18, 17, 0)
TODAY = datetime.date(2026, 2, 18)
OUTPUT_NAME = "comparison_realtime_0218PM.xlsx"


# ============================================
# ユーティリティ
# ============================================
def normalize_delivery_answer(answer):
    if not answer:
        return ("", "", "")
    answer = answer.strip()
    if answer in ("確認中", "欠品中", "日程調整中", "納品済み", "分納完了"):
        return (answer, "", "")
    m = re.match(r"(\d+/\d+)(出荷)(済み|済)?→(\d+/\d+)(着)(予定)?", answer)
    if m:
        return (f"{m.group(1)}→{m.group(4)}着", "出荷", "済み" if m.group(3) else "予定")
    m = re.match(r"(.+?)(配達|出荷)(予定|済み|済)", answer)
    if m:
        return (m.group(1), m.group(2), "済み" if m.group(3) == "済" else m.group(3))
    m = re.match(r"(.+?)(引取|作業)(予定|済み|済)", answer)
    if m:
        return (m.group(1), m.group(2), m.group(3))
    return (answer, "", "")


def compare_answers(vba, py):
    v = normalize_delivery_answer(vba)
    p = normalize_delivery_answer(py)
    if v[0] != p[0]:
        return (False, f"日付: VBA={v[0]} vs PY={p[0]}")
    if v[1] != p[1]:
        return (False, f"区分: VBA={v[1]} vs PY={p[1]}")
    return (True, "")


def normalize_product_name(name):
    if not name:
        return ""
    return name.replace("\u3000", " ").strip()[:30]


def normalize_customer_name(name):
    if not name:
        return ""
    s = re.sub(r"\s+", " ", name.replace("\u3000", " ")).strip()
    return s.replace("（株）", "(株)").replace("（有）", "(有)")


def extract_customer_name_from_filename(filename):
    name = filename.replace("納期回答書_", "").replace(".xlsx", "")
    name = re.sub(r"_\d{8}$", "", name)
    if name.endswith("様"):
        name = name[:-1]
    return normalize_customer_name(name)


# ============================================
# VBA読み込み
# ============================================
def read_vba_excel(file_path, customer_name):
    result = []
    try:
        wb = load_workbook(str(file_path), data_only=True)
    except Exception as e:
        print(f"  読込エラー: {file_path.name} - {e}")
        return result
    ws = wb.active
    header_row = None
    for r in range(1, min(20, ws.max_row + 1)):
        if ws.cell(row=r, column=1).value and str(ws.cell(row=r, column=1).value).strip() == "受注日":
            header_row = r
            break
    if not header_row:
        wb.close()
        return result
    for r in range(header_row + 1, ws.max_row + 1):
        if not ws.cell(row=r, column=1).value:
            continue
        up = ws.cell(row=r, column=7).value
        if up and "※" in str(up):
            break
        result.append({
            "顧客": customer_name,
            "品名": str(ws.cell(row=r, column=5).value or "").strip(),
            "品名_正規化": normalize_product_name(str(ws.cell(row=r, column=5).value or "")),
            "納期回答": str(ws.cell(row=r, column=9).value or "").strip(),
            "弊社注番": str(ws.cell(row=r, column=12).value or "").strip(),
            "ファイル": file_path.name,
        })
    wb.close()
    return result


def get_type_label(row, cache):
    dt = row.document_type.strip()
    if dt == "【受注】直送販売":
        st = row.storage_place.strip() or cache.storage.get(row.order_number, "")
        return "直送" if st == "転送中（直送用）" else "紐付き"
    elif dt == "【受注】在庫販売":
        return "在庫販売"
    return dt


def main():
    print("=" * 70)
    print(f"リアルタイム比較: {TODAY.strftime('%m/%d')} {'AM' if EXECUTION_TIME.hour < 15 else 'PM'}")
    print(f"  受注データ: {SOURCE_FILE.name}")
    print(f"  VBAフォルダ: {[f.name for f in VBA_FOLDERS]}")
    print(f"  execution_time: {EXECUTION_TIME}")
    print("=" * 70)

    # 受注データ
    print("\n受注データ読込中...")
    raw = load_source_file(str(SOURCE_FILE))
    cols = get_column_positions(raw)
    orders = []
    for i in get_data_rows_range(raw, cols):
        if is_data_row(raw, i, cols):
            orders.append(parse_order_row(raw, i, cols))
    print(f"  受注データ: {len(orders)}件")

    # マスター（ローカルコピー）
    print("マスター読込中...")
    mfg_wb = load_workbook(str(LOCAL_FOLDER / "メーカー一覧.xlsx"), data_only=True)
    cust_wb = load_workbook(str(LOCAL_FOLDER / "顧客マスター_v2.xlsm"), data_only=True)
    try:
        confirming_ws = cust_wb["確認中一覧"]
    except KeyError:
        confirming_ws = None

    cache = build_all_caches(mfg_wb, cust_wb, confirming_ws, raw, cols)
    branch = load_branch_settings(mfg_wb, raw, cols)
    holidays = load_holidays(mfg_wb)

    # VBA回答書（複数フォルダ対応、同一顧客は後勝ち）
    print("VBA回答書読込中...")
    customer_data = {}
    for folder in VBA_FOLDERS:
        if not folder.exists():
            print(f"  警告: {folder.name} が見つかりません")
            continue
        for f in sorted(folder.glob("*.xlsx")):
            if f.name.startswith("~$"):
                continue
            cname = extract_customer_name_from_filename(f.name)
            rows = read_vba_excel(f, cname)
            if rows:
                customer_data[normalize_customer_name(cname)] = rows

    vba_data = []
    for rows in customer_data.values():
        vba_data.extend(rows)
    print(f"  VBA回答書: {len(vba_data)}件")

    # 注番インデックス
    order_index = {}
    for row in orders:
        k = row.order_number.strip()
        order_index.setdefault(k, []).append(row)

    # 全件比較
    print("\n比較中...")
    results = []
    used = set()
    not_found = 0
    empty_l = 0

    for vba in vba_data:
        onum = vba["弊社注番"]
        if not onum:
            empty_l += 1
            not_found += 1
            continue

        candidates = order_index.get(onum, [])
        if not candidates:
            not_found += 1
            results.append({
                "伝票番号": onum, "明細": "?", "顧客": vba["顧客"],
                "品名": vba["品名"][:30], "VBA": vba["納期回答"],
                "Python": "(受注データなし)", "一致": None, "詳細": "SAP未検出",
                "伝票タイプ": "", "出荷ステータス": "", "指定納期": "", "受注納期": "",
            })
            continue

        # 品名マッチ
        vp = vba["品名_正規化"]
        match = None
        for row in candidates:
            dk = f"{row.order_number}|{row.detail_number}"
            if dk in used:
                continue
            if normalize_product_name(row.product_name) == vp:
                match = row
                break
        if not match:
            for row in candidates:
                dk = f"{row.order_number}|{row.detail_number}"
                if dk in used:
                    continue
                if normalize_product_name(row.product_name)[:20] == vp[:20] and vp[:20]:
                    match = row
                    break
        if not match:
            for row in candidates:
                dk = f"{row.order_number}|{row.detail_number}"
                if dk in used:
                    continue
                match = row
                break

        if not match:
            not_found += 1
            continue

        dk = f"{match.order_number}|{match.detail_number}"
        used.add(dk)

        # Python計算
        if not _pass_basic_filter(match):
            py_answer = "(フィルタ除外)"
        else:
            fd, himo, bc = _determine_flags(match, cache, {}, dk, "")
            rr, py_answer = build_report_row(
                match, cache, holidays, branch, EXECUTION_TIME, fd, TODAY
            )
            if fd and py_answer in ("確認中", "欠品中", "日程調整中"):
                py_answer = "納品済み"
            elif fd and "（欠品）" in py_answer:
                py_answer = "納品済み"
            if bc:
                py_answer = "納品済み"

        # 比較
        is_match, reason = compare_answers(vba["納期回答"], py_answer)

        sd = match.specified_delivery_date.strftime("%m/%d") if match.specified_delivery_date else ""
        od = match.order_delivery_date.strftime("%m/%d") if match.order_delivery_date else ""

        results.append({
            "伝票番号": match.order_number, "明細": match.detail_number,
            "顧客": vba["顧客"], "品名": vba["品名"][:30],
            "VBA": vba["納期回答"], "Python": py_answer,
            "一致": is_match, "詳細": reason,
            "伝票タイプ": get_type_label(match, cache),
            "出荷ステータス": match.ship_status,
            "指定納期": sd, "受注納期": od,
        })

    # 結果表示
    total = len(results)
    match_count = sum(1 for r in results if r["一致"] is True)
    mismatches = [r for r in results if r["一致"] is False]
    unresolved = sum(1 for r in results if r["一致"] is None)

    print()
    print("=" * 70)
    compared = match_count + len(mismatches)
    print(f"結果: 全{total}件  一致={match_count}  不一致={len(mismatches)}  "
          f"未検出={not_found}(L列空{empty_l})")
    if compared > 0:
        rate = match_count / compared * 100
        print(f"一致率: {rate:.1f}% ({match_count}/{compared})")
    print("=" * 70)

    if mismatches:
        print(f"\n--- 不一致 {len(mismatches)}件 ---")
        for r in mismatches:
            print(f"  {r['伝票番号']}|{r['明細']}  [{r['伝票タイプ']}] {r['出荷ステータス']}")
            print(f"    顧客: {r['顧客']}")
            print(f"    品名: {r['品名']}")
            print(f"    VBA: {r['VBA']}")
            print(f"    Python: {r['Python']}")
            print(f"    {r['詳細']}")
            print(f"    指定納期={r['指定納期']}  受注納期={r['受注納期']}")
            print()
    else:
        print("\n不一致なし！")

    # Excel出力
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
    MISMATCH_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    wb = Workbook()
    ws = wb.active
    ws.title = "比較結果"
    headers = [
        "伝票番号", "明細", "顧客", "品名", "VBA回答", "Python回答",
        "結果", "詳細", "伝票タイプ", "出荷ステータス", "指定納期", "受注納期",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for ri, r in enumerate(results, 2):
        if r["一致"] is True:
            label = "一致"
        elif r["一致"] is False:
            label = "不一致"
        else:
            label = "未検出"
        ws.cell(row=ri, column=1, value=r["伝票番号"])
        ws.cell(row=ri, column=2, value=r["明細"])
        ws.cell(row=ri, column=3, value=r["顧客"])
        ws.cell(row=ri, column=4, value=r["品名"])
        ws.cell(row=ri, column=5, value=r["VBA"])
        ws.cell(row=ri, column=6, value=r["Python"])
        ws.cell(row=ri, column=7, value=label)
        ws.cell(row=ri, column=8, value=r["詳細"])
        ws.cell(row=ri, column=9, value=r["伝票タイプ"])
        ws.cell(row=ri, column=10, value=r["出荷ステータス"])
        ws.cell(row=ri, column=11, value=r["指定納期"])
        ws.cell(row=ri, column=12, value=r["受注納期"])
        if r["一致"] is False:
            for c in range(1, 13):
                ws.cell(row=ri, column=c).fill = MISMATCH_FILL

    widths = [14, 6, 28, 32, 22, 22, 8, 35, 10, 12, 10, 10]
    for i, w in enumerate(widths):
        ws.column_dimensions[chr(65 + i)].width = w

    out = Path(__file__).parent / OUTPUT_NAME
    wb.save(str(out))
    print(f"\nExcel出力: {out}")

    mfg_wb.close()
    cust_wb.close()


if __name__ == "__main__":
    main()
