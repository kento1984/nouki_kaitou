"""納期整合性INFO 38件 + マスターデータINFO 16件のパターン別集計"""

import sys
from pathlib import Path
import re
from collections import defaultdict

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import load_workbook

def analyze_sheet(wb, sheet_keyword, category_name):
    """指定シートを分析"""
    target_sheet = None
    for sheet_name in wb.sheetnames:
        if sheet_keyword in sheet_name:
            target_sheet = sheet_name
            break

    if not target_sheet:
        print(f"  シートが見つかりません: {sheet_keyword}")
        return []

    ws = wb[target_sheet]

    # データを収集
    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and len(row) > 3 and row[2]:
            order_num = str(row[2]).strip()
            detail_num = str(row[3] or "").strip() if len(row) > 3 else ""
            customer = str(row[4] or "").strip()[:25] if len(row) > 4 else ""
            ship_to = str(row[5] or "").strip()[:25] if len(row) > 5 else ""
            product = str(row[6] or "").strip()[:30] if len(row) > 6 else ""
            problem = str(row[12] or "").strip() if len(row) > 12 else ""

            # 問題列が空なら他の列を探す
            if not problem:
                for col_idx in range(10, min(15, len(row))):
                    val = str(row[col_idx] or "").strip()
                    if val and len(val) > 5:
                        problem = val
                        break

            items.append({
                "order": order_num,
                "detail": detail_num,
                "customer": customer,
                "ship_to": ship_to,
                "product": product,
                "problem": problem,
                "category": category_name,
            })

    return items

def main():
    validation_path = Path(__file__).parent / "validation_result.xlsx"

    print("validation_result.xlsxを読み込み中...")
    wb = load_workbook(validation_path, read_only=True, data_only=True)
    print(f"シート一覧: {wb.sheetnames}\n")

    all_items = []

    # 納期整合性
    print("【納期整合性】を読み込み中...")
    items1 = analyze_sheet(wb, "納期整合性", "納期整合性")
    print(f"  {len(items1)}件")
    all_items.extend(items1)

    # マスターデータ
    print("【マスターデータ】を読み込み中...")
    items2 = analyze_sheet(wb, "マスターデータ", "マスターデータ")
    print(f"  {len(items2)}件")
    all_items.extend(items2)

    wb.close()

    # カテゴリ別・パターン別に集計
    for category_name in ["納期整合性", "マスターデータ"]:
        category_items = [i for i in all_items if i["category"] == category_name]

        print(f"\n{'='*100}")
        print(f"【{category_name}】INFO {len(category_items)}件")
        print(f"{'='*100}")

        # パターン別集計
        pattern_counts = defaultdict(list)
        for item in category_items:
            problem = item["problem"]
            # 数値を正規化
            normalized = re.sub(r"\d+日", "N日", problem)
            normalized = re.sub(r"\d{4}-\d{2}-\d{2}", "YYYY-MM-DD", normalized)
            normalized = re.sub(r"\d+件", "N件", normalized)
            pattern_counts[normalized].append(item)

        # 件数順にソート
        sorted_patterns = sorted(pattern_counts.items(), key=lambda x: -len(x[1]))

        for pattern, items in sorted_patterns:
            print(f"\n■ {pattern} ({len(items)}件)")
            print("-" * 80)

            # 代表例を3件まで表示
            for i, item in enumerate(items[:3]):
                print(f"  例{i+1}: {item['order']}|{item['detail']}")
                print(f"       顧客: {item['customer']}")
                if item['ship_to'] and item['ship_to'] != item['customer']:
                    print(f"       出荷先: {item['ship_to']}")
                print(f"       品名: {item['product']}")
                if item['problem'] != pattern and item['problem']:
                    actual = item['problem'][:60]
                    print(f"       問題: {actual}")

            if len(items) > 3:
                print(f"  ... 他 {len(items) - 3}件")

        # パターンサマリー
        print(f"\n--- {category_name} パターン別件数 ---")
        for pattern, items in sorted_patterns:
            short_pattern = pattern[:65] + "..." if len(pattern) > 65 else pattern
            print(f"  {len(items):3d}件: {short_pattern}")

    # 全体サマリー
    print(f"\n{'='*100}")
    print("全体サマリー")
    print(f"{'='*100}")
    print(f"納期整合性: {len(items1)}件")
    print(f"マスターデータ: {len(items2)}件")
    print(f"合計: {len(all_items)}件")

if __name__ == "__main__":
    main()
