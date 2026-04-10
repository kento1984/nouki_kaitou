"""確認中一覧のオーファンレコード調査スクリプト

使い方:
  python dev_scripts/check_orphans.py [送付履歴.xlsxのパス]

パス未指定時は共有フォルダのデフォルトパスを使用。
"""

import datetime
import sys
from pathlib import Path

from openpyxl import load_workbook

# デフォルトパス（共有フォルダ）
DEFAULT_PATH = r"\\flsv04\316京葉\納期回答書ツールフォルダ\送付履歴.xlsx"

# オーファン判定の基準日（新SAP切替日）
CUTOFF_DATE = datetime.date(2026, 4, 1)


def main() -> None:
    if len(sys.argv) > 1:
        path = Path(sys.argv[1])
    else:
        path = Path(DEFAULT_PATH)
        if not path.exists():
            # ローカルにあれば使う
            local = Path(__file__).resolve().parent.parent / "送付履歴.xlsx"
            if local.exists():
                path = local
            else:
                print(f"ファイルが見つかりません: {path}")
                print(f"使い方: python {sys.argv[0]} <送付履歴.xlsxのパス>")
                sys.exit(1)

    print(f"読み込み: {path}")
    wb = load_workbook(str(path), read_only=True, data_only=True)

    if "確認中一覧" not in wb.sheetnames:
        print("エラー: 確認中一覧シートが見つかりません")
        wb.close()
        sys.exit(1)

    ws = wb["確認中一覧"]

    # ヘッダー: A=送付日時, B=受注日, C=顧客名, D=受発注伝票, E=明細,
    #          F=メーカー名, G=品名, H=問合せ状況, I=ステータス, J=受注納期, K=送付者
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        order_num = str(row[3] or "").strip() if len(row) > 3 else ""
        if not order_num:
            continue

        sent_date_raw = row[0]
        order_date_raw = row[1]
        customer = str(row[2] or "").strip() if len(row) > 2 else ""
        detail = str(row[4] or "").strip() if len(row) > 4 else ""
        mfg = str(row[5] or "").strip() if len(row) > 5 else ""
        product = str(row[6] or "").strip() if len(row) > 6 else ""
        inquiry = str(row[7] or "").strip() if len(row) > 7 else ""
        status = str(row[8] or "").strip() if len(row) > 8 else ""

        # 受注日をdate化
        order_date = None
        if isinstance(order_date_raw, datetime.datetime):
            order_date = order_date_raw.date()
        elif isinstance(order_date_raw, datetime.date):
            order_date = order_date_raw

        # 送付日時をdate化
        sent_date = None
        if isinstance(sent_date_raw, datetime.datetime):
            sent_date = sent_date_raw.date()
        elif isinstance(sent_date_raw, datetime.date):
            sent_date = sent_date_raw

        records.append({
            "order_num": order_num,
            "detail": detail,
            "customer": customer,
            "mfg": mfg,
            "product": product,
            "order_date": order_date,
            "sent_date": sent_date,
            "inquiry": inquiry,
            "status": status,
        })

    wb.close()

    # --- 集計 ---
    total = len(records)
    orphans = [r for r in records if r["sent_date"] is not None and r["sent_date"] < CUTOFF_DATE]
    active = [r for r in records if r["inquiry"] not in ("除外", "済")]

    print(f"\n=== 確認中一覧 概要 ===")
    print(f"全レコード数:     {total}件")
    print(f"オーファン候補:   {len(orphans)}件 (送付日 < {CUTOFF_DATE})")
    print(f"  うちアクティブ: {len([r for r in orphans if r['inquiry'] not in ('除外', '済')])}件 (除外・済 以外)")

    if not orphans:
        print("\nオーファン候補はありません。")
        return

    # --- 注番先頭文字別集計 ---
    prefix_count: dict[str, int] = {}
    for r in orphans:
        # 先頭2文字を営業所コードとして使用
        prefix = r["order_num"][:2] if len(r["order_num"]) >= 2 else r["order_num"]
        prefix_count[prefix] = prefix_count.get(prefix, 0) + 1

    print(f"\n=== 注番プレフィックス別集計（オーファン候補） ===")
    for prefix in sorted(prefix_count, key=lambda p: -prefix_count[p]):
        print(f"  {prefix}: {prefix_count[prefix]}件")

    # --- 問合せ状況別集計 ---
    inquiry_count: dict[str, int] = {}
    for r in orphans:
        key = r["inquiry"] or "(空)"
        inquiry_count[key] = inquiry_count.get(key, 0) + 1

    print(f"\n=== 問合せ状況別（オーファン候補） ===")
    for k in sorted(inquiry_count, key=lambda x: -inquiry_count[x]):
        print(f"  {k}: {inquiry_count[k]}件")

    # --- 一覧表示 ---
    print(f"\n=== オーファン候補一覧 ===")
    print(f"{'注番':<16} {'明細':>4} {'送付日':>10} {'問合せ':>6} {'顧客名':<20} {'メーカー':<14} {'品名'}")
    print("-" * 100)
    for r in sorted(orphans, key=lambda x: (x["sent_date"] or datetime.date.min, x["order_num"])):
        sent = str(r["sent_date"]) if r["sent_date"] else "?"
        inq = r["inquiry"] or "-"
        # 品名は長すぎる場合truncate
        product = r["product"][:30] + "..." if len(r["product"]) > 30 else r["product"]
        print(f"{r['order_num']:<16} {r['detail']:>4} {sent:>10} {inq:>6} {r['customer']:<20} {r['mfg']:<14} {product}")


if __name__ == "__main__":
    main()
