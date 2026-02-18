"""4件の新規不一致をトレース"""

import sys
import io
import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from openpyxl import load_workbook

from nouki_kaitou.cache import build_all_caches
from nouki_kaitou.config import load_branch_settings, load_holidays
from nouki_kaitou.data_loader import (
    get_column_positions,
    get_data_rows_range,
    is_data_row,
    load_source_file,
    parse_order_row,
)
from nouki_kaitou.delivery_calc import calculate_delivery_date
from nouki_kaitou.customer import is_route_delivery


def main():
    tool_folder = Path(r"\\flsv04\316京葉\納期回答書ツールフォルダ")
    source_path = tool_folder / "受注一覧" / "17PM.xls"

    # 対象注番
    targets = [
        ("GL2C447202", "10"),  # VBA=2月17日出荷済み / PY=2月18日配達予定
        ("GL2Z446911", "20"),  # VBA=2月18日出荷予定 / PY=2月17日出荷済み
        ("GL2S447221", "10"),  # VBA=2月20日配達予定 / PY=2月19日出荷予定
        ("GL2V447251", "10"),  # VBA=2月17日出荷済み / PY=2月17日他拠点より出荷済み
    ]

    # ソースファイル読み込み
    source_data_raw = load_source_file(str(source_path))
    cols = get_column_positions(source_data_raw)

    # マスターファイル読み込み
    mfg_wb = load_workbook(str(tool_folder / "メーカー一覧.xlsx"), data_only=True)
    cust_wb = load_workbook(str(tool_folder / "顧客マスター_v2.xlsm"), data_only=True)

    try:
        confirming_ws = cust_wb["確認中一覧"]
    except KeyError:
        confirming_ws = None

    cache = build_all_caches(mfg_wb, cust_wb, confirming_ws, source_data_raw, cols)
    branch = load_branch_settings(mfg_wb, source_data_raw, cols)
    holidays = load_holidays(mfg_wb)

    today = datetime.date.today()
    execution_time = datetime.datetime.now()

    print("=" * 80)
    print(f"4件の新規不一致トレース")
    print(f"実行時刻: {execution_time}")
    print(f"今日: {today}")
    print(f"base_center: {branch.base_center}")
    print("=" * 80)

    # 対象注番を検索してトレース
    for order_no, detail_no in targets:
        for i in get_data_rows_range(source_data_raw, cols):
            if is_data_row(source_data_raw, i, cols):
                row = parse_order_row(source_data_raw, i, cols)
                if row.order_number == order_no and row.detail_number == detail_no:
                    print()
                    print("=" * 80)
                    print(f"【{order_no}|{detail_no}】")
                    print("=" * 80)

                    print(f"\n[SAPデータ]")
                    print(f"  伝票タイプ: {row.document_type}")
                    print(f"  出荷ステータス: {row.ship_status}")
                    print(f"  保管場所: {row.storage_place}")
                    print(f"  受注先: {row.customer_name}")
                    print(f"  出荷先: {row.ship_to_name}")
                    print(f"  指定納期: {row.specified_delivery_date}")
                    print(f"  受注納期: {row.order_delivery_date}")
                    print(f"  登録日: {row.registration_date}")
                    print(f"  時刻: {row.time_value}")

                    # 条件判定
                    print(f"\n[条件判定]")
                    print(f"  受注先=出荷先？: {row.customer_name.strip() == row.ship_to_name.strip()}")
                    print(f"  路線便？: {is_route_delivery(row.customer_name, cache)}")
                    print(f"  他拠点？: {row.storage_place != branch.base_center and row.storage_place != ''}")

                    # Python計算
                    result = calculate_delivery_date(
                        row, cache, holidays, branch, execution_time, today
                    )
                    print(f"\n[Python計算結果]")
                    print(f"  calculate_delivery_date: {result}")

                    break


if __name__ == "__main__":
    main()
