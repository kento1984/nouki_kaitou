"""納期回答計算の整合性検証スクリプト

16AM.xlsの全注番について納期計算を行い、
ロジック上おかしい・矛盾がある・実務的に不自然な結果を検出する。
"""

import datetime
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from nouki_kaitou.models import BranchSettings, CacheStore, OrderRow, HolidayMap
from nouki_kaitou.data_loader import (
    load_source_file, get_column_positions, parse_order_row,
    is_data_row, get_data_rows_range
)
from nouki_kaitou.cache import build_all_caches
from nouki_kaitou.config import load_branch_settings, load_holidays
from nouki_kaitou.delivery_calc import (
    calculate_delivery_date,
    extract_pickup_date,
    extract_arrival_date_from_internal,
)
from nouki_kaitou.customer import get_customer_delivery_days, is_route_delivery
from nouki_kaitou.manufacturer import get_delivery_days_to_add
from nouki_kaitou.confirming import get_confirmed_delivery_date
from nouki_kaitou.utils import is_december_31


@dataclass
class Issue:
    """検出した問題"""
    category: str           # 問題カテゴリ
    severity: str           # 重要度: ERROR / WARNING / INFO
    order_number: str       # 注番
    detail_number: str      # 明細
    customer_name: str      # 受注先
    ship_to_name: str       # 出荷先
    product_name: str       # 品名
    document_type: str      # 伝票タイプ
    ship_status: str        # 出荷ステータス
    specified_date: Optional[datetime.date]  # 指定納期
    order_date: Optional[datetime.date]      # 受注納期
    result: str             # 計算結果
    description: str        # 問題の説明
    comment_internal: str = ""  # コメント（社内）
    comment_external: str = ""  # コメント（社外）


def extract_date_from_result(result: str) -> Optional[datetime.date]:
    """結果文字列から日付を抽出"""
    # "2月19日出荷予定" → 2/19
    match = re.match(r"(\d{1,2})月(\d{1,2})日", result)
    if match:
        month = int(match.group(1))
        day = int(match.group(2))
        year = datetime.date.today().year
        try:
            d = datetime.date(year, month, day)
            # 180日以上過去なら翌年
            if d < datetime.date.today() and (datetime.date.today() - d).days > 180:
                d = datetime.date(year + 1, month, day)
            return d
        except ValueError:
            return None
    return None


def validate_order(
    row: OrderRow,
    result: str,
    cache: CacheStore,
    holidays: HolidayMap,
    branch: BranchSettings,
    today: datetime.date,
) -> list[Issue]:
    """1注文の検証を行い、問題リストを返す"""
    issues: list[Issue] = []

    def add_issue(category: str, severity: str, description: str):
        issues.append(Issue(
            category=category,
            severity=severity,
            order_number=row.order_number,
            detail_number=row.detail_number,
            customer_name=row.customer_name,
            ship_to_name=row.ship_to_name,
            product_name=row.product_name,
            document_type=row.document_type,
            ship_status=row.ship_status,
            specified_date=row.specified_delivery_date,
            order_date=row.order_delivery_date,
            result=result,
            description=description,
            comment_internal=row.comment_internal,
            comment_external=row.comment_external,
        ))

    # 結果から情報を抽出
    is_delivery = "配達" in result  # 配達予定/配達済み
    is_ship = "出荷" in result and "配達" not in result  # 出荷予定/出荷済み
    is_pickup = "引取" in result
    is_work = "作業" in result
    is_arrival = "着" in result and "→" in result  # M/D出荷→M/D着
    is_confirming = result in ("確認中", "日程調整中")
    is_past = "済み" in result or "済" in result
    is_future = "予定" in result
    result_date = extract_date_from_result(result)

    # フラグ
    is_stock = row.document_type == "【受注】在庫販売"
    is_direct = row.document_type == "【受注】直送販売"
    customer_diff_ship = row.customer_name != row.ship_to_name
    is_rosenbin = is_route_delivery(row.customer_name, cache)
    delivery_days = get_customer_delivery_days(row.customer_name, cache)
    has_delivery_days = bool(delivery_days)

    # コメント
    internal = row.comment_internal.strip()
    external = row.comment_external.strip()
    combined_comment = f"{external} {internal}".strip()

    # ============================================
    # 1. 配達/出荷の使い分けチェック
    # ============================================

    # 1-1. 受注先=出荷先なのに「出荷予定」（曜日制限・路線便以外で、在庫販売のみ）
    # 直送販売の場合は曜日制限で「出荷予定」になるケースがあるので除外
    if not customer_diff_ship and is_ship and not is_rosenbin and not has_delivery_days:
        if not is_pickup and not is_work and not is_arrival:
            if is_stock:  # 在庫販売のみチェック
                add_issue(
                    "配達/出荷の使い分け",
                    "WARNING",
                    f"在庫販売で受注先=出荷先だが「出荷」になっている（曜日制限なし、路線便なし）"
                )

    # 1-2. 受注先≠出荷先なのに「配達予定」
    if customer_diff_ship and is_delivery:
        add_issue(
            "配達/出荷の使い分け",
            "ERROR",
            f"受注先≠出荷先だが「配達」になっている（出荷先: {row.ship_to_name}）"
        )

    # 1-3. 路線便なのに「配達予定」
    if is_rosenbin and is_delivery:
        add_issue(
            "配達/出荷の使い分け",
            "WARNING",
            f"路線便顧客だが「配達」になっている"
        )

    # ============================================
    # 2. 日付の妥当性チェック
    # ============================================

    if result_date:
        # 2-1. 過去の日付なのに「予定」
        if result_date < today and is_future:
            add_issue(
                "日付の妥当性",
                "WARNING",
                f"結果日付({result_date})が過去だが「予定」になっている"
            )

        # 2-2. 未来の日付なのに「済み」
        if result_date > today and is_past:
            add_issue(
                "日付の妥当性",
                "ERROR",
                f"結果日付({result_date})が未来だが「済み」になっている"
            )

        # 2-3. 結果日付が異常に遠い（90日以上先）
        if result_date > today and (result_date - today).days > 90:
            add_issue(
                "日付の妥当性",
                "INFO",
                f"結果日付({result_date})が{(result_date - today).days}日後と遠い"
            )

        # 2-4. 結果日付が登録日より前
        if row.registration_date and result_date < row.registration_date:
            if not is_past:  # 済みなら問題ない
                add_issue(
                    "日付の妥当性",
                    "WARNING",
                    f"結果日付({result_date})が登録日({row.registration_date})より前"
                )

    # ============================================
    # 3. コメントと結果の整合性チェック
    # ============================================

    # 3-1. 引取キーワードがあるのに引取結果でない
    if ("引取" in combined_comment or "引き取り" in combined_comment) and not is_pickup:
        pickup_date = extract_pickup_date(combined_comment, today)
        if pickup_date:
            add_issue(
                "コメント整合性",
                "ERROR",
                f"コメントに引取日({pickup_date})があるが引取結果でない"
            )

    # 3-2. @@着日があるのに着日結果でない
    if ("@@" in internal or "＠＠" in internal) and not is_arrival:
        arrival_date = extract_arrival_date_from_internal(internal, today)
        if arrival_date:
            add_issue(
                "コメント整合性",
                "ERROR",
                f"コメントに@@着日({arrival_date})があるが着日結果でない"
            )

    # 3-3. &&作業があるのに作業結果でない
    if ("&&" in internal or "＆＆" in internal) and not is_work:
        add_issue(
            "コメント整合性",
            "ERROR",
            f"コメントに&&作業マーカーがあるが作業結果でない"
        )

    # 3-4. 送り状番号らしきもの（10桁以上の数字）がコメントにある
    tracking_match = re.search(r"\d{10,}", combined_comment)
    if tracking_match and row.ship_status == "未処理":
        add_issue(
            "コメント整合性",
            "INFO",
            f"コメントに送り状番号らしき数字({tracking_match.group()})があるが未処理"
        )

    # ============================================
    # 4. 伝票タイプと計算パスの整合性チェック
    # ============================================

    # 4-1. 直送販売で転送中なのに配達予定
    storage = row.storage_place.strip()
    if is_direct and storage == "転送中（直送用）" and is_delivery:
        add_issue(
            "伝票タイプ整合性",
            "ERROR",
            f"直送販売で転送中だが「配達」になっている"
        )

    # 4-2. 在庫販売で他拠点在庫なのに自社配達
    if is_stock and storage and storage != branch.base_center:
        if is_delivery and not is_confirming:
            add_issue(
                "伝票タイプ整合性",
                "WARNING",
                f"在庫販売で他拠点在庫({storage})だが「配達」になっている"
            )

    # ============================================
    # 5. 出荷ステータスと結果の整合性チェック
    # ============================================

    # 5-1. 処理完了なのに結果日付が30日以上先
    if row.ship_status == "処理完了" and result_date:
        if result_date > today and (result_date - today).days > 30:
            add_issue(
                "出荷ステータス整合性",
                "WARNING",
                f"処理完了だが結果日付が{(result_date - today).days}日後"
            )

    # 5-2. 未処理なのに「済み」
    if row.ship_status == "未処理" and is_past:
        add_issue(
            "出荷ステータス整合性",
            "WARNING",
            f"未処理だが「{result}」（済み）になっている"
        )

    # ============================================
    # 6. 受注納期/指定納期の整合性チェック
    # ============================================

    # 6-1. 受注納期=12/31で指定納期もあるのに「確認中」
    if row.order_delivery_date and is_december_31(row.order_delivery_date):
        if row.specified_delivery_date and not is_december_31(row.specified_delivery_date):
            if result == "確認中":
                add_issue(
                    "納期整合性",
                    "WARNING",
                    f"受注納期=12/31だが指定納期({row.specified_delivery_date})があるのに確認中"
                )

    # 6-2. 指定納期が過去なのに未処理で「予定」
    if row.specified_delivery_date and row.specified_delivery_date < today:
        if row.ship_status == "未処理" and is_future:
            days_past = (today - row.specified_delivery_date).days
            if days_past > 7:  # 1週間以上過去
                add_issue(
                    "納期整合性",
                    "WARNING",
                    f"指定納期({row.specified_delivery_date})が{days_past}日前だが未処理で予定"
                )

    # 6-3. 指定納期 > 受注納期（12/31以外で）
    if row.specified_delivery_date and row.order_delivery_date:
        if not is_december_31(row.order_delivery_date):
            if row.specified_delivery_date > row.order_delivery_date:
                add_issue(
                    "納期整合性",
                    "INFO",
                    f"指定納期({row.specified_delivery_date}) > 受注納期({row.order_delivery_date})"
                )

    # ============================================
    # 7. マスターデータ整合性チェック
    # ============================================

    # 7-1. 品目Groupがマスターにない
    if row.item_group_code:
        days = get_delivery_days_to_add(row.item_group_code, cache)
        if days == 2:  # デフォルト値
            if row.item_group_code not in cache.mfg_days:
                add_issue(
                    "マスターデータ",
                    "INFO",
                    f"品目Group({row.item_group_code})がメーカー一覧にない（デフォルト2日）"
                )
    else:
        add_issue(
            "マスターデータ",
            "INFO",
            f"品目Groupが空"
        )

    # 7-2. 顧客名がマスターにない
    if row.customer_name and row.customer_name not in cache.cust_days:
        add_issue(
            "マスターデータ",
            "INFO",
            f"顧客({row.customer_name})が顧客マスターにない"
        )

    # ============================================
    # 8. その他の異常チェック
    # ============================================

    # 8-1. 「日程調整中」だが指定納期がある
    if result == "日程調整中" and row.specified_delivery_date:
        if not is_december_31(row.specified_delivery_date):
            add_issue(
                "その他",
                "ERROR",
                f"日程調整中だが指定納期({row.specified_delivery_date})がある"
            )

    # 8-2. 「確認中」だが在庫販売
    if result == "確認中" and is_stock:
        add_issue(
            "その他",
            "WARNING",
            f"在庫販売なのに「確認中」（通常は日程調整中）"
        )

    # 8-3. 伝票タイプが想定外
    if row.document_type not in ("【受注】在庫販売", "【受注】直送販売"):
        add_issue(
            "その他",
            "INFO",
            f"伝票タイプが想定外: {row.document_type}"
        )

    return issues


def write_excel(issues: list[Issue], output_path: Path):
    """問題リストをExcelに出力"""
    wb = Workbook()

    # ヘッダースタイル
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 重要度別の色
    severity_fills = {
        "ERROR": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "WARNING": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "INFO": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    }

    # 罫線
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # カテゴリ別にグループ化
    categories = {}
    for issue in issues:
        if issue.category not in categories:
            categories[issue.category] = []
        categories[issue.category].append(issue)

    # 全件シート
    ws_all = wb.active
    ws_all.title = "全件"

    headers = [
        "重要度", "カテゴリ", "注番", "明細", "受注先", "出荷先",
        "品名", "伝票タイプ", "出荷ステータス", "指定納期", "受注納期",
        "計算結果", "問題の説明", "コメント（社内）", "コメント（社外）"
    ]

    # 列幅
    col_widths = [8, 18, 14, 6, 25, 25, 35, 15, 12, 12, 12, 18, 50, 30, 30]

    def setup_sheet(ws, data: list[Issue]):
        # ヘッダー
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col)].width = col_widths[col - 1]

        # データ
        for row_idx, issue in enumerate(data, 2):
            values = [
                issue.severity,
                issue.category,
                issue.order_number,
                issue.detail_number,
                issue.customer_name,
                issue.ship_to_name,
                issue.product_name,
                issue.document_type,
                issue.ship_status,
                str(issue.specified_date) if issue.specified_date else "",
                str(issue.order_date) if issue.order_date else "",
                issue.result,
                issue.description,
                issue.comment_internal,
                issue.comment_external,
            ]

            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border
                if col == 1:  # 重要度列
                    cell.fill = severity_fills.get(issue.severity, PatternFill())
                cell.alignment = Alignment(vertical="center", wrap_text=(col >= 13))

        # フィルター
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(data) + 1}"
        # ヘッダー行を固定
        ws.freeze_panes = "A2"

    # 全件シート
    setup_sheet(ws_all, issues)

    # カテゴリ別シート
    for category, cat_issues in sorted(categories.items()):
        # シート名を31文字以内に、禁止文字を置換
        sheet_name = category.replace("/", "_").replace("\\", "_").replace("*", "_")
        sheet_name = sheet_name.replace("?", "_").replace("[", "_").replace("]", "_")
        sheet_name = sheet_name[:31] if len(sheet_name) > 31 else sheet_name
        ws = wb.create_sheet(title=sheet_name)
        setup_sheet(ws, cat_issues)

    # サマリーシート
    ws_summary = wb.create_sheet(title="サマリー", index=0)
    ws_summary.cell(row=1, column=1, value="カテゴリ").font = header_font
    ws_summary.cell(row=1, column=1).fill = header_fill
    ws_summary.cell(row=1, column=2, value="ERROR").font = header_font
    ws_summary.cell(row=1, column=2).fill = header_fill
    ws_summary.cell(row=1, column=3, value="WARNING").font = header_font
    ws_summary.cell(row=1, column=3).fill = header_fill
    ws_summary.cell(row=1, column=4, value="INFO").font = header_font
    ws_summary.cell(row=1, column=4).fill = header_fill
    ws_summary.cell(row=1, column=5, value="合計").font = header_font
    ws_summary.cell(row=1, column=5).fill = header_fill

    row_idx = 2
    total_error = 0
    total_warning = 0
    total_info = 0

    for category in sorted(categories.keys()):
        cat_issues = categories[category]
        error_count = sum(1 for i in cat_issues if i.severity == "ERROR")
        warning_count = sum(1 for i in cat_issues if i.severity == "WARNING")
        info_count = sum(1 for i in cat_issues if i.severity == "INFO")

        ws_summary.cell(row=row_idx, column=1, value=category)
        ws_summary.cell(row=row_idx, column=2, value=error_count)
        ws_summary.cell(row=row_idx, column=3, value=warning_count)
        ws_summary.cell(row=row_idx, column=4, value=info_count)
        ws_summary.cell(row=row_idx, column=5, value=len(cat_issues))

        total_error += error_count
        total_warning += warning_count
        total_info += info_count
        row_idx += 1

    # 合計行
    ws_summary.cell(row=row_idx, column=1, value="合計").font = Font(bold=True)
    ws_summary.cell(row=row_idx, column=2, value=total_error).font = Font(bold=True)
    ws_summary.cell(row=row_idx, column=3, value=total_warning).font = Font(bold=True)
    ws_summary.cell(row=row_idx, column=4, value=total_info).font = Font(bold=True)
    ws_summary.cell(row=row_idx, column=5, value=len(issues)).font = Font(bold=True)

    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 10
    ws_summary.column_dimensions['C'].width = 10
    ws_summary.column_dimensions['D'].width = 10
    ws_summary.column_dimensions['E'].width = 10

    wb.save(output_path)


def main():
    # パラメータ
    source_file = r"\\flsv04\316京葉\納期回答書ツールフォルダ\受注一覧\16AM.xls"
    tool_folder = r"\\flsv04\316京葉\納期回答書ツールフォルダ"
    output_path = Path(__file__).parent / "validation_result.xlsx"

    today = datetime.date.today()
    execution_time = datetime.datetime.now()

    print("=" * 60)
    print("納期回答計算 整合性検証")
    print("=" * 60)
    print(f"対象ファイル: {source_file}")
    print(f"今日の日付: {today}")
    print()

    # ファイル読み込み
    print("ファイル読み込み中...")
    source_data = load_source_file(source_file)
    cols = get_column_positions(source_data)
    if cols is None:
        print("ERROR: 列位置を取得できませんでした")
        sys.exit(1)

    from openpyxl import load_workbook
    manufacturer_wb = load_workbook(Path(tool_folder) / "メーカー一覧.xlsx", data_only=True)
    customer_wb = load_workbook(Path(tool_folder) / "顧客マスター_v2.xlsm", data_only=True)

    try:
        history_wb = load_workbook(Path(tool_folder) / "送付履歴.xlsx", data_only=True)
        confirming_ws = history_wb["確認中一覧"] if "確認中一覧" in history_wb.sheetnames else None
    except Exception:
        confirming_ws = None

    branch = load_branch_settings(manufacturer_wb, source_data, cols)
    holidays = load_holidays(manufacturer_wb)
    cache = build_all_caches(manufacturer_wb, customer_wb, confirming_ws, source_data, cols)

    print(f"  営業所: {branch.name}")
    print(f"  祝日件数: {len(holidays)}")
    print()

    # 全注番を処理
    print("納期計算と検証中...")
    all_issues: list[Issue] = []
    order_count = 0
    error_count = 0

    for row_idx in get_data_rows_range(source_data, cols):
        if not is_data_row(source_data, row_idx, cols):
            continue

        order_row = parse_order_row(source_data, row_idx, cols)
        order_count += 1

        try:
            result = calculate_delivery_date(
                order_row, cache, holidays, branch, execution_time, today
            )
        except Exception as e:
            error_count += 1
            all_issues.append(Issue(
                category="計算エラー",
                severity="ERROR",
                order_number=order_row.order_number,
                detail_number=order_row.detail_number,
                customer_name=order_row.customer_name,
                ship_to_name=order_row.ship_to_name,
                product_name=order_row.product_name,
                document_type=order_row.document_type,
                ship_status=order_row.ship_status,
                specified_date=order_row.specified_delivery_date,
                order_date=order_row.order_delivery_date,
                result="(エラー)",
                description=f"計算中に例外発生: {e}",
            ))
            continue

        # 検証
        issues = validate_order(order_row, result, cache, holidays, branch, today)
        all_issues.extend(issues)

    print(f"  処理件数: {order_count}")
    print(f"  計算エラー: {error_count}")
    print(f"  検出した問題: {len(all_issues)}")
    print()

    # 重要度別集計
    error_issues = [i for i in all_issues if i.severity == "ERROR"]
    warning_issues = [i for i in all_issues if i.severity == "WARNING"]
    info_issues = [i for i in all_issues if i.severity == "INFO"]

    print("重要度別集計:")
    print(f"  ERROR:   {len(error_issues)}")
    print(f"  WARNING: {len(warning_issues)}")
    print(f"  INFO:    {len(info_issues)}")
    print()

    # カテゴリ別集計
    print("カテゴリ別集計:")
    categories = {}
    for issue in all_issues:
        if issue.category not in categories:
            categories[issue.category] = 0
        categories[issue.category] += 1

    for cat, count in sorted(categories.items(), key=lambda x: -x[1]):
        print(f"  {cat}: {count}")
    print()

    # Excel出力
    print(f"結果をExcelに出力中: {output_path}")
    write_excel(all_issues, output_path)
    print("完了!")


if __name__ == "__main__":
    main()
