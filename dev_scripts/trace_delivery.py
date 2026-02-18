"""GL2V446963|10のdelivery_calc.pyトレーススクリプト"""

import sys
import io
import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from openpyxl import load_workbook

from nouki_kaitou.cache import build_all_caches
from nouki_kaitou.config import load_branch_settings, load_holidays
from nouki_kaitou.data_loader import (
    get_column_positions,
    get_data_rows_range,
    is_data_row,
    load_source_file,
    parse_order_row,
)
from nouki_kaitou.history import load_delivery_history
from nouki_kaitou.delivery_calc import calculate_delivery_date
from nouki_kaitou.report_generator import build_report_row, _determine_flags
from nouki_kaitou.confirming import get_confirming_status
from nouki_kaitou.utils import is_december_31


def main():
    order_no = sys.argv[1] if len(sys.argv) > 1 else "GL2V446963"
    detail_no = sys.argv[2] if len(sys.argv) > 2 else "10"

    tool_folder = Path(r"\\flsv04\316京葉\納期回答書ツールフォルダ")
    source_path = tool_folder / "受注一覧" / "17PM.xls"

    print("=" * 80)
    print(f"トレース: {order_no}|{detail_no}")
    print("=" * 80)

    # ソースファイル読み込み
    print("\n1. ソースファイル読み込み...")
    source_data_raw = load_source_file(str(source_path))
    cols = get_column_positions(source_data_raw)

    # マスターファイル読み込み
    print("2. マスターファイル読み込み...")
    mfg_wb = load_workbook(str(tool_folder / "メーカー一覧.xlsx"), data_only=True)
    cust_wb = load_workbook(str(tool_folder / "顧客マスター_v2.xlsm"), data_only=True)

    try:
        confirming_ws = cust_wb["確認中一覧"]
    except KeyError:
        confirming_ws = None

    cache = build_all_caches(mfg_wb, cust_wb, confirming_ws, source_data_raw, cols)
    branch = load_branch_settings(mfg_wb, source_data_raw, cols)
    holidays = load_holidays(mfg_wb)

    # 送付履歴読み込み
    print("3. 送付履歴読み込み...")
    history_wb = load_workbook(str(tool_folder / "送付履歴.xlsx"), data_only=True, read_only=True)
    ws_history = history_wb["送付履歴"]
    ws_confirming = history_wb["確認中一覧"]

    today = datetime.date.today()
    sent_orders = load_delivery_history(ws_history, ws_confirming, cache, holidays, today)
    history_wb.close()

    # 対象注番を検索
    print(f"4. {order_no}|{detail_no} を検索...")
    target_row = None
    for i in get_data_rows_range(source_data_raw, cols):
        if is_data_row(source_data_raw, i, cols):
            row = parse_order_row(source_data_raw, i, cols)
            if row.order_number == order_no and row.detail_number == detail_no:
                target_row = row
                break

    if target_row is None:
        print(f"エラー: {order_no}|{detail_no} が見つかりません")
        return

    print()
    print("=" * 80)
    print("【SAPデータ】")
    print("=" * 80)
    print(f"  注番|明細: {target_row.order_number}|{target_row.detail_number}")
    print(f"  顧客: {target_row.customer_name}")
    print(f"  出荷先: {target_row.ship_to_name}")
    print(f"  品名: {target_row.product_name}")
    print(f"  伝票タイプ: {target_row.document_type}")
    print(f"  出荷ステータス: {target_row.ship_status}")
    print(f"  保管場所: {target_row.storage_place}")
    print(f"  指定納期: {target_row.specified_delivery_date}")
    print(f"  受注納期: {target_row.order_delivery_date}")

    print()
    print("=" * 80)
    print("【delivery_calc.py トレース】")
    print("=" * 80)

    # 実行時刻
    execution_time = datetime.datetime.now()
    print(f"\n  実行時刻: {execution_time}")

    # Step 1: 各チェック関数の判定
    print()
    print("【Step 1】各チェック関数の判定")
    print("-" * 60)

    # _check_specified_date
    spec_date = target_row.specified_delivery_date
    print(f"  指定納期 = {spec_date}")
    print(f"    12/31か？ {is_december_31(spec_date) if spec_date else 'None'}")
    print(f"    → _check_specified_date: {'スキップ' if spec_date is None or is_december_31(spec_date) else '処理'}")

    # _check_stock_completed
    print()
    print(f"  伝票タイプ = {target_row.document_type}")
    print(f"    在庫販売か？ {target_row.document_type == '【受注】在庫販売'}")
    print(f"    → _check_stock_completed: {'処理' if target_row.document_type == '【受注】在庫販売' else 'スキップ（直送販売）'}")

    # _check_himozuki_completed
    print()
    storage = target_row.storage_place.strip()
    if not storage:
        storage = cache.storage.get(target_row.order_number, "")
    print(f"  保管場所 = '{storage}'")
    print(f"    転送中（直送用）か？ {storage == '転送中（直送用）'}")

    himozuki_check = (
        target_row.document_type == "【受注】直送販売"
        and target_row.ship_status == "処理完了"
        and storage != "転送中（直送用）"
    )
    print(f"    → _check_himozuki_completed: {'処理' if himozuki_check else 'スキップ（転送中のため）'}")

    # _check_dec31
    print()
    order_date = target_row.order_delivery_date
    print(f"  受注納期 = {order_date}")
    print(f"    12/31か？ {is_december_31(order_date) if order_date else 'None'}")
    print(f"    → _check_dec31: {'処理' if order_date and is_december_31(order_date) else 'スキップ'}")

    # 実際に計算
    print()
    print("【Step 2】calculate_delivery_date 実行")
    print("-" * 60)

    result = calculate_delivery_date(
        target_row, cache, holidays, branch, execution_time, today
    )
    print(f"  結果: {result}")

    print()
    print("=" * 80)
    print("【report_generator.py トレース】")
    print("=" * 80)

    # 送付履歴チェック
    history_key = f"{target_row.order_number}|{target_row.detail_number}"
    print(f"\n  送付履歴キー: {history_key}")
    print(f"  送付履歴に存在: {history_key in sent_orders}")
    if history_key in sent_orders:
        print(f"  前回ステータス: {sent_orders[history_key]}")

    # _determine_flags
    print()
    print("【_determine_flags】")
    print("-" * 60)

    # 紐付き判定の詳細
    is_himozuki = False
    if target_row.document_type.strip() == "【受注】直送販売":
        storage_for_flag = target_row.storage_place.strip()
        if not storage_for_flag:
            storage_for_flag = cache.storage.get(target_row.order_number, "")
        print(f"  直送販売: True")
        print(f"  保管場所: '{storage_for_flag}'")
        print(f"  転送中でない: {storage_for_flag != '転送中（直送用）'}")
        if storage_for_flag != "転送中（直送用）":
            is_himozuki = True
    else:
        print(f"  直送販売: False")

    print(f"  → is_himozuki = {is_himozuki}")

    # 確認中一覧のステータス
    confirming_status = get_confirming_status(
        target_row.order_number, target_row.detail_number, cache
    )
    print(f"  確認中一覧ステータス: {confirming_status or 'なし'}")
    is_bunno_in_confirming = (confirming_status == "分納")
    print(f"  is_bunno_in_confirming: {is_bunno_in_confirming}")

    # force_delivered判定
    print()
    print("【force_delivered判定】")
    previous_status = sent_orders.get(history_key, "")
    force_delivered = False

    print(f"  ship_status == '処理完了': {target_row.ship_status == '処理完了'}")
    print(f"  送付履歴になし: {not sent_orders.get(history_key)}")
    print(f"  is_himozuki: {is_himozuki}")
    print(f"  is_bunno_in_confirming: {is_bunno_in_confirming}")

    if target_row.ship_status == "処理完了":
        if not sent_orders.get(history_key):
            if not is_himozuki and not is_bunno_in_confirming:
                force_delivered = True
                print(f"  → 条件: 処理完了 + 送付履歴なし + 紐付きFalse + 分納False")
        elif previous_status == "確認中":
            if not is_himozuki and not is_bunno_in_confirming:
                force_delivered = True
                print(f"  → 条件: 処理完了 + 前回確認中 + 紐付きFalse + 分納False")

    print(f"  → force_delivered = {force_delivered}")

    # build_report_row (force_delivered=True)
    print()
    print("【build_report_row with force_delivered=True】")
    print("-" * 60)

    report_row, delivery_status = build_report_row(
        target_row, cache, holidays, branch, execution_time,
        force_delivered=True, today=today
    )
    print(f"  delivery_status (before override): {delivery_status}")

    # forceDelivered時の上書き
    if force_delivered and delivery_status in ("確認中", "欠品中", "日程調整中"):
        final_status = "納品済み"
        print(f"  → 納品済み上書き適用")
    elif force_delivered and "（欠品）" in delivery_status:
        final_status = "納品済み"
        print(f"  → 納品済み上書き適用（欠品）")
    else:
        final_status = delivery_status

    print(f"  final_status: {final_status}")

    print()
    print("=" * 80)
    print("【結論】")
    print("=" * 80)
    print()
    print(f"  保管場所 = '{storage}' → 転送中（直送用）")
    print(f"  is_himozuki = {is_himozuki} (転送中のため紐付き判定はFalse)")
    print()
    print(f"  delivery_calc.py → '{result}'")
    print(f"    理由: 受注納期12/31 + 確認中一覧に確定日なし → _check_dec31 → '確認中'")
    print()
    print(f"  force_delivered = {force_delivered}")
    if force_delivered:
        print(f"    理由: 処理完了 + 送付履歴なし + 紐付きFalse")
        print(f"    → '確認中' は '納品済み' に上書きされるべき")
    else:
        print(f"    → '確認中' のまま")


if __name__ == "__main__":
    main()
