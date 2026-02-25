"""VBA vs Python 1週間比較スクリプト

2/10〜2/13（3営業日×AM/PM＝6回分）のVBA版回答書とPython版計算結果を突き合わせ、
納期回答の不一致を検出・判定する。

方針:
- VBA回答書の各行について、同じ受注データからPythonで再計算して比較
- L列（弊社注番）で受注データと突き合わせ（品名マッチより確実）
- 送付履歴は使わない — 過去時点の履歴を再現できないため
- execution_time / today は各実行日時に合わせて設定
"""

import sys
import io
import datetime
import re
from pathlib import Path
from typing import Optional

# パッケージの親ディレクトリをsys.pathに追加
# dev_scripts/ → nouki_kaitou/ → Users/kento.kashiwabara/
sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from nouki_kaitou.cache import build_all_caches
from nouki_kaitou.config import load_branch_settings, load_holidays
from nouki_kaitou.data_loader import (
    get_column_positions,
    get_data_rows_range,
    is_data_row,
    load_source_file,
    parse_order_row,
)
from nouki_kaitou.models import BranchSettings, OrderRow
from nouki_kaitou.report_generator import (
    build_report_row,
    _determine_flags,
    _pass_basic_filter,
)


# ============================================
# 共通ユーティリティ（compare_vba_python.pyから流用）
# ============================================

def normalize_delivery_answer(answer: str) -> tuple[str, str, str]:
    """納期回答を正規化して(日付部分, 配達/出荷, 予定/済み)に分解"""
    if not answer:
        return ("", "", "")

    answer = answer.strip()

    # 特殊ケース
    if answer in ("確認中", "欠品中", "日程調整中", "納品済み", "分納完了"):
        return (answer, "", "")

    # @@着日指定パターン
    match = re.match(r"(\d+/\d+)(出荷)(済み|済)?→(\d+/\d+)(着)(予定)?", answer)
    if match:
        ship_date = match.group(1)
        arrival_date = match.group(4)
        tense = "済み" if match.group(3) else "予定"
        return (f"{ship_date}→{arrival_date}着", "出荷", tense)

    # 通常パターン
    match = re.match(r"(.+?)(配達|出荷)(予定|済み|済)", answer)
    if match:
        date_part = match.group(1)
        delivery_type = match.group(2)
        tense = match.group(3)
        if tense == "済":
            tense = "済み"
        return (date_part, delivery_type, tense)

    # 引取パターン
    match = re.match(r"(.+?)(引取)(予定|済み|済)", answer)
    if match:
        return (match.group(1), "引取", match.group(3))

    # 作業パターン
    match = re.match(r"(.+?)(作業)(予定|済み|済)", answer)
    if match:
        return (match.group(1), "作業", match.group(3))

    return (answer, "", "")


def compare_answers(vba_answer: str, py_answer: str) -> tuple[bool, str]:
    """2つの納期回答を比較。予定/済みの違いは無視。"""
    vba_norm = normalize_delivery_answer(vba_answer)
    py_norm = normalize_delivery_answer(py_answer)

    # 日付部分が異なる
    if vba_norm[0] != py_norm[0]:
        return (False, f"日付: VBA={vba_norm[0]} vs PY={py_norm[0]}")

    # 配達/出荷の区別が異なる
    if vba_norm[1] != py_norm[1]:
        return (False, f"区分: VBA={vba_norm[1]} vs PY={py_norm[1]}")

    return (True, "")


def normalize_product_name(name: str) -> str:
    """品名を正規化（比較用）"""
    if not name:
        return ""
    s = name.replace("\u3000", " ").strip()
    return s[:30]


def normalize_quantity(value) -> str:
    """数量を正規化"""
    if value is None:
        return ""
    s = str(value).strip().replace(",", "")
    try:
        return str(int(float(s)))
    except ValueError:
        return s


def parse_date_cell(val) -> Optional[datetime.date]:
    """セルの日付値をパース"""
    if val is None:
        return None
    if isinstance(val, datetime.datetime):
        return val.date()
    if isinstance(val, datetime.date):
        return val
    s = str(val).strip()
    for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S"]:
        try:
            return datetime.datetime.strptime(s.split()[0], fmt.split()[0]).date()
        except ValueError:
            continue
    return None


def normalize_customer_name(name: str) -> str:
    """顧客名を正規化（比較用）"""
    if not name:
        return ""
    s = name.replace("\u3000", " ")
    s = re.sub(r"\s+", " ", s)
    s = s.strip()
    s = s.replace("（株）", "(株)")
    s = s.replace("（有）", "(有)")
    return s


def extract_customer_name_from_filename(filename: str) -> str:
    """ファイル名から顧客名を抽出"""
    name = filename.replace("納期回答書_", "").replace(".xlsx", "")
    name = re.sub(r"_\d{8}$", "", name)
    if name.endswith("様"):
        name = name[:-1]
    return normalize_customer_name(name)


# ============================================
# VBA回答書の読み込み（L列=弊社注番を追加）
# ============================================

def read_vba_excel(file_path: Path, customer_name: str) -> list[dict]:
    """VBA生成のExcelファイルを読み込み。L列（弊社注番）も取得。"""
    result = []

    try:
        wb = load_workbook(str(file_path), data_only=True)
    except Exception as e:
        print(f"  読み込みエラー: {file_path.name} - {e}")
        return result

    ws = wb.active

    # ヘッダー行を探す（A列が「受注日」の行）
    header_row = None
    for row_idx in range(1, min(20, ws.max_row + 1)):
        cell_val = ws.cell(row=row_idx, column=1).value
        if cell_val and str(cell_val).strip() == "受注日":
            header_row = row_idx
            break

    if header_row is None:
        print(f"  ヘッダー行が見つかりません: {file_path.name}")
        wb.close()
        return result

    # 列インデックス（A=1から）
    col_order_date = 1       # A: 受注日
    col_manufacturer = 4     # D: メーカー名
    col_product = 5          # E: 品名
    col_quantity = 6         # F: 数量
    col_unit_price = 7       # G: 単価
    col_amount = 8           # H: 金額
    col_delivery = 9         # I: 納期回答
    col_order_number = 12    # L: 弊社注番

    for row_idx in range(header_row + 1, ws.max_row + 1):
        order_date_val = ws.cell(row=row_idx, column=col_order_date).value
        if not order_date_val:
            continue

        # フッター行の判定
        unit_price_val = ws.cell(row=row_idx, column=col_unit_price).value
        if unit_price_val and "※" in str(unit_price_val):
            break

        order_date = parse_date_cell(order_date_val)
        manufacturer = str(ws.cell(row=row_idx, column=col_manufacturer).value or "").strip()
        product = str(ws.cell(row=row_idx, column=col_product).value or "").strip()
        quantity = str(ws.cell(row=row_idx, column=col_quantity).value or "").strip()
        delivery = str(ws.cell(row=row_idx, column=col_delivery).value or "").strip()

        # L列: 弊社注番
        order_number_val = ws.cell(row=row_idx, column=col_order_number).value
        order_number = str(order_number_val or "").strip()

        result.append({
            "顧客": customer_name,
            "受注日": order_date,
            "メーカー": manufacturer,
            "品名": product,
            "品名_正規化": normalize_product_name(product),
            "数量": normalize_quantity(quantity),
            "納期回答": delivery,
            "弊社注番": order_number,
            "ファイル": file_path.name,
        })

    wb.close()
    return result


# ============================================
# 実行回の定義
# ============================================

TOOL_FOLDER = Path(r"\\flsv04\316京葉\納期回答書ツールフォルダ")
SOURCE_FOLDER = TOOL_FOLDER / "受注一覧"
VBA_BASE_FOLDER = TOOL_FOLDER / "納期回答書"

# 実行回とデータのマッピング
RUNS = [
    {
        "label": "2/10 AM",
        "source_file": "10AM.XLS",
        "vba_folders": ["2月10日(火)_①回目"],
        "execution_time": datetime.datetime(2026, 2, 10, 12, 0),
        "today": datetime.date(2026, 2, 10),
    },
    {
        "label": "2/10 PM",
        "source_file": "10PM.XLS",
        "vba_folders": ["2月10日(火)_②回目"],
        "execution_time": datetime.datetime(2026, 2, 10, 17, 0),
        "today": datetime.date(2026, 2, 10),
    },
    {
        "label": "2/12 AM",
        "source_file": "12AM.XLS",
        "vba_folders": ["2月12日(木)_①回目"],
        "execution_time": datetime.datetime(2026, 2, 12, 12, 0),
        "today": datetime.date(2026, 2, 12),
    },
    {
        "label": "2/12 PM",
        "source_file": "12PM.XLS",
        "vba_folders": ["2月12日(木)_②回目"],
        "execution_time": datetime.datetime(2026, 2, 12, 17, 0),
        "today": datetime.date(2026, 2, 12),
    },
    {
        "label": "2/13 AM",
        "source_file": "13AM.XLS",
        "vba_folders": ["2月13日(金)_①回目"],
        "execution_time": datetime.datetime(2026, 2, 13, 12, 0),
        "today": datetime.date(2026, 2, 13),
    },
    {
        "label": "2/13 PM",
        "source_file": "13PM.XLS",
        # ②③は個別再実行（各1件）、⑤が本番（36件）。同一顧客は後のフォルダで上書き
        "vba_folders": [
            "2月13日(金)_②回目",
            "2月13日(金)_③回目",
            "2月13日(金)_⑤回目",
        ],
        "execution_time": datetime.datetime(2026, 2, 13, 17, 0),
        "today": datetime.date(2026, 2, 13),
    },
]


# ============================================
# VBAデータ読み込み（複数フォルダ対応）
# ============================================

def load_vba_data(vba_folders: list[str]) -> list[dict]:
    """複数のVBA回答書フォルダからデータを読み込み。
    同一顧客が複数フォルダにある場合は後のフォルダで上書き。
    """
    # 顧客名→行リスト のマップで管理（上書き用）
    customer_data: dict[str, list[dict]] = {}

    for folder_name in vba_folders:
        folder_path = VBA_BASE_FOLDER / folder_name
        if not folder_path.exists():
            print(f"  警告: フォルダが見つかりません: {folder_name}")
            continue

        vba_files = list(folder_path.glob("*.xlsx"))
        for vba_file in vba_files:
            if vba_file.name.startswith("~$"):
                continue
            customer_name = extract_customer_name_from_filename(vba_file.name)
            file_data = read_vba_excel(vba_file, customer_name)
            if file_data:
                # 同一顧客は上書き（後のフォルダ優先）
                norm_name = normalize_customer_name(customer_name)
                customer_data[norm_name] = file_data

    # 全顧客のデータをフラットに
    result = []
    for rows in customer_data.values():
        result.extend(rows)
    return result


# ============================================
# 受注データの読み込みとOrderRow変換
# ============================================

def load_orders(source_file: str) -> tuple[list[OrderRow], object, object]:
    """受注データとraw dataを読み込み。

    Returns:
        (orders, source_data_raw, cols)
    """
    source_path = SOURCE_FOLDER / source_file
    source_data_raw = load_source_file(str(source_path))
    cols = get_column_positions(source_data_raw)

    if cols is None:
        print(f"  エラー: ヘッダー行の列位置を検出できませんでした: {source_file}")
        return [], source_data_raw, cols

    orders: list[OrderRow] = []
    for i in get_data_rows_range(source_data_raw, cols):
        if is_data_row(source_data_raw, i, cols):
            orders.append(parse_order_row(source_data_raw, i, cols))

    return orders, source_data_raw, cols


# ============================================
# マッチングとPython計算
# ============================================

def build_order_index(orders: list[OrderRow]) -> dict[str, list[OrderRow]]:
    """注番でインデックスを構築。注番→明細リスト。"""
    index: dict[str, list[OrderRow]] = {}
    for row in orders:
        key = row.order_number.strip()
        if key not in index:
            index[key] = []
        index[key].append(row)
    return index


def match_vba_to_python(
    vba_item: dict,
    order_index: dict[str, list[OrderRow]],
    used_details: set[str],
) -> Optional[OrderRow]:
    """VBA行に対応するOrderRowを見つける。

    弊社注番で注番を特定し、品名で明細を特定。
    used_detailsで同一明細の重複マッチを防止。
    """
    order_number = vba_item["弊社注番"]
    if not order_number:
        return None

    candidates = order_index.get(order_number, [])
    if not candidates:
        return None

    vba_product = vba_item["品名_正規化"]

    # 品名マッチで明細を特定
    best_match = None
    for row in candidates:
        detail_key = f"{row.order_number}|{row.detail_number}"
        if detail_key in used_details:
            continue

        py_product = normalize_product_name(row.product_name)
        if py_product == vba_product:
            best_match = row
            break

    # 完全一致がなければ先頭一致で試行
    if best_match is None:
        for row in candidates:
            detail_key = f"{row.order_number}|{row.detail_number}"
            if detail_key in used_details:
                continue

            py_product = normalize_product_name(row.product_name)
            # 先頭20文字で比較
            if py_product[:20] == vba_product[:20] and vba_product[:20]:
                best_match = row
                break

    if best_match is not None:
        detail_key = f"{best_match.order_number}|{best_match.detail_number}"
        used_details.add(detail_key)

    return best_match


def compute_python_answer(
    row: OrderRow,
    cache,
    holidays,
    branch,
    execution_time: datetime.datetime,
    today: datetime.date,
) -> str:
    """Pythonで納期回答を計算する。送付履歴なし（sent_orders={}）。"""
    if not _pass_basic_filter(row):
        return ""

    history_key = f"{row.order_number}|{row.detail_number}"

    # フラグ算出（sent_orders={}で呼ぶ）
    force_delivered, is_himozuki, is_bunno_completed = _determine_flags(
        row, cache, {}, history_key, ""
    )

    # 納期計算
    report_row, delivery_status = build_report_row(
        row, cache, holidays, branch, execution_time,
        force_delivered, today
    )

    # force_delivered時の上書き
    if force_delivered and delivery_status in ("確認中", "欠品中", "日程調整中"):
        delivery_status = "納品済み"
    elif force_delivered and "（欠品）" in delivery_status:
        delivery_status = "納品済み"

    # 分納完了
    if is_bunno_completed:
        delivery_status = "納品済み"

    return delivery_status


# ============================================
# 比較ルールに基づく除外判定
# ============================================

SKIP_STATUSES = {"確認中", "欠品中", "日程調整中"}


def should_skip_comparison(vba_answer: str, py_answer: str) -> bool:
    """VBA/Pythonどちらかが除外対象ステータスなら比較しない。"""
    vba_norm = normalize_delivery_answer(vba_answer)
    py_norm = normalize_delivery_answer(py_answer)

    if vba_norm[0] in SKIP_STATUSES:
        return True
    if py_norm[0] in SKIP_STATUSES:
        return True

    # 分納も除外
    if vba_answer.strip() == "分納" or py_answer.strip() == "分納":
        return True

    return False


def classify_mismatch(vba_answer: str, py_answer: str, reason: str) -> str:
    """不一致の種別を判定。"""
    if "日付" in reason:
        vba_norm = normalize_delivery_answer(vba_answer)
        py_norm = normalize_delivery_answer(py_answer)
        # 納品済み vs 具体的日付
        if vba_norm[0] == "納品済み" or py_norm[0] == "納品済み":
            return "納品済みvs日付"
        return "日付不一致"
    if "区分" in reason:
        return "配達/出荷不一致"
    return "その他"


# ============================================
# 1回の実行を処理
# ============================================

def process_one_run(
    run_config: dict,
    cache,
    holidays,
    branch,
) -> tuple[list[dict], int]:
    """1回の実行を処理して(比較結果リスト, 未検出件数)を返す。"""
    label = run_config["label"]
    execution_time = run_config["execution_time"]
    today = run_config["today"]

    print(f"\n{'='*60}")
    print(f"実行回: {label}")
    print(f"  受注データ: {run_config['source_file']}")
    print(f"  VBAフォルダ: {run_config['vba_folders']}")
    print(f"  execution_time: {execution_time}")
    print(f"{'='*60}")

    # 受注データ読み込み
    orders, source_data_raw, cols = load_orders(run_config["source_file"])
    if not orders:
        print(f"  受注データなし。スキップ。")
        return [], 0
    print(f"  受注データ: {len(orders)}件")

    # 注番インデックス構築
    order_index = build_order_index(orders)

    # VBA回答書読み込み
    vba_data = load_vba_data(run_config["vba_folders"])
    print(f"  VBA回答書: {len(vba_data)}件")

    if not vba_data:
        print(f"  VBAデータなし。スキップ。")
        return [], 0

    # 比較
    results = []
    used_details: set[str] = set()
    matched = 0
    skipped = 0
    not_found = 0

    for vba_item in vba_data:
        row = match_vba_to_python(vba_item, order_index, used_details)

        if row is None:
            not_found += 1
            if not_found <= 3:
                print(f"  マッチなし: 注番={vba_item['弊社注番']} 品名={vba_item['品名'][:20]}")
            continue

        # Python計算
        py_answer = compute_python_answer(
            row, cache, holidays, branch, execution_time, today
        )

        vba_answer = vba_item["納期回答"]

        # 除外判定
        if should_skip_comparison(vba_answer, py_answer):
            skipped += 1
            result_entry = _build_result_entry(
                label, row, vba_item, vba_answer, py_answer,
                is_match=None, reason="除外", mismatch_type="除外"
            )
            results.append(result_entry)
            continue

        # 比較
        is_match, reason = compare_answers(vba_answer, py_answer)
        matched += 1

        if is_match:
            result_entry = _build_result_entry(
                label, row, vba_item, vba_answer, py_answer,
                is_match=True, reason="", mismatch_type=""
            )
        else:
            mismatch_type = classify_mismatch(vba_answer, py_answer, reason)
            result_entry = _build_result_entry(
                label, row, vba_item, vba_answer, py_answer,
                is_match=False, reason=reason, mismatch_type=mismatch_type
            )

        results.append(result_entry)

    mismatch_count = sum(1 for r in results if r["一致"] is False)
    empty_l = sum(1 for v in vba_data if not v["弊社注番"])
    print(f"  比較: マッチ={matched}, 除外={skipped}, 未検出={not_found}(L列空{empty_l}), 不一致={mismatch_count}")

    return results, not_found


def _build_result_entry(
    label: str,
    row: OrderRow,
    vba_item: dict,
    vba_answer: str,
    py_answer: str,
    is_match: Optional[bool],
    reason: str,
    mismatch_type: str,
) -> dict:
    """比較結果の1行を構築。"""
    # 伝票タイプの簡略化
    doc_type = row.document_type.strip()
    if doc_type == "【受注】直送販売":
        storage = row.storage_place.strip()
        if storage == "転送中（直送用）":
            type_label = "直送"
        else:
            type_label = "紐付き"
    elif doc_type == "【受注】在庫販売":
        type_label = "在庫販売"
    else:
        type_label = doc_type

    # 指定納期・受注納期の表示
    spec_date = ""
    if row.specified_delivery_date:
        spec_date = row.specified_delivery_date.strftime("%m/%d")
    order_date = ""
    if row.order_delivery_date:
        order_date = row.order_delivery_date.strftime("%m/%d")

    return {
        "実行回": label,
        "受注日": row.registration_date,
        "伝票番号": row.order_number,
        "明細": row.detail_number,
        "顧客": vba_item["顧客"],
        "品名": vba_item["品名"][:30],
        "VBA回答": vba_answer,
        "Python回答": py_answer,
        "一致": is_match,
        "不一致種別": mismatch_type,
        "不一致詳細": reason,
        "伝票タイプ": type_label,
        "出荷ステータス": row.ship_status,
        "指定納期": spec_date,
        "受注納期": order_date,
        "判定": "",  # 手動判定用（空白）
    }


# ============================================
# Excel出力
# ============================================

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
MISMATCH_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
SKIP_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")


def write_header_row(ws, row_idx: int, headers: list[str]):
    """ヘッダー行を書き込み。"""
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def write_output(all_results: list[dict], not_found_map: dict[str, int], output_path: Path):
    """Excel 3シートに出力。not_found_map: 実行回ラベル→未検出件数。"""
    wb = Workbook()

    # --- シート1: 不一致一覧 ---
    ws_diff = wb.active
    ws_diff.title = "不一致一覧"

    diff_headers = [
        "実行回", "受注日", "伝票番号", "明細", "顧客", "品名",
        "VBA回答", "Python回答", "不一致種別", "不一致詳細",
        "伝票タイプ", "出荷ステータス", "指定納期", "受注納期", "判定",
    ]
    write_header_row(ws_diff, 1, diff_headers)

    mismatches = [r for r in all_results if r["一致"] is False]
    for row_idx, r in enumerate(mismatches, 2):
        ws_diff.cell(row=row_idx, column=1, value=r["実行回"])
        ws_diff.cell(row=row_idx, column=2, value=r["受注日"])
        ws_diff.cell(row=row_idx, column=3, value=r["伝票番号"])
        ws_diff.cell(row=row_idx, column=4, value=r["明細"])
        ws_diff.cell(row=row_idx, column=5, value=r["顧客"])
        ws_diff.cell(row=row_idx, column=6, value=r["品名"])
        ws_diff.cell(row=row_idx, column=7, value=r["VBA回答"])
        ws_diff.cell(row=row_idx, column=8, value=r["Python回答"])
        ws_diff.cell(row=row_idx, column=9, value=r["不一致種別"])
        ws_diff.cell(row=row_idx, column=10, value=r["不一致詳細"])
        ws_diff.cell(row=row_idx, column=11, value=r["伝票タイプ"])
        ws_diff.cell(row=row_idx, column=12, value=r["出荷ステータス"])
        ws_diff.cell(row=row_idx, column=13, value=r["指定納期"])
        ws_diff.cell(row=row_idx, column=14, value=r["受注納期"])
        ws_diff.cell(row=row_idx, column=15, value=r["判定"])
        # 行に色を付ける
        for col in range(1, 16):
            ws_diff.cell(row=row_idx, column=col).fill = MISMATCH_FILL

    # 列幅調整
    col_widths_diff = [10, 12, 12, 6, 25, 30, 20, 20, 15, 30, 10, 12, 10, 10, 10]
    for i, w in enumerate(col_widths_diff):
        col_letter = chr(65 + i)
        ws_diff.column_dimensions[col_letter].width = w

    # --- シート2: サマリー ---
    ws_summary = wb.create_sheet("サマリー")

    summary_headers = [
        "実行回", "VBA件数", "比較対象", "除外", "未検出",
        "一致", "不一致", "一致率",
        "日付不一致", "配達/出荷不一致", "納品済みvs日付", "その他",
    ]
    write_header_row(ws_summary, 1, summary_headers)

    # 実行回ごとに集計
    run_labels = [r["label"] for r in RUNS]
    row_idx = 2
    total_compared = 0
    total_match = 0
    total_mismatch = 0

    for label in run_labels:
        run_results = [r for r in all_results if r["実行回"] == label]
        not_found_count = not_found_map.get(label, 0)
        vba_count = len(run_results) + not_found_count
        excluded = sum(1 for r in run_results if r["一致"] is None)
        compared = sum(1 for r in run_results if r["一致"] is not None)
        match_count = sum(1 for r in run_results if r["一致"] is True)
        mismatch_count = sum(1 for r in run_results if r["一致"] is False)

        total_compared += compared
        total_match += match_count
        total_mismatch += mismatch_count

        match_rate = f"{match_count / compared * 100:.1f}%" if compared > 0 else "N/A"

        # 不一致内訳
        date_diff = sum(1 for r in run_results if r["不一致種別"] == "日付不一致")
        type_diff = sum(1 for r in run_results if r["不一致種別"] == "配達/出荷不一致")
        delivered_diff = sum(1 for r in run_results if r["不一致種別"] == "納品済みvs日付")
        other_diff = sum(1 for r in run_results if r["不一致種別"] == "その他")

        ws_summary.cell(row=row_idx, column=1, value=label)
        ws_summary.cell(row=row_idx, column=2, value=vba_count)
        ws_summary.cell(row=row_idx, column=3, value=compared)
        ws_summary.cell(row=row_idx, column=4, value=excluded)
        ws_summary.cell(row=row_idx, column=5, value=not_found_count)
        ws_summary.cell(row=row_idx, column=6, value=match_count)
        ws_summary.cell(row=row_idx, column=7, value=mismatch_count)
        ws_summary.cell(row=row_idx, column=8, value=match_rate)
        ws_summary.cell(row=row_idx, column=9, value=date_diff)
        ws_summary.cell(row=row_idx, column=10, value=type_diff)
        ws_summary.cell(row=row_idx, column=11, value=delivered_diff)
        ws_summary.cell(row=row_idx, column=12, value=other_diff)
        row_idx += 1

    # 合計行
    total_rate = f"{total_match / total_compared * 100:.1f}%" if total_compared > 0 else "N/A"
    ws_summary.cell(row=row_idx, column=1, value="合計")
    ws_summary.cell(row=row_idx, column=3, value=total_compared)
    ws_summary.cell(row=row_idx, column=6, value=total_match)
    ws_summary.cell(row=row_idx, column=7, value=total_mismatch)
    ws_summary.cell(row=row_idx, column=8, value=total_rate)
    for col in range(1, 13):
        ws_summary.cell(row=row_idx, column=col).font = Font(bold=True)

    # 列幅
    summary_widths = [10, 10, 10, 8, 8, 8, 8, 10, 12, 16, 16, 8]
    for i, w in enumerate(summary_widths):
        col_letter = chr(65 + i)
        ws_summary.column_dimensions[col_letter].width = w

    # --- シート3: 全件 ---
    ws_all = wb.create_sheet("全件")

    all_headers = [
        "実行回", "受注日", "伝票番号", "明細", "顧客", "品名",
        "VBA回答", "Python回答", "結果", "不一致種別", "不一致詳細",
        "伝票タイプ", "出荷ステータス", "指定納期", "受注納期",
    ]
    write_header_row(ws_all, 1, all_headers)

    for row_idx, r in enumerate(all_results, 2):
        if r["一致"] is True:
            result_label = "一致"
        elif r["一致"] is False:
            result_label = "不一致"
        else:
            result_label = "除外"

        ws_all.cell(row=row_idx, column=1, value=r["実行回"])
        ws_all.cell(row=row_idx, column=2, value=r["受注日"])
        ws_all.cell(row=row_idx, column=3, value=r["伝票番号"])
        ws_all.cell(row=row_idx, column=4, value=r["明細"])
        ws_all.cell(row=row_idx, column=5, value=r["顧客"])
        ws_all.cell(row=row_idx, column=6, value=r["品名"])
        ws_all.cell(row=row_idx, column=7, value=r["VBA回答"])
        ws_all.cell(row=row_idx, column=8, value=r["Python回答"])
        ws_all.cell(row=row_idx, column=9, value=result_label)
        ws_all.cell(row=row_idx, column=10, value=r["不一致種別"])
        ws_all.cell(row=row_idx, column=11, value=r["不一致詳細"])
        ws_all.cell(row=row_idx, column=12, value=r["伝票タイプ"])
        ws_all.cell(row=row_idx, column=13, value=r["出荷ステータス"])
        ws_all.cell(row=row_idx, column=14, value=r["指定納期"])
        ws_all.cell(row=row_idx, column=15, value=r["受注納期"])

        # 不一致行に色
        if r["一致"] is False:
            for col in range(1, 16):
                ws_all.cell(row=row_idx, column=col).fill = MISMATCH_FILL
        elif r["一致"] is None:
            for col in range(1, 16):
                ws_all.cell(row=row_idx, column=col).fill = SKIP_FILL

    # 列幅
    all_widths = [10, 12, 12, 6, 25, 30, 20, 20, 8, 15, 30, 10, 12, 10, 10]
    for i, w in enumerate(all_widths):
        col_letter = chr(65 + i)
        ws_all.column_dimensions[col_letter].width = w

    wb.save(str(output_path))
    print(f"\n結果を保存しました: {output_path}")


# ============================================
# メイン
# ============================================

def main():
    print("=" * 80)
    print("VBA vs Python 1週間比較検証（2/10〜2/13）")
    print("=" * 80)

    # マスターファイル読み込み（全実行回で共通）
    print("\nマスターファイルを読み込み中...")
    mfg_path = TOOL_FOLDER / "メーカー一覧.xlsx"
    cust_path = TOOL_FOLDER / "顧客マスター_v2.xlsm"

    mfg_wb = load_workbook(str(mfg_path), data_only=True)
    cust_wb = load_workbook(str(cust_path), data_only=True)

    try:
        confirming_ws = cust_wb["確認中一覧"]
    except KeyError:
        confirming_ws = None

    # キャッシュ構築に必要なsource_data — 最初の実行回のデータで構築
    # （ストレージキャッシュは実行回ごとに更新する必要があるが、
    #  メーカー・顧客キャッシュは共通）
    first_source = SOURCE_FOLDER / RUNS[0]["source_file"]
    first_raw = load_source_file(str(first_source))
    first_cols = get_column_positions(first_raw)

    cache = build_all_caches(mfg_wb, cust_wb, confirming_ws, first_raw, first_cols)
    branch = load_branch_settings(mfg_wb, first_raw, first_cols)
    holidays = load_holidays(mfg_wb)

    print(f"  メーカーキャッシュ: {len(cache.mfg_name)}件")
    print(f"  顧客キャッシュ: {len(cache.cust_days)}件")
    print(f"  確認中キャッシュ: {len(cache.confirm)}件")

    # 各実行回を処理
    all_results: list[dict] = []
    not_found_map: dict[str, int] = {}

    for run_config in RUNS:
        # ストレージキャッシュは実行回ごとに更新
        source_path = SOURCE_FOLDER / run_config["source_file"]
        raw = load_source_file(str(source_path))
        cols = get_column_positions(raw)
        if cols is not None:
            from nouki_kaitou.cache import build_storage_cache
            cache.storage = build_storage_cache(raw, cols)

        results, not_found = process_one_run(run_config, cache, holidays, branch)
        all_results.extend(results)
        not_found_map[run_config["label"]] = not_found

    # 結果出力
    output_path = Path(__file__).parent / "comparison_week_result.xlsx"
    write_output(all_results, not_found_map, output_path)

    # サマリー表示
    total = len(all_results)
    mismatches = [r for r in all_results if r["一致"] is False]
    matches = [r for r in all_results if r["一致"] is True]
    excluded = [r for r in all_results if r["一致"] is None]

    print(f"\n{'='*60}")
    print(f"最終結果サマリー")
    print(f"{'='*60}")
    print(f"  全件: {total}")
    print(f"  一致: {len(matches)}")
    print(f"  不一致: {len(mismatches)}")
    print(f"  除外: {len(excluded)}")

    if matches or mismatches:
        compared = len(matches) + len(mismatches)
        rate = len(matches) / compared * 100
        print(f"  一致率: {rate:.1f}% ({len(matches)}/{compared})")

    if mismatches:
        print(f"\n--- 不一致の先頭10件 ---")
        for r in mismatches[:10]:
            print(f"  [{r['実行回']}] {r['伝票番号']}|{r['明細']}")
            print(f"    顧客: {r['顧客']}")
            print(f"    品名: {r['品名']}")
            print(f"    VBA: {r['VBA回答']}  Python: {r['Python回答']}")
            print(f"    種別: {r['不一致種別']}  {r['不一致詳細']}")
            print()

    mfg_wb.close()
    cust_wb.close()


if __name__ == "__main__":
    main()
