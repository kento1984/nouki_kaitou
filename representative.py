"""担当者マスター参照モジュール

VBAの以下の関数を移植:
- IsSplitByRep (L7614): 担当者マスター登録有無判定
- GetRepList (L7635): 担当者名リスト取得
- GetRepEmailAddresses (L7673): 担当者メールアドレス取得
- ShouldIncludeForRep (L7719): 担当者フィルタ判定
- ParseRepNames (L7559): 担当者名分割（区切り文字対応）
- ContainsRep (L7595): 担当者名含有判定
"""

from __future__ import annotations

from typing import TYPE_CHECKING

if TYPE_CHECKING:
    import openpyxl.worksheet.worksheet as ws_type


# ============================================
# VBA: ParseRepNames (L7559-7590)
# 担当者名文字列を分割してリストにする
# ============================================
def parse_rep_names(cell_value: str) -> list[str]:
    """担当者名文字列を分割してリストにする。

    区切り文字: 「、」「・」「,」
    「様」も区切り文字として扱う（「柏原様首藤様」→ ["柏原", "首藤"]）

    Args:
        cell_value: セルの値（例: "田中、鈴木"、"柏原様首藤様"）

    Returns:
        担当者名のリスト
    """
    text = cell_value.strip()
    if not text:
        return []

    # 区切り文字を半角カンマに統一
    text = text.replace("、", ",")
    text = text.replace("・", ",")
    # 「様」も区切り文字として扱う
    text = text.replace("様", ",")

    parts = text.split(",")
    return [p.strip() for p in parts if p.strip()]


# ============================================
# VBA: ContainsRep (L7595-7608)
# リスト内に指定の担当者名が含まれるか判定
# ============================================
def contains_rep(rep_names: list[str], target_rep: str) -> bool:
    """リスト内に指定の担当者名が含まれるかを判定する。

    Args:
        rep_names: 担当者名リスト
        target_rep: 検索対象の担当者名

    Returns:
        True = 含まれている
    """
    return target_rep in rep_names


# ============================================
# VBA: IsSplitByRep (L7614-7630)
# 担当者マスターにその顧客が存在するか判定
# ============================================
def is_split_by_rep(
    customer_name: str,
    rep_master_ws: object,
) -> bool:
    """担当者マスターにその顧客が登録されているかを判定する。

    登録されていれば担当者別分割送信が有効。

    Args:
        customer_name: 顧客名
        rep_master_ws: 担当者マスターシート

    Returns:
        True = 担当者別分割あり
    """
    if rep_master_ws is None:
        return False

    for row in rep_master_ws.iter_rows(min_row=2, values_only=True):
        name = str(row[0]).strip() if row[0] else ""
        if name == customer_name:
            return True

    return False


# ============================================
# VBA: GetRepList (L7635-7667)
# 指定顧客の担当者名リストを取得（重複排除）
# ============================================
def get_rep_list(
    customer_name: str,
    rep_master_ws: object,
) -> list[str]:
    """指定顧客の担当者名リストを取得する。

    末尾の「様」は除去し、重複を排除。

    Args:
        customer_name: 顧客名
        rep_master_ws: 担当者マスターシート

    Returns:
        担当者名のリスト（重複なし、順序保持）
    """
    result: list[str] = []
    seen: set[str] = set()

    if rep_master_ws is None:
        return result

    for row in rep_master_ws.iter_rows(min_row=2, values_only=True):
        name = str(row[0]).strip() if row[0] else ""
        if name != customer_name:
            continue

        rep_name = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        # 末尾の「様」を除去
        if len(rep_name) > 1 and rep_name.endswith("様"):
            rep_name = rep_name[:-1]

        if rep_name and rep_name not in seen:
            seen.add(rep_name)
            result.append(rep_name)

    return result


# ============================================
# VBA: GetRepEmailAddresses (L7673-7711)
# 担当者のメールアドレスを取得
# ============================================
def get_rep_email_addresses(
    customer_name: str,
    rep_name: str,
    rep_master_ws: object,
) -> str:
    """担当者マスターから指定顧客・担当者のメールアドレスを取得する。

    C列以降のメールアドレスを「; 」区切りで返す。

    Args:
        customer_name: 顧客名
        rep_name: 担当者名
        rep_master_ws: 担当者マスターシート

    Returns:
        メールアドレス（「; 」区切り）。見つからなければ空文字。
    """
    if rep_master_ws is None:
        return ""

    for row in rep_master_ws.iter_rows(min_row=2, values_only=True):
        name = str(row[0]).strip() if row[0] else ""
        if name != customer_name:
            continue

        master_rep = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        # 末尾の「様」を除去
        if len(master_rep) > 1 and master_rep.endswith("様"):
            master_rep = master_rep[:-1]

        if master_rep != rep_name:
            continue

        # C列（0-indexed: 2）以降
        emails: list[str] = []
        for j in range(2, len(row)):
            addr = str(row[j]).strip() if row[j] else ""
            if addr:
                emails.append(addr)

        return "; ".join(emails)

    return ""


# ============================================
# VBA: ShouldIncludeForRep (L7719-7748)
# 担当者フィルタ判定
# ============================================
def should_include_for_rep(
    cell_rep_value: str,
    rep_name: str,
    registered_rep_list: list[str],
) -> bool:
    """伝票の担当者が指定担当者のフィルタに合致するか判定する。

    - rep_name="" → 常にTrue（フィルタなし）
    - rep_name="__OTHER__" → 空欄 or 未登録担当者がいればTrue
    - rep_name=具体名 → セルにその名前が含まれていればTrue

    Args:
        cell_rep_value: セルの担当者名値（例: "田中、鈴木"）
        rep_name: フィルタ対象の担当者名
        registered_rep_list: 登録済み担当者名リスト

    Returns:
        True = フィルタに合致
    """
    # フィルタなし
    if not rep_name:
        return True

    cell_reps = parse_rep_names(cell_rep_value)

    if rep_name == "__OTHER__":
        # 空欄 → その他に含める
        if not cell_reps:
            return True

        # 未登録の担当者が1人でもいれば含める
        for rep in cell_reps:
            if not contains_rep(registered_rep_list, rep):
                return True
        return False

    # 特定担当者
    return contains_rep(cell_reps, rep_name)
