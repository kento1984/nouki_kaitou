"""TWF2026 社内手配管理表（チェックリスト台帳）ジェネレーター

納期回答書ツールとは独立した別スクリプト。main.py の本番フローには
組み込まない（手動実行のワンショットツール）。

目的:
    東京ウェルディングフェスタ（TWF2026）の展示会受注を、担当者が
    SAP上で1件ずつ手配（メーカー発注／在庫受注の計上）していくための
    チェックリスト台帳を Excel（.xlsx）で生成する。納期は納期回答書側で
    管理するため、この台帳には載せない。

データソース:
    SAP受注一覧（0612pm.txt 等）。TWF判定・お客様名抽出は twf.py を流用。
    受注はもう増えない前提（展示会終了済み）。

作り:
    - テーブル（ListObject）＋オートフィルタ＋ステータスのドロップダウン。
    - スライサーは Python では作らない（openpyxl 非対応）。人が後から手で
      挿入できるよう、テーブルの左（A列）と上（行1〜7）に余白を確保する。
    - 既存台帳から手入力（ステータス・備考）を注番+明細キーで引き継ぐ。
    - 出力はタイムスタンプ付き別名。正本は直接上書きしない。

使い方:
    python twf_ledger.py <SAP受注一覧.txt> [--carryover <既存台帳.xlsx>]
                                          [--outdir <出力先>]
"""

from __future__ import annotations

import argparse
import datetime
import sys
from dataclasses import dataclass
from pathlib import Path

# リポジトリ直下から `python twf_ledger.py` でも動くようにパッケージ親をパス追加
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

from nouki_kaitou.cache import build_all_caches, build_manufacturer_cache
from nouki_kaitou.config import load_branch_settings, load_holidays
from nouki_kaitou.data_loader import (
    get_column_positions,
    get_data_rows_range,
    is_data_row,
    load_source_file,
    parse_order_row,
)
from nouki_kaitou.history import CONFIRMING_SHEET_NAME
from nouki_kaitou.models import OrderRow
from nouki_kaitou.report_generator import build_report_row
from nouki_kaitou.twf import (
    build_twf_info_map,
    collect_twf_orders,
    parse_twf_comment,
    twf_sort_key,
)
from nouki_kaitou.utils import is_file_open, normalize_item_group_code

# メーカー名が空（在庫販売）で品目Groupマスターでも引けなかったときの表示。
# 空欄だとデータ抜けに見えるため、マスター未登録であることを明示する。
MANUFACTURER_UNKNOWN = "（要確認）"

# ============================================
# ステータス選択肢（ドロップダウン）
# ============================================
STATUS_CHOICES = [
    "未着手",
    "メーカー手配済み",
    "在庫計上済み",
    "保留",
    "その他",
]
STATUS_DEFAULT = "未着手"

# 回答納期の計算がこの割合以上の明細で失敗したら、設定ミス・マスター不整合・
# 実装バグの疑いとして取込を中止する（全件空欄を「成功」に見せないため）。
DELIVERY_FAIL_ABORT_RATIO = 0.5

# ============================================
# 手配区分の判定（直送 / 紐付き / 在庫販売）
# delivery_calc.py と同じ判定基準（保管場所「転送中（直送用）」=直送）
# ============================================
TENSOUCHU = "転送中（直送用）"


def classify_tehai(row: OrderRow) -> str:
    """伝票タイプ・保管場所から手配区分を返す。"""
    dt = row.document_type
    sp = row.storage_place.strip()
    if dt == "【受注】在庫販売":
        return "在庫販売"
    if dt == "【受注】直送販売":
        return "直送" if sp == TENSOUCHU else "紐付き"
    # 想定外の伝票タイプはそのまま表示（目視確認用）
    return dt.replace("【受注】", "").strip() or "不明"


TWF_NO_WIDTH = 6
"""TWF No. の標準桁数（実データの98%が6桁。台帳表示のゼロ埋め基準）。"""


def format_twf_no(number: str) -> str:
    """台帳表示用にTWF No.を桁揃えする（台帳のみの整形。回答書には適用しない）。

    純粋な数字で6桁未満のものだけ6桁ゼロ埋め（例「0014」→「000014」）。
    ？を含む番号（既に6桁ぶん埋まっている）・「不明」・空・6桁以上はそのまま。
    """
    s = number.strip()
    if s.isdigit() and len(s) < TWF_NO_WIDTH:
        return s.zfill(TWF_NO_WIDTH)
    return s


def parse_quantity(value: str) -> int | float | str | None:
    """数量を数値型に変換する（Excelの「文字列保存」エラーマーク対策）。

    - 「10」→ int 10、「1.00」「800.00」→ 整数値なら int、「1.5」→ float
    - 桁区切りカンマ（「1,000」）は除去して数値化
    - 空欄 → None（セルは空欄）
    - 非数値（「未定」等）→ そのまま文字列で返す（落とさない）
    """
    s = str(value).strip().replace(",", "")
    if not s:
        return None
    try:
        f = float(s)
    except ValueError:
        return str(value).strip()  # 非数値は文字列のまま
    return int(f) if f.is_integer() else f


def resolve_manufacturer(
    row: OrderRow, mfg_name: dict[str, str] | None
) -> str:
    """台帳に載せるメーカー名を返す。

    SAP「名称」列を最優先。在庫販売は仕様上ここが空になるため、空のときだけ
    品目Groupマスター（メーカー一覧.xlsx）で補完する。直送・紐付きは名称が
    入るので実質そのまま。補完できないときは「（要確認）」を返す
    （空欄だとデータ抜けに見えるため）。

    Args:
        mfg_name: 品目GroupCode → メーカー名 の辞書。None ならマスター未提供と
            みなし、補完せず空欄のまま返す。
    """
    name = row.manufacturer_name.strip()
    if name:
        return name
    if mfg_name is None:
        return ""  # マスター未提供なら補完判断できないので空のまま
    code = normalize_item_group_code(row.item_group_code)
    backfilled = mfg_name.get(code, "").strip() if code else ""
    return backfilled or MANUFACTURER_UNKNOWN


# ============================================
# 台帳の1行
# ============================================
@dataclass
class LedgerRow:
    twf_no: str = ""        # TWF No.（番号のみ。「不明」もあり得る）
    order_number: str = ""  # 注番
    detail_number: str = "" # 明細
    ship_dealer: str = ""   # 受注先（販売店＝SAP取引先。customer_name由来）
    customer: str = ""      # ユーザー名（コメント由来のエンドユーザー名）
    manufacturer: str = ""  # メーカー名
    product: str = ""       # 品名
    quantity: str = ""      # 数量
    tehai: str = ""         # 手配区分（直送/紐付き/在庫販売）
    delivery_answer: str = ""  # 回答納期（納期回答書と同じ算出。delivery_ctx指定時のみ）
    status: str = STATUS_DEFAULT  # ステータス（手入力・引き継ぎ対象）
    note: str = ""          # 備考（手入力・引き継ぎ対象）
    rep_name: str = ""      # 担当者（マツモト担当者名。スライサー絞り込み用・最右端）


# 台帳の列定義（表示順・ヘッダー・幅）
# (属性名, ヘッダー, 幅)
# 受注先（販売店）を注番の近くに追加、ユーザー名と並べて役割を区別。
# 担当者はスライサーで絞る前提のためほぼ見ない情報として最右端に配置。
COLUMNS: list[tuple[str, str, int]] = [
    ("twf_no", "TWF No.", 9),
    ("order_number", "注番", 13),
    ("detail_number", "明細", 6),
    ("ship_dealer", "受注先（販売店）", 28),
    ("customer", "ユーザー名", 24),
    ("manufacturer", "メーカー名", 22),
    ("product", "品名", 32),
    ("quantity", "数量", 8),
    ("tehai", "手配区分", 10),
    ("status", "ステータス", 14),
    ("note", "備考", 26),
    ("rep_name", "担当者", 12),
]


# ============================================
# 台帳行の構築
# ============================================
def build_ledger_rows(
    orders: list[OrderRow],
    mfg_name: dict[str, str] | None = None,
    delivery_ctx: tuple | None = None,
) -> list[LedgerRow]:
    """全受注からTWF展示会受注を抽出し、台帳行に変換する。

    - TWF判定・注番伝播は collect_twf_orders に従う。
    - お客様名・TWF No. はコメント由来（parse_twf_comment）。入れ忘れ明細は
      build_twf_info_map で同一注番から番号・お客様名を引き継ぐ。
    - メーカー名は SAP「名称」優先、空（在庫販売）のときだけ品目Groupマスター
      （mfg_name）で補完。補完不可は「（要確認）」（resolve_manufacturer）。
    - 並び順は twf_sort_key（TWF No.昇順 → 注番→明細順、番号なし・不明は末尾）。

    Args:
        delivery_ctx: build_delivery_context() の戻り値
            (CacheStore, holidays, branch)。指定すると各明細に納期回答書と同じ
            回答納期（delivery_answer）を入れる。None なら回答納期は空のまま。
    """
    twf_orders, _ = collect_twf_orders(orders)
    info_map = build_twf_info_map(orders)

    cache = holidays = branch = None
    execution_time = today = None
    if delivery_ctx is not None:
        cache, holidays, branch = delivery_ctx
        # P1-1: 実行時刻・基準日は取込開始時に1回だけ固定し、全明細で共有する。
        # 行ごとに datetime.now() を取ると、締切時刻付近・日跨ぎ・紐付き処理完了
        # （execution_time.hour で分岐）で明細間の回答がブレるため。
        execution_time = datetime.datetime.now()
        today = datetime.date.today()
    delivery_attempts = 0
    delivery_errors = 0
    delivery_fail_samples: list[str] = []

    rows: list[LedgerRow] = []
    for o in orders:
        onum = o.order_number.strip()
        if onum not in twf_orders:
            continue

        # 明細固有のTWF記載を優先、なければ注番の代表情報を引き継ぐ
        info = parse_twf_comment(o.comment_detail) or info_map.get(onum)
        twf_no = format_twf_no(info.number) if info else ""
        customer = info.customer if info else ""

        # 回答納期（納期回答書と同じ build_report_row 経由。分納/欠品の上書き込み）。
        # 1明細の不正データで取込全体を止めないよう行単位で握るが、系統的失敗は
        # 後段の失敗率チェック（P1-4）で検知して中止する。
        delivery_answer = ""
        if cache is not None:
            delivery_attempts += 1
            try:
                _, delivery_answer = build_report_row(
                    o, cache, holidays, branch, execution_time, False, today
                )
            except Exception as e:  # noqa: BLE001 行データ起因を握る。系統失敗は下で停止
                delivery_errors += 1
                delivery_answer = ""
                if len(delivery_fail_samples) < 10:
                    delivery_fail_samples.append(
                        f"{onum}|{o.detail_number.strip()}:{type(e).__name__}"
                    )

        rows.append(
            LedgerRow(
                twf_no=twf_no,
                order_number=onum,
                detail_number=o.detail_number.strip(),
                ship_dealer=o.customer_name.strip(),
                customer=customer,
                manufacturer=resolve_manufacturer(o, mfg_name),
                product=o.product_name.strip(),
                quantity=o.quantity.strip(),
                tehai=classify_tehai(o),
                delivery_answer=delivery_answer,
                rep_name=o.rep_name.strip(),
            )
        )

    # P1-4: 系統的失敗（全件/高率）は設定ミス・マスター不整合・実装バグの疑い。
    # 全件空欄でも「成功」に見えるのを防ぐため、高率なら停止して気づけるようにする。
    if delivery_attempts and delivery_errors:
        ratio = delivery_errors / delivery_attempts
        sample = "  ".join(delivery_fail_samples)
        summary = (f"{delivery_errors}/{delivery_attempts}明細"
                   f"（{ratio:.0%}）  例: {sample}")
        if ratio >= DELIVERY_FAIL_ABORT_RATIO:
            raise SystemExit(
                "⚠ 回答納期の計算が大量に失敗しました（設定ミス・マスター不整合・"
                f"実装バグの疑い）。取込を中止します。\n  {summary}"
            )
        print(f"⚠ 回答納期の計算に失敗: {summary}（該当明細のみ空欄で続行）")

    rows.sort(
        key=lambda r: twf_sort_key(r.twf_no, r.order_number, r.detail_number)
    )
    return rows


# ============================================
# 既存台帳からの手入力引き継ぎ（注番+明細キー）
# ============================================
def _carryover_key(order_number: str, detail_number: str) -> str:
    return f"{order_number.strip()}|{detail_number.strip()}"


def read_existing_status(path: str | Path) -> dict[str, tuple[str, str]]:
    """既存台帳xlsxから (注番|明細) → (ステータス, 備考) を読み取る。

    ヘッダー行は「注番」「ステータス」を含むセルから動的に検出するため、
    多少レイアウトが変わっても引き継げる。見つからなければ空dict。
    """
    path = Path(path)
    if not path.exists():
        return {}

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # ヘッダー行とキー列を検出
    header_row = None
    col_idx: dict[str, int] = {}
    for r in range(1, min(ws.max_row, 30) + 1):
        labels = {
            str(ws.cell(r, c).value).strip(): c
            for c in range(1, ws.max_column + 1)
            if ws.cell(r, c).value is not None
        }
        if "注番" in labels and "ステータス" in labels:
            header_row = r
            col_idx = labels
            break
    if header_row is None:
        return {}

    c_order = col_idx.get("注番")
    c_detail = col_idx.get("明細")
    c_status = col_idx.get("ステータス")
    c_note = col_idx.get("備考")
    if c_order is None or c_detail is None or c_status is None:
        return {}

    result: dict[str, tuple[str, str]] = {}
    for r in range(header_row + 1, ws.max_row + 1):
        onum = ws.cell(r, c_order).value
        dnum = ws.cell(r, c_detail).value
        if onum is None:
            continue
        status = ws.cell(r, c_status).value if c_status else None
        note = ws.cell(r, c_note).value if c_note else None
        key = _carryover_key(str(onum), str(dnum) if dnum is not None else "")
        result[key] = (
            str(status).strip() if status is not None else "",
            str(note).strip() if note is not None else "",
        )
    return result


def apply_carryover(
    rows: list[LedgerRow], existing: dict[str, tuple[str, str]]
) -> int:
    """既存台帳の手入力（ステータス・備考）を新しい行に引き継ぐ。

    Returns:
        引き継いだ行数。
    """
    applied = 0
    for row in rows:
        key = _carryover_key(row.order_number, row.detail_number)
        if key in existing:
            status, note = existing[key]
            if status:
                row.status = status
            if note:
                row.note = note
            applied += 1
    return applied


# ============================================
# Excel 出力
# ============================================
# レイアウト: 左にA列の余白、上に行1〜7の余白を確保（スライサー手動挿入用）。
# テーブルはB列・8行目ヘッダーから開始する。
GUTTER_COL = 1          # A列（左余白）
TABLE_START_COL = 2     # B列
TITLE_ROW = 2
SLICER_BAND_ROWS = (4, 7)   # スライサー設置エリア（手動挿入用の空白帯）
HEADER_ROW = 8
DATA_START_ROW = 9

_THIN = Side(style="thin", color="B0B0B0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_FILL = PatternFill("solid", fgColor="305496")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=16, color="305496")
_NOTE_FONT = Font(italic=True, size=9, color="808080")
_SLICER_FILL = PatternFill("solid", fgColor="F2F2F2")


def write_ledger(rows: list[LedgerRow], out_path: str | Path) -> Path:
    """台帳をxlsxとして書き出す。"""
    out_path = Path(out_path)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TWF手配管理"

    last_col = TABLE_START_COL + len(COLUMNS) - 1
    last_col_letter = get_column_letter(last_col)

    # --- タイトル ---
    ws.cell(TITLE_ROW, TABLE_START_COL, "TWF2026 社内手配管理表").font = _TITLE_FONT
    ws.cell(
        TITLE_ROW + 1,
        TABLE_START_COL,
        f"展示会受注の社内手配チェックリスト（全{len(rows)}明細）"
        f"／生成: {datetime.datetime.now():%Y-%m-%d %H:%M}",
    ).font = _NOTE_FONT

    # --- スライサー設置エリア（手動挿入用の空白帯） ---
    band_top, band_bottom = SLICER_BAND_ROWS
    for r in range(band_top, band_bottom + 1):
        for c in range(TABLE_START_COL, last_col + 1):
            ws.cell(r, c).fill = _SLICER_FILL
    ws.cell(
        band_top,
        TABLE_START_COL,
        "▼ スライサー設置エリア：［挿入］→［スライサー］で 担当者／手配区分／"
        "ステータス を追加し、この灰色の帯に配置してください",
    ).font = _NOTE_FONT

    # --- ヘッダー行 ---
    for i, (_, header, _) in enumerate(COLUMNS):
        c = TABLE_START_COL + i
        cell = ws.cell(HEADER_ROW, c, header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # --- データ行 ---
    for ridx, row in enumerate(rows):
        r = DATA_START_ROW + ridx
        for i, (attr, _, _) in enumerate(COLUMNS):
            c = TABLE_START_COL + i
            # 数量は数値型で書き込む（文字列保存のエラーマーク対策）
            value = (
                parse_quantity(row.quantity)
                if attr == "quantity"
                else getattr(row, attr)
            )
            cell = ws.cell(r, c, value)
            cell.border = _BORDER
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=(attr in ("ship_dealer", "customer", "manufacturer",
                                    "product", "note")),
            )

    last_data_row = DATA_START_ROW + len(rows) - 1
    # データが0件でもテーブルには最低1行必要なのでヘッダー直下までを範囲にする
    table_last_row = max(last_data_row, DATA_START_ROW)

    # --- 列幅・余白列 ---
    ws.column_dimensions[get_column_letter(GUTTER_COL)].width = 2.5
    for i, (_, _, width) in enumerate(COLUMNS):
        ws.column_dimensions[get_column_letter(TABLE_START_COL + i)].width = width

    # --- テーブル（ListObject）＋オートフィルタ ---
    start_letter = get_column_letter(TABLE_START_COL)
    ref = f"{start_letter}{HEADER_ROW}:{last_col_letter}{table_last_row}"
    table = Table(displayName="TWFArrangeLedger", ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showRowStripes=True,
        showColumnStripes=False,
        showFirstColumn=False,
        showLastColumn=False,
    )
    ws.add_table(table)

    # --- ステータス列のドロップダウン（入力規則） ---
    status_col = TABLE_START_COL + [a for a, _, _ in COLUMNS].index("status")
    status_letter = get_column_letter(status_col)
    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(STATUS_CHOICES) + '"',
        allow_blank=True,
    )
    dv.error = "リストから選択してください"
    dv.errorTitle = "入力エラー"
    dv.prompt = "手配の進捗を選択"
    dv.promptTitle = "ステータス"
    if last_data_row >= DATA_START_ROW:
        dv.add(f"{status_letter}{DATA_START_ROW}:{status_letter}{last_data_row}")
        ws.add_data_validation(dv)

    # --- ヘッダー行を固定（スクロールしても見出しが残る） ---
    ws.freeze_panes = f"A{DATA_START_ROW}"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


# ============================================
# 受注一覧の読込
# ============================================
def load_orders_with_meta(
    source_path: str | Path,
) -> tuple[list[OrderRow], list[list[str]], dict[str, int], int]:
    """受注一覧を読み、(orders, source_data, cols, header_row_idx) を返す。

    回答納期の計算（build_all_caches / load_branch_settings）には素データと
    列位置・ヘッダー行インデックスが要るため、これらも一緒に返す。
    """
    data = load_source_file(source_path)
    result = get_column_positions(data)
    if result is None:
        raise ValueError("ヘッダー行を検出できませんでした。")
    cols, header_row_idx = result

    orders: list[OrderRow] = []
    for i in get_data_rows_range(data, cols, header_row_idx):
        if is_data_row(data, i, cols):
            orders.append(parse_order_row(data, i, cols))
    return orders, data, cols, header_row_idx


def load_orders(source_path: str | Path) -> list[OrderRow]:
    """SAP受注一覧ファイルを読み込んで OrderRow のリストにする。"""
    return load_orders_with_meta(source_path)[0]


# ============================================
# メーカー一覧マスターの読込・探索（在庫販売のメーカー名補完用）
# ============================================
MANUFACTURER_MASTER_NAME = "メーカー一覧.xlsx"


def load_manufacturer_map(path: str | Path) -> dict[str, str]:
    """メーカー一覧.xlsx から 品目GroupCode → メーカー名 の辞書を読み込む。

    納期回答書ツールと同じ build_manufacturer_cache を流用する。
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    mfg_name, _ = build_manufacturer_cache(wb)
    wb.close()
    return mfg_name


def find_manufacturer_master(
    source_path: str | Path, explicit: str | Path | None = None
) -> Path | None:
    """メーカー一覧.xlsx の場所を解決する。

    優先順位: 明示指定 → 受注ファイルと同じフォルダ → その親フォルダ
    （納期回答書ツールの配置 受注一覧\\ の親にメーカー一覧.xlsx がある形に対応）。
    見つからなければ None。
    """
    if explicit:
        p = Path(explicit)
        return p if p.exists() else None
    base = Path(source_path).resolve().parent
    for cand in (base / MANUFACTURER_MASTER_NAME,
                 base.parent / MANUFACTURER_MASTER_NAME):
        if cand.exists():
            return cand
    return None


# ============================================
# 回答納期コンテキスト（納期回答書と同じマスター・キャッシュを組む）
# ============================================
# 回答納期マスターの既定の置き場所。twf_ledger.py 自身が nouki_kaitou 直下に
# あるため、その隣＝納期回答書ツールが読むのと同じ3マスターを指す。
DELIVERY_MASTER_DIR_DEFAULT = Path(__file__).resolve().parent
CUSTOMER_MASTER_NAME = "顧客マスター_v2.xlsm"
HISTORY_NAME = "送付履歴.xlsx"


def _safe_load_workbook(
    path: str | Path, *, read_only: bool = False, data_only: bool = True
) -> object | None:
    """マスターを「あれば読む」。無い/使用中/壊れている → None。"""
    path = Path(path)
    if not path.exists():
        return None
    if is_file_open(path):
        print(f"  警告: {path.name} が使用中のため読み飛ばします")
        return None
    try:
        return openpyxl.load_workbook(
            str(path), data_only=data_only, read_only=read_only
        )
    except Exception as e:  # 壊れたxlsx・想定外フォーマット等
        print(f"  警告: {path.name} を読めませんでした（{e}）")
        return None


def _close_workbook(wb: object | None) -> None:
    """workbook を閉じる（None・close不可でも例外を出さない）。"""
    if wb is None:
        return
    try:
        wb.close()  # type: ignore[attr-defined]
    except Exception:
        pass


def build_delivery_context(
    source_data: list[list[str]],
    cols: dict[str, int],
    header_row_idx: int = 4,
    masters_dir: str | Path | None = None,
) -> tuple | None:
    """回答納期計算用の (CacheStore, holidays, branch) を組む。出せないとき None。

    マスターは masters_dir（既定: twf_ledger.py と同じ nouki_kaitou フォルダ）から
    読む。劣化方針は2段階:

    - メーカー一覧.xlsx と 顧客マスター_v2.xlsm は **回答納期に必須**。どちらかが
      無い/使用中/壊れていると、配送加算日数・配送曜日・路線便・祝日が効かず
      「正しく見える間違った日付」が出てしまう。そのため None を返して
      **全明細の回答納期を空欄にする**（嘘の日付を出さない＝P1-3）。取込自体は
      止めない（status/note・他列は通常どおり）。
    - 送付履歴.xlsx だけは optional。無い/古いと **受注納期=12/31 の明細のみ**
      「確認中」（在庫販売は「日程調整中」）になり、それ以外の納期は正しく出る。

    Returns:
        (CacheStore, holidays, branch)、または必須マスター欠落時 None。
    """
    masters_dir = Path(masters_dir) if masters_dir else DELIVERY_MASTER_DIR_DEFAULT

    mfg_wb = _safe_load_workbook(masters_dir / MANUFACTURER_MASTER_NAME)
    cust_wb = _safe_load_workbook(masters_dir / CUSTOMER_MASTER_NAME)

    # P1-3: メーカー一覧・顧客マスターは納期必須。欠けたら回答納期は出さない。
    if mfg_wb is None or cust_wb is None:
        missing = []
        if mfg_wb is None:
            missing.append(MANUFACTURER_MASTER_NAME)
        if cust_wb is None:
            missing.append(CUSTOMER_MASTER_NAME)
        _close_workbook(mfg_wb)  # 片方だけ開けていたら閉じる
        _close_workbook(cust_wb)
        print(
            "\n" + "=" * 64 + "\n"
            f"⚠⚠ 回答納期を計算できません: {' / '.join(missing)} が読めません\n"
            "   （無い/使用中/壊れている）。全明細の回答納期を空欄にします。\n"
            "   メーカー一覧・顧客マスターは納期計算に必須です（近似値の間違った\n"
            "   日付を出さないため、あえて空欄にしています）。\n"
            f"   置き場所: {masters_dir}\n"
            + "=" * 64 + "\n"
        )
        return None

    # 送付履歴は optional（確認中一覧シートだけ使う）。
    history_wb = _safe_load_workbook(
        masters_dir / HISTORY_NAME, read_only=True, data_only=False
    )
    try:
        confirming_ws = None
        if history_wb is not None:
            try:
                confirming_ws = history_wb[CONFIRMING_SHEET_NAME]
            except KeyError:
                print(f"  警告: {HISTORY_NAME} に「{CONFIRMING_SHEET_NAME}」シートが無く、"
                      f"12/31案件は「確認中」になります")

        print(
            "回答納期マスター: メーカー一覧=OK / 顧客マスター=OK / "
            f"送付履歴(確認中)={'OK' if confirming_ws is not None else '無（12/31案件は確認中）'}"
            f"  （元: {masters_dir}）"
        )

        cache = build_all_caches(mfg_wb, cust_wb, confirming_ws, source_data, cols)
        holidays = load_holidays(mfg_wb)
        branch = load_branch_settings(mfg_wb, source_data, cols, header_row_idx)
    finally:
        # P2-2: キャッシュ構築後は値を取り込み済み。ファイルハンドルを保持しない
        # （Windowsで顧客マスター等を開きっぱなしにすると他者のロック要因になる）。
        _close_workbook(mfg_wb)
        _close_workbook(cust_wb)
        _close_workbook(history_wb)

    # P1-3(追加): ファイルは開けても必須シート欠落/実質空だと build_xxx_cache は
    # 例外を出さず空辞書を返す。その場合 get_delivery_days_to_add は既定2日・
    # 路線便なし等の「正しく見える間違った日付」を出してしまう。嘘の日付を出さない
    # ため、必須キャッシュ（メーカー名/メーカー日数・顧客曜日等）が空なら全件空欄に倒す。
    # ※ cust_pattern は旧フォーマットの正常マスターでも空のため判定に使わない。
    mfg_ok = bool(cache.mfg_name)
    cust_ok = bool(cache.cust_days or cache.cust_route or cache.cust_retention)
    if not (mfg_ok and cust_ok):
        broken = []
        if not mfg_ok:
            broken.append(MANUFACTURER_MASTER_NAME)
        if not cust_ok:
            broken.append(CUSTOMER_MASTER_NAME)
        print(
            "\n" + "=" * 64 + "\n"
            f"⚠⚠ 回答納期を計算できません: {' / '.join(broken)} は開けましたが\n"
            "   必須シート/データが空です（シート名違い・空ファイル・破損等）。\n"
            "   近似値の間違った日付を出さないため、全明細の回答納期を空欄にします。\n"
            f"   置き場所: {masters_dir}\n"
            + "=" * 64 + "\n"
        )
        return None

    return cache, holidays, branch


# ============================================
# CLI
# ============================================
def generate(
    source_path: str | Path,
    carryover_path: str | Path | None = None,
    out_dir: str | Path | None = None,
    makers_path: str | Path | None = None,
) -> Path:
    """受注一覧から台帳を生成し、保存先パスを返す。"""
    orders = load_orders(source_path)

    # メーカー一覧マスター（在庫販売のメーカー名補完用）
    master_path = find_manufacturer_master(source_path, makers_path)
    mfg_name: dict[str, str] | None = None
    if master_path is not None:
        mfg_name = load_manufacturer_map(master_path)
        print(f"メーカー補完マスター: {master_path.name}（{len(mfg_name)}件）")
    else:
        print("メーカー補完マスター: 見つからず（在庫販売のメーカー欄は空のまま）")

    rows = build_ledger_rows(orders, mfg_name)

    # 補完結果のサマリ
    if mfg_name is not None:
        backfilled = sum(
            1 for r in rows
            if r.tehai == "在庫販売" and r.manufacturer not in ("", MANUFACTURER_UNKNOWN)
        )
        unknown = sum(1 for r in rows if r.manufacturer == MANUFACTURER_UNKNOWN)
        print(f"在庫販売のメーカー補完: {backfilled}件 / "
              f"{MANUFACTURER_UNKNOWN} {unknown}件")

    if carryover_path:
        existing = read_existing_status(carryover_path)
        applied = apply_carryover(rows, existing)
        print(f"引き継ぎ: 既存台帳 {Path(carryover_path).name} から "
              f"{applied}/{len(rows)} 明細のステータス・備考を反映")

    out_dir = Path(out_dir) if out_dir else Path(source_path).resolve().parent
    stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M")
    out_path = out_dir / f"TWF2026社内手配管理表_{stamp}.xlsx"
    return write_ledger(rows, out_path)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="TWF2026 社内手配管理表ジェネレーター"
    )
    parser.add_argument("source", help="SAP受注一覧ファイル（.txt/.XLS）")
    parser.add_argument(
        "--carryover", help="既存台帳xlsx（ステータス・備考を引き継ぐ）"
    )
    parser.add_argument("--outdir", help="出力先フォルダ（既定: ソースと同じ場所）")
    parser.add_argument(
        "--makers",
        help=f"メーカー一覧.xlsx のパス（既定: ソースと同じ/親フォルダの "
             f"{MANUFACTURER_MASTER_NAME} を自動探索）",
    )
    args = parser.parse_args()

    orders = load_orders(args.source)
    twf_orders, _ = collect_twf_orders(orders)
    print(f"受注 {len(orders)} 明細中、TWF注番 {len(twf_orders)} 件を検出")

    out_path = generate(args.source, args.carryover, args.outdir, args.makers)
    print(f"生成しました: {out_path}")


if __name__ == "__main__":
    main()
