"""2件の再確認: VBAセル直読み + SAPデータ直読み"""

import sys
import io
import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from openpyxl import load_workbook

from nouki_kaitou.data_loader import (
    get_column_positions,
    get_data_rows_range,
    is_data_row,
    load_source_file,
    parse_order_row,
)


def main():
    tool_folder = Path(r"\\flsv04\316京葉\納期回答書ツールフォルダ")
    source_path = tool_folder / "受注一覧" / "17PM.xls"
    vba_folder = tool_folder / "納期回答書" / "2月17日(火)_②回目"

    # =============================================
    # 1. GL2Z446911|20: VBA回答書の実際のセル値を直接読む
    # =============================================
    print("=" * 80)
    print("【1】GL2Z446911|20: VBA回答書のセル値を直接読み取り")
    print("=" * 80)

    # まずSAPデータから顧客名を取得
    print("\nSAPデータから顧客名を取得中...")
    source_data_raw = load_source_file(str(source_path))
    cols = get_column_positions(source_data_raw)

    target1_customer = None
    target1_row = None
    target2_row = None

    for i in get_data_rows_range(source_data_raw, cols):
        if is_data_row(source_data_raw, i, cols):
            row = parse_order_row(source_data_raw, i, cols)
            if row.order_number == "GL2Z446911" and row.detail_number == "20":
                target1_customer = row.customer_name
                target1_row = row
            if row.order_number == "GL2S447221" and row.detail_number == "10":
                target2_row = row

    if target1_row:
        print(f"  GL2Z446911|20 の顧客名: [{target1_row.customer_name}]")
        print(f"  GL2Z446911|20 の出荷先: [{target1_row.ship_to_name}]")
        print(f"  品名: [{target1_row.product_name}]")
        print(f"  伝票タイプ: [{target1_row.document_type}]")
        print(f"  保管場所: [{target1_row.storage_place}]")
        print(f"  出荷ステータス: [{target1_row.ship_status}]")
        print(f"  指定納期: [{target1_row.specified_delivery_date}]")
        print(f"  受注納期: [{target1_row.order_delivery_date}]")
        print(f"  登録日時: [{target1_row.registration_date}] [{target1_row.time_value}]")
        print(f"  数量: [{target1_row.quantity}]")
        print(f"  コメント社内: [{target1_row.comment_internal}]")

    # VBA回答書フォルダ内の全ファイルを検索
    print(f"\nVBA回答書フォルダ: {vba_folder}")
    vba_files = list(vba_folder.glob("*.xlsx"))
    print(f"ファイル数: {len(vba_files)}")

    # 全ファイルを探索してGL2Z446911|20の顧客に該当するファイルを探す
    # ファイル名から顧客名を推測するのではなく、全ファイルの全行をチェック
    found_vba = False

    for vba_file in vba_files:
        if vba_file.name.startswith("~$"):
            continue
        try:
            wb = load_workbook(str(vba_file), data_only=True)
        except Exception as e:
            print(f"  スキップ: {vba_file.name} ({e})")
            continue

        ws = wb.active

        # ヘッダー行を探す
        header_row = None
        for row_idx in range(1, min(20, ws.max_row + 1)):
            cell_val = ws.cell(row=row_idx, column=1).value
            if cell_val and str(cell_val).strip() == "受注日":
                header_row = row_idx
                break

        if header_row is None:
            wb.close()
            continue

        # GL2Z446911|20は送料行なので、品名が「送料」を含む行を探す
        # まず、このファイルが対象顧客のものか確認
        # 顧客名はファイル名に含まれている
        import re
        file_customer = vba_file.name.replace("納期回答書_", "").replace(".xlsx", "")
        file_customer = re.sub(r"_\d{8}$", "", file_customer)

        # 品名列(5)で「送料」を含む行をチェック + 全行の情報を出力
        for row_idx in range(header_row + 1, ws.max_row + 1):
            order_date_val = ws.cell(row=row_idx, column=1).value
            if not order_date_val:
                continue

            # フッター判定
            unit_price_val = ws.cell(row=row_idx, column=7).value
            if unit_price_val and "※" in str(unit_price_val):
                break

            product = str(ws.cell(row=row_idx, column=5).value or "").strip()
            # GL2Z446911|20は送料行
            if target1_row and target1_row.product_name.strip()[:15] in product[:15]:
                delivery_cell = ws.cell(row=row_idx, column=9)
                delivery_val = delivery_cell.value

                print(f"\n  ★ VBA回答書で発見!")
                print(f"  ファイル: {vba_file.name}")
                print(f"  行番号: {row_idx}")
                print(f"  品名(セル値): [{product}]")
                print(f"  納期回答(column=9):")
                print(f"    value = [{delivery_val}]")
                print(f"    type  = {type(delivery_val)}")
                print(f"    repr  = {repr(delivery_val)}")

                # 周辺のセルも全部読む
                print(f"\n  この行の全セル値:")
                for col_idx in range(1, 12):
                    cv = ws.cell(row=row_idx, column=col_idx).value
                    print(f"    列{col_idx}: [{cv}] (type={type(cv).__name__})")

                found_vba = True

        wb.close()

    if not found_vba:
        # 送料行はVBA回答書に含まれない可能性もある
        # 顧客名で該当ファイルを探して、全行を出力する
        print("\n  送料行がVBA回答書に見つかりません。")
        print("  顧客名でファイルを検索中...")

        if target1_customer:
            # 顧客名の一部でファイルを探す
            customer_short = target1_customer.strip().replace("\u3000", " ")[:10]
            for vba_file in vba_files:
                if vba_file.name.startswith("~$"):
                    continue
                if customer_short[:6] in vba_file.name:
                    print(f"\n  候補ファイル: {vba_file.name}")
                    try:
                        wb = load_workbook(str(vba_file), data_only=True)
                        ws = wb.active
                        header_row = None
                        for row_idx in range(1, min(20, ws.max_row + 1)):
                            cell_val = ws.cell(row=row_idx, column=1).value
                            if cell_val and str(cell_val).strip() == "受注日":
                                header_row = row_idx
                                break

                        if header_row:
                            print(f"  ヘッダー行: {header_row}")
                            print(f"  全データ行:")
                            for row_idx in range(header_row + 1, ws.max_row + 1):
                                order_date_val = ws.cell(row=row_idx, column=1).value
                                if not order_date_val:
                                    continue
                                unit_price_val = ws.cell(row=row_idx, column=7).value
                                if unit_price_val and "※" in str(unit_price_val):
                                    break
                                product = str(ws.cell(row=row_idx, column=5).value or "")
                                delivery = str(ws.cell(row=row_idx, column=9).value or "")
                                print(f"    行{row_idx}: 品名=[{product}] 納期回答=[{delivery}]")
                        wb.close()
                    except Exception as e:
                        print(f"    エラー: {e}")

    # =============================================
    # 2. GL2S447221|10: 受注先と出荷先の実際の値
    # =============================================
    print("\n" + "=" * 80)
    print("【2】GL2S447221|10: 受注先と出荷先の実際の値")
    print("=" * 80)

    if target2_row:
        print(f"\n  customer_name (受注先):")
        print(f"    値:   [{target2_row.customer_name}]")
        print(f"    repr: {repr(target2_row.customer_name)}")
        print(f"    len:  {len(target2_row.customer_name)}")
        print(f"    各文字: ", end="")
        for i, c in enumerate(target2_row.customer_name):
            print(f"[{i}:{c}(U+{ord(c):04X})]", end=" ")
        print()

        print(f"\n  ship_to_name (出荷先):")
        print(f"    値:   [{target2_row.ship_to_name}]")
        print(f"    repr: {repr(target2_row.ship_to_name)}")
        print(f"    len:  {len(target2_row.ship_to_name)}")
        print(f"    各文字: ", end="")
        for i, c in enumerate(target2_row.ship_to_name):
            print(f"[{i}:{c}(U+{ord(c):04X})]", end=" ")
        print()

        # strip後の比較
        cust = target2_row.customer_name.strip()
        ship = target2_row.ship_to_name.strip()
        print(f"\n  strip後の比較:")
        print(f"    customer_name.strip(): [{cust}] (len={len(cust)})")
        print(f"    ship_to_name.strip():  [{ship}] (len={len(ship)})")
        print(f"    一致するか: {cust == ship}")

        # 文字ごとの比較
        if cust != ship:
            print(f"\n  文字ごとの差異:")
            max_len = max(len(cust), len(ship))
            for i in range(max_len):
                c1 = cust[i] if i < len(cust) else "(なし)"
                c2 = ship[i] if i < len(ship) else "(なし)"
                marker = " ★" if c1 != c2 else ""
                if i < len(cust) and i < len(ship):
                    print(f"    位置{i}: customer=[{c1}(U+{ord(c1):04X})] ship_to=[{c2}(U+{ord(c2):04X})]{marker}")
                else:
                    print(f"    位置{i}: customer=[{c1}] ship_to=[{c2}]{marker}")

        # その他の情報も出力
        print(f"\n  追加情報:")
        print(f"    品名: [{target2_row.product_name}]")
        print(f"    伝票タイプ: [{target2_row.document_type}]")
        print(f"    保管場所: [{target2_row.storage_place}]")
        print(f"    出荷ステータス: [{target2_row.ship_status}]")
        print(f"    指定納期: [{target2_row.specified_delivery_date}]")
        print(f"    受注納期: [{target2_row.order_delivery_date}]")
        print(f"    登録日時: [{target2_row.registration_date}] [{target2_row.time_value}]")
        print(f"    コメント社内: [{target2_row.comment_internal}]")
    else:
        print("  GL2S447221|10 がSAPデータ内に見つかりません")


if __name__ == "__main__":
    main()
